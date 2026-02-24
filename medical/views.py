"""
Medical Views
Views für das Medical-Modul (Rettungsdienst & BTM)
"""

from django.views.generic import ListView, DetailView, CreateView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required
from django.db.models import Q, F, Sum
from django.db import models
from django.utils import timezone
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.urls import reverse_lazy
from django.http import HttpResponse, JsonResponse
from datetime import timedelta
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill
import logging

logger = logging.getLogger(__name__)

from .models import (
    MedicalItem,
    MedicalItemMaster,
    MedicalDeviceInstance,
    MedicalStockMovement,
    MedicalBatch,
    TemperatureLog,
    BTMApprovalStatus,
)
from documents.models import Document, DocumentCategory, DocumentType, DocumentStatus, AccessLevel
from django.contrib.contenttypes.models import ContentType
from inventory_base.models import StockMovementType
from .forms import (
    MedicalItemMasterForm,
    MedicalDeviceInstanceForm,
    MedicalStockMovementForm,
    MedicalBatchForm,
)
from django.views.generic import TemplateView, DeleteView, View
from django.urls import reverse
from inventory_base.models import Category
from permissions.constants import Roles


# ============================================================================
# DASHBOARD
# ============================================================================

class MedicalDashboardView(LoginRequiredMixin, TemplateView):
    """
    Rettungsdienst-Dashboard mit Übersicht und KPIs
    """
    template_name = 'medical/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()

        # KPIs - Optimized with single aggregation query where possible
        from django.db.models import Count, Q

        # Stammdaten (Masters) Statistiken
        master_stats = MedicalItemMaster.objects.filter(is_active=True).aggregate(
            total=Count('id'),
            btm_count=Count('id', filter=Q(is_btm=True)),
        )

        context['total_items'] = master_stats['total']
        context['total_masters'] = master_stats['total']
        context['btm_items'] = master_stats['btm_count']

        # Medizintechnik-Instanzen
        device_stats = MedicalDeviceInstance.objects.filter(is_active=True).aggregate(
            total=Count('id'),
            operational=Count('id', filter=Q(is_operational=True)),
        )
        context['total_devices'] = device_stats['total']
        context['operational_devices'] = device_stats['operational']

        # Wartungsbedürftige Geräte über DeviceInstance
        context['maintenance_due_count'] = MedicalDeviceInstance.objects.filter(
            master__requires_maintenance=True,
            next_maintenance_date__isnull=False,
            next_maintenance_date__lte=today,
            is_operational=True
        ).count()

        # Low-Stock wird über Chargen berechnet (Artikel mit wenig Restbestand)
        # TODO: Dies könnte über aggregierte Batch-Mengen optimiert werden
        context['low_stock_count'] = 0  # Temporär deaktiviert, benötigt komplexere Batch-Aggregation

        # Ablaufende Chargen (nächste 30 Tage)
        threshold_30 = today + timedelta(days=30)
        context['expiring_batches_30'] = MedicalBatch.objects.filter(
            expiry_date__lte=threshold_30,
            expiry_date__gte=today,
            quantity_remaining__gt=0,
            is_recalled=False
        ).count()

        # Ausstehende BTM-Freigaben
        context['pending_btm_approvals'] = MedicalStockMovement.objects.filter(
            requires_approval=True,
            approval_status=BTMApprovalStatus.PENDING
        ).count()

        # Kühlketten-Unterbrechungen
        context['cold_chain_breaks'] = MedicalBatch.objects.filter(
            cold_chain_break=True,
            quantity_remaining__gt=0
        ).count()

        # Kritische Alerts
        critical_alerts = []

        # BTM-Freigaben
        if context['pending_btm_approvals'] > 0:
            critical_alerts.append({
                'type': 'btm',
                'severity': 'high',
                'icon': '☢️',
                'title': f'{context["pending_btm_approvals"]} BTM-Freigabe(n) ausstehend',
                'url': '/medical/btm/approvals/',
                'color': 'red'
            })

        # Niedrige Bestände
        if context['low_stock_count'] > 0:
            critical_alerts.append({
                'type': 'stock',
                'severity': 'medium',
                'icon': '📦',
                'title': f'{context["low_stock_count"]} Artikel mit niedrigem Bestand',
                'url': '/medical/low-stock/',
                'color': 'amber'
            })

        # Ablaufende Chargen
        if context['expiring_batches_30'] > 0:
            critical_alerts.append({
                'type': 'expiry',
                'severity': 'medium',
                'icon': '⏰',
                'title': f'{context["expiring_batches_30"]} Charge(n) laufen in 30 Tagen ab',
                'url': '/medical/expiring-batches/',
                'color': 'orange'
            })

        # Kühlketten-Unterbrechungen
        if context['cold_chain_breaks'] > 0:
            critical_alerts.append({
                'type': 'cold_chain',
                'severity': 'high',
                'icon': '❄️',
                'title': f'{context["cold_chain_breaks"]} Kühlketten-Unterbrechung(en)',
                'url': '/medical/cold-chain/',
                'color': 'blue'
            })

        # Wartungen fällig
        if context['maintenance_due_count'] > 0:
            critical_alerts.append({
                'type': 'maintenance',
                'severity': 'medium',
                'icon': '🔧',
                'title': f'{context["maintenance_due_count"]} Wartung(en) fällig',
                'url': '/medical/maintenance-due/',
                'color': 'purple'
            })

        context['critical_alerts'] = critical_alerts

        # Letzte Lagerbewegungen (Top 5)
        context['recent_movements'] = MedicalStockMovement.objects.select_related(
            'item', 'from_location', 'to_location', 'created_by'
        ).order_by('-movement_date')[:5]

        # Ablaufende Chargen (Top 5)
        context['expiring_batches'] = MedicalBatch.objects.filter(
            expiry_date__isnull=False,
            expiry_date__gte=today,
            quantity_remaining__gt=0,
            is_recalled=False
        ).select_related('master', 'location').order_by('expiry_date')[:5]

        # BTM-Artikel mit niedrigem Bestand
        # TODO: Über Batch-Aggregation implementieren
        context['btm_low_stock'] = []  # Temporär deaktiviert

        # Temperatur-Anomalien (letzte 24h)
        last_24h = timezone.now() - timedelta(hours=24)
        context['temp_anomalies'] = TemperatureLog.objects.filter(
            is_within_range=False,
            measured_at__gte=last_24h
        ).select_related('batch__item', 'location').order_by('-measured_at')[:10]

        # Modul-Info
        context['current_module'] = 'medical'

        return context


# ============================================================================
# MEDICAL ITEM VIEWS
# ============================================================================

# ============================================================================
# LEGACY ITEM VIEWS ENTFERNT
# Das alte MedicalItem-System wurde durch MedicalItemMaster ersetzt.
# Siehe MedicalItemMaster* Views weiter unten.
# ============================================================================




# ============================================================================
# BTM-BEREICH
# ============================================================================

class BTMItemListView(LoginRequiredMixin, ListView):
    """
    Liste aller BTM-Artikel (Betäubungsmittel)
    Besondere Sicherheitsansicht
    """
    model = MedicalItemMaster
    template_name = 'medical/btm_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        return MedicalItemMaster.objects.filter(
            is_active=True,
            is_btm=True
        ).select_related('category', 'supplier').order_by('name')


class BTMApprovalListView(LoginRequiredMixin, ListView):
    """
    Liste aller ausstehenden BTM-Freigaben
    """
    model = MedicalStockMovement
    template_name = 'medical/btm_approvals.html'
    context_object_name = 'movements'
    paginate_by = 50

    def get_queryset(self):
        return MedicalStockMovement.objects.filter(
            requires_approval=True,
            approval_status=BTMApprovalStatus.PENDING
        ).select_related(
            'item',
            'from_location',
            'to_location',
            'created_by',
        ).order_by('-created_at')


@login_required
def approve_btm_movement(request, pk):
    """
    BTM-Bewegung freigeben (Vier-Augen-Prinzip)
    """
    movement = get_object_or_404(MedicalStockMovement, pk=pk)

    if not movement.requires_approval:
        messages.error(request, _('Diese Bewegung erfordert keine Freigabe.'))
        return redirect('medical:btm_approvals')

    if movement.approval_status != BTMApprovalStatus.PENDING:
        messages.error(request, _('Diese Bewegung wurde bereits bearbeitet.'))
        return redirect('medical:btm_approvals')

    try:
        movement.approve(request.user)
        messages.success(
            request,
            _('BTM-Bewegung wurde freigegeben. Bestand wurde aktualisiert.')
        )
    except ValueError as e:
        messages.error(request, str(e))

    return redirect('medical:btm_approvals')


@login_required
def reject_btm_movement(request, pk):
    """
    BTM-Bewegung ablehnen
    """
    movement = get_object_or_404(MedicalStockMovement, pk=pk)

    if not movement.requires_approval:
        messages.error(request, _('Diese Bewegung erfordert keine Freigabe.'))
        return redirect('medical:btm_approvals')

    if movement.approval_status != BTMApprovalStatus.PENDING:
        messages.error(request, _('Diese Bewegung wurde bereits bearbeitet.'))
        return redirect('medical:btm_approvals')

    # Ablehnungsgrund aus POST
    reason = request.POST.get('reason', _('Keine Begründung angegeben'))

    movement.reject(request.user, reason)
    messages.warning(
        request,
        _('BTM-Bewegung wurde abgelehnt.')
    )

    return redirect('medical:btm_approvals')


# ============================================================================
# LOW STOCK & EXPIRING
# ============================================================================

class LowStockListView(LoginRequiredMixin, ListView):
    """
    Liste der Artikel mit niedrigem Bestand
    HINWEIS: Temporär deaktiviert - benötigt Umstellung auf Batch-basierte Bestandsberechnung
    """
    model = MedicalItemMaster
    template_name = 'medical/low_stock_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        # TODO: Bestandsberechnung über Batches implementieren
        # Temporär: Zeige alle Masters mit min_quantity Einstellung
        return MedicalItemMaster.objects.filter(
            is_active=True,
            min_quantity__isnull=False
        ).select_related('category', 'supplier')


class ExpiringBatchesListView(LoginRequiredMixin, ListView):
    """
    Liste der ablaufenden Chargen (nächste 90 Tage)
    """
    model = MedicalBatch
    template_name = 'medical/expiring_batches_list.html'
    context_object_name = 'batches'
    paginate_by = 50

    def get_queryset(self):
        today = timezone.now().date()
        threshold = today + timedelta(days=90)

        return MedicalBatch.objects.filter(
            expiry_date__isnull=False,
            expiry_date__lte=threshold,
            expiry_date__gte=today,
            quantity_remaining__gt=0,
            is_recalled=False
        ).select_related('item', 'location').order_by('expiry_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()

        # Statistiken für die Kacheln
        all_expiring = self.get_queryset()

        # Nächste 7 Tage
        context['expiring_7days'] = all_expiring.filter(
            expiry_date__lte=today + timedelta(days=7)
        ).count()

        # Nächste 14 Tage
        context['expiring_14days'] = all_expiring.filter(
            expiry_date__lte=today + timedelta(days=14)
        ).count()

        # Nächste 30 Tage
        context['expiring_30days'] = all_expiring.filter(
            expiry_date__lte=today + timedelta(days=30)
        ).count()

        # Nächste 90 Tage (alle)
        context['expiring_90days'] = all_expiring.count()

        # Bereits abgelaufen (sollte eigentlich nicht in dieser Liste sein, aber zur Sicherheit)
        context['expired_count'] = MedicalBatch.objects.filter(
            expiry_date__lt=today,
            quantity_remaining__gt=0,
            is_recalled=False
        ).count()

        return context


# ============================================================================
# KÜHLKETTE
# ============================================================================

class ColdChainItemsListView(LoginRequiredMixin, ListView):
    """
    Liste aller Artikel mit Kühlketten-Anforderung
    """
    model = MedicalItemMaster
    template_name = 'medical/cold_chain_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        return MedicalItemMaster.objects.filter(
            is_active=True,
            requires_cold_chain=True
        ).select_related('category', 'supplier').order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Chargen mit Kühlketten-Unterbrechung
        context['broken_cold_chain_batches'] = MedicalBatch.objects.filter(
            cold_chain_break=True,
            quantity_remaining__gt=0
        ).select_related('master', 'location')

        # Temperatur-Logs (letzte 24h)
        from datetime import timedelta
        since = timezone.now() - timedelta(hours=24)

        # QuerySet für Statistiken (ohne Slice)
        temp_logs_qs = TemperatureLog.objects.filter(measured_at__gte=since)

        # Statistiken
        context['monitored_items'] = self.get_queryset().count()
        context['ok_count'] = temp_logs_qs.filter(is_within_range=True).count()
        context['warnings_24h'] = temp_logs_qs.filter(is_within_range=False).count()
        context['critical_alerts'] = context['broken_cold_chain_batches'].count()
        context['broken_chains'] = context['critical_alerts']
        context['last_update'] = timezone.now()

        # Logs für Anzeige
        context['temperature_logs'] = temp_logs_qs.select_related(
            'batch__master', 'location', 'measured_by'
        ).order_by('-measured_at')[:50]

        return context


class ColdChainAPIDocsView(LoginRequiredMixin, TemplateView):
    """
    API Dokumentation für IoT-Temperatursensoren
    """
    template_name = 'medical/cold_chain_api_docs.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Kühlpflichtige Artikel
        context['items'] = MedicalItemMaster.objects.filter(
            is_active=True,
            requires_cold_chain=True
        ).select_related('category', 'supplier').order_by('name')

        # Chargen mit Kühlketten-Unterbrechung
        context['broken_cold_chain_batches'] = MedicalBatch.objects.filter(
            cold_chain_break=True,
            quantity_remaining__gt=0
        ).select_related('master', 'location')

        # Temperatur-Logs (letzte 24h)
        from datetime import timedelta
        since = timezone.now() - timedelta(hours=24)

        # QuerySet für Statistiken (ohne Slice)
        temp_logs_qs = TemperatureLog.objects.filter(measured_at__gte=since)

        # Statistiken
        context['monitored_items'] = context['items'].count()
        context['ok_count'] = temp_logs_qs.filter(is_within_range=True).count()
        context['warnings_24h'] = temp_logs_qs.filter(is_within_range=False).count()
        context['critical_alerts'] = context['broken_cold_chain_batches'].count()
        context['broken_chains'] = context['critical_alerts']
        context['last_update'] = timezone.now()

        # Logs für Anzeige (mit Slice)
        context['temperature_logs'] = temp_logs_qs.select_related(
            'batch__item', 'location', 'measured_by'
        ).order_by('-measured_at')[:50]

        # API Dokumentation
        from rest_framework.authtoken.models import Token

        # Beispiel API-Tokens (für Dokumentation)
        context['api_base_url'] = self.request.build_absolute_uri('/api/medical/')

        # Prüfe ob User API-Token hat
        try:
            token, created = Token.objects.get_or_create(user=self.request.user)
            context['user_api_token'] = token.key
        except:
            context['user_api_token'] = None

        return context


# ============================================================================
# WARTUNG
# ============================================================================

class MaintenanceDueListView(LoginRequiredMixin, ListView):
    """
    Liste der Medizingeräte mit fälliger Wartung
    Zeigt konkrete Geräte-Instanzen, die Wartung benötigen
    """
    model = MedicalDeviceInstance
    template_name = 'medical/maintenance_due_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        today = timezone.now().date()

        return MedicalDeviceInstance.objects.filter(
            master__requires_maintenance=True,
            next_maintenance_date__isnull=False,
            next_maintenance_date__lte=today,
            is_operational=True
        ).select_related('master', 'location').order_by('next_maintenance_date')


# ============================================================================
# STOCK MOVEMENT VIEWS
# ============================================================================

class StockMovementListView(LoginRequiredMixin, ListView):
    """
    Liste aller Lagerbewegungen
    """
    model = MedicalStockMovement
    template_name = 'medical/movement_list.html'
    context_object_name = 'movements'
    paginate_by = 50

    def get_queryset(self):
        queryset = MedicalStockMovement.objects.select_related(
            'item',
            'from_location',
            'to_location',
            'approved_by',
        )

        # Bewegungstyp-Filter
        movement_type = self.request.GET.get('movement_type')
        if movement_type:
            queryset = queryset.filter(movement_type=movement_type)

        # Artikel-Filter
        item = self.request.GET.get('item')
        if item:
            queryset = queryset.filter(item_id=item)

        # BTM-Filter
        show_btm_only = self.request.GET.get('btm_only')
        if show_btm_only:
            queryset = queryset.filter(requires_approval=True)

        # Freigabe-Status-Filter
        approval_status = self.request.GET.get('approval_status')
        if approval_status:
            queryset = queryset.filter(approval_status=approval_status)

        return queryset.order_by('-movement_date')


class StockMovementDetailView(LoginRequiredMixin, DetailView):
    """
    Detail-Ansicht für Lagerbewegung
    """
    model = MedicalStockMovement
    template_name = 'medical/movement_detail.html'
    context_object_name = 'movement'

    def get_queryset(self):
        return MedicalStockMovement.objects.select_related(
            'item',
            'from_location',
            'to_location',
            'created_by',
            'updated_by',
            'administered_by',
            'approved_by',
        )


# ============================================================================
# BATCH VIEWS
# ============================================================================

class BatchListView(LoginRequiredMixin, ListView):
    """
    Liste aller Chargen
    """
    model = MedicalBatch
    template_name = 'medical/batch_list.html'
    context_object_name = 'batches'
    paginate_by = 50

    def get_queryset(self):
        from datetime import timedelta
        from django.utils import timezone

        queryset = MedicalBatch.objects.select_related(
            'master',
            'item',
            'location',
        )

        # Nur aktive Chargen (nicht aufgebraucht)
        show_depleted = self.request.GET.get('show_depleted')
        if not show_depleted:
            queryset = queryset.filter(quantity_remaining__gt=0)

        # Nur nicht zurückgerufene Chargen
        show_recalled = self.request.GET.get('show_recalled')
        if not show_recalled:
            queryset = queryset.filter(is_recalled=False)

        # Artikel-Filter (unterstützt beide Felder)
        item = self.request.GET.get('item')
        if item:
            queryset = queryset.filter(
                Q(master_id=item) | Q(item_id=item)
            )

        # Suche
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(batch_number__icontains=search) |
                Q(master__name__icontains=search) |
                Q(item__name__icontains=search)
            )

        # Status-Filter
        status_filters = self.request.GET.getlist('status')
        if status_filters:
            status_q = Q()
            today = timezone.now().date()

            if 'active' in status_filters:
                # Aktiv: nicht abgelaufen, Bestand > 0
                status_q |= Q(expiry_date__gte=today, quantity_remaining__gt=0)

            if 'expiring' in status_filters:
                # Bald abgelaufen: läuft in den nächsten 30 Tagen ab
                expiring_date = today + timedelta(days=30)
                status_q |= Q(expiry_date__gte=today, expiry_date__lte=expiring_date, quantity_remaining__gt=0)

            if 'expired' in status_filters:
                # Abgelaufen
                status_q |= Q(expiry_date__lt=today)

            if 'empty' in status_filters:
                # Aufgebraucht
                status_q |= Q(quantity_remaining=0)

            queryset = queryset.filter(status_q)

        # Ablaufdatum-Filter
        expiry_until = self.request.GET.get('expiry_until')
        if expiry_until:
            queryset = queryset.filter(expiry_date__lte=expiry_until)

        # BTM-Filter
        btm_only = self.request.GET.get('btm_only')
        if btm_only:
            queryset = queryset.filter(
                Q(master__is_btm=True) | Q(item__is_btm=True)
            )

        # Kühlketten-Filter
        cold_chain = self.request.GET.get('cold_chain')
        if cold_chain:
            queryset = queryset.filter(
                Q(master__requires_cold_chain=True) | Q(item__requires_cold_chain=True)
            )

        return queryset.order_by('expiry_date', 'received_date')

    def get_context_data(self, **kwargs):
        from datetime import timedelta
        from django.utils import timezone

        context = super().get_context_data(**kwargs)
        today = timezone.now().date()
        expiring_date = today + timedelta(days=30)

        # Statistiken
        all_batches = MedicalBatch.objects.filter(is_recalled=False)
        context['expiring_count'] = all_batches.filter(
            expiry_date__gte=today,
            expiry_date__lte=expiring_date,
            quantity_remaining__gt=0
        ).count()
        context['expired_count'] = all_batches.filter(expiry_date__lt=today).count()
        context['cold_chain_breaks'] = all_batches.filter(cold_chain_break=True).count()

        # Artikel für Filter-Dropdown
        context['items'] = MedicalItemMaster.objects.filter(is_active=True).order_by('name')

        return context


class RecalledBatchesListView(LoginRequiredMixin, ListView):
    """
    Liste aller zurückgerufenen Chargen
    """
    model = MedicalBatch
    template_name = 'medical/recalled_batches_list.html'
    context_object_name = 'batches'
    paginate_by = 50

    def get_queryset(self):
        return MedicalBatch.objects.filter(
            is_recalled=True
        ).select_related('item', 'location').order_by('-recall_date')


# ============================================================================
# STOCK MOVEMENTS (LAGERBEWEGUNGEN)
# ============================================================================

class StockMovementListView(LoginRequiredMixin, ListView):
    """Liste aller Lagerbewegungen"""
    model = MedicalStockMovement
    template_name = 'medical/stock_movement_list.html'
    context_object_name = 'movements'
    paginate_by = 50

    def get_queryset(self):
        queryset = MedicalStockMovement.objects.select_related(
            'item', 'to_location', 'from_location', 'created_by'
        ).order_by('-movement_date')

        # Filter nach Artikel
        item_id = self.request.GET.get('item')
        if item_id:
            queryset = queryset.filter(item_id=item_id)

        # Filter nach Bewegungstyp
        movement_type = self.request.GET.get('type')
        if movement_type:
            queryset = queryset.filter(movement_type=movement_type)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_filter'] = self.request.GET.get('item', '')
        context['type_filter'] = self.request.GET.get('type', '')
        return context


class StockMovementCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Neue Lagerbewegung erstellen (Wareneingang, Warenausgang, etc.)"""
    model = MedicalStockMovement
    form_class = MedicalStockMovementForm
    template_name = 'medical/stock_movement_form.html'
    success_url = reverse_lazy('medical:stock_movements')
    permission_required = 'medical.add_medicalstockmovement'

    def get_initial(self):
        initial = super().get_initial()
        # Wenn item_id in URL, vorausfüllen
        item_id = self.request.GET.get('item_id')
        if item_id:
            initial['item'] = item_id
        return initial

    def form_valid(self, form):
        # Audit-Felder setzen
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user

        # BTM-Prüfung: Wenn Artikel BTM ist, Freigabe erforderlich
        if form.instance.item.is_btm:
            form.instance.requires_approval = True
            form.instance.approval_status = BTMApprovalStatus.PENDING

        response = super().form_valid(form)

        messages.success(
            self.request,
            f'Lagerbewegung ({form.instance.get_movement_type_display()}) für {form.instance.item.name} wurde erfolgreich erstellt.'
        )

        return response


class StockMovementDetailView(LoginRequiredMixin, DetailView):
    """Detail-Ansicht einer Lagerbewegung"""
    model = MedicalStockMovement
    template_name = 'medical/stock_movement_detail.html'
    context_object_name = 'movement'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Verknüpfte Dokumente laden
        from documents.models import Document
        from django.contrib.contenttypes.models import ContentType

        content_type = ContentType.objects.get_for_model(MedicalStockMovement)
        documents = Document.objects.filter(
            related_content_type=content_type,
            related_object_id=self.object.pk
        ).select_related('category', 'created_by').order_by('-created_at')

        context['documents'] = documents
        return context


class QuickIncomingView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Schnell-Wareneingang (vorbefüllt)"""
    model = MedicalStockMovement
    form_class = MedicalStockMovementForm
    template_name = 'medical/quick_incoming.html'
    success_url = reverse_lazy('medical:stock_movements')
    permission_required = 'medical.add_medicalstockmovement'

    def get_initial(self):
        initial = super().get_initial()
        initial['movement_type'] = 'incoming'
        # Wenn item_id in URL, vorausfüllen
        item_id = self.request.GET.get('item_id')
        if item_id:
            initial['item'] = item_id
        return initial

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user

        if form.instance.item.is_btm:
            form.instance.requires_approval = True
            form.instance.approval_status = BTMApprovalStatus.PENDING

        response = super().form_valid(form)

        # Lieferschein-Upload verarbeiten
        if 'delivery_note' in self.request.FILES:
            delivery_file = self.request.FILES['delivery_note']

            # Dokument erstellen und mit StockMovement verknüpfen
            from documents.models import Document, DocumentCategory, DocumentType, DocumentStatus, AccessLevel
            from django.contrib.contenttypes.models import ContentType

            # Kategorie für Lieferscheine
            category, _ = DocumentCategory.objects.get_or_create(
                name='Lieferscheine',
                defaults={'description': 'Lieferscheine und Wareneingangsdokumente'}
            )

            # Content Type für Generic FK
            content_type = ContentType.objects.get_for_model(MedicalStockMovement)

            try:
                # Titel generieren
                title = f"Lieferschein - {form.instance.item.name} - {form.instance.movement_date.strftime('%d.%m.%Y')}"
                if form.instance.reference_number:
                    title = f"Lieferschein {form.instance.reference_number} - {form.instance.item.name}"

                document = Document.objects.create(
                    title=title,
                    category=category,
                    document_type=DocumentType.INVOICE,  # Lieferschein als Invoice-Typ
                    status=DocumentStatus.ACTIVE,
                    file=delivery_file,
                    description=f'Lieferschein für Wareneingang ({form.instance.quantity} {form.instance.unit})',
                    access_level=AccessLevel.INTERNAL,
                    related_content_type=content_type,
                    related_object_id=form.instance.pk,
                    created_by=self.request.user,
                    updated_by=self.request.user
                )

                # MIME-Type ermitteln
                import magic
                try:
                    mime = magic.Magic(mime=True)
                    document.mime_type = mime.from_buffer(delivery_file.read(1024))
                    delivery_file.seek(0)
                    document.save()
                except:
                    pass

                # OCR-Task triggern
                try:
                    from documents.tasks import process_document_ocr
                    process_document_ocr.delay(document.id)
                    logger.info(f'OCR-Task für Lieferschein #{document.id} gestartet')
                except Exception as e:
                    logger.warning(f'OCR-Task konnte nicht gestartet werden: {e}')

                messages.success(self.request, f'Lieferschein "{document.title}" wurde hochgeladen.')

            except Exception as e:
                logger.error(f'Fehler beim Upload des Lieferscheins: {e}')
                messages.warning(self.request, f'Lieferschein konnte nicht hochgeladen werden: {e}')

        messages.success(self.request, f'Wareneingang für {form.instance.item.name} erfolgreich erfasst.')
        return response


class QuickOutgoingView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Schnell-Warenausgang (vorbefüllt)"""
    model = MedicalStockMovement
    form_class = MedicalStockMovementForm
    template_name = 'medical/quick_outgoing.html'
    success_url = reverse_lazy('medical:stock_movements')
    permission_required = 'medical.add_medicalstockmovement'

    def get_initial(self):
        initial = super().get_initial()
        initial['movement_type'] = 'outgoing'
        # Wenn item_id in URL, vorausfüllen
        item_id = self.request.GET.get('item_id')
        if item_id:
            initial['item'] = item_id
        return initial

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user

        if form.instance.item.is_btm:
            form.instance.requires_approval = True
            form.instance.approval_status = BTMApprovalStatus.PENDING

        response = super().form_valid(form)
        messages.success(self.request, f'Warenausgang für {form.instance.item.name} erfolgreich erfasst.')
        return response


class DisposalView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Entsorgung (MHD abgelaufen, beschädigt, etc.)"""
    model = MedicalStockMovement
    form_class = MedicalStockMovementForm
    template_name = 'medical/disposal_form.html'
    success_url = reverse_lazy('medical:stock_movements')
    permission_required = 'medical.add_medicalstockmovement'

    def get_initial(self):
        initial = super().get_initial()
        initial['movement_type'] = 'disposal'
        item_id = self.request.GET.get('item_id')
        if item_id:
            initial['item'] = item_id
        return initial

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user

        # BTM-Entsorgung erfordert IMMER Freigabe (Vier-Augen-Prinzip)
        if form.instance.item.is_btm:
            form.instance.requires_approval = True
            form.instance.approval_status = BTMApprovalStatus.PENDING
            messages.warning(
                self.request,
                f'BTM-Entsorgung für {form.instance.item.name} wurde zur Freigabe vorgelegt (Vier-Augen-Prinzip).'
            )
        else:
            messages.success(self.request, f'Entsorgung von {form.instance.item.name} wurde erfasst.')

        return super().form_valid(form)


# ============================================================================
# CATEGORY MANAGEMENT
# ============================================================================

class CategoryListView(LoginRequiredMixin, ListView):
    """Liste aller Kategorien"""
    model = Category
    template_name = 'medical/category_list.html'
    context_object_name = 'categories'

    def get_queryset(self):
        return Category.objects.filter(is_active=True).order_by('tree_id', 'lft')


class CategoryCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Neue Kategorie erstellen"""
    model = Category
    template_name = 'medical/category_form.html'
    fields = ['name', 'parent', 'code', 'description']
    success_url = reverse_lazy('medical:category_list')
    permission_required = 'inventory_base.add_category'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, f'Kategorie "{form.instance.name}" wurde erstellt.')
        return super().form_valid(form)


class CategoryUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Kategorie bearbeiten"""
    model = Category
    template_name = 'medical/category_form.html'
    fields = ['name', 'parent', 'code', 'description', 'is_active']
    success_url = reverse_lazy('medical:category_list')
    permission_required = 'inventory_base.change_category'

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, f'Kategorie "{form.instance.name}" wurde aktualisiert.')
        return super().form_valid(form)


class CategoryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Kategorie löschen"""
    model = Category
    template_name = 'medical/category_confirm_delete.html'
    success_url = reverse_lazy('medical:category_list')
    permission_required = 'inventory_base.delete_category'

    def delete(self, request, *args, **kwargs):
        category = self.get_object()
        messages.success(request, f'Kategorie "{category.name}" wurde gelöscht.')
        return super().delete(request, *args, **kwargs)


# ============================================================================
# QUICK STOCK MOVEMENT
# ============================================================================

class QuickStockMovementView(LoginRequiredMixin, TemplateView):
    """
    Schnellerfassung von Lagerbewegungen mit Artikelsuche/-auswahl
    """
    template_name = 'medical/quick_stock_movement.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Alle aktiven Masters für die Auswahl (als Liste für JSON)
        items = MedicalItemMaster.objects.filter(is_active=True).order_by('name')
        context['items'] = items

        # Items als JSON für JavaScript (json_script template tag handles the safety)
        items_data = [{
            'id': item.id,
            'name': item.name,
            'quantity': '0',  # TODO: Über Batches aggregieren
            'unit': item.unit or '',
            'is_btm': item.is_btm
        } for item in items]
        context['items_json'] = items_data

        # Bewegungstypen für Dropdown
        from inventory_base.models import StockMovementType
        context['movement_types'] = StockMovementType.choices

        # Lagerorte für die Dropdowns
        from locations.models import Location
        context['locations'] = Location.objects.filter(is_active=True).order_by('name')

        return context

    def post(self, request, *args, **kwargs):
        """Verarbeite die Lagerbewegung"""
        form = MedicalStockMovementForm(request.POST)

        if form.is_valid():
            movement = form.save(commit=False)
            movement.created_by = request.user
            movement.updated_by = request.user

            # BTM-Prüfung
            if movement.item.is_btm:
                movement.requires_approval = True
                movement.approval_status = BTMApprovalStatus.PENDING
                messages.warning(
                    request,
                    f'BTM-Bewegung für {movement.item.name} wurde zur Freigabe vorgelegt (Vier-Augen-Prinzip).'
                )
            else:
                messages.success(
                    request,
                    f'Lagerbewegung ({movement.get_movement_type_display()}) für {movement.item.name} wurde erfolgreich erfasst.'
                )

            movement.save()

            # Wenn Wareneingang mit Chargen-Info: Charge anlegen
            if movement.movement_type == StockMovementType.INCOMING and movement.batch_number and movement.expiry_date:
                from datetime import date
                batch = MedicalBatch.objects.create(
                    master=movement.item,
                    batch_number=movement.batch_number,
                    received_date=date.today(),
                    expiry_date=movement.expiry_date,
                    quantity_received=movement.quantity,
                    quantity_remaining=movement.quantity,
                    location=movement.to_location,
                    notes=f'Erstellt aus Lagerbewegung #{movement.id}'
                )
                messages.success(
                    request,
                    f'Charge "{batch.batch_number}" wurde automatisch angelegt.'
                )

            # Bleibe auf der Seite für weitere Erfassungen
            return redirect('medical:quick_stock_movement')

        # Bei Fehlern: Zeige Formular mit Fehlern
        context = self.get_context_data()
        context['form'] = form
        return self.render_to_response(context)


class BatchDetailView(LoginRequiredMixin, DetailView):
    """
    Detail-Ansicht einer Charge mit Bewegungshistorie
    """
    model = MedicalBatch
    template_name = 'medical/batch_detail.html'
    context_object_name = 'batch'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        batch = self.object

        # Lagerbewegungen dieser Charge
        movements = MedicalStockMovement.objects.filter(
            batch_number=batch.batch_number
        ).select_related(
            'item', 'from_location', 'to_location', 'created_by'
        ).order_by('-movement_date')
        context['movements'] = movements

        # Temperatur-Logs für diese Charge
        temp_logs = TemperatureLog.objects.filter(
            batch=batch
        ).order_by('-measured_at')[:50]
        context['temp_logs'] = temp_logs

        # Statistiken
        context['total_movements'] = movements.count()
        context['incoming_quantity'] = movements.filter(
            movement_type=StockMovementType.INCOMING
        ).aggregate(total=Sum('quantity'))['total'] or 0
        context['outgoing_quantity'] = movements.filter(
            movement_type=StockMovementType.OUTGOING
        ).aggregate(total=Sum('quantity'))['total'] or 0

        return context


class UniversalBatchCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """
    Universelle Chargen-Erstellung ohne vorherige Artikel-Auswahl
    """
    model = MedicalBatch
    form_class = MedicalBatchForm
    template_name = 'medical/batch_form_universal.html'
    permission_required = 'medical.add_medicalbatch'

    def form_valid(self, form):
        # Master-Artikel setzen
        form.instance.master = form.cleaned_data.get('master_article')
        form.instance.item = None  # Legacy-Feld bleibt leer

        # Restmenge initial auf Eingangsmenge setzen
        form.instance.quantity_remaining = form.cleaned_data['quantity_received']

        messages.success(self.request, f'Charge "{form.instance.batch_number}" wurde erfolgreich angelegt.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('medical:batch_list')



# Legacy BatchCreateView wurde entfernt - verwenden Sie UniversalBatchCreateView


@login_required
def master_batches_json(request, pk):
    """JSON-API: Gibt alle Chargen eines MedicalItemMaster zurück"""
    master = get_object_or_404(MedicalItemMaster, pk=pk)
    batches = MedicalBatch.objects.filter(
        master=master,
        quantity_remaining__gt=0,
    ).select_related('location').order_by('expiry_date')

    data = []
    for batch in batches:
        data.append({
            'id': batch.pk,
            'batch_number': batch.batch_number,
            'expiry_date': batch.expiry_date.isoformat() if batch.expiry_date else '',
            'quantity': str(batch.quantity_remaining),
            'unit': master.unit,
            'location_id': batch.location_id,
            'location_name': str(batch.location) if batch.location else '',
        })

    return JsonResponse(data, safe=False)



# ============================================================================
# DOCUMENT MANAGEMENT
# ============================================================================

@login_required
def item_documents_list(request, item_pk):
    """
    Liste aller Dokumente eines Medical Items (AJAX)
    """
    item = get_object_or_404(MedicalItem, pk=item_pk)
    content_type = ContentType.objects.get_for_model(MedicalItem)

    documents = Document.objects.filter(
        related_content_type=content_type,
        related_object_id=item.pk
    ).select_related('category', 'created_by').order_by('-created_at')

    documents_data = []
    for doc in documents:
        documents_data.append({
            'id': doc.id,
            'title': doc.title,
            'document_number': doc.document_number,
            'document_type': doc.get_document_type_display(),
            'status': doc.get_status_display(),
            'status_color': {
                'draft': 'gray',
                'review': 'yellow',
                'approved': 'green',
                'active': 'blue',
                'superseded': 'orange',
                'expired': 'red',
                'archived': 'gray'
            }.get(doc.status, 'gray'),
            'file_name': doc.file.name.split('/')[-1] if doc.file else '',
            'file_size': doc.get_file_size_display(),
            'file_extension': doc.get_file_extension(),
            'created_at': doc.created_at.strftime('%d.%m.%Y %H:%M'),
            'created_by': doc.created_by.get_full_name(),
            'download_url': reverse_lazy('medical:document_download', kwargs={'item_pk': item_pk, 'doc_pk': doc.pk}),
            'delete_url': reverse_lazy('medical:document_delete', kwargs={'item_pk': item_pk, 'doc_pk': doc.pk}),
        })

    return JsonResponse({'documents': documents_data})


@login_required
def item_document_upload(request, item_pk):
    """
    Dokument zu Medical Item hochladen
    """
    item = get_object_or_404(MedicalItem, pk=item_pk)

    if request.method == 'POST':
        if 'file' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'Keine Datei ausgewählt'}, status=400)

        file = request.FILES['file']
        title = request.POST.get('title', file.name)
        document_type = request.POST.get('document_type', DocumentType.OTHER)
        description = request.POST.get('description', '')
        access_level = request.POST.get('access_level', AccessLevel.INTERNAL)

        # Standard-Kategorie für Medical Items (oder erstelle sie)
        category, _ = DocumentCategory.objects.get_or_create(
            name='Medizinische Artikel',
            defaults={'description': 'Dokumente zu medizinischen Artikeln'}
        )

        # Content Type für Generic FK
        content_type = ContentType.objects.get_for_model(MedicalItem)

        try:
            # Dokument erstellen
            document = Document.objects.create(
                title=title,
                category=category,
                document_type=document_type,
                status=DocumentStatus.ACTIVE,
                file=file,
                description=description,
                access_level=access_level,
                related_content_type=content_type,
                related_object_id=item.pk,
                created_by=request.user,
                updated_by=request.user
            )

            # MIME-Type ermitteln
            import magic
            try:
                mime = magic.Magic(mime=True)
                document.mime_type = mime.from_buffer(file.read(1024))
                file.seek(0)
                document.save()
            except:
                pass

            # Trigger asynchrone OCR-Verarbeitung
            try:
                from documents.tasks import process_document_ocr
                process_document_ocr.delay(document.id)
                logger.info(f'OCR-Task gestartet für Dokument #{document.id}')
            except Exception as e:
                logger.warning(f'OCR-Task konnte nicht gestartet werden: {e}')

            messages.success(request, f'Dokument "{document.title}" wurde erfolgreich hochgeladen.')

            return JsonResponse({
                'success': True,
                'document': {
                    'id': document.id,
                    'title': document.title,
                    'document_number': document.document_number,
                }
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    return JsonResponse({'success': False, 'error': 'Nur POST erlaubt'}, status=405)


@login_required
def item_document_download(request, item_pk, doc_pk):
    """
    Dokument herunterladen und Zugriff protokollieren
    """
    item = get_object_or_404(MedicalItem, pk=item_pk)
    document = get_object_or_404(Document, pk=doc_pk)

    # Zugriff prüfen (Generic FK)
    content_type = ContentType.objects.get_for_model(MedicalItem)
    if document.related_content_type != content_type or document.related_object_id != item.pk:
        messages.error(request, 'Dokument gehört nicht zu diesem Artikel')
        return redirect('medical:item_detail', pk=item_pk)

    # Zugriff protokollieren
    from documents.models import DocumentAccess
    DocumentAccess.objects.create(
        document=document,
        user=request.user,
        access_type='download',
        ip_address=request.META.get('REMOTE_ADDR'),
        user_agent=request.META.get('HTTP_USER_AGENT', '')[:500]
    )

    # Download-Counter erhöhen
    document.increment_download_count()

    # Datei-Response
    from django.http import FileResponse
    response = FileResponse(document.file.open('rb'))
    response['Content-Disposition'] = f'attachment; filename="{document.file.name.split("/")[-1]}"'
    return response


@login_required
def item_document_delete(request, item_pk, doc_pk):
    """
    Dokument löschen
    """
    item = get_object_or_404(MedicalItem, pk=item_pk)
    document = get_object_or_404(Document, pk=doc_pk)

    # Zugriff prüfen
    content_type = ContentType.objects.get_for_model(MedicalItem)
    if document.related_content_type != content_type or document.related_object_id != item.pk:
        messages.error(request, 'Dokument gehört nicht zu diesem Artikel')
        return redirect('medical:item_detail', pk=item_pk)

    if request.method == 'POST':
        # Zugriff protokollieren
        from documents.models import DocumentAccess
        DocumentAccess.objects.create(
            document=document,
            user=request.user,
            access_type='delete',
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:500]
        )

        title = document.title
        document.delete()
        messages.success(request, f'Dokument "{title}" wurde gelöscht.')

        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'error': 'Nur POST erlaubt'}, status=405)


# ============================================================================
# ARTIKEL-STAMMDATEN (MASTER DATA)
# ============================================================================

class MedicalItemMasterListView(LoginRequiredMixin, ListView):
    """
    Liste aller Artikel-Stammdaten (ohne konkrete Lagerorte)
    """
    model = MedicalItemMaster
    template_name = 'medical/master_list.html'
    context_object_name = 'masters'
    paginate_by = 50

    def get_queryset(self):
        queryset = MedicalItemMaster.objects.filter(is_active=True).select_related(
            'category',
            'supplier',
        )

        # Suchfilter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(master_number__icontains=search) |
                Q(description__icontains=search) |
                Q(active_ingredient__icontains=search) |
                Q(pzn__icontains=search) |
                Q(atc_code__icontains=search)
            )

        # Typ-Filter
        item_type = self.request.GET.get('item_type')
        if item_type:
            queryset = queryset.filter(item_type=item_type)

        # Kategorie-Filter
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category_id=category)

        # BTM-Filter
        show_btm_only = self.request.GET.get('btm_only')
        if show_btm_only:
            queryset = queryset.filter(is_btm=True)

        # Medizinprodukt-Filter
        show_devices_only = self.request.GET.get('devices_only')
        if show_devices_only:
            queryset = queryset.filter(is_medical_device=True)

        return queryset.order_by('master_number')


class MedicalItemMasterDetailView(LoginRequiredMixin, DetailView):
    """
    Detail-Ansicht für Artikel-Stammdaten mit Chargen/Instanzen
    """
    model = MedicalItemMaster
    template_name = 'medical/master_detail.html'
    context_object_name = 'master'

    def get_queryset(self):
        return MedicalItemMaster.objects.select_related(
            'category',
            'supplier',
            'created_by',
            'updated_by',
        ).prefetch_related(
            'batches',
            'device_instances',
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        master = self.object

        # Gesamtbestand berechnen
        context['total_stock'] = master.get_total_stock()

        # Aktive Chargen (für Verbrauchsmaterial)
        if not master.is_medical_device:
            context['active_batches'] = master.batches.filter(
                quantity_remaining__gt=0
            ).select_related('location').order_by('expiry_date')

        # Geräte-Instanzen (für Medizintechnik)
        else:
            context['device_instances'] = master.device_instances.filter(
                is_active=True
            ).select_related('location').order_by('inventory_number')

            # Wartungsstatistiken
            today = timezone.now().date()
            context['maintenance_due_count'] = master.device_instances.filter(
                is_active=True,
                next_maintenance_date__isnull=False,
                next_maintenance_date__lte=today
            ).count()

            context['inspection_due_count'] = master.device_instances.filter(
                is_active=True,
                next_inspection_date__isnull=False,
                next_inspection_date__lte=today
            ).count()

        # Bestellung-generieren Button: Zeige wenn Mindestbestand unterschritten
        # und User Procurement-Berechtigung hat
        total_stock = context['total_stock']
        if master.min_quantity and total_stock < master.min_quantity:
            has_procurement_access = (
                self.request.user.is_superuser or
                self.request.user.has_perm('procurement.add_purchaseorder')
            )
            context['show_order_button'] = has_procurement_access
            context['order_deficit'] = master.min_quantity - total_stock
        else:
            context['show_order_button'] = False

        return context


class MedicalItemMasterCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """
    Erstellen neuer Artikel-Stammdaten
    """
    model = MedicalItemMaster
    form_class = MedicalItemMasterForm
    template_name = 'medical/master_form.html'
    success_url = reverse_lazy('medical:master_list')
    permission_required = 'medical.add_medicalitemmaster'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, f'Artikel-Stammdaten "{form.instance.name}" wurden erfolgreich erstellt.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Neue Artikel-Stammdaten'
        context['submit_text'] = 'Stammdaten anlegen'
        return context


class MedicalItemMasterUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """
    Bearbeiten von Artikel-Stammdaten
    """
    model = MedicalItemMaster
    form_class = MedicalItemMasterForm
    template_name = 'medical/master_form.html'
    permission_required = 'medical.change_medicalitemmaster'

    def get_success_url(self):
        return reverse_lazy('medical:master_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, f'Artikel-Stammdaten "{form.instance.name}" wurden erfolgreich aktualisiert.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Bearbeiten: {self.object.name}'
        context['submit_text'] = 'Änderungen speichern'
        return context


class MedicalItemMasterDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """
    Löschen von Artikel-Stammdaten (Soft-Delete via is_active)
    """
    model = MedicalItemMaster
    template_name = 'medical/master_confirm_delete.html'
    success_url = reverse_lazy('medical:master_list')
    permission_required = 'medical.delete_medicalitemmaster'

    def delete(self, request, *args, **kwargs):
        master = self.get_object()

        # Prüfen ob Chargen oder Instanzen vorhanden sind
        has_batches = master.batches.filter(quantity_remaining__gt=0).exists()
        has_instances = master.device_instances.filter(is_active=True).exists()

        if has_batches or has_instances:
            messages.error(
                request,
                f'Artikel-Stammdaten "{master.name}" können nicht gelöscht werden. '
                'Es existieren noch aktive Chargen oder Geräte-Instanzen.'
            )
            return redirect('medical:master_detail', pk=master.pk)

        # Soft-Delete
        master.is_active = False
        master.save()

        messages.success(request, f'Artikel-Stammdaten "{master.name}" wurden deaktiviert.')
        return redirect(self.success_url)


@login_required
def master_qrcode_view(request, pk):
    """
    QR-Code für Artikel-Stammdaten generieren und anzeigen
    """
    master = get_object_or_404(MedicalItemMaster, pk=pk)

    # QR-Code als SVG generieren
    qr_svg = master.generate_qr_code()

    # Als HttpResponse mit SVG-Content-Type
    response = HttpResponse(qr_svg, content_type='image/svg+xml')
    response['Content-Disposition'] = f'inline; filename="qrcode_master_{master.master_number}.svg"'

    return response


@login_required
def master_barcode_view(request, pk):
    """
    Barcode für Artikel-Stammdaten generieren und anzeigen
    """
    master = get_object_or_404(MedicalItemMaster, pk=pk)

    # Barcode als SVG generieren
    barcode_svg = master.generate_barcode()

    # Als HttpResponse mit SVG-Content-Type
    response = HttpResponse(barcode_svg, content_type='image/svg+xml')
    response['Content-Disposition'] = f'inline; filename="barcode_master_{master.master_number}.svg"'

    return response


# ============================================================================
# MEDIZINTECHNIK-INSTANZEN (DEVICE INSTANCES)
# ============================================================================

class MedicalDeviceInstanceListView(LoginRequiredMixin, ListView):
    """
    Liste aller Medizintechnik-Instanzen (einzelne Geräte)
    """
    model = MedicalDeviceInstance
    template_name = 'medical/device_list.html'
    context_object_name = 'devices'
    paginate_by = 50

    def get_queryset(self):
        queryset = MedicalDeviceInstance.objects.filter(is_active=True).select_related(
            'master',
            'location',
        )

        # Suchfilter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(inventory_number__icontains=search) |
                Q(serial_number__icontains=search) |
                Q(master__name__icontains=search) |
                Q(notes__icontains=search)
            )

        # Master-Filter
        master = self.request.GET.get('master')
        if master:
            queryset = queryset.filter(master_id=master)

        # Lagerort-Filter
        location = self.request.GET.get('location')
        if location:
            queryset = queryset.filter(location_id=location)

        # Zustandsfilter
        condition = self.request.GET.get('condition')
        if condition:
            queryset = queryset.filter(condition=condition)

        # Nur einsatzbereite Geräte
        operational_only = self.request.GET.get('operational_only')
        if operational_only:
            queryset = queryset.filter(is_operational=True)

        # Wartung fällig
        maintenance_due = self.request.GET.get('maintenance_due')
        if maintenance_due:
            today = timezone.now().date()
            queryset = queryset.filter(
                next_maintenance_date__isnull=False,
                next_maintenance_date__lte=today
            )

        return queryset.order_by('inventory_number')


class MedicalDeviceInstanceDetailView(LoginRequiredMixin, DetailView):
    """
    Detail-Ansicht einer Medizintechnik-Instanz
    """
    model = MedicalDeviceInstance
    template_name = 'medical/device_detail.html'
    context_object_name = 'device'

    def get_queryset(self):
        return MedicalDeviceInstance.objects.select_related(
            'master',
            'location',
            'created_by',
            'updated_by',
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        device = self.object
        today = timezone.now().date()

        # Wartungsstatus
        if device.next_maintenance_date:
            days_until_maintenance = (device.next_maintenance_date - today).days
            context['days_until_maintenance'] = days_until_maintenance
            context['maintenance_overdue'] = days_until_maintenance < 0

        # Prüfungsstatus
        if device.next_inspection_date:
            days_until_inspection = (device.next_inspection_date - today).days
            context['days_until_inspection'] = days_until_inspection
            context['inspection_overdue'] = days_until_inspection < 0

        return context


class MedicalDeviceInstanceCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """
    Neue Medizintechnik-Instanz erstellen
    """
    model = MedicalDeviceInstance
    form_class = MedicalDeviceInstanceForm
    template_name = 'medical/device_form.html'
    success_url = reverse_lazy('medical:device_list')
    permission_required = 'medical.add_medicaldeviceinstance'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user

        # Nächste Wartung berechnen (falls nicht manuell gesetzt)
        if not form.instance.next_maintenance_date and form.instance.master.requires_maintenance:
            if form.instance.master.maintenance_interval_months:
                from dateutil.relativedelta import relativedelta
                if form.instance.last_maintenance_date:
                    form.instance.next_maintenance_date = form.instance.last_maintenance_date + relativedelta(
                        months=form.instance.master.maintenance_interval_months
                    )
                elif form.instance.commissioning_date:
                    form.instance.next_maintenance_date = form.instance.commissioning_date + relativedelta(
                        months=form.instance.master.maintenance_interval_months
                    )

        messages.success(
            self.request,
            f'Medizintechnik-Instanz "{form.instance.inventory_number}" wurde erfolgreich erstellt.'
        )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Neue Medizintechnik-Instanz'
        context['submit_text'] = 'Gerät anlegen'

        # Master vorausfüllen (falls in URL)
        master_id = self.request.GET.get('master_id')
        if master_id:
            try:
                context['preselected_master'] = MedicalItemMaster.objects.get(pk=master_id)
            except MedicalItemMaster.DoesNotExist:
                pass

        return context


class MedicalDeviceInstanceUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """
    Medizintechnik-Instanz bearbeiten
    """
    model = MedicalDeviceInstance
    form_class = MedicalDeviceInstanceForm
    template_name = 'medical/device_form.html'
    permission_required = 'medical.change_medicaldeviceinstance'

    def get_success_url(self):
        return reverse_lazy('medical:device_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(
            self.request,
            f'Medizintechnik-Instanz "{form.instance.inventory_number}" wurde erfolgreich aktualisiert.'
        )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Bearbeiten: {self.object.inventory_number}'
        context['submit_text'] = 'Änderungen speichern'
        return context


class MedicalDeviceInstanceDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """
    Medizintechnik-Instanz löschen (Soft-Delete via is_active)
    """
    model = MedicalDeviceInstance
    template_name = 'medical/device_confirm_delete.html'
    success_url = reverse_lazy('medical:device_list')
    permission_required = 'medical.delete_medicaldeviceinstance'

    def delete(self, request, *args, **kwargs):
        device = self.get_object()

        # Soft-Delete
        device.is_active = False
        device.decommissioned_date = timezone.now().date()
        device.save()

        messages.success(request, f'Medizintechnik-Instanz "{device.inventory_number}" wurde ausgemustert.')
        return redirect(self.success_url)


@login_required
def device_qrcode_view(request, pk):
    """
    QR-Code für Medizintechnik-Instanz generieren und anzeigen
    """
    device = get_object_or_404(MedicalDeviceInstance, pk=pk)

    # QR-Code als SVG generieren
    qr_svg = device.generate_qr_code()

    # Als HttpResponse mit SVG-Content-Type
    response = HttpResponse(qr_svg, content_type='image/svg+xml')
    response['Content-Disposition'] = f'inline; filename="qrcode_device_{device.inventory_number}.svg"'

    return response


@login_required
def device_barcode_view(request, pk):
    """
    Barcode für Medizintechnik-Instanz generieren und anzeigen
    """
    device = get_object_or_404(MedicalDeviceInstance, pk=pk)

    # Barcode als SVG generieren
    barcode_svg = device.generate_barcode()

    # Als HttpResponse mit SVG-Content-Type
    response = HttpResponse(barcode_svg, content_type='image/svg+xml')
    response['Content-Disposition'] = f'inline; filename="barcode_device_{device.inventory_number}.svg"'

    return response


@login_required
def device_batch_qrcodes(request):
    """
    Batch-Generierung von QR-Codes für mehrere Medizintechnik-Instanzen
    """
    device_ids = request.GET.getlist('device_ids')

    if not device_ids:
        messages.error(request, 'Keine Geräte ausgewählt.')
        return redirect('medical:device_list')

    devices = MedicalDeviceInstance.objects.filter(id__in=device_ids, is_active=True)

    if not devices:
        messages.error(request, 'Keine gültigen Geräte gefunden.')
        return redirect('medical:device_list')

    # HTML-Template für Batch-QR-Codes
    from django.template.loader import render_to_string

    html_content = render_to_string('medical/device_batch_qrcodes.html', {
        'devices': devices,
    })

    return HttpResponse(html_content)


@login_required
def device_batch_barcodes(request):
    """
    Batch-Generierung von Barcodes für mehrere Medizintechnik-Instanzen
    """
    device_ids = request.GET.getlist('device_ids')

    if not device_ids:
        messages.error(request, 'Keine Geräte ausgewählt.')
        return redirect('medical:device_list')

    devices = MedicalDeviceInstance.objects.filter(id__in=device_ids, is_active=True)

    if not devices:
        messages.error(request, 'Keine gültigen Geräte gefunden.')
        return redirect('medical:device_list')

    # HTML-Template für Batch-Barcodes
    from django.template.loader import render_to_string

    html_content = render_to_string('medical/device_batch_barcodes.html', {
        'devices': devices,
    })

    return HttpResponse(html_content)


@login_required
def device_print_labels(request):
    """
    Druckbare Etiketten für Medizintechnik-Instanzen generieren (QR + Barcode kombiniert)
    Optimiert für Etikettendrucker (z.B. Brother QL-Serie, Dymo, Zebra)
    """
    device_ids = request.GET.getlist('device_ids')

    if not device_ids:
        messages.error(request, 'Keine Geräte ausgewählt.')
        return redirect('medical:device_list')

    devices = MedicalDeviceInstance.objects.filter(
        id__in=device_ids,
        is_active=True
    ).select_related('master', 'location')

    if not devices:
        messages.error(request, 'Keine gültigen Geräte gefunden.')
        return redirect('medical:device_list')

    # Format-Auswahl (kann später per GET-Parameter gesteuert werden)
    label_format = request.GET.get('format', 'default')  # default, small, large

    # HTML-Template für druckbare Etiketten
    from django.template.loader import render_to_string

    html_content = render_to_string('medical/device_print_labels.html', {
        'devices': devices,
        'label_format': label_format,
    })

    return HttpResponse(html_content)


# ============================================================================
# IMPORT/EXPORT
# ============================================================================

class ImportExportView(LoginRequiredMixin, TemplateView):
    """
    Import/Export-Verwaltung für Rettungsdienst-Daten
    """
    template_name = 'medical/import_export.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Statistiken für Export
        context['total_masters'] = MedicalItemMaster.objects.filter(is_active=True).count()
        context['total_devices'] = MedicalDeviceInstance.objects.filter(is_active=True).count()
        context['total_batches'] = MedicalBatch.objects.filter(is_recalled=False).count()
        context['total_movements'] = MedicalStockMovement.objects.count()

        return context


@login_required
def export_masters(request):
    """
    Export Medical Item Masters als Excel
    """
    masters = MedicalItemMaster.objects.filter(is_active=True).select_related(
        'category', 'manufacturer', 'created_by', 'updated_by'
    ).order_by('master_number')

    # Excel-Workbook erstellen
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stammdaten"

    # Header
    headers = [
        'Stammdatennummer', 'Name', 'Kategorie', 'PZN', 'ATC-Code',
        'Interne Bestellnr.', 'Externe Bestellnr.', 'Hersteller',
        'Wirkstoff', 'Einheit', 'Stärke', 'Darreichungsform',
        'BTM', 'Kühlpflichtig', 'Beschreibung', 'Erstellt am', 'Aktualisiert am'
    ]

    # Header-Styling
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Daten
    for row_num, master in enumerate(masters, 2):
        ws.cell(row=row_num, column=1, value=master.master_number)
        ws.cell(row=row_num, column=2, value=master.name)
        ws.cell(row=row_num, column=3, value=master.category.name if master.category else '')
        ws.cell(row=row_num, column=4, value=master.pzn)
        ws.cell(row=row_num, column=5, value=master.atc_code)
        ws.cell(row=row_num, column=6, value=master.internal_order_number)
        ws.cell(row=row_num, column=7, value=master.external_order_number)
        ws.cell(row=row_num, column=8, value=master.manufacturer.name if master.manufacturer else '')
        ws.cell(row=row_num, column=9, value=master.active_ingredient)
        ws.cell(row=row_num, column=10, value=master.unit)
        ws.cell(row=row_num, column=11, value=master.strength)
        ws.cell(row=row_num, column=12, value=master.dosage_form)
        ws.cell(row=row_num, column=13, value='Ja' if master.is_btm else 'Nein')
        ws.cell(row=row_num, column=14, value='Ja' if master.requires_cooling else 'Nein')
        ws.cell(row=row_num, column=15, value=master.description)
        ws.cell(row=row_num, column=16, value=master.created_at.strftime('%d.%m.%Y %H:%M') if master.created_at else '')
        ws.cell(row=row_num, column=17, value=master.updated_at.strftime('%d.%m.%Y %H:%M') if master.updated_at else '')

    # Spaltenbreite anpassen
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=medical_stammdaten_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response


@login_required
def export_devices(request):
    """
    Export Medical Device Instances als Excel
    """
    devices = MedicalDeviceInstance.objects.filter(is_active=True).select_related(
        'master', 'location', 'created_by'
    ).order_by('inventory_number')

    # Excel-Workbook erstellen
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Medizintechnik"

    # Header
    headers = [
        'Inventarnummer', 'Stammdatennummer', 'Name', 'Seriennummer',
        'Standort', 'Status', 'Nächste Prüfung', 'Nächste Wartung',
        'Anschaffungsdatum', 'Anschaffungspreis', 'Notizen', 'Erstellt am'
    ]

    # Header-Styling
    header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Daten
    for row_num, device in enumerate(devices, 2):
        ws.cell(row=row_num, column=1, value=device.inventory_number)
        ws.cell(row=row_num, column=2, value=device.master.master_number if device.master else '')
        ws.cell(row=row_num, column=3, value=device.master.name if device.master else '')
        ws.cell(row=row_num, column=4, value=device.serial_number)
        ws.cell(row=row_num, column=5, value=device.location.name if device.location else '')
        ws.cell(row=row_num, column=6, value=device.get_status_display())
        ws.cell(row=row_num, column=7, value=device.next_inspection_date.strftime('%d.%m.%Y') if device.next_inspection_date else '')
        ws.cell(row=row_num, column=8, value=device.next_maintenance_date.strftime('%d.%m.%Y') if device.next_maintenance_date else '')
        ws.cell(row=row_num, column=9, value=device.acquisition_date.strftime('%d.%m.%Y') if device.acquisition_date else '')
        ws.cell(row=row_num, column=10, value=f'{device.acquisition_price:.2f}' if device.acquisition_price else '')
        ws.cell(row=row_num, column=11, value=device.notes)
        ws.cell(row=row_num, column=12, value=device.created_at.strftime('%d.%m.%Y %H:%M') if device.created_at else '')

    # Spaltenbreite anpassen
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=medical_medizintechnik_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response


@login_required
def export_batches(request):
    """
    Export Medical Batches als Excel
    """
    batches = MedicalBatch.objects.filter(is_recalled=False).select_related(
        'master', 'item', 'location'
    ).order_by('-expiry_date')

    # Excel-Workbook erstellen
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chargen"

    # Header
    headers = [
        'Chargennummer', 'Artikel', 'Standort', 'Menge (Gesamt)',
        'Menge (Verbleibend)', 'Einheit', 'Verfallsdatum',
        'Erstellt am', 'Status'
    ]

    # Header-Styling
    header_fill = PatternFill(start_color='9966FF', end_color='9966FF', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Daten
    today = timezone.now().date()
    for row_num, batch in enumerate(batches, 2):
        article = batch.master if batch.master else batch.item

        # Status ermitteln
        if batch.quantity_remaining <= 0:
            status = 'Leer'
        elif batch.expiry_date < today:
            status = 'Abgelaufen'
        elif batch.expiry_date <= today + timedelta(days=30):
            status = 'Bald ablaufend'
        else:
            status = 'Aktiv'

        ws.cell(row=row_num, column=1, value=batch.batch_number)
        ws.cell(row=row_num, column=2, value=article.name if article else '')
        ws.cell(row=row_num, column=3, value=batch.location.name if batch.location else '')
        ws.cell(row=row_num, column=4, value=f'{batch.quantity:.2f}')
        ws.cell(row=row_num, column=5, value=f'{batch.quantity_remaining:.2f}')
        ws.cell(row=row_num, column=6, value=batch.unit)
        ws.cell(row=row_num, column=7, value=batch.expiry_date.strftime('%d.%m.%Y') if batch.expiry_date else '')
        ws.cell(row=row_num, column=8, value=batch.created_at.strftime('%d.%m.%Y %H:%M') if batch.created_at else '')
        ws.cell(row=row_num, column=9, value=status)

    # Spaltenbreite anpassen
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=medical_chargen_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response


@login_required
def export_movements(request):
    """
    Export Stock Movements als Excel
    """
    movements = MedicalStockMovement.objects.select_related(
        'item', 'from_location', 'to_location', 'created_by'
    ).order_by('-movement_date')[:1000]  # Limit auf 1000 neueste

    # Excel-Workbook erstellen
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lagerbewegungen"

    # Header
    headers = [
        'Datum', 'Typ', 'Artikel', 'Menge', 'Einheit',
        'Von', 'Nach', 'Chargennummer', 'Referenz',
        'Notizen', 'Erstellt von'
    ]

    # Header-Styling
    header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    header_font = Font(bold=True, color='000000')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Daten
    for row_num, movement in enumerate(movements, 2):
        ws.cell(row=row_num, column=1, value=movement.movement_date.strftime('%d.%m.%Y %H:%M') if movement.movement_date else '')
        ws.cell(row=row_num, column=2, value=movement.get_movement_type_display())
        ws.cell(row=row_num, column=3, value=movement.item.name if movement.item else '')
        ws.cell(row=row_num, column=4, value=f'{movement.quantity:.2f}')
        ws.cell(row=row_num, column=5, value=movement.unit)
        ws.cell(row=row_num, column=6, value=movement.from_location.name if movement.from_location else '')
        ws.cell(row=row_num, column=7, value=movement.to_location.name if movement.to_location else '')
        ws.cell(row=row_num, column=8, value=movement.batch_number)
        ws.cell(row=row_num, column=9, value=movement.reference_number)
        ws.cell(row=row_num, column=10, value=movement.notes)
        ws.cell(row=row_num, column=11, value=movement.created_by.get_full_name() if movement.created_by else '')

    # Spaltenbreite anpassen
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=medical_bewegungen_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response


# ============================================================================
# EXCEL TEMPLATE DOWNLOADS
# ============================================================================

@login_required
def template_masters(request):
    """
    CSV-Vorlage für Stammdaten-Import
    """
    # CSV mit UTF-8 BOM für Excel-Kompatibilität
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)

    # Header
    headers = [
        'Stammdatennummer*', 'Name*', 'Kategorie-ID', 'PZN', 'ATC-Code',
        'Interne Bestellnr.', 'Externe Bestellnr.', 'Hersteller-ID',
        'Wirkstoff', 'Einheit*', 'Stärke', 'Darreichungsform',
        'BTM (Ja/Nein)', 'Kühlpflichtig (Ja/Nein)', 'Beschreibung'
    ]
    writer.writerow(headers)

    # Beispielzeile
    example_data = [
        'MED-001', 'Ibuprofen 600mg', '1', '12345678', 'M01AE01',
        'IBO-600', 'EXT-IBO-600', '1',
        'Ibuprofen', 'Stück', '600mg', 'Filmtablette',
        'Nein', 'Nein', 'Beispiel: Schmerzmittel - diese Zeile löschen!'
    ]
    writer.writerow(example_data)

    # Leere Zeile
    writer.writerow([])

    # Hinweise als Kommentare
    writer.writerow(['# HINWEISE:'])
    writer.writerow(['# Felder mit * sind Pflichtfelder'])
    writer.writerow(['# Kategorie-ID und Hersteller-ID müssen existierende IDs sein'])
    writer.writerow(['# BTM und Kühlpflichtig: Schreiben Sie "Ja" oder "Nein"'])
    writer.writerow(['# Zeile 2 ist ein Beispiel und wird beim Import ignoriert'])
    writer.writerow(['# Löschen Sie die Beispielzeile und diese Hinweise vor dem Import'])

    # Response mit UTF-8 BOM
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename=vorlage_stammdaten_import.csv'
    response.write('\ufeff')  # UTF-8 BOM
    response.write(output.getvalue())

    return response


@login_required
def template_devices(request):
    """
    CSV-Vorlage für Medizintechnik-Import
    """
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)

    # Header
    headers = [
        'Inventarnummer*', 'Stammdatennummer*', 'Seriennummer', 'Standort-ID',
        'Status', 'Nächste Prüfung (TT.MM.JJJJ)', 'Nächste Wartung (TT.MM.JJJJ)',
        'Anschaffungsdatum (TT.MM.JJJJ)', 'Anschaffungspreis', 'Notizen'
    ]
    writer.writerow(headers)

    # Beispielzeile
    example_data = [
        'DEF-001', 'MED-001', 'SN123456', '1',
        'operational', '31.12.2025', '30.06.2025',
        '01.01.2024', '1500.00', 'Beispiel: Defibrillator - diese Zeile löschen!'
    ]
    writer.writerow(example_data)

    # Leere Zeile
    writer.writerow([])

    # Hinweise
    writer.writerow(['# HINWEISE:'])
    writer.writerow(['# Felder mit * sind Pflichtfelder'])
    writer.writerow(['# Stammdatennummer muss in Stammdaten existieren'])
    writer.writerow(['# Status-Werte: operational, maintenance, defect, out_of_service'])
    writer.writerow(['# Datumsformat: TT.MM.JJJJ (z.B. 31.12.2025)'])
    writer.writerow(['# Zeile 2 ist ein Beispiel - löschen vor Import'])

    # Response
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename=vorlage_medizintechnik_import.csv'
    response.write('\ufeff')  # UTF-8 BOM
    response.write(output.getvalue())

    return response


@login_required
def template_batches(request):
    """
    CSV-Vorlage für Chargen-Import
    """
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)

    # Header
    headers = [
        'Chargennummer*', 'Stammdatennummer*', 'Standort-ID*',
        'Menge*', 'Einheit*', 'Verfallsdatum* (TT.MM.JJJJ)'
    ]
    writer.writerow(headers)

    # Beispielzeile
    example_data = [
        'CH-2025-001', 'MED-001', '1',
        '100', 'Stück', '31.12.2026'
    ]
    writer.writerow(example_data)

    # Leere Zeile
    writer.writerow([])

    # Hinweise
    writer.writerow(['# HINWEISE:'])
    writer.writerow(['# Alle Felder sind Pflichtfelder'])
    writer.writerow(['# Stammdatennummer muss existieren'])
    writer.writerow(['# Standort-ID muss existieren'])
    writer.writerow(['# Datumsformat: TT.MM.JJJJ (z.B. 31.12.2026)'])
    writer.writerow(['# Zeile 2 ist ein Beispiel - löschen vor Import'])

    # Response
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename=vorlage_chargen_import.csv'
    response.write('\ufeff')  # UTF-8 BOM
    response.write(output.getvalue())

    return response


# ============================================================================
# IMPORT FUNCTIONS
# ============================================================================

@login_required
def import_masters(request):
    """
    Import Medical Item Masters aus CSV
    """
    if request.method != 'POST':
        messages.error(request, 'Nur POST-Anfragen erlaubt.')
        return redirect('medical:import_export')

    if 'file' not in request.FILES:
        messages.error(request, 'Keine Datei hochgeladen.')
        return redirect('medical:import_export')

    uploaded_file = request.FILES['file']

    try:
        # CSV-Datei lesen (UTF-8 mit BOM)
        decoded_file = uploaded_file.read().decode('utf-8-sig')
        csv_reader = csv.reader(io.StringIO(decoded_file), delimiter=';')

        success_count = 0
        error_count = 0
        errors = []

        # Header überspringen
        next(csv_reader, None)

        # Zeilen durchgehen
        for row_num, row in enumerate(csv_reader, start=2):
            # Leere Zeilen und Kommentare überspringen
            if not row or not row[0] or row[0].startswith('#'):
                continue

            # Beispielzeile überspringen
            if 'Beispiel' in str(row[-1] if row else ''):
                continue

            try:
                # Daten extrahieren (mit Index-Safety)
                master_number = row[0] if len(row) > 0 else None
                name = row[1] if len(row) > 1 else None
                category_id = row[2] if len(row) > 2 else None
                pzn = row[3] if len(row) > 3 else None
                atc_code = row[4] if len(row) > 4 else None
                internal_order_number = row[5] if len(row) > 5 else None
                external_order_number = row[6] if len(row) > 6 else None
                manufacturer_id = row[7] if len(row) > 7 else None
                active_ingredient = row[8] if len(row) > 8 else None
                unit = row[9] if len(row) > 9 else None
                strength = row[10] if len(row) > 10 else None
                dosage_form = row[11] if len(row) > 11 else None
                is_btm = str(row[12] if len(row) > 12 else '').lower() in ['ja', 'yes', '1', 'true']
                requires_cooling = str(row[13] if len(row) > 13 else '').lower() in ['ja', 'yes', '1', 'true']
                description = row[14] if len(row) > 14 else None

                # Validierung
                if not master_number or not name or not unit:
                    errors.append(f"Zeile {row_num}: Pflichtfelder fehlen (Stammdatennummer, Name, Einheit)")
                    error_count += 1
                    continue

                # Duplikatprüfung
                if MedicalItemMaster.objects.filter(master_number=master_number).exists():
                    errors.append(f"Zeile {row_num}: Stammdatennummer '{master_number}' existiert bereits")
                    error_count += 1
                    continue

                # Kategorie laden (optional)
                category = None
                if category_id and category_id.strip():
                    try:
                        category = Category.objects.get(id=int(category_id))
                    except (Category.DoesNotExist, ValueError):
                        errors.append(f"Zeile {row_num}: Kategorie-ID {category_id} nicht gefunden")
                        error_count += 1
                        continue

                # Hersteller laden (optional)
                from inventory_base.models import Manufacturer
                manufacturer = None
                if manufacturer_id and manufacturer_id.strip():
                    try:
                        manufacturer = Manufacturer.objects.get(id=int(manufacturer_id))
                    except (Manufacturer.DoesNotExist, ValueError):
                        pass  # Optional, kein Fehler

                # Master erstellen
                master = MedicalItemMaster.objects.create(
                    master_number=master_number,
                    name=name,
                    category=category,
                    pzn=pzn or None,
                    atc_code=atc_code or None,
                    internal_order_number=internal_order_number or None,
                    external_order_number=external_order_number or None,
                    manufacturer=manufacturer,
                    active_ingredient=active_ingredient or None,
                    unit=unit,
                    strength=strength or None,
                    dosage_form=dosage_form or None,
                    is_btm=is_btm,
                    requires_cooling=requires_cooling,
                    description=description or None,
                    created_by=request.user,
                    updated_by=request.user
                )

                success_count += 1

            except Exception as e:
                errors.append(f"Zeile {row_num}: Fehler - {str(e)}")
                error_count += 1
                logger.error(f"Import error in row {row_num}: {e}")

        # Erfolgsmeldung
        if success_count > 0:
            messages.success(request, f"✅ {success_count} Stammdaten erfolgreich importiert.")

        if error_count > 0:
            messages.warning(request, f"⚠️ {error_count} Fehler beim Import. Details:")
            for error in errors[:10]:  # Max 10 Fehler anzeigen
                messages.error(request, error)
            if len(errors) > 10:
                messages.error(request, f"... und {len(errors) - 10} weitere Fehler")

    except Exception as e:
        messages.error(request, f"Fehler beim Lesen der CSV-Datei: {str(e)}")
        logger.error(f"CSV import error: {e}")

    return redirect('medical:import_export')


@login_required
@login_required
def import_devices(request):
    """Import Medical Device Instances aus CSV"""
    if request.method != 'POST':
        messages.error(request, 'Nur POST-Anfragen erlaubt.')
        return redirect('medical:import_export')
    if 'file' not in request.FILES:
        messages.error(request, 'Keine Datei hochgeladen.')
        return redirect('medical:import_export')
    uploaded_file = request.FILES['file']
    try:
        decoded_file = uploaded_file.read().decode('utf-8-sig')
        csv_reader = csv.reader(io.StringIO(decoded_file), delimiter=';')
        next(csv_reader, None)
        success_count, error_count, errors = 0, 0, []
        for row_num, row in enumerate(csv_reader, start=2):
            if not row or not row[0] or row[0].startswith('#') or 'Beispiel' in str(row[-1] if row else ''):
                continue
            try:
                inventory_number = row[0] if len(row) > 0 else None
                master_number = row[1] if len(row) > 1 else None
                if not inventory_number or not master_number:
                    errors.append(f"Zeile {row_num}: Pflichtfelder fehlen")
                    error_count += 1
                    continue
                if MedicalDeviceInstance.objects.filter(inventory_number=inventory_number).exists():
                    errors.append(f"Zeile {row_num}: Inventarnummer '{inventory_number}' existiert bereits")
                    error_count += 1
                    continue
                try:
                    master = MedicalItemMaster.objects.get(master_number=master_number)
                except MedicalItemMaster.DoesNotExist:
                    errors.append(f"Zeile {row_num}: Stammdatennummer '{master_number}' nicht gefunden")
                    error_count += 1
                    continue
                from locations.models import Location
                from datetime import datetime
                location = None
                if len(row) > 3 and row[3] and row[3].strip():
                    try:
                        location = Location.objects.get(id=int(row[3]))
                    except:
                        pass
                next_inspection_date = next_maintenance_date = acquisition_date = None
                if len(row) > 5 and row[5]:
                    try:
                        next_inspection_date = datetime.strptime(str(row[5]), '%d.%m.%Y').date()
                    except:
                        pass
                if len(row) > 6 and row[6]:
                    try:
                        next_maintenance_date = datetime.strptime(str(row[6]), '%d.%m.%Y').date()
                    except:
                        pass
                if len(row) > 7 and row[7]:
                    try:
                        acquisition_date = datetime.strptime(str(row[7]), '%d.%m.%Y').date()
                    except:
                        pass
                MedicalDeviceInstance.objects.create(
                    inventory_number=inventory_number,
                    master=master,
                    serial_number=row[2] if len(row) > 2 else None,
                    location=location,
                    status=row[4] if len(row) > 4 and row[4] else 'operational',
                    next_inspection_date=next_inspection_date,
                    next_maintenance_date=next_maintenance_date,
                    acquisition_date=acquisition_date,
                    acquisition_price=row[8] if len(row) > 8 else None,
                    notes=row[9] if len(row) > 9 else None,
                    created_by=request.user
                )
                success_count += 1
            except Exception as e:
                errors.append(f"Zeile {row_num}: Fehler - {str(e)}")
                error_count += 1
                logger.error(f"Device import error in row {row_num}: {e}")
        if success_count > 0:
            messages.success(request, f"✅ {success_count} Geräte erfolgreich importiert.")
        if error_count > 0:
            messages.warning(request, f"⚠️ {error_count} Fehler beim Import.")
            for error in errors[:10]:
                messages.error(request, error)
            if len(errors) > 10:
                messages.error(request, f"... und {len(errors) - 10} weitere Fehler")
    except Exception as e:
        messages.error(request, f"Fehler beim Lesen der CSV-Datei: {str(e)}")
        logger.error(f"CSV import error: {e}")
    return redirect('medical:import_export')


@login_required
@login_required
@login_required
def import_batches(request):
    """
    Import Medical Batches aus CSV
    """
    if request.method != 'POST':
        messages.error(request, 'Nur POST-Anfragen erlaubt.')
        return redirect('medical:import_export')

    if 'file' not in request.FILES:
        messages.error(request, 'Keine Datei hochgeladen.')
        return redirect('medical:import_export')

    uploaded_file = request.FILES['file']

    try:
        # CSV-Datei lesen (UTF-8 mit BOM)
        decoded_file = uploaded_file.read().decode('utf-8-sig')
        csv_reader = csv.reader(io.StringIO(decoded_file), delimiter=';')

        success_count = 0
        error_count = 0
        errors = []

        # Header überspringen
        next(csv_reader, None)

        # Zeilen durchgehen
        for row_num, row in enumerate(csv_reader, start=2):
            # Leere Zeilen und Kommentare überspringen
            if not row or not row[0] or row[0].startswith('#'):
                continue

            # Beispielzeile überspringen
            if 'Beispiel' in str(row[-1] if row else ''):
                continue

            try:
                # Daten extrahieren (mit Index-Safety)
                batch_number = row[0] if len(row) > 0 else None
                master_number = row[1] if len(row) > 1 else None
                location_id = row[2] if len(row) > 2 else None
                quantity = row[3] if len(row) > 3 else None
                unit = row[4] if len(row) > 4 else None
                expiry_date_str = row[5] if len(row) > 5 else None

                # Validierung
                if not all([batch_number, master_number, location_id, quantity, unit, expiry_date_str]):
                    errors.append(f"Zeile {row_num}: Alle Felder sind Pflichtfelder")
                    error_count += 1
                    continue

                # Duplikatprüfung
                if MedicalBatch.objects.filter(batch_number=batch_number).exists():
                    errors.append(f"Zeile {row_num}: Chargennummer '{batch_number}' existiert bereits")
                    error_count += 1
                    continue

                # Master laden
                try:
                    master = MedicalItemMaster.objects.get(master_number=master_number)
                except MedicalItemMaster.DoesNotExist:
                    errors.append(f"Zeile {row_num}: Stammdatennummer '{master_number}' nicht gefunden")
                    error_count += 1
                    continue

                # Standort laden
                from locations.models import Location
                try:
                    location = Location.objects.get(id=int(location_id))
                except (Location.DoesNotExist, ValueError):
                    errors.append(f"Zeile {row_num}: Standort-ID {location_id} nicht gefunden")
                    error_count += 1
                    continue

                # Datumskonvertierung
                from datetime import datetime
                try:
                    expiry_date = datetime.strptime(str(expiry_date_str), '%d.%m.%Y').date()
                except:
                    errors.append(f"Zeile {row_num}: Ungültiges Datumsformat (erwartet: TT.MM.JJJJ)")
                    error_count += 1
                    continue

                # Batch erstellen
                batch = MedicalBatch.objects.create(
                    batch_number=batch_number,
                    master=master,
                    location=location,
                    quantity=float(quantity),
                    quantity_remaining=float(quantity),
                    unit=unit,
                    expiry_date=expiry_date,
                    created_by=request.user
                )

                success_count += 1

            except Exception as e:
                errors.append(f"Zeile {row_num}: Fehler - {str(e)}")
                error_count += 1
                logger.error(f"Batch import error in row {row_num}: {e}")

        # Erfolgsmeldung
        if success_count > 0:
            messages.success(request, f"✅ {success_count} Chargen erfolgreich importiert.")

        if error_count > 0:
            messages.warning(request, f"⚠️ {error_count} Fehler beim Import.")
            for error in errors[:10]:
                messages.error(request, error)
            if len(errors) > 10:
                messages.error(request, f"... und {len(errors) - 10} weitere Fehler")

    except Exception as e:
        messages.error(request, f"Fehler beim Lesen der CSV-Datei: {str(e)}")
        logger.error(f"CSV import error: {e}")

    return redirect('medical:import_export')


    return redirect('medical:import_export')


# ============================================================================
# INVENTUR VIEWS (aus inventory_views.py importiert)
# ============================================================================

from .inventory_views import (
    MedicalInventoryListView,
    MedicalInventoryCreateView,
    MedicalInventoryDetailView,
    MedicalInventoryStartView,
    MedicalInventoryCountingView,
    MedicalInventoryCompleteView,
    MedicalInventoryApproveView,
    MedicalInventoryItemUpdateView,
    MedicalInventoryProgressView,
    MedicalInventoryExportView,
)


# ============================================================================
# BESTELLUNG AUS MINDESTBESTAND GENERIEREN
# ============================================================================

class GenerateOrderFromMasterView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    Generiert eine Bestellanforderung (PurchaseOrder) aus einem MedicalItemMaster,
    wenn der Mindestbestand unterschritten ist.
    """
    permission_required = 'procurement.add_purchaseorder'

    def post(self, request, pk):
        from procurement.models import PurchaseOrder, OrderItem, OrderStatus
        from django.contrib.contenttypes.models import ContentType

        master = get_object_or_404(MedicalItemMaster, pk=pk)

        total_stock = master.get_total_stock()
        if not master.min_quantity or total_stock >= master.min_quantity:
            messages.warning(request, 'Mindestbestand ist nicht unterschritten.')
            return redirect('medical:master_detail', pk=pk)

        deficit = master.min_quantity - total_stock

        # PurchaseOrder erstellen
        order = PurchaseOrder()
        order.title = f'Nachbestellung: {master.name}'
        order.status = OrderStatus.DRAFT
        order.requested_date = timezone.now().date()
        order.created_by = request.user
        order.updated_by = request.user

        # Lieferant uebernehmen falls vorhanden
        if master.supplier:
            order.supplier = master.supplier

        # Ansprechpartner setzen (Person des aktuellen Users)
        if hasattr(request.user, 'person') and request.user.person:
            order.requested_by = request.user.person

        order.save()

        # OrderItem mit Rueckverweis auf MedicalItemMaster
        ct = ContentType.objects.get_for_model(MedicalItemMaster)
        OrderItem.objects.create(
            purchase_order=order,
            item_name=master.name,
            quantity=deficit,
            unit=master.unit or 'Stück',
            unit_price=0,
            inventory_content_type=ct,
            inventory_object_id=master.pk,
        )

        # Gesamtsumme berechnen
        order.gesamtsumme_netto = 0
        order.save(update_fields=['gesamtsumme_netto'])

        messages.success(
            request,
            f'Bestellanforderung {order.order_number} wurde erstellt. '
            f'Bitte vervollstaendigen Sie die Bestellung.'
        )
        return redirect('procurement:order_update', pk=order.pk)
