from django import forms
from django.shortcuts import render, redirect
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DetailView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.urls import reverse_lazy
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from django.db.models import Count, Q
from django.http import HttpResponse
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import csv
import io
import logging

logger = logging.getLogger(__name__)

from .models import (
    DivingItemMaster,
    DivingDeviceInstance,
    DivingItem,
    DivingStockMovement,
    DivingServiceLog,
    DivingServiceType,
    DivingServiceAssignment,
    DivingServiceRecord,
    InspectionStatus,
    DivingItemType,
    EquipmentType,
    BottleType,
    DivingBatch,
)
from personnel.models import Person, Qualification, DutyHoursEntry, DutyHoursRequirement
from django.db.models import Sum
from .forms import (
    DivingItemMasterForm,
    DivingDeviceInstanceForm,
    DivingStockMovementForm,
    DivingServiceLogForm,
    DivingServiceTypeForm,
    DivingServiceAssignmentForm,
    DivingServiceRecordForm,
    DynamicServiceRecordForm,
    EquipmentTypeForm,
    BottleTypeForm,
    DivingBatchForm,
)


# ============================================================================
# STAMMDATEN VIEWS (MASTER DATA)
# ============================================================================

class DivingMasterListView(LoginRequiredMixin, ListView):
    """Liste aller Tauchausrüstungs-Stammdaten"""
    model = DivingItemMaster
    template_name = 'diving/master_list.html'
    context_object_name = 'masters'
    paginate_by = 50

    def get_queryset(self):
        queryset = DivingItemMaster.objects.filter(is_active=True)
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(master_number__icontains=search) |
                Q(manufacturer__icontains=search)
            )
        return queryset.order_by('master_number')


class DivingMasterCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'diving.add_divingitemmaster'
    """Neue Stammdaten anlegen"""
    model = DivingItemMaster
    form_class = DivingItemMasterForm
    template_name = 'diving/master_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Neue Stammdaten anlegen'
        context['submit_text'] = 'Speichern'
        return context

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Stammdaten erfolgreich erstellt.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('diving:master_detail', kwargs={'pk': self.object.pk})


class DivingMasterUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'diving.change_divingitemmaster'
    """Stammdaten bearbeiten"""
    model = DivingItemMaster
    form_class = DivingItemMasterForm
    template_name = 'diving/master_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Stammdaten bearbeiten'
        context['submit_text'] = 'Aktualisieren'
        return context

    def form_valid(self, form):
        messages.success(self.request, _('Stammdaten erfolgreich aktualisiert.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('diving:master_detail', kwargs={'pk': self.object.pk})


class DivingMasterDetailView(LoginRequiredMixin, DetailView):
    """Detailansicht Stammdaten"""
    model = DivingItemMaster
    template_name = 'diving/master_detail.html'
    context_object_name = 'master'


class DivingDeviceListView(LoginRequiredMixin, ListView):
    """Liste aller Tauchgeräte-Instanzen"""
    model = DivingDeviceInstance
    template_name = 'diving/device_list.html'
    context_object_name = 'devices'
    paginate_by = 50

    def get_queryset(self):
        queryset = DivingDeviceInstance.objects.filter(is_active=True).select_related('master', 'location')
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(inventory_number__icontains=search) |
                Q(serial_number__icontains=search) |
                Q(master__name__icontains=search)
            )
        return queryset.order_by('inventory_number')


class DivingDeviceCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'diving.add_divingdeviceinstance'
    """Neue Geräte-Instanz anlegen"""
    model = DivingDeviceInstance
    form_class = DivingDeviceInstanceForm
    template_name = 'diving/device_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Neue Geräte-Instanz anlegen'
        context['submit_text'] = 'Gerät speichern'
        return context

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, _('Gerät erfolgreich erstellt.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('diving:device_detail', kwargs={'pk': self.object.pk})


class DivingDeviceUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'diving.change_divingdeviceinstance'
    """Geräte-Instanz bearbeiten"""
    model = DivingDeviceInstance
    form_class = DivingDeviceInstanceForm
    template_name = 'diving/device_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Geräte-Instanz bearbeiten'
        context['submit_text'] = 'Änderungen speichern'
        return context

    def form_valid(self, form):
        messages.success(self.request, _('Gerät erfolgreich aktualisiert.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('diving:device_detail', kwargs={'pk': self.object.pk})


class DivingDeviceDetailView(LoginRequiredMixin, DetailView):
    """Detailansicht Geräte-Instanz"""
    model = DivingDeviceInstance
    template_name = 'diving/device_detail.html'
    context_object_name = 'device'


# ============================================================================
# DASHBOARD
# ============================================================================

class DivingDashboardView(LoginRequiredMixin, TemplateView):
    """Dashboard für Taucher"""
    template_name = 'diving/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()

        # Master Data & Device Instance KPIs
        context['total_masters'] = DivingItemMaster.objects.filter(is_active=True).count()
        context['total_devices'] = DivingDeviceInstance.objects.filter(is_active=True).count()
        context['operational_devices'] = DivingDeviceInstance.objects.filter(is_active=True, is_operational=True).count()

        # Legacy KPIs
        context['total_items'] = DivingItem.objects.filter(is_active=True).count()

        # Prüfung fällig (innerhalb 30 Tage)
        context['inspection_due_count'] = DivingItem.objects.filter(
            is_active=True,
            next_service_date__isnull=False,
            next_service_date__lte=today + timedelta(days=30)
        ).count()

        # Ausgemustert
        context['retired_count'] = DivingItem.objects.filter(
            Q(inspection_status=InspectionStatus.CONDEMNED) | Q(retirement_date__isnull=False)
        ).count()

        # Prüfungstypen (nicht verfügbar in diving)
        context['inspection_types_count'] = 0

        # Letzte Bewegungen
        context['recent_movements'] = DivingStockMovement.objects.select_related(
            'item',
            'created_by',
            'service_technician'
        ).order_by('-movement_date', '-created_at')[:10]

        # Kritische Prüfungen (überfällig >30 Tage)
        context['overdue_inspections'] = DivingItem.objects.filter(
            is_active=True,
            next_service_date__isnull=False,
            next_service_date__lt=today - timedelta(days=30)
        ).select_related('last_service_technician')[:5]

        # Prüfungen demnächst fällig (nächste 14 Tage)
        context['upcoming_inspections'] = DivingItem.objects.filter(
            is_active=True,
            next_service_date__isnull=False,
            next_service_date__gte=today,
            next_service_date__lte=today + timedelta(days=14)
        ).order_by('next_service_date')[:5]

        # Ausrüstung kurz vor Aussonderung (nicht verfügbar in diving - kein max_service_life_years Feld)
        context['items_near_retirement'] = []

        # Letzte Prüfungen (Service Logs)
        context['recent_inspections'] = DivingServiceLog.objects.select_related(
            'item',
            'performed_by'
        ).order_by('-service_date')[:10]

        # Statistiken nach Typ
        context['stats_by_type'] = DivingItem.objects.filter(
            is_active=True
        ).values('item_type').annotate(
            count=Count('id')
        ).order_by('-count')[:5]

        # ========================================================================
        # PERSONAL: Qualifikationen & Pflichtstunden
        # ========================================================================

        # Personen mit Taucher-Funktion finden
        diving_personnel = Person.objects.filter(
            is_active=True,
            functions__related_module='diving',
            functions__is_active=True
        ).distinct().prefetch_related('qualifications', 'duty_hours')

        # Ablaufende Qualifikationen (nächste 60 Tage oder bereits abgelaufen)
        expiring_qualifications = []
        for person in diving_personnel:
            for qualification in person.qualifications.filter(is_active=True, expiry_date__isnull=False):
                days_until_expiry = (qualification.expiry_date - today).days
                if days_until_expiry <= 60:  # Läuft in 60 Tagen oder weniger ab
                    expiring_qualifications.append({
                        'person': person,
                        'qualification': qualification,
                        'days_until_expiry': days_until_expiry,
                        'is_expired': days_until_expiry < 0,
                    })

        # Nach Dringlichkeit sortieren (abgelaufen zuerst, dann nach Tagen)
        expiring_qualifications.sort(key=lambda x: (not x['is_expired'], x['days_until_expiry']))
        context['expiring_qualifications'] = expiring_qualifications[:10]

        # Fehlende Pflichtstunden für aktuelles Jahr
        current_year = today.year
        missing_duty_hours = []

        for person in diving_personnel:
            # Alle Pflichtstunden-Kategorien finden, die für diese Person relevant sind
            # (über Funktionen, die dem diving-Modul zugeordnet sind)
            required_categories = set()
            for function in person.functions.filter(related_module='diving', is_active=True):
                for category in function.required_duty_hour_categories.all():
                    required_categories.add(category)

            # Für jede Kategorie prüfen, ob Anforderungen erfüllt sind
            for category in required_categories:
                # Anforderung für dieses Jahr finden
                try:
                    requirement = DutyHoursRequirement.objects.get(
                        category=category,
                        year=current_year,
                        is_active=True
                    )

                    # Geleistete Stunden berechnen
                    completed_hours = DutyHoursEntry.objects.filter(
                        person=person,
                        category=category,
                        date__year=current_year
                    ).aggregate(total=Sum('hours'))['total'] or 0

                    # Fehlende Stunden berechnen
                    missing_hours = requirement.required_hours - completed_hours

                    if missing_hours > 0:
                        missing_duty_hours.append({
                            'person': person,
                            'category': category,
                            'required_hours': requirement.required_hours,
                            'completed_hours': completed_hours,
                            'missing_hours': missing_hours,
                        })
                except DutyHoursRequirement.DoesNotExist:
                    # Keine Anforderung für dieses Jahr definiert
                    pass

        # Nach fehlenden Stunden sortieren (meiste fehlende Stunden zuerst)
        missing_duty_hours.sort(key=lambda x: x['missing_hours'], reverse=True)
        context['missing_duty_hours'] = missing_duty_hours[:10]

        return context


# ============================================================================
# ITEM VIEWS
# ============================================================================

class DivingItemListView(LoginRequiredMixin, ListView):
    """Liste aller Tauchausrüstung"""
    model = DivingItem
    template_name = 'diving/items.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        queryset = DivingItem.objects.filter(is_active=True).select_related(
            'category',
            'location',
            'last_service_technician'
        )

        # Suchfilter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(serial_number__icontains=search) |
                Q(manufacturer__icontains=search) |
                Q(item_number__icontains=search) |
                Q(description__icontains=search)
            )

        # Typ-Filter
        item_type = self.request.GET.get('type')
        if item_type:
            queryset = queryset.filter(item_type=item_type)

        # Kategorie-Filter
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category_id=category)

        # Lagerort-Filter
        location = self.request.GET.get('location')
        if location:
            queryset = queryset.filter(location_id=location)

        # Prüfstatus-Filter
        inspection_status = self.request.GET.get('inspection_status')
        if inspection_status:
            queryset = queryset.filter(inspection_status=inspection_status)

        # TÜV-Prüfung fällig
        if self.request.GET.get('tuv_due'):
            today = timezone.now().date()
            queryset = queryset.filter(
                next_tuv_inspection__isnull=False,
                next_tuv_inspection__lte=today
            )

        # Wartung fällig
        if self.request.GET.get('service_due'):
            today = timezone.now().date()
            queryset = queryset.filter(
                next_service_date__isnull=False,
                next_service_date__lte=today
            )

        # Gastyp-Filter
        gas_type = self.request.GET.get('gas_type')
        if gas_type:
            queryset = queryset.filter(current_gas_type=gas_type)

        # Sortierung
        sort = self.request.GET.get('sort', 'item_type')
        if sort:
            queryset = queryset.order_by(sort, 'name')
        else:
            queryset = queryset.order_by('item_type', 'name')

        return queryset

    def get_context_data(self, **kwargs):
        from inventory_base.models import Category
        from locations.models import Location
        from .models import GasType

        context = super().get_context_data(**kwargs)
        context['categories'] = Category.objects.filter(is_active=True).order_by('name')
        context['locations'] = Location.objects.filter(is_active=True).order_by('name')
        context['item_types'] = DivingItemType.choices
        context['inspection_statuses'] = InspectionStatus.choices
        context['gas_types'] = GasType.choices
        return context


class DivingItemCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'diving.add_divingitem'
    """Neue Tauchausrüstung anlegen"""
    model = DivingItem
    template_name = 'diving/item_form.html'
    fields = [
        'name',
        'item_type',
        'item_number',
        'manufacturer',
        'manufacturer_part_number',
        'manufacturing_date',
        'description',
        'category',
        'unit',
        'min_quantity',
        'serial_number',
        'tank_type',
        'volume_liters',
        'working_pressure_bar',
        'test_pressure_bar',
        'service_interval_months',
        'max_depth_m',
        'weight_kg',
        'size',
        'material',
        'service_manual',
        'notes',
    ]

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Tailwind CSS Styling
        for field_name, field in form.fields.items():
            if isinstance(field.widget, forms.CheckboxInput):
                field.widget.attrs.update({'class': 'rounded border-gray-300 text-blue-600 focus:ring-blue-500'})
            elif isinstance(field.widget, forms.Textarea):
                field.widget.attrs.update({
                    'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500',
                    'rows': 4
                })
            elif isinstance(field.widget, forms.FileInput):
                field.widget.attrs.update({
                    'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500',
                    'accept': '.pdf,.jpg,.jpeg,.png,.tiff,.bmp'
                })
            else:
                field.widget.attrs.update({
                    'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500'
                })

        # Labels und Help-Texte
        if 'service_manual' in form.fields:
            form.fields['service_manual'].label = 'Serviceanleitung / Handbuch'
            form.fields['service_manual'].help_text = 'PDF oder Bild-Datei (wird automatisch per OCR durchsuchbar gemacht)'

        if 'manufacturing_date' in form.fields:
            form.fields['manufacturing_date'].widget.attrs.update({'type': 'date'})

        return form

    def form_valid(self, form):
        from locations.models import Location
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        form.instance.inspection_status = InspectionStatus.NOT_DUE
        # Setze Standardlagerort (erster aktiver Lagerort) wenn nicht gesetzt
        if not hasattr(form.instance, 'location') or form.instance.location is None:
            default_location = Location.objects.filter(is_active=True).first()
            if default_location:
                form.instance.location = default_location
        # Bestand initial auf 0 setzen (wird über Wareneingang erhöht)
        form.instance.quantity = 0
        messages.success(self.request, _('Tauchausrüstung erfolgreich erstellt.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('diving:item_detail', kwargs={'pk': self.object.pk})


class DivingItemUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'diving.change_divingitem'
    """Tauchausrüstung bearbeiten"""
    model = DivingItem
    template_name = 'diving/item_form.html'
    fields = [
        'name',
        'item_type',
        'item_number',
        'manufacturer',
        'manufacturer_part_number',
        'manufacturing_date',
        'description',
        'category',
        'unit',
        'min_quantity',
        'serial_number',
        'tank_type',
        'volume_liters',
        'working_pressure_bar',
        'test_pressure_bar',
        'service_interval_months',
        'max_depth_m',
        'weight_kg',
        'size',
        'material',
        'service_manual',
        'notes',
    ]

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Tailwind CSS Styling
        for field_name, field in form.fields.items():
            if isinstance(field.widget, forms.CheckboxInput):
                field.widget.attrs.update({'class': 'rounded border-gray-300 text-blue-600 focus:ring-blue-500'})
            elif isinstance(field.widget, forms.Textarea):
                field.widget.attrs.update({
                    'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500',
                    'rows': 4
                })
            elif isinstance(field.widget, forms.FileInput):
                field.widget.attrs.update({
                    'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500',
                    'accept': '.pdf,.jpg,.jpeg,.png,.tiff,.bmp'
                })
            else:
                field.widget.attrs.update({
                    'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500'
                })

        # Labels und Help-Texte
        if 'service_manual' in form.fields:
            form.fields['service_manual'].label = 'Serviceanleitung / Handbuch'
            form.fields['service_manual'].help_text = 'PDF oder Bild-Datei (wird automatisch per OCR durchsuchbar gemacht)'

        if 'manufacturing_date' in form.fields:
            form.fields['manufacturing_date'].widget.attrs.update({'type': 'date'})

        return form

    def form_valid(self, form):
        messages.success(self.request, _('Tauchausrüstung erfolgreich aktualisiert.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('diving:item_detail', kwargs={'pk': self.object.pk})


# ============================================================================
# LAGERBEWEGUNGEN (STOCK MOVEMENTS)
# ============================================================================

class DivingMovementListView(LoginRequiredMixin, ListView):
    """Liste aller Lagerbewegungen"""
    model = DivingStockMovement
    template_name = 'diving/movement_list.html'
    context_object_name = 'movements'
    paginate_by = 50

    def get_queryset(self):
        queryset = DivingStockMovement.objects.select_related(
            'item',
            'created_by',
            'service_technician'
        )

        # Typ-Filter
        movement_type = self.request.GET.get('type')
        if movement_type:
            queryset = queryset.filter(movement_type=movement_type)

        # Item-Filter
        item_id = self.request.GET.get('item')
        if item_id:
            queryset = queryset.filter(item_id=item_id)

        # Gasfüllung-Filter
        gas_filled = self.request.GET.get('gas_filled')
        if gas_filled == 'yes':
            queryset = queryset.filter(gas_filled=True)
        elif gas_filled == 'no':
            queryset = queryset.filter(gas_filled=False)

        # Service-Filter
        service_performed = self.request.GET.get('service')
        if service_performed == 'yes':
            queryset = queryset.filter(service_performed=True)

        # Tauchgang-Filter
        dive_date = self.request.GET.get('dive')
        if dive_date == 'yes':
            queryset = queryset.filter(dive_date__isnull=False)

        return queryset.order_by('-movement_date', '-created_at')


class DivingMovementDetailView(LoginRequiredMixin, DetailView):
    """Detailansicht einer Lagerbewegung"""
    model = DivingStockMovement
    template_name = 'diving/movement_detail.html'
    context_object_name = 'movement'


class DivingMovementCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'diving.add_divingstockmovement'
    """Neue Lagerbewegung erstellen"""
    model = DivingStockMovement
    form_class = DivingStockMovementForm
    template_name = 'diving/movement_form.html'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Lagerbewegung wurde erfolgreich erstellt.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('diving:movement_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        import json
        context = super().get_context_data(**kwargs)
        context['title'] = 'Neue Lagerbewegung'
        context['submit_text'] = 'Bewegung speichern'
        item_locations = dict(
            DivingItem.objects.filter(location__isnull=False)
            .values_list('pk', 'location_id')
        )
        context['item_location_map'] = json.dumps(item_locations)
        return context


# ============================================================================
# WARTUNGSVERWALTUNG (SERVICE MANAGEMENT)
# ============================================================================

class DivingServiceListView(LoginRequiredMixin, ListView):
    """Liste aller Service-Protokolle"""
    model = DivingServiceLog
    template_name = 'diving/service_list.html'
    context_object_name = 'services'
    paginate_by = 50

    def get_queryset(self):
        return DivingServiceLog.objects.select_related(
            'device',
            'technician'
        ).order_by('-service_date')


class DivingServiceLogCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'diving.add_divingservicelog'
    """Service-Protokoll erstellen"""
    model = DivingServiceLog
    form_class = DivingServiceLogForm
    template_name = 'diving/service_log_form.html'
    success_url = reverse_lazy('diving:inspection_list')

    def form_valid(self, form):
        # Device-Status aktualisieren, wenn Service durchgeführt wurde
        device = form.instance.device
        if form.instance.passed:
            device.last_service_date = form.instance.service_date
            device.next_service_date = form.instance.next_service_due
            device.last_service_technician = form.instance.technician
            device.save()

        messages.success(
            self.request,
            f'Service-Protokoll für {device.inventory_number} erfolgreich erstellt.'
        )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        import json
        context = super().get_context_data(**kwargs)
        context['title'] = 'Service dokumentieren'
        context['submit_text'] = 'Protokoll speichern'

        # Service Types mit Checklisten als JSON für JavaScript
        service_types_data = {}
        for st in DivingServiceType.objects.filter(is_active=True):
            service_types_data[str(st.pk)] = {
                'name': st.name,
                'checklist': st.checklist or [],
                'custom_fields': st.custom_fields or [],
            }
        context['service_types_json'] = json.dumps(service_types_data)

        return context


class DivingServiceManagementView(LoginRequiredMixin, TemplateView):
    """Zentrale Wartungsverwaltung"""
    template_name = 'diving/service_management.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()

        # Service-Statistiken
        context['total_services'] = DivingServiceLog.objects.count()
        context['services_this_month'] = DivingServiceLog.objects.filter(
            service_date__year=today.year,
            service_date__month=today.month
        ).count()

        # Überfällige Wartungen
        context['overdue_services'] = DivingDeviceInstance.objects.filter(
            is_operational=True,
            next_service_date__isnull=False,
            next_service_date__lt=today
        ).count()

        # Demnächst fällige Wartungen (nächste 30 Tage)
        context['upcoming_services'] = DivingDeviceInstance.objects.filter(
            is_operational=True,
            next_service_date__isnull=False,
            next_service_date__gte=today,
            next_service_date__lte=today + timedelta(days=30)
        ).count()

        # Letzte Service-Logs
        context['recent_services'] = DivingServiceLog.objects.select_related(
            'device',
            'technician'
        ).order_by('-service_date')[:10]

        # Devices mit überfälliger Wartung
        context['items_overdue'] = DivingDeviceInstance.objects.filter(
            is_operational=True,
            next_service_date__isnull=False,
            next_service_date__lt=today
        ).select_related('master', 'last_service_technician')[:10]

        # Devices mit demnächst fälliger Wartung
        context['items_upcoming'] = DivingDeviceInstance.objects.filter(
            is_operational=True,
            next_service_date__isnull=False,
            next_service_date__gte=today,
            next_service_date__lte=today + timedelta(days=30)
        ).select_related('master', 'last_service_technician').order_by('next_service_date')[:10]

        # Wartungstypen (nur wenn das neue System verwendet wird)
        context['service_types'] = DivingServiceType.objects.filter(
            is_active=True
        ).annotate(
            assignment_count=Count('item_assignments')
        ).order_by('name')

        context['total_types'] = context['service_types'].count()
        context['total_assignments'] = DivingServiceAssignment.objects.filter(
            is_active=True
        ).count()
        context['total_records'] = DivingServiceRecord.objects.count()

        return context


class InspectionDueListView(LoginRequiredMixin, ListView):
    """Liste aller Geräte mit fälligen oder überfälligen Prüfungen"""
    model = DivingDeviceInstance
    template_name = 'diving/inspection_due_list.html'
    context_object_name = 'devices'
    paginate_by = 50

    def get_queryset(self):
        today = timezone.now().date()
        # Alle Geräte mit fälliger oder überfälliger Wartung (innerhalb 60 Tage oder überfällig)
        return DivingDeviceInstance.objects.filter(
            is_active=True,
            next_service_date__isnull=False,
            next_service_date__lte=today + timedelta(days=60)
        ).select_related('master', 'last_service_technician').order_by('next_service_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()

        # Überfällige Prüfungen
        context['overdue_count'] = DivingDeviceInstance.objects.filter(
            is_active=True,
            next_service_date__isnull=False,
            next_service_date__lt=today
        ).count()

        # Fällig in den nächsten 30 Tagen
        context['due_soon_count'] = DivingDeviceInstance.objects.filter(
            is_active=True,
            next_service_date__isnull=False,
            next_service_date__gte=today,
            next_service_date__lte=today + timedelta(days=30)
        ).count()

        # Fällig in 31-60 Tagen
        context['upcoming_count'] = DivingDeviceInstance.objects.filter(
            is_active=True,
            next_service_date__isnull=False,
            next_service_date__gt=today + timedelta(days=30),
            next_service_date__lte=today + timedelta(days=60)
        ).count()

        context['today'] = today
        return context


class ServiceRecordCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'diving.add_servicerecord'
    """
    Erstellt ein Wartungsprotokoll mit dynamischer Checkliste.
    Basiert auf einer ServiceAssignment (Verknüpfung Item <-> ServiceType).
    """
    model = DivingServiceRecord
    form_class = DynamicServiceRecordForm
    template_name = 'diving/service_record_form.html'
    success_url = reverse_lazy('diving:service_management')

    def get_assignment(self):
        """Lädt die ServiceAssignment basierend auf URL-Parameter."""
        assignment_pk = self.kwargs.get('assignment_pk')
        if assignment_pk:
            return DivingServiceAssignment.objects.select_related(
                'item', 'service_type'
            ).get(pk=assignment_pk)
        return None

    def get_form_kwargs(self):
        """Übergibt die Assignment an das Form."""
        kwargs = super().get_form_kwargs()
        kwargs['assignment'] = self.get_assignment()
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        assignment = self.get_assignment()

        if assignment:
            context['assignment'] = assignment
            context['device'] = assignment.device
            context['service_type'] = assignment.service_type
            context['title'] = f'Wartung: {assignment.service_type.name}'
            context['checklist_items'] = self.get_form().checklist_items
            context['custom_fields'] = self.get_form().custom_fields_config

        context['submit_text'] = 'Wartungsprotokoll speichern'
        return context

    def form_valid(self, form):
        """Speichert das Protokoll und aktualisiert die Assignment."""
        assignment = self.get_assignment()
        if not assignment:
            messages.error(self.request, 'Keine gültige Wartungszuweisung gefunden.')
            return self.form_invalid(form)

        response = super().form_valid(form)

        messages.success(
            self.request,
            f'Wartungsprotokoll für {assignment.device.inventory_number} ({assignment.service_type.name}) '
            f'erfolgreich erstellt.'
        )
        return response

    def get_success_url(self):
        """Redirect zur Service-Management Seite."""
        return reverse_lazy('diving:service_management')


class AssignServiceTypeView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'diving.add_serviceassignment'
    """
    Weist einem Gerät einen ServiceType zu (erstellt eine ServiceAssignment).
    Ermöglicht das Verknüpfen von Geräten mit Wartungsarten.
    """
    model = DivingServiceAssignment
    form_class = DivingServiceAssignmentForm
    template_name = 'diving/service_assignment_form.html'

    def get_initial(self):
        """Setzt das Gerät aus der URL."""
        initial = super().get_initial()
        device_pk = self.kwargs.get('device_pk')
        if device_pk:
            initial['device'] = device_pk
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        device_pk = self.kwargs.get('device_pk')
        if device_pk:
            device = DivingDeviceInstance.objects.get(pk=device_pk)
            context['device'] = device
            context['title'] = f'Wartungsart zuweisen: {device.inventory_number}'
            # Bereits zugewiesene ServiceTypes
            context['existing_assignments'] = device.service_assignments.select_related('service_type').all()
        else:
            context['title'] = 'Wartungsart zuweisen'
        return context

    def form_valid(self, form):
        """Setzt created_by und speichert."""
        form.instance.created_by = self.request.user
        response = super().form_valid(form)

        messages.success(
            self.request,
            f'Wartungsart "{form.instance.service_type.name}" wurde '
            f'{form.instance.device.inventory_number} zugewiesen.'
        )
        return response

    def get_success_url(self):
        """Redirect zurück zum Gerät oder Service Management."""
        device_pk = self.kwargs.get('device_pk')
        if device_pk:
            return reverse_lazy('diving:device_detail', kwargs={'pk': device_pk})
        return reverse_lazy('diving:service_management')


class DeviceServiceAssignmentsView(LoginRequiredMixin, TemplateView):
    """
    Zeigt alle Wartungszuweisungen für ein Gerät an.
    Ermöglicht das Starten einer Wartung (ServiceRecord erstellen).
    """
    template_name = 'diving/device_service_assignments.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        device_pk = self.kwargs.get('pk')
        device = DivingDeviceInstance.objects.get(pk=device_pk)

        context['device'] = device
        context['assignments'] = device.service_assignments.select_related(
            'service_type'
        ).prefetch_related('service_records').order_by('service_type__name')

        # Verfügbare ServiceTypes die noch nicht zugewiesen sind
        assigned_type_ids = device.service_assignments.values_list('service_type_id', flat=True)
        context['available_service_types'] = DivingServiceType.objects.filter(
            is_active=True
        ).exclude(id__in=assigned_type_ids)

        context['title'] = f'Wartungen für {device.inventory_number}'
        return context


# ============================================================================
# SERVICE TYPE MANAGEMENT (WARTUNGSTYP-VERWALTUNG)
# ============================================================================

class ServiceTypeListView(LoginRequiredMixin, ListView):
    """Liste aller Wartungstypen"""
    model = DivingServiceType
    template_name = 'diving/service_type_list.html'
    context_object_name = 'service_types'

    def get_queryset(self):
        return DivingServiceType.objects.filter(is_active=True).order_by('name')


class ServiceTypeCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'diving.add_servicetype'
    """Wartungstyp erstellen"""
    model = DivingServiceType
    form_class = DivingServiceTypeForm
    template_name = 'diving/service_type_form.html'
    success_url = reverse_lazy('diving:service_type_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Neuen Wartungstyp anlegen'
        context['submit_text'] = 'Wartungstyp speichern'
        return context

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, _('Wartungstyp erfolgreich erstellt.'))
        return super().form_valid(form)


class ServiceTypeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'diving.change_servicetype'
    """Wartungstyp bearbeiten"""
    model = DivingServiceType
    form_class = DivingServiceTypeForm
    template_name = 'diving/service_type_form.html'
    success_url = reverse_lazy('diving:service_type_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Wartungstyp bearbeiten'
        context['submit_text'] = 'Änderungen speichern'
        return context

    def form_valid(self, form):
        messages.success(self.request, _('Wartungstyp erfolgreich aktualisiert.'))
        return super().form_valid(form)


class ServiceTypeDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    permission_required = 'diving.delete_servicetype'
    """Wartungstyp löschen"""
    model = DivingServiceType
    template_name = 'diving/service_type_confirm_delete.html'
    success_url = reverse_lazy('diving:service_type_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, _('Wartungstyp erfolgreich gelöscht.'))
        return super().delete(request, *args, **kwargs)


# ============================================================================
# TYPE MANAGEMENT VIEWS (TYPENVERWALTUNG)
# ============================================================================

class TypeManagementView(LoginRequiredMixin, TemplateView):
    """Übersicht über alle verwaltbaren Typen"""
    template_name = 'diving/type_management.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['equipment_types'] = EquipmentType.objects.all().order_by('sort_order', 'name')
        context['bottle_types'] = BottleType.objects.all().order_by('sort_order', 'name')
        context['equipment_types_count'] = context['equipment_types'].count()
        context['bottle_types_count'] = context['bottle_types'].count()
        return context


# Equipment Type Views
class EquipmentTypeListView(LoginRequiredMixin, ListView):
    """Liste aller Ausrüstungstypen"""
    model = EquipmentType
    template_name = 'diving/equipment_type_list.html'
    context_object_name = 'equipment_types'

    def get_queryset(self):
        return EquipmentType.objects.all().order_by('sort_order', 'name')


class EquipmentTypeCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'diving.add_equipmenttype'
    """Ausrüstungstyp erstellen"""
    model = EquipmentType
    form_class = EquipmentTypeForm
    template_name = 'diving/equipment_type_form.html'
    success_url = reverse_lazy('diving:type_management')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, _('Ausrüstungstyp erfolgreich erstellt.'))
        return super().form_valid(form)


class EquipmentTypeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'diving.change_equipmenttype'
    """Ausrüstungstyp bearbeiten"""
    model = EquipmentType
    form_class = EquipmentTypeForm
    template_name = 'diving/equipment_type_form.html'
    success_url = reverse_lazy('diving:type_management')

    def form_valid(self, form):
        messages.success(self.request, _('Ausrüstungstyp erfolgreich aktualisiert.'))
        return super().form_valid(form)


class EquipmentTypeDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    permission_required = 'diving.delete_equipmenttype'
    """Ausrüstungstyp löschen"""
    model = EquipmentType
    template_name = 'diving/equipment_type_confirm_delete.html'
    success_url = reverse_lazy('diving:type_management')

    def delete(self, request, *args, **kwargs):
        messages.success(request, _('Ausrüstungstyp erfolgreich gelöscht.'))
        return super().delete(request, *args, **kwargs)


# Bottle Type Views
class BottleTypeListView(LoginRequiredMixin, ListView):
    """Liste aller Flaschentypen"""
    model = BottleType
    template_name = 'diving/bottle_type_list.html'
    context_object_name = 'bottle_types'

    def get_queryset(self):
        return BottleType.objects.all().order_by('sort_order', 'name')


class BottleTypeCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'diving.add_bottletype'
    """Flaschentyp erstellen"""
    model = BottleType
    form_class = BottleTypeForm
    template_name = 'diving/bottle_type_form.html'
    success_url = reverse_lazy('diving:type_management')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, _('Flaschentyp erfolgreich erstellt.'))
        return super().form_valid(form)


class BottleTypeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'diving.change_bottletype'
    """Flaschentyp bearbeiten"""
    model = BottleType
    form_class = BottleTypeForm
    template_name = 'diving/bottle_type_form.html'
    success_url = reverse_lazy('diving:type_management')

    def form_valid(self, form):
        messages.success(self.request, _('Flaschentyp erfolgreich aktualisiert.'))
        return super().form_valid(form)


class BottleTypeDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    permission_required = 'diving.delete_bottletype'
    """Flaschentyp löschen"""
    model = BottleType
    template_name = 'diving/bottle_type_confirm_delete.html'
    success_url = reverse_lazy('diving:type_management')

    def delete(self, request, *args, **kwargs):
        messages.success(request, _('Flaschentyp erfolgreich gelöscht.'))
        return super().delete(request, *args, **kwargs)


# ============================================================================
# QR-CODE AND BARCODE VIEWS
# ============================================================================

class DivingMasterQRCodeView(LoginRequiredMixin, DetailView):
    """QR-Code für Tauchausrüstungs-Stammdaten"""
    model = DivingItemMaster

    def get(self, request, *args, **kwargs):
        master = self.get_object()
        import qrcode
        import qrcode.image.svg
        from io import BytesIO
        from django.http import HttpResponse

        # Nur URL kodieren für maximale Scanner-Kompatibilität
        qr_data = f"/diving/masters/{master.pk}/"

        # QR-Code generieren (SVG)
        factory = qrcode.image.svg.SvgPathImage
        img = qrcode.make(qr_data, image_factory=factory, box_size=10)

        # Als SVG zurückgeben
        stream = BytesIO()
        img.save(stream)
        response = HttpResponse(stream.getvalue(), content_type='image/svg+xml')
        response['Content-Disposition'] = f'inline; filename="qr_{master.master_number}.svg"'
        return response


class DivingMasterBarcodeView(LoginRequiredMixin, DetailView):
    """Barcode für Tauchausrüstungs-Stammdaten"""
    model = DivingItemMaster

    def get(self, request, *args, **kwargs):
        master = self.get_object()
        import barcode
        from barcode.writer import SVGWriter
        from io import BytesIO
        from django.http import HttpResponse

        # Barcode generieren (Code128)
        code128 = barcode.get_barcode_class('code128')
        barcode_instance = code128(master.master_number, writer=SVGWriter())

        # Als SVG zurückgeben
        stream = BytesIO()
        barcode_instance.write(stream)
        response = HttpResponse(stream.getvalue(), content_type='image/svg+xml')
        response['Content-Disposition'] = f'inline; filename="barcode_{master.master_number}.svg"'
        return response


class DivingDeviceQRCodeView(LoginRequiredMixin, DetailView):
    """QR-Code für Tauchgerät"""
    model = DivingDeviceInstance

    def get(self, request, *args, **kwargs):
        device = self.get_object()
        import qrcode
        import qrcode.image.svg
        from io import BytesIO
        from django.http import HttpResponse

        # Nur URL kodieren für maximale Scanner-Kompatibilität
        qr_data = f"/diving/device/{device.pk}/"

        # QR-Code generieren (SVG)
        factory = qrcode.image.svg.SvgPathImage
        img = qrcode.make(qr_data, image_factory=factory, box_size=10)

        # Als SVG zurückgeben
        stream = BytesIO()
        img.save(stream)
        response = HttpResponse(stream.getvalue(), content_type='image/svg+xml')
        response['Content-Disposition'] = f'inline; filename="qr_{device.inventory_number}.svg"'
        return response


class DivingDeviceBarcodeView(LoginRequiredMixin, DetailView):
    """Barcode für Tauchgerät"""
    model = DivingDeviceInstance

    def get(self, request, *args, **kwargs):
        device = self.get_object()
        import barcode
        from barcode.writer import SVGWriter
        from io import BytesIO
        from django.http import HttpResponse

        # Barcode generieren (Code128)
        code128 = barcode.get_barcode_class('code128')
        barcode_instance = code128(device.inventory_number, writer=SVGWriter())

        # Als SVG zurückgeben
        stream = BytesIO()
        barcode_instance.write(stream)
        response = HttpResponse(stream.getvalue(), content_type='image/svg+xml')
        response['Content-Disposition'] = f'inline; filename="barcode_{device.inventory_number}.svg"'
        return response


# ============================================================================
# BATCH MANAGEMENT VIEWS (CHARGENVERWALTUNG)
# ============================================================================

class DivingBatchListView(LoginRequiredMixin, ListView):
    """Liste aller Chargen für einen Master"""
    model = DivingBatch
    template_name = 'diving/batch_list.html'
    context_object_name = 'batches'
    paginate_by = 20

    def get_queryset(self):
        master_pk = self.kwargs.get('master_pk')
        return DivingBatch.objects.filter(master_id=master_pk, is_active=True).order_by('-received_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        master_pk = self.kwargs.get('master_pk')
        context['master'] = DivingItemMaster.objects.get(pk=master_pk)
        return context


class DivingBatchCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'diving.add_divingbatch'
    """Neue Charge anlegen"""
    model = DivingBatch
    form_class = DivingBatchForm
    template_name = 'diving/batch_form.html'

    def get_initial(self):
        initial = super().get_initial()
        master_pk = self.kwargs.get('master_pk')
        if master_pk:
            initial['master'] = master_pk
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Neue Charge anlegen'
        context['submit_text'] = 'Charge speichern'
        master_pk = self.kwargs.get('master_pk')
        if master_pk:
            context['master'] = DivingItemMaster.objects.get(pk=master_pk)
        return context

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, f'Charge {form.instance.batch_number} wurde erfolgreich angelegt.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('diving:master_detail', kwargs={'pk': self.object.master.pk})


class DivingBatchUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'diving.change_divingbatch'
    """Charge bearbeiten"""
    model = DivingBatch
    form_class = DivingBatchForm
    template_name = 'diving/batch_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Charge bearbeiten'
        context['submit_text'] = 'Änderungen speichern'
        context['master'] = self.object.master
        return context

    def form_valid(self, form):
        messages.success(self.request, f'Charge {form.instance.batch_number} wurde aktualisiert.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('diving:master_detail', kwargs={'pk': self.object.master.pk})


class DivingBatchDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    permission_required = 'diving.delete_divingbatch'
    """Charge löschen"""
    model = DivingBatch
    template_name = 'diving/batch_confirm_delete.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['master'] = self.object.master
        return context

    def delete(self, request, *args, **kwargs):
        batch = self.get_object()
        messages.success(request, f'Charge {batch.batch_number} wurde gelöscht.')
        return super().delete(request, *args, **kwargs)

    def get_success_url(self):
        return reverse_lazy('diving:master_detail', kwargs={'pk': self.object.master.pk})


# ============================================================================
# IMPORT/EXPORT
# ============================================================================

class ImportExportView(LoginRequiredMixin, TemplateView):
    """Import/Export Übersicht"""
    template_name = 'diving/import_export.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_masters'] = DivingItemMaster.objects.filter(is_active=True).count()
        context['total_devices'] = DivingDeviceInstance.objects.filter(is_active=True).count()
        context['total_movements'] = DivingStockMovement.objects.count()
        context['total_services'] = DivingServiceLog.objects.count()
        return context


# ============================================================================
# EXPORT FUNCTIONS (Excel)
# ============================================================================

@login_required
def export_masters(request):
    """Export Diving Masters als Excel"""
    masters = DivingItemMaster.objects.filter(is_active=True).select_related(
        'created_by'
    ).order_by('master_number')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stammdaten"

    # Header
    headers = [
        'Stammdaten-Nr.', 'Name', 'Typ', 'Hersteller', 'Modell',
        'Flaschentyp', 'Volumen (L)', 'Nenndruck (bar)', 'Gewicht (kg)',
        'Zertifizierungsnummer', 'Beschreibung', 'Erstellt am', 'Aktualisiert am'
    ]

    # Header-Styling
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Daten
    for row_num, master in enumerate(masters, start=2):
        ws.cell(row=row_num, column=1, value=master.master_number)
        ws.cell(row=row_num, column=2, value=master.name)
        ws.cell(row=row_num, column=3, value=master.get_item_type_display() if hasattr(master, 'get_item_type_display') else master.item_type)
        ws.cell(row=row_num, column=4, value=master.manufacturer)
        ws.cell(row=row_num, column=5, value=master.model)
        ws.cell(row=row_num, column=6, value=master.get_tank_type_display() if master.tank_type and hasattr(master, 'get_tank_type_display') else master.tank_type if master.tank_type else '')
        ws.cell(row=row_num, column=7, value=float(master.volume_liters) if master.volume_liters else '')
        ws.cell(row=row_num, column=8, value=master.working_pressure_bar if master.working_pressure_bar else '')
        ws.cell(row=row_num, column=9, value=float(master.weight_kg) if hasattr(master, 'weight_kg') and master.weight_kg else '')
        ws.cell(row=row_num, column=10, value=master.certification_number)
        ws.cell(row=row_num, column=11, value=master.description)
        ws.cell(row=row_num, column=12, value=master.created_at.strftime('%d.%m.%Y %H:%M') if master.created_at else '')
        ws.cell(row=row_num, column=13, value=master.updated_at.strftime('%d.%m.%Y %H:%M') if master.updated_at else '')

    # Spaltenbreiten
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=taucher_stammdaten_export.xlsx'
    wb.save(response)
    return response


@login_required
def export_devices(request):
    """Export Diving Devices als Excel"""
    devices = DivingDeviceInstance.objects.filter(is_active=True).select_related(
        'master', 'location', 'assigned_vehicle', 'assigned_to', 'last_service_technician', 'created_by'
    ).order_by('inventory_number')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Geräte-Instanzen"

    # Header
    headers = [
        'Inventarnummer', 'Seriennummer', 'Stammdaten', 'Lagerort', 'Fahrzeug', 'Zugeordnet an',
        'Zustand', 'Einsatzbereit', 'Herstellungsdatum', 'Letzte Wartung', 'Nächste Wartung',
        'Letzte Prüfer', 'Notizen', 'Erstellt am', 'Aktualisiert am'
    ]

    # Header-Styling
    header_fill = PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Daten
    for row_num, device in enumerate(devices, start=2):
        ws.cell(row=row_num, column=1, value=device.inventory_number)
        ws.cell(row=row_num, column=2, value=device.serial_number)
        ws.cell(row=row_num, column=3, value=device.master.name if device.master else '')
        ws.cell(row=row_num, column=4, value=device.location.name if device.location else '')
        ws.cell(row=row_num, column=5, value=str(device.assigned_vehicle) if device.assigned_vehicle else '')
        ws.cell(row=row_num, column=6, value=device.assigned_to.get_full_name() if device.assigned_to else '')
        ws.cell(row=row_num, column=7, value=device.get_condition_display() if device.condition and hasattr(device, 'get_condition_display') else device.condition if device.condition else '')
        ws.cell(row=row_num, column=8, value='Ja' if device.is_operational else 'Nein')
        ws.cell(row=row_num, column=9, value=device.manufacturing_date.strftime('%d.%m.%Y') if device.manufacturing_date else '')
        ws.cell(row=row_num, column=10, value=device.last_service_date.strftime('%d.%m.%Y') if hasattr(device, 'last_service_date') and device.last_service_date else '')
        ws.cell(row=row_num, column=11, value=device.next_service_date.strftime('%d.%m.%Y') if hasattr(device, 'next_service_date') and device.next_service_date else '')
        ws.cell(row=row_num, column=12, value=device.last_service_technician.get_full_name() if hasattr(device, 'last_service_technician') and device.last_service_technician else '')
        ws.cell(row=row_num, column=13, value=device.condition_notes)
        ws.cell(row=row_num, column=14, value=device.created_at.strftime('%d.%m.%Y %H:%M') if device.created_at else '')
        ws.cell(row=row_num, column=15, value=device.updated_at.strftime('%d.%m.%Y %H:%M') if device.updated_at else '')

    # Spaltenbreiten
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=taucher_geraete_export.xlsx'
    wb.save(response)
    return response


@login_required
def export_services(request):
    """Export Diving Service Logs als Excel"""
    services = DivingServiceLog.objects.select_related(
        'device', 'device__master', 'technician'
    ).order_by('-service_date')[:1000]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wartungsprotokolle"

    # Header
    headers = [
        'ID', 'Wartungsdatum', 'Gerät', 'Techniker', 'Bestanden', 'Nächste Wartung',
        'Feststellungen', 'Durchgeführte Arbeiten', 'Ersetzte Teile', 'Erstellt am'
    ]

    # Header-Styling
    header_fill = PatternFill(start_color='9C27B0', end_color='9C27B0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Daten
    for row_num, service in enumerate(services, start=2):
        ws.cell(row=row_num, column=1, value=service.id)
        ws.cell(row=row_num, column=2, value=service.service_date.strftime('%d.%m.%Y') if service.service_date else '')
        ws.cell(row=row_num, column=3, value=str(service.device) if service.device else '')
        ws.cell(row=row_num, column=4, value=service.technician.get_full_name() if service.technician else '')
        ws.cell(row=row_num, column=5, value='Ja' if service.passed else 'Nein')
        ws.cell(row=row_num, column=6, value=service.next_service_due.strftime('%d.%m.%Y') if service.next_service_due else '')
        ws.cell(row=row_num, column=7, value=service.findings)
        ws.cell(row=row_num, column=8, value=service.work_performed)
        ws.cell(row=row_num, column=9, value=service.parts_replaced)
        ws.cell(row=row_num, column=10, value=service.created_at.strftime('%d.%m.%Y %H:%M') if service.created_at else '')

    # Spaltenbreiten
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=taucher_wartungen_export.xlsx'
    wb.save(response)
    return response


# ============================================================================
# TEMPLATE DOWNLOAD FUNCTIONS (CSV)
# ============================================================================

@login_required
def template_masters(request):
    """CSV-Vorlage für Stammdaten-Import"""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)

    # Header
    headers = [
        'Stammdaten-Nr.*', 'Name*', 'Typ*', 'Hersteller', 'Modell',
        'Flaschentyp', 'Volumen (L)', 'Nenndruck (bar)',
        'Zertifizierungsnummer', 'Beschreibung'
    ]
    writer.writerow(headers)

    # Beispielzeile
    example_data = [
        'DIV-2025-001', 'Atemregler Scubapro MK25 EVO', 'regulator', 'Scubapro', 'MK25 EVO',
        '', '', '', 'CE-12345', 'Beispiel: Hochleistungsatemregler - diese Zeile löschen!'
    ]
    writer.writerow(example_data)

    # Hinweise
    writer.writerow([])
    writer.writerow(['# HINWEISE:'])
    writer.writerow(['# Felder mit * sind Pflichtfelder'])
    writer.writerow(['# Typ: regulator, bcd, wetsuit, drysuit, fins, mask, tank, computer, light, other'])
    writer.writerow(['# Flaschentyp: steel, aluminum, composite'])

    # Response
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename=vorlage_taucher_stammdaten_import.csv'
    response.write('\ufeff')  # UTF-8 BOM
    response.write(output.getvalue())
    return response


@login_required
def template_devices(request):
    """CSV-Vorlage für Geräte-Import"""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)

    # Header
    headers = [
        'Inventarnummer*', 'Seriennummer*', 'Stammdaten-Nr.*', 'Lagerort-ID*',
        'Herstellungsdatum', 'Zustand', 'Einsatzbereit (Ja/Nein)', 'Notizen'
    ]
    writer.writerow(headers)

    # Beispielzeile
    example_data = [
        'INV-DIV-001', 'SN123456', 'DIV-2025-001', '1',
        '2024-01-15', 'good', 'Ja', 'Beispiel: Neues Gerät - diese Zeile löschen!'
    ]
    writer.writerow(example_data)

    # Hinweise
    writer.writerow([])
    writer.writerow(['# HINWEISE:'])
    writer.writerow(['# Felder mit * sind Pflichtfelder'])
    writer.writerow(['# Zustand: excellent, good, fair, poor, damaged'])
    writer.writerow(['# Einsatzbereit: Ja oder Nein'])
    writer.writerow(['# Lagerort-ID muss existieren'])
    writer.writerow(['# Stammdaten-Nr. muss existieren'])

    # Response
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename=vorlage_taucher_geraete_import.csv'
    response.write('\ufeff')  # UTF-8 BOM
    response.write(output.getvalue())
    return response


# ============================================================================
# IMPORT FUNCTIONS (CSV)
# ============================================================================

@login_required
def import_masters(request):
    """Import Diving Masters aus CSV"""
    if request.method != 'POST':
        messages.error(request, 'Nur POST-Anfragen erlaubt.')
        return redirect('diving:import_export')

    if 'file' not in request.FILES:
        messages.error(request, 'Keine Datei hochgeladen.')
        return redirect('diving:import_export')

    uploaded_file = request.FILES['file']

    try:
        # CSV-Datei lesen (UTF-8 mit BOM)
        decoded_file = uploaded_file.read().decode('utf-8-sig')
        csv_reader = csv.reader(io.StringIO(decoded_file), delimiter=';')

        success_count = 0
        error_count = 0
        errors = []

        # Header überspringen
        next(csv_reader, None)

        # Zeilen durchgehen
        for row_num, row in enumerate(csv_reader, start=2):
            # Leere Zeilen und Kommentare überspringen
            if not row or not row[0] or row[0].startswith('#'):
                continue

            # Beispielzeile überspringen
            if 'Beispiel' in str(row[-1] if row else ''):
                continue

            try:
                # Daten extrahieren (mit Index-Safety)
                master_number = row[0] if len(row) > 0 else None
                name = row[1] if len(row) > 1 else None
                item_type = row[2] if len(row) > 2 else None
                manufacturer = row[3] if len(row) > 3 else ''
                model = row[4] if len(row) > 4 else ''
                tank_type = row[5] if len(row) > 5 else ''
                volume_liters = row[6] if len(row) > 6 else None
                working_pressure_bar = row[7] if len(row) > 7 else None
                certification_number = row[8] if len(row) > 8 else ''
                description = row[9] if len(row) > 9 else ''

                # Validierung
                if not all([master_number, name, item_type]):
                    errors.append(f"Zeile {row_num}: Pflichtfelder fehlen")
                    error_count += 1
                    continue

                # Duplikatprüfung
                if DivingItemMaster.objects.filter(master_number=master_number).exists():
                    errors.append(f"Zeile {row_num}: Stammdaten-Nr. '{master_number}' existiert bereits")
                    error_count += 1
                    continue

                # Master erstellen
                master = DivingItemMaster.objects.create(
                    master_number=master_number,
                    name=name,
                    item_type=item_type,
                    manufacturer=manufacturer,
                    model=model,
                    tank_type=tank_type if tank_type else '',
                    volume_liters=float(volume_liters) if volume_liters else None,
                    working_pressure_bar=int(working_pressure_bar) if working_pressure_bar else None,
                    certification_number=certification_number,
                    description=description,
                    created_by=request.user
                )

                success_count += 1

            except Exception as e:
                errors.append(f"Zeile {row_num}: Fehler - {str(e)}")
                error_count += 1
                logger.error(f"Master import error in row {row_num}: {e}")

        # Erfolgsmeldung
        if success_count > 0:
            messages.success(request, f"✅ {success_count} Stammdaten erfolgreich importiert.")

        if error_count > 0:
            messages.warning(request, f"⚠️ {error_count} Fehler beim Import.")
            for error in errors[:10]:
                messages.error(request, error)
            if len(errors) > 10:
                messages.error(request, f"... und {len(errors) - 10} weitere Fehler")

    except Exception as e:
        messages.error(request, f"Fehler beim Lesen der CSV-Datei: {str(e)}")
        logger.error(f"CSV import error: {e}")

    return redirect('diving:import_export')


@login_required
def import_devices(request):
    """Import Diving Devices aus CSV"""
    if request.method != 'POST':
        messages.error(request, 'Nur POST-Anfragen erlaubt.')
        return redirect('diving:import_export')

    if 'file' not in request.FILES:
        messages.error(request, 'Keine Datei hochgeladen.')
        return redirect('diving:import_export')

    uploaded_file = request.FILES['file']

    try:
        # CSV-Datei lesen (UTF-8 mit BOM)
        decoded_file = uploaded_file.read().decode('utf-8-sig')
        csv_reader = csv.reader(io.StringIO(decoded_file), delimiter=';')

        success_count = 0
        error_count = 0
        errors = []

        # Header überspringen
        next(csv_reader, None)

        # Zeilen durchgehen
        for row_num, row in enumerate(csv_reader, start=2):
            # Leere Zeilen und Kommentare überspringen
            if not row or not row[0] or row[0].startswith('#'):
                continue

            # Beispielzeile überspringen
            if 'Beispiel' in str(row[-1] if row else ''):
                continue

            try:
                # Daten extrahieren (mit Index-Safety)
                inventory_number = row[0] if len(row) > 0 else None
                serial_number = row[1] if len(row) > 1 else None
                master_number = row[2] if len(row) > 2 else None
                location_id = row[3] if len(row) > 3 else None
                manufacturing_date_str = row[4] if len(row) > 4 else ''
                condition = row[5] if len(row) > 5 else 'good'
                is_operational_str = row[6] if len(row) > 6 else 'Ja'
                notes = row[7] if len(row) > 7 else ''

                # Validierung
                if not all([inventory_number, serial_number, master_number, location_id]):
                    errors.append(f"Zeile {row_num}: Pflichtfelder fehlen")
                    error_count += 1
                    continue

                # Duplikatprüfung
                if DivingDeviceInstance.objects.filter(inventory_number=inventory_number).exists():
                    errors.append(f"Zeile {row_num}: Inventarnummer '{inventory_number}' existiert bereits")
                    error_count += 1
                    continue

                # Master laden
                try:
                    master = DivingItemMaster.objects.get(master_number=master_number)
                except DivingItemMaster.DoesNotExist:
                    errors.append(f"Zeile {row_num}: Stammdaten-Nr. {master_number} nicht gefunden")
                    error_count += 1
                    continue

                # Standort laden
                from locations.models import Location
                try:
                    location = Location.objects.get(id=int(location_id))
                except (Location.DoesNotExist, ValueError):
                    errors.append(f"Zeile {row_num}: Standort-ID {location_id} nicht gefunden")
                    error_count += 1
                    continue

                # Datum parsen
                manufacturing_date = None
                if manufacturing_date_str:
                    from datetime import datetime
                    try:
                        manufacturing_date = datetime.strptime(manufacturing_date_str, '%Y-%m-%d').date()
                    except ValueError:
                        try:
                            manufacturing_date = datetime.strptime(manufacturing_date_str, '%d.%m.%Y').date()
                        except ValueError:
                            pass

                # Boolean-Konvertierung
                is_operational = is_operational_str.lower() in ['ja', 'yes', 'true', '1']

                # Device erstellen
                device = DivingDeviceInstance.objects.create(
                    inventory_number=inventory_number,
                    serial_number=serial_number,
                    master=master,
                    location=location,
                    manufacturing_date=manufacturing_date,
                    condition=condition,
                    is_operational=is_operational,
                    condition_notes=notes,
                    created_by=request.user
                )

                success_count += 1

            except Exception as e:
                errors.append(f"Zeile {row_num}: Fehler - {str(e)}")
                error_count += 1
                logger.error(f"Device import error in row {row_num}: {e}")

        # Erfolgsmeldung
        if success_count > 0:
            messages.success(request, f"✅ {success_count} Geräte erfolgreich importiert.")

        if error_count > 0:
            messages.warning(request, f"⚠️ {error_count} Fehler beim Import.")
            for error in errors[:10]:
                messages.error(request, error)
            if len(errors) > 10:
                messages.error(request, f"... und {len(errors) - 10} weitere Fehler")

    except Exception as e:
        messages.error(request, f"Fehler beim Lesen der CSV-Datei: {str(e)}")
        logger.error(f"CSV import error: {e}")

    return redirect('diving:import_export')
