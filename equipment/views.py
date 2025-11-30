"""
Equipment Views
Views für Ausrüstung & Geräte-Verwaltung
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.utils.translation import gettext_lazy as _
from django.db import models
from django.db.models import Q, Count, F, Sum
from django.utils import timezone
from datetime import timedelta

from .models import (
    EquipmentItemMaster,
    EquipmentDeviceInstance,
    EquipmentItem,
    EquipmentStockMovement,
    InspectionType,
    EquipmentInspectionAssignment,
    InspectionRecord,
    MaintenanceType,
    MasterMaintenanceAssignment,
    EquipmentMaintenanceAssignment,
    MaintenanceRecord,
    EquipmentBatch,
)
from personnel.models import Person, Qualification, DutyHoursEntry, DutyHoursRequirement
from .forms import (
    EquipmentItemMasterForm,
    EquipmentDeviceInstanceForm,
    EquipmentItemForm,
    EquipmentStockMovementForm,
    InspectionTypeForm,
    EquipmentInspectionAssignmentForm,
    InspectionRecordForm,
    MaintenanceTypeForm,
    MasterMaintenanceAssignmentForm,
    MaintenanceAssignmentForm,
    MaintenanceRecordForm,
    EquipmentBatchForm,
)
from .utils import process_equipment_manual
from inventory_base.models import Category
from locations.models import Location


# ============================================================================
# DASHBOARD
# ============================================================================

class EquipmentDashboardView(LoginRequiredMixin, TemplateView):
    """Dashboard für Ausrüstung & Geräte"""
    template_name = 'equipment/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()

        # KPIs - NEW: Basierend auf Device Instances
        context['total_masters'] = EquipmentItemMaster.objects.filter(is_active=True).count()
        context['total_devices'] = EquipmentDeviceInstance.objects.filter(is_active=True).count()
        context['operational_devices'] = EquipmentDeviceInstance.objects.filter(
            is_active=True,
            is_operational=True
        ).count()

        # Legacy items
        context['total_items'] = EquipmentItem.objects.filter(is_active=True).count()

        # Wartung fällig (Device Instances)
        context['maintenance_due_count'] = EquipmentDeviceInstance.objects.filter(
            is_active=True,
            next_maintenance_date__isnull=False,
            next_maintenance_date__lte=today
        ).count()

        # Prüfung fällig (Device Instances)
        context['inspection_due_count'] = EquipmentDeviceInstance.objects.filter(
            is_active=True,
            next_inspection_date__isnull=False,
            next_inspection_date__lte=today
        ).count()

        # Defekte Geräte (Device Instances)
        context['defective_count'] = EquipmentDeviceInstance.objects.filter(
            is_active=True,
            equipment_status='defective'
        ).count()

        # Letzte Bewegungen
        context['recent_movements'] = EquipmentStockMovement.objects.select_related(
            'item',
            'vehicle',
            'person',
            'created_by'
        ).order_by('-movement_date')[:10]

        # Kritische Wartungen (überfällig) - Device Instances
        context['overdue_maintenance'] = EquipmentDeviceInstance.objects.filter(
            is_active=True,
            next_maintenance_date__isnull=False,
            next_maintenance_date__lt=today - timedelta(days=30)
        ).select_related('master', 'assigned_vehicle')[:5]

        # Kritische Prüfungen (überfällig) - Device Instances
        context['overdue_inspections'] = EquipmentDeviceInstance.objects.filter(
            is_active=True,
            next_inspection_date__isnull=False,
            next_inspection_date__lt=today - timedelta(days=30)
        ).select_related('master', 'assigned_vehicle')[:5]

        # ========================================================================
        # PERSONAL: Qualifikationen & Pflichtstunden
        # ========================================================================

        # Personen mit Ausrüstungs-Funktion finden
        equipment_personnel = Person.objects.filter(
            is_active=True,
            functions__related_module='equipment',
            functions__is_active=True
        ).distinct().prefetch_related('qualifications', 'duty_hours')

        # Ablaufende Qualifikationen (nächste 60 Tage oder bereits abgelaufen)
        expiring_qualifications = []
        for person in equipment_personnel:
            for qualification in person.qualifications.filter(is_active=True, expiry_date__isnull=False):
                days_until_expiry = (qualification.expiry_date - today).days
                if days_until_expiry <= 60:  # Läuft in 60 Tagen oder weniger ab
                    expiring_qualifications.append({
                        'person': person,
                        'qualification': qualification,
                        'days_until_expiry': days_until_expiry,
                        'is_expired': days_until_expiry < 0,
                    })

        # Nach Dringlichkeit sortieren (abgelaufen zuerst, dann nach Tagen)
        expiring_qualifications.sort(key=lambda x: (not x['is_expired'], x['days_until_expiry']))
        context['expiring_qualifications'] = expiring_qualifications[:10]

        # Fehlende Pflichtstunden für aktuelles Jahr
        current_year = today.year
        missing_duty_hours = []

        for person in equipment_personnel:
            # Alle Pflichtstunden-Kategorien finden, die für diese Person relevant sind
            required_categories = set()
            for function in person.functions.filter(related_module='equipment', is_active=True):
                for category in function.required_duty_hour_categories.all():
                    required_categories.add(category)

            # Für jede Kategorie prüfen, ob Anforderungen erfüllt sind
            for category in required_categories:
                # Anforderung für dieses Jahr finden
                try:
                    requirement = DutyHoursRequirement.objects.get(
                        category=category,
                        year=current_year,
                        is_active=True
                    )

                    # Geleistete Stunden berechnen
                    completed_hours = DutyHoursEntry.objects.filter(
                        person=person,
                        category=category,
                        date__year=current_year
                    ).aggregate(total=Sum('hours'))['total'] or 0

                    # Fehlende Stunden berechnen
                    missing_hours = requirement.required_hours - completed_hours

                    if missing_hours > 0:
                        missing_duty_hours.append({
                            'person': person,
                            'category': category,
                            'required_hours': requirement.required_hours,
                            'completed_hours': completed_hours,
                            'missing_hours': missing_hours,
                        })
                except DutyHoursRequirement.DoesNotExist:
                    # Keine Anforderung für dieses Jahr definiert
                    pass

        # Nach fehlenden Stunden sortieren (meiste fehlende Stunden zuerst)
        missing_duty_hours.sort(key=lambda x: x['missing_hours'], reverse=True)
        context['missing_duty_hours'] = missing_duty_hours[:10]

        return context


# ============================================================================
# EQUIPMENT MASTER DATA VIEWS (STAMMDATEN)
# ============================================================================

class EquipmentMasterListView(LoginRequiredMixin, ListView):
    """Liste aller Ausrüstungs-Stammdaten"""
    model = EquipmentItemMaster
    template_name = 'equipment/master_list.html'
    context_object_name = 'masters'
    paginate_by = 50

    def get_queryset(self):
        queryset = EquipmentItemMaster.objects.filter(is_active=True).select_related(
            'category',
            'supplier',
        )

        # Suchfilter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(master_number__icontains=search) |
                Q(name__icontains=search) |
                Q(manufacturer__icontains=search) |
                Q(model__icontains=search) |
                Q(description__icontains=search)
            )

        # Typ-Filter
        equipment_type = self.request.GET.get('equipment_type')
        if equipment_type:
            queryset = queryset.filter(equipment_type=equipment_type)

        # Kategorie-Filter
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category_id=category)

        return queryset.order_by('master_number')


class EquipmentMasterDetailView(LoginRequiredMixin, DetailView):
    """Detailansicht Ausrüstungs-Stammdaten"""
    model = EquipmentItemMaster
    template_name = 'equipment/master_detail.html'
    context_object_name = 'master'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Geräte-Instanzen dieses Masters
        context['device_instances'] = self.object.device_instances.filter(
            is_active=True
        ).select_related('location', 'assigned_vehicle')
        return context


class EquipmentMasterCreateView(LoginRequiredMixin, CreateView):
    """Neue Ausrüstungs-Stammdaten anlegen"""
    model = EquipmentItemMaster
    form_class = EquipmentItemMasterForm
    template_name = 'equipment/master_form.html'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Stammdaten wurden erfolgreich erstellt.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('equipment:master_detail', kwargs={'pk': self.object.pk})


class EquipmentMasterUpdateView(LoginRequiredMixin, UpdateView):
    """Ausrüstungs-Stammdaten bearbeiten"""
    model = EquipmentItemMaster
    form_class = EquipmentItemMasterForm
    template_name = 'equipment/master_form.html'

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Stammdaten wurden erfolgreich aktualisiert.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('equipment:master_detail', kwargs={'pk': self.object.pk})


class EquipmentMasterDeleteView(LoginRequiredMixin, DeleteView):
    """Ausrüstungs-Stammdaten löschen"""
    model = EquipmentItemMaster
    template_name = 'equipment/master_confirm_delete.html'
    success_url = reverse_lazy('equipment:master_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, _('Stammdaten wurden erfolgreich gelöscht.'))
        return super().delete(request, *args, **kwargs)


# QR-Code & Barcode für Masters
class EquipmentMasterQRCodeView(LoginRequiredMixin, DetailView):
    """QR-Code für Ausrüstungs-Stammdaten"""
    model = EquipmentItemMaster

    def get(self, request, *args, **kwargs):
        master = self.get_object()
        import qrcode
        import qrcode.image.svg
        from io import BytesIO
        from django.http import HttpResponse

        # Nur URL kodieren für maximale Scanner-Kompatibilität
        qr_data = f"/equipment/masters/{master.pk}/"

        # QR-Code generieren (SVG)
        factory = qrcode.image.svg.SvgPathImage
        img = qrcode.make(qr_data, image_factory=factory, box_size=10)

        # Als SVG zurückgeben
        stream = BytesIO()
        img.save(stream)
        response = HttpResponse(stream.getvalue(), content_type='image/svg+xml')
        response['Content-Disposition'] = f'inline; filename="qr_{master.master_number}.svg"'
        return response


class EquipmentMasterBarcodeView(LoginRequiredMixin, DetailView):
    """Barcode für Ausrüstungs-Stammdaten"""
    model = EquipmentItemMaster

    def get(self, request, *args, **kwargs):
        master = self.get_object()
        import barcode
        from barcode.writer import SVGWriter
        from io import BytesIO
        from django.http import HttpResponse

        # Barcode generieren (Code128)
        code128 = barcode.get_barcode_class('code128')
        barcode_instance = code128(master.master_number, writer=SVGWriter())

        # Als SVG zurückgeben
        stream = BytesIO()
        barcode_instance.write(stream)
        response = HttpResponse(stream.getvalue(), content_type='image/svg+xml')
        response['Content-Disposition'] = f'inline; filename="barcode_{master.master_number}.svg"'
        return response


# ============================================================================
# EQUIPMENT DEVICE INSTANCE VIEWS (GERÄTE MIT INVENTARNUMMER)
# ============================================================================

class EquipmentDeviceListView(LoginRequiredMixin, ListView):
    """Liste aller Ausrüstungs-Geräte (Instanzen)"""
    model = EquipmentDeviceInstance
    template_name = 'equipment/device_list.html'
    context_object_name = 'devices'
    paginate_by = 50

    def get_queryset(self):
        queryset = EquipmentDeviceInstance.objects.filter(is_active=True).select_related(
            'master',
            'location',
            'assigned_vehicle',
        )

        # Suchfilter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(inventory_number__icontains=search) |
                Q(serial_number__icontains=search) |
                Q(master__name__icontains=search) |
                Q(master__manufacturer__icontains=search)
            )

        # Master-Filter
        master_id = self.request.GET.get('master')
        if master_id:
            queryset = queryset.filter(master_id=master_id)

        # Status-Filter
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(equipment_status=status)

        # Operational-Filter
        operational = self.request.GET.get('operational')
        if operational == 'yes':
            queryset = queryset.filter(is_operational=True)
        elif operational == 'no':
            queryset = queryset.filter(is_operational=False)

        # Fahrzeug-Filter
        vehicle_id = self.request.GET.get('vehicle')
        if vehicle_id:
            queryset = queryset.filter(assigned_vehicle_id=vehicle_id)

        # Wartung fällig
        if self.request.GET.get('maintenance_due'):
            today = timezone.now().date()
            queryset = queryset.filter(
                next_maintenance_date__isnull=False,
                next_maintenance_date__lte=today
            )

        return queryset.order_by('inventory_number')


class EquipmentDeviceDetailView(LoginRequiredMixin, DetailView):
    """Detailansicht Ausrüstungs-Gerät"""
    model = EquipmentDeviceInstance
    template_name = 'equipment/device_detail.html'
    context_object_name = 'device'


class EquipmentDeviceCreateView(LoginRequiredMixin, CreateView):
    """Neues Ausrüstungs-Gerät anlegen"""
    model = EquipmentDeviceInstance
    form_class = EquipmentDeviceInstanceForm
    template_name = 'equipment/device_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Neues Gerät anlegen'
        context['submit_text'] = 'Gerät erstellen'
        return context

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Gerät wurde erfolgreich erstellt.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('equipment:device_detail', kwargs={'pk': self.object.pk})


class EquipmentDeviceUpdateView(LoginRequiredMixin, UpdateView):
    """Ausrüstungs-Gerät bearbeiten"""
    model = EquipmentDeviceInstance
    form_class = EquipmentDeviceInstanceForm
    template_name = 'equipment/device_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Gerät bearbeiten'
        context['submit_text'] = 'Änderungen speichern'
        return context

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Gerät wurde erfolgreich aktualisiert.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('equipment:device_detail', kwargs={'pk': self.object.pk})


# QR-Code & Barcode für Device Instances
class EquipmentDeviceQRCodeView(LoginRequiredMixin, DetailView):
    """QR-Code für Ausrüstungs-Gerät"""
    model = EquipmentDeviceInstance

    def get(self, request, *args, **kwargs):
        device = self.get_object()
        import qrcode
        import qrcode.image.svg
        from io import BytesIO
        from django.http import HttpResponse

        # Nur URL kodieren für maximale Scanner-Kompatibilität
        qr_data = f"/equipment/device/{device.pk}/"

        # QR-Code generieren (SVG)
        factory = qrcode.image.svg.SvgPathImage
        img = qrcode.make(qr_data, image_factory=factory, box_size=10)

        # Als SVG zurückgeben
        stream = BytesIO()
        img.save(stream)
        response = HttpResponse(stream.getvalue(), content_type='image/svg+xml')
        response['Content-Disposition'] = f'inline; filename="qr_{device.inventory_number}.svg"'
        return response


class EquipmentDeviceBarcodeView(LoginRequiredMixin, DetailView):
    """Barcode für Ausrüstungs-Gerät"""
    model = EquipmentDeviceInstance

    def get(self, request, *args, **kwargs):
        device = self.get_object()
        import barcode
        from barcode.writer import SVGWriter
        from io import BytesIO
        from django.http import HttpResponse

        # Barcode generieren (Code128)
        code128 = barcode.get_barcode_class('code128')
        barcode_instance = code128(device.inventory_number, writer=SVGWriter())

        # Als SVG zurückgeben
        stream = BytesIO()
        barcode_instance.write(stream)
        response = HttpResponse(stream.getvalue(), content_type='image/svg+xml')
        response['Content-Disposition'] = f'inline; filename="barcode_{device.inventory_number}.svg"'
        return response


# ============================================================================
# EQUIPMENT ITEM VIEWS
# ============================================================================

class EquipmentItemListView(LoginRequiredMixin, ListView):
    """Liste aller Ausrüstung/Geräte"""
    model = EquipmentItem
    template_name = 'equipment/item_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        queryset = EquipmentItem.objects.filter(is_active=True).select_related(
            'category',
            'supplier',
            'assigned_vehicle',
            'location'
        )

        # Suchfilter (inkl. OCR-Text aus Handbüchern)
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(item_number__icontains=search) |
                Q(name__icontains=search) |
                Q(serial_number__icontains=search) |
                Q(manufacturer__icontains=search) |
                Q(description__icontains=search) |
                Q(manual_ocr_text__icontains=search)  # Durchsucht OCR-Text aus Handbüchern
            )

        # Typ-Filter
        equipment_type = self.request.GET.get('type')
        if equipment_type:
            queryset = queryset.filter(equipment_type=equipment_type)

        # Kategorie-Filter
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category_id=category)

        # Lagerort-Filter
        location = self.request.GET.get('location')
        if location:
            queryset = queryset.filter(location_id=location)

        # Status-Filter
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(equipment_status=status)

        # Fahrzeug-Filter
        vehicle_id = self.request.GET.get('vehicle')
        if vehicle_id == 'unassigned':
            queryset = queryset.filter(assigned_vehicle__isnull=True)
        elif vehicle_id:
            queryset = queryset.filter(assigned_vehicle_id=vehicle_id)

        # Wartung fällig
        if self.request.GET.get('maintenance_due'):
            today = timezone.now().date()
            queryset = queryset.filter(
                requires_maintenance=True,
                next_maintenance_date__isnull=False,
                next_maintenance_date__lte=today
            )

        # Prüfung fällig
        if self.request.GET.get('inspection_due'):
            today = timezone.now().date()
            queryset = queryset.filter(
                requires_inspection=True,
                next_inspection_date__isnull=False,
                next_inspection_date__lte=today
            )

        # Zertifizierung erforderlich
        if self.request.GET.get('certification_required'):
            queryset = queryset.filter(requires_certification=True)

        # Sortierung
        sort = self.request.GET.get('sort', 'equipment_type')
        if sort:
            queryset = queryset.order_by(sort, 'name')
        else:
            queryset = queryset.order_by('equipment_type', 'name')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_count'] = self.get_queryset().count()
        context['operational_count'] = self.get_queryset().filter(equipment_status='operational').count()
        context['defective_count'] = self.get_queryset().filter(equipment_status='defective').count()
        context['categories'] = Category.objects.filter(is_active=True).order_by('name')
        context['locations'] = Location.objects.filter(is_active=True).order_by('name')
        return context


class EquipmentItemDetailView(LoginRequiredMixin, DetailView):
    """Detailansicht eines Geräts"""
    model = EquipmentItem
    template_name = 'equipment/item_detail.html'
    context_object_name = 'item'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['movements'] = self.object.stock_movements.select_related(
            'vehicle',
            'person'
        ).order_by('-movement_date')[:10]
        return context


class EquipmentItemCreateView(LoginRequiredMixin, CreateView):
    """Neues Gerät anlegen"""
    model = EquipmentItem
    form_class = EquipmentItemForm
    template_name = 'equipment/item_form.html'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user

        # Setze Standardlagerort (erster aktiver Lagerort) wenn nicht gesetzt
        if not hasattr(form.instance, 'location') or form.instance.location is None:
            default_location = Location.objects.filter(is_active=True).first()
            if default_location:
                form.instance.location = default_location

        # Bestand initial auf 0 setzen (wird über Wareneingang erhöht)
        form.instance.quantity = 0

        response = super().form_valid(form)

        # OCR-Verarbeitung des Handbuchs starten (wenn hochgeladen)
        if form.instance.manual_document:
            from django.core.management import call_command
            from threading import Thread

            # OCR in Hintergrund-Thread ausführen
            def process_ocr():
                process_equipment_manual(form.instance)

            thread = Thread(target=process_ocr)
            thread.daemon = True
            thread.start()

            messages.success(self.request, _('Gerät wurde erfolgreich erstellt. Das Handbuch wird im Hintergrund verarbeitet.'))
        else:
            messages.success(self.request, _('Gerät wurde erfolgreich erstellt.'))

        return response

    def get_success_url(self):
        return reverse_lazy('equipment:item_detail', kwargs={'pk': self.object.pk})


class EquipmentItemUpdateView(LoginRequiredMixin, UpdateView):
    """Gerät bearbeiten"""
    model = EquipmentItem
    form_class = EquipmentItemForm
    template_name = 'equipment/item_form.html'

    def form_valid(self, form):
        form.instance.updated_by = self.request.user

        # Prüfen ob neues Handbuch hochgeladen wurde
        manual_changed = 'manual_document' in form.changed_data

        response = super().form_valid(form)

        # OCR-Verarbeitung des Handbuchs starten (wenn neu hochgeladen)
        if manual_changed and form.instance.manual_document:
            from threading import Thread

            def process_ocr():
                process_equipment_manual(form.instance)

            thread = Thread(target=process_ocr)
            thread.daemon = True
            thread.start()

            messages.success(self.request, _('Gerät wurde erfolgreich aktualisiert. Das Handbuch wird im Hintergrund verarbeitet.'))
        else:
            messages.success(self.request, _('Gerät wurde erfolgreich aktualisiert.'))

        return response

    def get_success_url(self):
        return reverse_lazy('equipment:item_detail', kwargs={'pk': self.object.pk})


class EquipmentItemDeleteView(LoginRequiredMixin, DeleteView):
    """Gerät löschen"""
    model = EquipmentItem
    template_name = 'equipment/item_confirm_delete.html'
    success_url = reverse_lazy('equipment:item_list')

    def form_valid(self, form):
        item = self.get_object()
        item_name = str(item)

        # Delete related stock movements first
        movements_count = 0
        if hasattr(item, 'stock_movements'):
            movements_count = item.stock_movements.count()
            item.stock_movements.all().delete()

        response = super().form_valid(form)

        msg = f'Gerät "{item_name}" wurde gelöscht'
        if movements_count:
            msg += f' (inkl. {movements_count} Bewegungen)'
        msg += '.'
        messages.success(self.request, msg)
        return response


# ============================================================================
# STOCK MOVEMENT VIEWS
# ============================================================================

class EquipmentStockMovementListView(LoginRequiredMixin, ListView):
    """Liste aller Lagerbewegungen"""
    model = EquipmentStockMovement
    template_name = 'equipment/movement_list.html'
    context_object_name = 'movements'
    paginate_by = 50

    def get_queryset(self):
        queryset = EquipmentStockMovement.objects.select_related(
            'item',
            'vehicle',
            'person',
            'created_by'
        )

        # Typ-Filter
        movement_type = self.request.GET.get('type')
        if movement_type:
            queryset = queryset.filter(movement_type=movement_type)

        # Fahrzeug-Filter
        vehicle_id = self.request.GET.get('vehicle')
        if vehicle_id:
            queryset = queryset.filter(vehicle_id=vehicle_id)

        # Defekt-Filter
        defective = self.request.GET.get('defective')
        if defective == 'yes':
            queryset = queryset.filter(is_defective=True)

        return queryset.order_by('-movement_date')


class EquipmentStockMovementDetailView(LoginRequiredMixin, DetailView):
    """Detailansicht einer Lagerbewegung"""
    model = EquipmentStockMovement
    template_name = 'equipment/movement_detail.html'
    context_object_name = 'movement'


class EquipmentStockMovementCreateView(LoginRequiredMixin, CreateView):
    """Neue Lagerbewegung erstellen"""
    model = EquipmentStockMovement
    form_class = EquipmentStockMovementForm
    template_name = 'equipment/movement_form.html'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Lagerbewegung wurde erfolgreich erstellt.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('equipment:movement_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Neue Lagerbewegung'
        context['submit_text'] = 'Bewegung speichern'
        return context


# ============================================================================
# MAINTENANCE & INSPECTION VIEWS
# ============================================================================

class MaintenanceManagementView(LoginRequiredMixin, TemplateView):
    """Zentrale Verwaltung für Wartungen und Prüfungen"""
    template_name = 'equipment/maintenance_management.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()

        # Wartungsdaten
        context['maintenance_due_count'] = EquipmentItem.objects.filter(
            is_active=True,
            requires_maintenance=True,
            next_maintenance_date__isnull=False,
            next_maintenance_date__lte=today
        ).count()

        # Prüfungsdaten
        context['inspection_due_count'] = EquipmentItem.objects.filter(
            is_active=True,
            requires_inspection=True,
            next_inspection_date__isnull=False,
            next_inspection_date__lte=today
        ).count()

        return context


class MaintenanceListView(LoginRequiredMixin, ListView):
    """Liste aller wartungspflichtigen Geräte"""
    model = EquipmentItem
    template_name = 'equipment/maintenance_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        return EquipmentItem.objects.filter(
            requires_maintenance=True
        ).select_related(
            'assigned_vehicle'
        ).order_by('next_maintenance_date')


class MaintenanceDueView(LoginRequiredMixin, ListView):
    """Fällige Wartungen"""
    model = EquipmentItem
    template_name = 'equipment/maintenance_due.html'
    context_object_name = 'items'

    def get_queryset(self):
        today = timezone.now().date()

        return EquipmentItem.objects.filter(
            requires_maintenance=True,
            next_maintenance_date__lte=today
        ).select_related(
            'assigned_vehicle'
        ).order_by('next_maintenance_date')


class InspectionListView(LoginRequiredMixin, ListView):
    """Liste aller prüfpflichtigen Geräte"""
    model = EquipmentItem
    template_name = 'equipment/inspection_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        return EquipmentItem.objects.filter(
            requires_inspection=True
        ).select_related(
            'assigned_vehicle'
        ).order_by('next_inspection_date')


class InspectionDueView(LoginRequiredMixin, ListView):
    """Fällige Prüfungen"""
    model = EquipmentItem
    template_name = 'equipment/inspection_due.html'
    context_object_name = 'items'

    def get_queryset(self):
        today = timezone.now().date()

        return EquipmentItem.objects.filter(
            requires_inspection=True,
            next_inspection_date__lte=today
        ).select_related(
            'assigned_vehicle'
        ).order_by('next_inspection_date')


# ============================================================================
# VEHICLE EQUIPMENT VIEWS
# ============================================================================

class VehicleEquipmentListView(LoginRequiredMixin, ListView):
    """Ausrüstung eines bestimmten Fahrzeugs"""
    model = EquipmentItem
    template_name = 'equipment/vehicle_equipment.html'
    context_object_name = 'items'

    def get_queryset(self):
        vehicle_id = self.kwargs.get('vehicle_id')
        return EquipmentItem.objects.filter(
            assigned_vehicle_id=vehicle_id
        ).order_by('equipment_type', 'name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        from vehicles.models import Vehicle
        context['vehicle'] = get_object_or_404(Vehicle, pk=self.kwargs.get('vehicle_id'))

        return context


# ============================================================================
# DEFECTIVE EQUIPMENT VIEWS
# ============================================================================

class DefectiveEquipmentListView(LoginRequiredMixin, ListView):
    """Liste aller defekten Geräte"""
    model = EquipmentItem
    template_name = 'equipment/defective_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        return EquipmentItem.objects.filter(
            equipment_status='defective'
        ).select_related(
            'assigned_vehicle'
        ).order_by('-updated_at')


# ============================================================================
# CATEGORY VIEWS
# ============================================================================

class CategoryListView(LoginRequiredMixin, ListView):
    """Liste aller Kategorien"""
    model = Category
    template_name = 'equipment/category_list.html'
    context_object_name = 'categories'

    def get_queryset(self):
        return Category.objects.filter(is_active=True, module='equipment').order_by('tree_id', 'lft')


class CategoryCreateView(LoginRequiredMixin, CreateView):
    """Neue Kategorie erstellen"""
    model = Category
    template_name = 'equipment/category_form.html'
    fields = ['name', 'parent', 'code', 'description']
    success_url = reverse_lazy('equipment:category_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        form.instance.module = 'equipment'
        messages.success(self.request, f'Kategorie "{form.instance.name}" wurde erstellt.')
        return super().form_valid(form)


class CategoryUpdateView(LoginRequiredMixin, UpdateView):
    """Kategorie bearbeiten"""
    model = Category
    template_name = 'equipment/category_form.html'
    fields = ['name', 'parent', 'code', 'description', 'is_active']
    success_url = reverse_lazy('equipment:category_list')

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, f'Kategorie "{form.instance.name}" wurde aktualisiert.')
        return super().form_valid(form)


class CategoryDeleteView(LoginRequiredMixin, DeleteView):
    """Kategorie löschen"""
    model = Category
    template_name = 'equipment/category_confirm_delete.html'
    success_url = reverse_lazy('equipment:category_list')

    def form_valid(self, form):
        category = self.get_object()
        try:
            response = super().form_valid(form)
            messages.success(self.request, f'Kategorie "{category.name}" wurde gelöscht.')
            return response
        except models.ProtectedError:
            messages.error(
                self.request,
                f'Kategorie "{category.name}" kann nicht gelöscht werden, da sie noch von Artikeln verwendet wird.'
            )
            return redirect('equipment:category_list')


# ============================================================================
# INSPECTION TYPE VIEWS
# ============================================================================

class InspectionTypeListView(LoginRequiredMixin, ListView):
    """Liste aller Prüfungsarten"""
    model = InspectionType
    template_name = 'equipment/inspection_type_list.html'
    context_object_name = 'inspection_types'
    paginate_by = 20

    def get_queryset(self):
        queryset = InspectionType.objects.filter(is_active=True)

        # Suchfilter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(description__icontains=search) |
                Q(inspection_standard__icontains=search)
            )

        # Filter nach Zertifizierung
        requires_certification = self.request.GET.get('requires_certification')
        if requires_certification == 'yes':
            queryset = queryset.filter(requires_certification=True)
        elif requires_certification == 'no':
            queryset = queryset.filter(requires_certification=False)

        # Filter nach externer Prüfung
        requires_external = self.request.GET.get('requires_external')
        if requires_external == 'yes':
            queryset = queryset.filter(requires_external_service=True)
        elif requires_external == 'no':
            queryset = queryset.filter(requires_external_service=False)

        return queryset.order_by('name')


class InspectionTypeCreateView(LoginRequiredMixin, CreateView):
    """Neue Prüfungsart erstellen"""
    model = InspectionType
    form_class = InspectionTypeForm
    template_name = 'equipment/inspection_type_form.html'
    success_url = reverse_lazy('equipment:inspection_type_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, f'Prüfungsart "{form.instance.name}" wurde erstellt.')
        return super().form_valid(form)


class InspectionTypeUpdateView(LoginRequiredMixin, UpdateView):
    """Prüfungsart bearbeiten"""
    model = InspectionType
    form_class = InspectionTypeForm
    template_name = 'equipment/inspection_type_form.html'
    success_url = reverse_lazy('equipment:inspection_type_list')

    def form_valid(self, form):
        messages.success(self.request, f'Prüfungsart "{form.instance.name}" wurde aktualisiert.')
        return super().form_valid(form)


class InspectionTypeDeleteView(LoginRequiredMixin, DeleteView):
    """Prüfungsart löschen"""
    model = InspectionType
    template_name = 'equipment/inspection_type_confirm_delete.html'
    success_url = reverse_lazy('equipment:inspection_type_list')

    def delete(self, request, *args, **kwargs):
        inspection_type = self.get_object()
        messages.success(request, f'Prüfungsart "{inspection_type.name}" wurde gelöscht.')
        return super().delete(request, *args, **kwargs)


# ============================================================================
# EQUIPMENT INSPECTION ASSIGNMENT VIEWS
# ============================================================================

class EquipmentInspectionAssignmentListView(LoginRequiredMixin, ListView):
    """Liste aller Prüfungszuweisungen"""
    model = EquipmentInspectionAssignment
    template_name = 'equipment/inspection_assignment_list.html'
    context_object_name = 'assignments'
    paginate_by = 50

    def get_queryset(self):
        queryset = EquipmentInspectionAssignment.objects.select_related(
            'device',
            'device__master',
            'inspection_type'
        ).filter(is_active=True)

        # Suchfilter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(device__inventory_number__icontains=search) |
                Q(device__master__name__icontains=search) |
                Q(inspection_type__name__icontains=search)
            )

        # Filter nach Prüfungsart
        inspection_type_id = self.request.GET.get('inspection_type')
        if inspection_type_id:
            queryset = queryset.filter(inspection_type_id=inspection_type_id)

        # Filter nach Status
        status = self.request.GET.get('status')
        today = timezone.now().date()

        if status == 'due':
            queryset = queryset.filter(next_inspection_date__lte=today)
        elif status == 'upcoming':
            queryset = queryset.filter(
                next_inspection_date__gt=today,
                next_inspection_date__lte=today + timedelta(days=30)
            )
        elif status == 'ok':
            queryset = queryset.filter(next_inspection_date__gt=today + timedelta(days=30))
        elif status == 'no_date':
            queryset = queryset.filter(next_inspection_date__isnull=True)

        return queryset.order_by('next_inspection_date', 'device__inventory_number')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['inspection_types'] = InspectionType.objects.filter(is_active=True).order_by('name')
        return context


class EquipmentInspectionAssignmentCreateView(LoginRequiredMixin, CreateView):
    """Neue Prüfungszuweisung erstellen"""
    model = EquipmentInspectionAssignment
    form_class = EquipmentInspectionAssignmentForm
    template_name = 'equipment/inspection_assignment_form.html'
    success_url = reverse_lazy('equipment:inspection_assignment_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(
            self.request,
            f'Prüfung "{form.instance.inspection_type.name}" wurde {form.instance.device.inventory_number} zugewiesen.'
        )
        return super().form_valid(form)


class EquipmentInspectionAssignmentUpdateView(LoginRequiredMixin, UpdateView):
    """Prüfungszuweisung bearbeiten"""
    model = EquipmentInspectionAssignment
    form_class = EquipmentInspectionAssignmentForm
    template_name = 'equipment/inspection_assignment_form.html'
    success_url = reverse_lazy('equipment:inspection_assignment_list')

    def form_valid(self, form):
        messages.success(self.request, 'Prüfungszuweisung wurde aktualisiert.')
        return super().form_valid(form)


class EquipmentInspectionAssignmentDeleteView(LoginRequiredMixin, DeleteView):
    """Prüfungszuweisung löschen"""
    model = EquipmentInspectionAssignment
    template_name = 'equipment/inspection_assignment_confirm_delete.html'
    success_url = reverse_lazy('equipment:inspection_assignment_list')

    def delete(self, request, *args, **kwargs):
        assignment = self.get_object()
        messages.success(
            request,
            f'Prüfungszuweisung für {assignment.device.inventory_number} wurde gelöscht.'
        )
        return super().delete(request, *args, **kwargs)


# ============================================================================
# INSPECTION RECORD VIEWS
# ============================================================================

class InspectionRecordListView(LoginRequiredMixin, ListView):
    """Liste aller Prüfprotokolle"""
    model = InspectionRecord
    template_name = 'equipment/inspection_record_list.html'
    context_object_name = 'records'
    paginate_by = 50

    def get_queryset(self):
        queryset = InspectionRecord.objects.select_related(
            'device',
            'device__master',
            'inspection_type',
            'inspector'
        )

        # Suchfilter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(device__inventory_number__icontains=search) |
                Q(device__master__name__icontains=search) |
                Q(inspection_type__name__icontains=search) |
                Q(inspector__first_name__icontains=search) |
                Q(inspector__last_name__icontains=search)
            )

        # Filter nach Ergebnis
        result = self.request.GET.get('result')
        if result:
            queryset = queryset.filter(result=result)

        # Filter nach Prüfer
        inspector_id = self.request.GET.get('inspector')
        if inspector_id:
            queryset = queryset.filter(inspector_id=inspector_id)

        return queryset.order_by('-inspection_date')


class InspectionRecordCreateView(LoginRequiredMixin, CreateView):
    """Neues Prüfprotokoll erstellen"""
    model = InspectionRecord
    form_class = InspectionRecordForm
    template_name = 'equipment/inspection_record_form.html'
    success_url = reverse_lazy('equipment:inspection_record_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Artikelnummer aus GET-Parameter (für vorausgefülltes Formular)
        context['scanned_item_number'] = self.request.GET.get('item_number', '')
        # Liste aller aktiven Geräte für Dropdown
        context['devices'] = EquipmentDeviceInstance.objects.select_related(
            'master'
        ).filter(is_active=True).order_by('inventory_number')
        return context

    def form_valid(self, form):
        messages.success(
            self.request,
            f'Prüfprotokoll für {form.instance.device.inventory_number} wurde erstellt.'
        )
        return super().form_valid(form)


class InspectionRecordDetailView(LoginRequiredMixin, DetailView):
    """Prüfprotokoll-Detailansicht"""
    model = InspectionRecord
    template_name = 'equipment/inspection_record_detail.html'
    context_object_name = 'record'


@login_required
def get_inspection_assignments_by_item(request):
    """
    API-Endpoint: Gibt Prüfungszuweisungen für einen Artikel zurück
    GET parameter: item_number
    """
    item_number = request.GET.get('item_number', '').strip()

    if not item_number:
        return JsonResponse({'error': 'Keine Artikelnummer angegeben'}, status=400)

    # Suche Artikel mit dieser Nummer
    try:
        item = EquipmentItem.objects.get(item_number=item_number, is_active=True)
    except EquipmentItem.DoesNotExist:
        return JsonResponse({'error': f'Artikel mit Nummer {item_number} nicht gefunden'}, status=404)

    # Finde alle aktiven Prüfungszuweisungen für diesen Artikel
    assignments = EquipmentInspectionAssignment.objects.filter(
        item=item,
        is_active=True
    ).select_related('inspection_type')

    # Baue Antwort
    data = {
        'item': {
            'id': item.id,
            'name': item.name,
            'item_number': item.item_number,
        },
        'assignments': [
            {
                'id': assignment.id,
                'inspection_type': assignment.inspection_type.name,
                'next_due_date': assignment.next_due_date.isoformat() if assignment.next_due_date else None,
                'display_text': f"{item.name} - {assignment.inspection_type.name}"
            }
            for assignment in assignments
        ]
    }

    return JsonResponse(data)


@login_required
def get_inspection_types_by_device(request):
    """
    API-Endpoint: Gibt verfügbare Prüfungsarten für ein Gerät zurück
    Basierend auf dem Gerätetyp (Master)
    GET parameter: device_id
    """
    device_id = request.GET.get('device_id', '').strip()

    if not device_id:
        return JsonResponse({'error': 'Keine Geräte-ID angegeben'}, status=400)

    # Suche Gerät mit dieser ID
    try:
        device = EquipmentDeviceInstance.objects.select_related('master').get(id=device_id, is_active=True)
    except EquipmentDeviceInstance.DoesNotExist:
        return JsonResponse({'error': f'Gerät mit ID {device_id} nicht gefunden'}, status=404)

    # Finde alle aktiven Prüfungsarten
    # Alle Prüfungsarten, die entweder für alle Gerätetypen gelten oder für diesen spezifischen Typ
    inspection_types = InspectionType.objects.filter(is_active=True)

    # Baue Antwort mit allen passenden Prüfungsarten
    inspection_type_list = []
    for it in inspection_types:
        # Prüfe ob applicable_equipment_types leer ist (= für alle) oder den Gerätetyp enthält
        if not it.applicable_equipment_types or device.master.equipment_type in it.applicable_equipment_types:
            inspection_type_list.append({
                'id': it.id,
                'name': it.name,
                'interval_months': it.interval_months,
                'display_text': f"{it.name}" + (f" ({it.interval_months} Monate)" if it.interval_months else "")
            })

    data = {
        'device': {
            'id': device.id,
            'name': f"{device.inventory_number} - {device.master.name}",
            'inventory_number': device.inventory_number,
            'equipment_type': device.master.equipment_type,
        },
        'inspection_types': inspection_type_list
    }

    return JsonResponse(data)


@login_required
def get_inspection_type_checklist(request):
    """
    API-Endpoint: Gibt Checkliste und custom_fields für eine Prüfungsart zurück
    GET parameter: inspection_type_id
    """
    inspection_type_id = request.GET.get('inspection_type_id', '').strip()

    if not inspection_type_id:
        return JsonResponse({'error': 'Keine Prüfungsart-ID angegeben'}, status=400)

    # Suche Prüfungsart
    try:
        inspection_type = InspectionType.objects.get(id=inspection_type_id, is_active=True)
    except InspectionType.DoesNotExist:
        return JsonResponse({'error': 'Prüfungsart nicht gefunden'}, status=404)

    # Baue Antwort
    data = {
        'inspection_type': {
            'id': inspection_type.id,
            'name': inspection_type.name,
            'interval_months': inspection_type.interval_months,
        },
        'checklist': inspection_type.checklist or [],
        'custom_fields': inspection_type.custom_fields or [],
    }

    return JsonResponse(data)


# ===================================================================
# WARTUNGSTYPEN VIEWS
# ===================================================================

class MaintenanceTypeListView(LoginRequiredMixin, ListView):
    """Liste aller Wartungsarten"""
    model = MaintenanceType
    template_name = 'equipment/maintenance_type_list.html'
    context_object_name = 'maintenance_types'
    paginate_by = 20

    def get_queryset(self):
        queryset = MaintenanceType.objects.filter(is_active=True)

        # Suchfilter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(description__icontains=search) |
                Q(maintenance_standard__icontains=search)
            )

        # Filter nach Befähigung
        requires_certification = self.request.GET.get('requires_certification')
        if requires_certification == 'yes':
            queryset = queryset.filter(requires_certification=True)
        elif requires_certification == 'no':
            queryset = queryset.filter(requires_certification=False)

        # Filter nach externer Wartung
        requires_external = self.request.GET.get('requires_external')
        if requires_external == 'yes':
            queryset = queryset.filter(requires_external_service=True)
        elif requires_external == 'no':
            queryset = queryset.filter(requires_external_service=False)

        return queryset.order_by('name')


class MaintenanceTypeCreateView(LoginRequiredMixin, CreateView):
    """Neue Wartungsart erstellen"""
    model = MaintenanceType
    form_class = MaintenanceTypeForm
    template_name = 'equipment/maintenance_type_form.html'
    success_url = reverse_lazy('equipment:maintenance_type_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(
            self.request,
            f'Wartungsart "{form.instance.name}" wurde erstellt.'
        )
        return super().form_valid(form)


class MaintenanceTypeUpdateView(LoginRequiredMixin, UpdateView):
    """Wartungsart bearbeiten"""
    model = MaintenanceType
    form_class = MaintenanceTypeForm
    template_name = 'equipment/maintenance_type_form.html'
    success_url = reverse_lazy('equipment:maintenance_type_list')

    def form_valid(self, form):
        messages.success(
            self.request,
            f'Wartungsart "{form.instance.name}" wurde aktualisiert.'
        )
        return super().form_valid(form)


class MaintenanceTypeDeleteView(LoginRequiredMixin, DeleteView):
    """Wartungsart löschen"""
    model = MaintenanceType
    template_name = 'equipment/maintenance_type_confirm_delete.html'
    success_url = reverse_lazy('equipment:maintenance_type_list')

    def delete(self, request, *args, **kwargs):
        maintenance_type = self.get_object()
        messages.success(
            self.request,
            f'Wartungsart "{maintenance_type.name}" wurde gelöscht.'
        )
        return super().delete(request, *args, **kwargs)


# ===================================================================
# WARTUNGSZUWEISUNGEN VIEWS
# ===================================================================

class MaintenanceAssignmentListView(LoginRequiredMixin, ListView):
    """Liste aller Stammdaten-Wartungszuweisungen"""
    model = MasterMaintenanceAssignment
    template_name = 'equipment/maintenance_assignment_list.html'
    context_object_name = 'assignments'
    paginate_by = 20

    def get_queryset(self):
        queryset = MasterMaintenanceAssignment.objects.select_related(
            'master', 'maintenance_type', 'created_by'
        ).annotate(
            device_count=Count('master__device_instances', filter=Q(master__device_instances__is_active=True))
        )

        # Suchfilter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(master__name__icontains=search) |
                Q(master__master_number__icontains=search) |
                Q(maintenance_type__name__icontains=search)
            )

        # Filter nach Wartungsart
        maintenance_type_id = self.request.GET.get('maintenance_type')
        if maintenance_type_id:
            queryset = queryset.filter(maintenance_type_id=maintenance_type_id)

        # Filter nach Aktiv-Status
        is_active = self.request.GET.get('is_active')
        if is_active == '1':
            queryset = queryset.filter(is_active=True)
        elif is_active == '0':
            queryset = queryset.filter(is_active=False)

        return queryset.order_by('master__name', 'maintenance_type__name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['maintenance_types'] = MaintenanceType.objects.filter(is_active=True)
        return context


class MaintenanceAssignmentCreateView(LoginRequiredMixin, CreateView):
    """Neue Wartungszuweisung erstellen - für Artikelstammdaten"""
    model = MasterMaintenanceAssignment
    form_class = MasterMaintenanceAssignmentForm
    template_name = 'equipment/maintenance_assignment_form.html'
    success_url = reverse_lazy('equipment:maintenance_assignment_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(
            self.request,
            f'Wartung "{form.instance.maintenance_type.name}" wurde den Stammdaten "{form.instance.master.name}" zugewiesen.'
        )
        return super().form_valid(form)


class MaintenanceAssignmentUpdateView(LoginRequiredMixin, UpdateView):
    """Wartungszuweisung bearbeiten - für Artikelstammdaten"""
    model = MasterMaintenanceAssignment
    form_class = MasterMaintenanceAssignmentForm
    template_name = 'equipment/maintenance_assignment_form.html'
    success_url = reverse_lazy('equipment:maintenance_assignment_list')

    def form_valid(self, form):
        messages.success(
            self.request,
            f'Wartungszuweisung wurde aktualisiert.'
        )
        return super().form_valid(form)


class MaintenanceAssignmentDeleteView(LoginRequiredMixin, DeleteView):
    """Wartungszuweisung löschen - für Artikelstammdaten"""
    model = MasterMaintenanceAssignment
    template_name = 'equipment/maintenance_assignment_confirm_delete.html'
    success_url = reverse_lazy('equipment:maintenance_assignment_list')

    def delete(self, request, *args, **kwargs):
        assignment = self.get_object()
        messages.success(
            self.request,
            f'Wartungszuweisung für "{assignment.master.name}" wurde gelöscht.'
        )
        return super().delete(request, *args, **kwargs)


# ===================================================================
# STAMMDATEN-WARTUNGSZUWEISUNGEN VIEWS
# ===================================================================

class MasterMaintenanceAssignmentListView(LoginRequiredMixin, ListView):
    """Liste aller Stammdaten-Wartungszuweisungen"""
    model = MasterMaintenanceAssignment
    template_name = 'equipment/master_maintenance_assignment_list.html'
    context_object_name = 'assignments'
    paginate_by = 20

    def get_queryset(self):
        queryset = MasterMaintenanceAssignment.objects.select_related(
            'master',
            'maintenance_type',
            'created_by'
        ).annotate(
            device_count=Count('master__device_instances', filter=Q(master__device_instances__is_active=True))
        )

        # Suchfilter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(master__name__icontains=search) |
                Q(master__master_number__icontains=search) |
                Q(maintenance_type__name__icontains=search)
            )

        # Filter nach Wartungsart
        maintenance_type_id = self.request.GET.get('maintenance_type')
        if maintenance_type_id:
            queryset = queryset.filter(maintenance_type_id=maintenance_type_id)

        # Filter nach Aktiv-Status
        is_active = self.request.GET.get('is_active')
        if is_active == '1':
            queryset = queryset.filter(is_active=True)
        elif is_active == '0':
            queryset = queryset.filter(is_active=False)

        return queryset.order_by('master__name', 'maintenance_type__name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['maintenance_types'] = MaintenanceType.objects.filter(is_active=True)
        return context


class MasterMaintenanceAssignmentCreateView(LoginRequiredMixin, CreateView):
    """Neue Stammdaten-Wartungszuweisung erstellen"""
    model = MasterMaintenanceAssignment
    form_class = MasterMaintenanceAssignmentForm
    template_name = 'equipment/master_maintenance_assignment_form.html'
    success_url = reverse_lazy('equipment:master_maintenance_assignment_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(
            self.request,
            f'Wartung "{form.instance.maintenance_type.name}" wurde den Stammdaten "{form.instance.master.name}" zugewiesen.'
        )
        return super().form_valid(form)


class MasterMaintenanceAssignmentUpdateView(LoginRequiredMixin, UpdateView):
    """Stammdaten-Wartungszuweisung bearbeiten"""
    model = MasterMaintenanceAssignment
    form_class = MasterMaintenanceAssignmentForm
    template_name = 'equipment/master_maintenance_assignment_form.html'
    success_url = reverse_lazy('equipment:master_maintenance_assignment_list')

    def form_valid(self, form):
        messages.success(
            self.request,
            f'Stammdaten-Wartungszuweisung wurde aktualisiert.'
        )
        return super().form_valid(form)


class MasterMaintenanceAssignmentDeleteView(LoginRequiredMixin, DeleteView):
    """Stammdaten-Wartungszuweisung löschen"""
    model = MasterMaintenanceAssignment
    template_name = 'equipment/master_maintenance_assignment_confirm_delete.html'
    success_url = reverse_lazy('equipment:master_maintenance_assignment_list')

    def delete(self, request, *args, **kwargs):
        assignment = self.get_object()
        messages.success(
            self.request,
            f'Wartungszuweisung für "{assignment.master.name}" wurde gelöscht.'
        )
        return super().delete(request, *args, **kwargs)


class MasterMaintenanceAssignmentDetailView(LoginRequiredMixin, DetailView):
    """Detail-Ansicht einer Stammdaten-Wartungszuweisung"""
    model = MasterMaintenanceAssignment
    template_name = 'equipment/master_maintenance_assignment_detail.html'
    context_object_name = 'assignment'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Alle zugehörigen Geräte anzeigen
        context['devices'] = self.object.master.device_instances.filter(
            is_active=True
        ).select_related('location', 'assigned_vehicle')
        return context


# ===================================================================
# WARTUNGSPROTOKOLLE VIEWS
# ===================================================================

class MaintenanceRecordListView(LoginRequiredMixin, ListView):
    """Liste aller Wartungsprotokolle"""
    model = MaintenanceRecord
    template_name = 'equipment/maintenance_record_list.html'
    context_object_name = 'records'
    paginate_by = 20

    def get_queryset(self):
        queryset = MaintenanceRecord.objects.select_related(
            'assignment__device',
            'assignment__device__master',
            'assignment__maintenance_type',
            'technician'
        )

        # Suchfilter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(assignment__device__inventory_number__icontains=search) |
                Q(assignment__device__master__name__icontains=search) |
                Q(assignment__maintenance_type__name__icontains=search) |
                Q(technician__first_name__icontains=search) |
                Q(technician__last_name__icontains=search)
            )

        # Filter nach Status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        # Filter nach Techniker
        technician_id = self.request.GET.get('technician')
        if technician_id:
            queryset = queryset.filter(technician_id=technician_id)

        return queryset.order_by('-maintenance_date')


class MaintenanceRecordCreateView(LoginRequiredMixin, CreateView):
    """Neues Wartungsprotokoll erstellen"""
    model = MaintenanceRecord
    form_class = MaintenanceRecordForm
    template_name = 'equipment/maintenance_record_form.html'
    success_url = reverse_lazy('equipment:maintenance_record_list')

    def form_valid(self, form):
        messages.success(
            self.request,
            f'Wartungsprotokoll für {form.instance.assignment.device.inventory_number} wurde erstellt.'
        )
        return super().form_valid(form)


class MaintenanceRecordDetailView(LoginRequiredMixin, DetailView):
    """Wartungsprotokoll-Detailansicht"""
    model = MaintenanceRecord
    template_name = 'equipment/maintenance_record_detail.html'
    context_object_name = 'record'


# ============================================================================
# IMPORT/EXPORT
# ============================================================================

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import csv
import io
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
import logging

logger = logging.getLogger(__name__)


class ImportExportView(LoginRequiredMixin, TemplateView):
    """Import/Export-Übersicht für Equipment"""
    template_name = 'equipment/import_export.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import EquipmentItemMaster, EquipmentDeviceInstance, EquipmentStockMovement
        from inventory_base.models import Category

        context['total_masters'] = EquipmentItemMaster.objects.filter(is_active=True).count()
        context['total_devices'] = EquipmentDeviceInstance.objects.filter(is_active=True).count()
        context['total_categories'] = Category.objects.filter(is_active=True).count()
        context['total_movements'] = EquipmentStockMovement.objects.count()
        return context


# ============================================================================
# EXPORT FUNCTIONS (Excel)
# ============================================================================

@login_required
def export_categories(request):
    """Export Equipment Categories als Excel"""
    from inventory_base.models import Category
    
    categories = Category.objects.filter(is_active=True).order_by('name')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kategorien"

    headers = [
        'ID', 'Name', 'Code', 'Beschreibung', 'Übergeordnete Kategorie',
        'Sortierung', 'Erstellt am', 'Aktualisiert am'
    ]

    header_fill = PatternFill(start_color='9C27B0', end_color='9C27B0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_num, category in enumerate(categories, start=2):
        ws.cell(row=row_num, column=1, value=category.id)
        ws.cell(row=row_num, column=2, value=category.name)
        ws.cell(row=row_num, column=3, value=category.code)
        ws.cell(row=row_num, column=4, value=category.description)
        ws.cell(row=row_num, column=5, value=category.parent.name if category.parent else '')
        ws.cell(row=row_num, column=6, value=category.lft)
        ws.cell(row=row_num, column=7, value=category.created_at.strftime('%d.%m.%Y %H:%M') if category.created_at else '')
        ws.cell(row=row_num, column=8, value=category.updated_at.strftime('%d.%m.%Y %H:%M') if category.updated_at else '')

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="equipment_kategorien.xlsx"'
    wb.save(response)
    return response


@login_required
def export_masters(request):
    """Export Equipment Master Data als Excel"""
    from .models import EquipmentItemMaster
    
    masters = EquipmentItemMaster.objects.filter(is_active=True).select_related(
        'category'
    ).order_by('name')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stammdaten"

    headers = [
        'Name', 'Beschreibung', 'Kategorie-ID', 'Kategorie', 'Hersteller',
        'Modell', 'Typ', 'Barcode', 'Wartungsintervall (Tage)', 'Prüfintervall (Tage)',
        'Lebensdauer (Jahre)', 'Notizen', 'Erstellt am', 'Aktualisiert am'
    ]

    header_fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_num, master in enumerate(masters, start=2):
        ws.cell(row=row_num, column=1, value=master.name)
        ws.cell(row=row_num, column=2, value=master.description)
        ws.cell(row=row_num, column=3, value=master.category.id if master.category else '')
        ws.cell(row=row_num, column=4, value=master.category.name if master.category else '')
        ws.cell(row=row_num, column=5, value=master.manufacturer)
        ws.cell(row=row_num, column=6, value=master.model_number)
        ws.cell(row=row_num, column=7, value=master.equipment_type)
        ws.cell(row=row_num, column=8, value=master.barcode)
        ws.cell(row=row_num, column=9, value=master.maintenance_interval_days if master.maintenance_interval_days else '')
        ws.cell(row=row_num, column=10, value=master.inspection_interval_days if master.inspection_interval_days else '')
        ws.cell(row=row_num, column=11, value=master.lifespan_years if master.lifespan_years else '')
        ws.cell(row=row_num, column=12, value=master.notes)
        ws.cell(row=row_num, column=13, value=master.created_at.strftime('%d.%m.%Y %H:%M') if master.created_at else '')
        ws.cell(row=row_num, column=14, value=master.updated_at.strftime('%d.%m.%Y %H:%M') if master.updated_at else '')

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="equipment_stammdaten.xlsx"'
    wb.save(response)
    return response


@login_required
def export_devices(request):
    """Export Equipment Devices als Excel"""
    from .models import EquipmentDeviceInstance
    
    devices = EquipmentDeviceInstance.objects.filter(is_active=True).select_related(
        'master', 'location', 'assigned_vehicle'
    ).order_by('inventory_number')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Geräte-Instanzen"

    headers = [
        'Inventarnummer', 'Stammdaten (Master)', 'Seriennummer', 'Anschaffungsdatum',
        'Anschaffungspreis', 'Lagerort', 'Zugewiesenes Fahrzeug', 'Status',
        'Condition', 'Nächste Wartung', 'Nächste Prüfung', 'Notizen',
        'Erstellt am', 'Aktualisiert am'
    ]

    header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_num, device in enumerate(devices, start=2):
        ws.cell(row=row_num, column=1, value=device.inventory_number)
        ws.cell(row=row_num, column=2, value=device.master.name if device.master else '')
        ws.cell(row=row_num, column=3, value=device.serial_number)
        ws.cell(row=row_num, column=4, value=device.purchase_date.strftime('%d.%m.%Y') if device.purchase_date else '')
        ws.cell(row=row_num, column=5, value=float(device.purchase_price) if device.purchase_price else '')
        ws.cell(row=row_num, column=6, value=device.location.name if device.location else '')
        ws.cell(row=row_num, column=7, value=device.assigned_vehicle.license_plate if device.assigned_vehicle else '')
        ws.cell(row=row_num, column=8, value=device.get_status_display())
        ws.cell(row=row_num, column=9, value=device.get_condition_display())
        ws.cell(row=row_num, column=10, value=device.next_maintenance_date.strftime('%d.%m.%Y') if device.next_maintenance_date else '')
        ws.cell(row=row_num, column=11, value=device.next_inspection_date.strftime('%d.%m.%Y') if device.next_inspection_date else '')
        ws.cell(row=row_num, column=12, value=device.notes)
        ws.cell(row=row_num, column=13, value=device.created_at.strftime('%d.%m.%Y %H:%M') if device.created_at else '')
        ws.cell(row=row_num, column=14, value=device.updated_at.strftime('%d.%m.%Y %H:%M') if device.updated_at else '')

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="equipment_geraete.xlsx"'
    wb.save(response)
    return response


@login_required
def export_movements(request):
    """Export Equipment Stock Movements als Excel (letzte 1000)"""
    from .models import EquipmentStockMovement
    
    movements = EquipmentStockMovement.objects.select_related(
        'item', 'from_location', 'to_location', 'vehicle', 'person', 'created_by'
    ).order_by('-movement_date')[:1000]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bewegungen"

    headers = [
        'Datum', 'Typ', 'Gerät', 'Menge', 'Einheit',
        'Von Lagerort', 'Nach Lagerort', 'Fahrzeug', 'Person',
        'Defekt', 'Notizen', 'Erstellt von', 'Erstellt am'
    ]

    header_fill = PatternFill(start_color='3F51B5', end_color='3F51B5', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_num, movement in enumerate(movements, start=2):
        ws.cell(row=row_num, column=1, value=movement.movement_date.strftime('%d.%m.%Y') if movement.movement_date else '')
        ws.cell(row=row_num, column=2, value=movement.get_movement_type_display())
        ws.cell(row=row_num, column=3, value=movement.item.name)
        ws.cell(row=row_num, column=4, value=float(movement.quantity))
        ws.cell(row=row_num, column=5, value=movement.get_unit_display())
        ws.cell(row=row_num, column=6, value=movement.from_location.name if movement.from_location else '')
        ws.cell(row=row_num, column=7, value=movement.to_location.name if movement.to_location else '')
        ws.cell(row=row_num, column=8, value=movement.vehicle.license_plate if movement.vehicle else '')
        ws.cell(row=row_num, column=9, value=str(movement.person) if movement.person else '')
        ws.cell(row=row_num, column=10, value='Ja' if movement.is_defective else 'Nein')
        ws.cell(row=row_num, column=11, value=movement.notes)
        ws.cell(row=row_num, column=12, value=str(movement.created_by) if movement.created_by else '')
        ws.cell(row=row_num, column=13, value=movement.created_at.strftime('%d.%m.%Y %H:%M') if movement.created_at else '')

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="equipment_bewegungen.xlsx"'
    wb.save(response)
    return response


# ============================================================================
# TEMPLATE FUNCTIONS (CSV)
# ============================================================================

@login_required
def template_categories(request):
    """CSV-Vorlage für Kategorien-Import"""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')

    writer.writerow(['# Equipment Kategorien Import-Vorlage'])
    writer.writerow(['# Pflichtfelder sind mit * markiert'])
    writer.writerow(['# Zeichenkodierung: UTF-8 mit BOM'])
    writer.writerow(['# Trennzeichen: Semikolon (;)'])
    writer.writerow([])

    writer.writerow(['Name*', 'Code', 'Beschreibung', 'Übergeordnete Kategorie (Name)'])
    writer.writerow(['Pumpen', 'PUMP', 'Alle Arten von Pumpen', ''])
    writer.writerow(['Tragkraftspritzen', 'PUMP-TS', 'Tragkraftspritzen', 'Pumpen'])

    response = HttpResponse(output.getvalue().encode('utf-8-sig'), content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="equipment_kategorien_vorlage.csv"'
    return response


@login_required
def template_masters(request):
    """CSV-Vorlage für Stammdaten-Import"""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')

    writer.writerow(['# Equipment Stammdaten Import-Vorlage'])
    writer.writerow(['# Pflichtfelder sind mit * markiert'])
    writer.writerow(['# Zeichenkodierung: UTF-8 mit BOM'])
    writer.writerow(['# Trennzeichen: Semikolon (;)'])
    writer.writerow([])

    writer.writerow([
        'Name*', 'Beschreibung', 'Kategorie-ID', 'Hersteller', 'Modell',
        'Typ', 'Barcode', 'Wartungsintervall (Tage)', 'Prüfintervall (Tage)',
        'Lebensdauer (Jahre)', 'Notizen'
    ])
    writer.writerow([
        'Tragkraftspritze TS 8/8', 'Tragkraftspritze mit 800l/min', '1',
        'Rosenbauer', 'TS 8/8', 'Pumpe', '', '365', '365', '20',
        'Standard-Tragkraftspritze'
    ])

    response = HttpResponse(output.getvalue().encode('utf-8-sig'), content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="equipment_stammdaten_vorlage.csv"'
    return response


@login_required
def template_devices(request):
    """CSV-Vorlage für Geräte-Import"""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')

    writer.writerow(['# Equipment Geräte-Instanzen Import-Vorlage'])
    writer.writerow(['# Pflichtfelder sind mit * markiert'])
    writer.writerow(['# Zeichenkodierung: UTF-8 mit BOM'])
    writer.writerow(['# Trennzeichen: Semikolon (;)'])
    writer.writerow(['# Datumsformat: TT.MM.JJJJ'])
    writer.writerow([])

    writer.writerow([
        'Inventarnummer*', 'Stammdaten (Name)*', 'Seriennummer', 'Anschaffungsdatum',
        'Anschaffungspreis', 'Lagerort-ID', 'Status', 'Condition', 'Notizen'
    ])
    writer.writerow([
        'P-2025-001', 'Tragkraftspritze TS 8/8', 'SN123456', '01.01.2025',
        '5000', '1', 'operational', 'good', 'Neue Pumpe 2025'
    ])

    response = HttpResponse(output.getvalue().encode('utf-8-sig'), content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="equipment_geraete_vorlage.csv"'
    return response


# ============================================================================
# IMPORT FUNCTIONS (CSV)
# ============================================================================

@login_required
def import_categories(request):
    """Import von Equipment-Kategorien aus CSV"""
    from inventory_base.models import Category
    
    if request.method != 'POST' or not request.FILES.get('file'):
        messages.error(request, 'Keine Datei hochgeladen.')
        return redirect('equipment:import_export')

    csv_file = request.FILES['file']

    if not csv_file.name.endswith('.csv'):
        messages.error(request, 'Bitte eine CSV-Datei hochladen.')
        return redirect('equipment:import_export')

    try:
        decoded_file = csv_file.read().decode('utf-8-sig')
        reader = csv.reader(io.StringIO(decoded_file), delimiter=';')

        success_count = 0
        error_count = 0
        errors = []

        for row_num, row in enumerate(reader, start=1):
            if not row or (row[0] and row[0].startswith('#')) or row_num <= 6:
                continue

            try:
                name = row[0] if len(row) > 0 else None

                if not name:
                    errors.append(f'Zeile {row_num}: Name fehlt')
                    error_count += 1
                    continue

                if Category.objects.filter(name=name, is_active=True).exists():
                    errors.append(f'Zeile {row_num}: Kategorie "{name}" existiert bereits')
                    error_count += 1
                    continue

                parent = None
                if len(row) > 3 and row[3]:
                    try:
                        parent = Category.objects.get(name=row[3], is_active=True)
                    except Category.DoesNotExist:
                        errors.append(f'Zeile {row_num}: Übergeordnete Kategorie "{row[3]}" nicht gefunden')
                        error_count += 1
                        continue

                category = Category(
                    name=name,
                    code=row[1] if len(row) > 1 else '',
                    description=row[2] if len(row) > 2 else '',
                    parent=parent,
                    created_by=request.user,
                    updated_by=request.user
                )
                category.save()
                success_count += 1

            except Exception as e:
                errors.append(f'Zeile {row_num}: {str(e)}')
                error_count += 1
                logger.error(f"Category import error in row {row_num}: {e}")

        if success_count > 0:
            messages.success(request, f'{success_count} Kategorien erfolgreich importiert.')
        if error_count > 0:
            messages.warning(request, f'{error_count} Fehler beim Import.')
            for error in errors[:10]:
                messages.error(request, error)

    except Exception as e:
        messages.error(request, f'Fehler beim Lesen der CSV-Datei: {str(e)}')
        logger.error(f"CSV import error: {e}")

    return redirect('equipment:import_export')


@login_required
def import_masters(request):
    """Import von Equipment-Stammdaten aus CSV"""
    from .models import EquipmentItemMaster
    from inventory_base.models import Category
    
    if request.method != 'POST' or not request.FILES.get('file'):
        messages.error(request, 'Keine Datei hochgeladen.')
        return redirect('equipment:import_export')

    csv_file = request.FILES['file']

    if not csv_file.name.endswith('.csv'):
        messages.error(request, 'Bitte eine CSV-Datei hochladen.')
        return redirect('equipment:import_export')

    try:
        decoded_file = csv_file.read().decode('utf-8-sig')
        reader = csv.reader(io.StringIO(decoded_file), delimiter=';')

        success_count = 0
        error_count = 0
        errors = []

        for row_num, row in enumerate(reader, start=1):
            if not row or (row[0] and row[0].startswith('#')) or row_num <= 6:
                continue

            try:
                name = row[0] if len(row) > 0 else None

                if not name:
                    errors.append(f'Zeile {row_num}: Name fehlt')
                    error_count += 1
                    continue

                if EquipmentItemMaster.objects.filter(name=name).exists():
                    errors.append(f'Zeile {row_num}: Stammdaten "{name}" existieren bereits')
                    error_count += 1
                    continue

                category = None
                if len(row) > 2 and row[2]:
                    try:
                        category = Category.objects.get(id=int(row[2]), is_active=True)
                    except (Category.DoesNotExist, ValueError):
                        errors.append(f'Zeile {row_num}: Kategorie-ID {row[2]} nicht gefunden')
                        error_count += 1
                        continue

                master = EquipmentItemMaster(
                    name=name,
                    description=row[1] if len(row) > 1 else '',
                    category=category,
                    manufacturer=row[3] if len(row) > 3 else '',
                    model_number=row[4] if len(row) > 4 else '',
                    equipment_type=row[5] if len(row) > 5 else '',
                    barcode=row[6] if len(row) > 6 else '',
                    maintenance_interval_days=int(row[7]) if len(row) > 7 and row[7] else None,
                    inspection_interval_days=int(row[8]) if len(row) > 8 and row[8] else None,
                    lifespan_years=int(row[9]) if len(row) > 9 and row[9] else None,
                    notes=row[10] if len(row) > 10 else '',
                    created_by=request.user,
                    updated_by=request.user
                )
                master.save()
                success_count += 1

            except Exception as e:
                errors.append(f'Zeile {row_num}: {str(e)}')
                error_count += 1
                logger.error(f"Master import error in row {row_num}: {e}")

        if success_count > 0:
            messages.success(request, f'{success_count} Stammdaten erfolgreich importiert.')
        if error_count > 0:
            messages.warning(request, f'{error_count} Fehler beim Import.')
            for error in errors[:10]:
                messages.error(request, error)

    except Exception as e:
        messages.error(request, f'Fehler beim Lesen der CSV-Datei: {str(e)}')
        logger.error(f"CSV import error: {e}")

    return redirect('equipment:import_export')


@login_required
def import_devices(request):
    """Import von Equipment-Geräten aus CSV"""
    from .models import EquipmentItemMaster, EquipmentDeviceInstance
    from locations.models import Location
    from datetime import datetime
    
    if request.method != 'POST' or not request.FILES.get('file'):
        messages.error(request, 'Keine Datei hochgeladen.')
        return redirect('equipment:import_export')

    csv_file = request.FILES['file']

    if not csv_file.name.endswith('.csv'):
        messages.error(request, 'Bitte eine CSV-Datei hochladen.')
        return redirect('equipment:import_export')

    try:
        decoded_file = csv_file.read().decode('utf-8-sig')
        reader = csv.reader(io.StringIO(decoded_file), delimiter=';')

        success_count = 0
        error_count = 0
        errors = []

        for row_num, row in enumerate(reader, start=1):
            if not row or (row[0] and row[0].startswith('#')) or row_num <= 7:
                continue

            try:
                inventory_number = row[0] if len(row) > 0 else None
                master_name = row[1] if len(row) > 1 else None

                if not all([inventory_number, master_name]):
                    errors.append(f'Zeile {row_num}: Pflichtfelder fehlen')
                    error_count += 1
                    continue

                if EquipmentDeviceInstance.objects.filter(inventory_number=inventory_number).exists():
                    errors.append(f'Zeile {row_num}: Gerät {inventory_number} existiert bereits')
                    error_count += 1
                    continue

                try:
                    master = EquipmentItemMaster.objects.get(name=master_name, is_active=True)
                except EquipmentItemMaster.DoesNotExist:
                    errors.append(f'Zeile {row_num}: Stammdaten "{master_name}" nicht gefunden')
                    error_count += 1
                    continue

                location = None
                if len(row) > 5 and row[5]:
                    try:
                        location = Location.objects.get(id=int(row[5]), is_active=True)
                    except (Location.DoesNotExist, ValueError):
                        errors.append(f'Zeile {row_num}: Lagerort-ID {row[5]} nicht gefunden')
                        error_count += 1
                        continue

                purchase_date = None
                if len(row) > 3 and row[3]:
                    try:
                        purchase_date = datetime.strptime(row[3], '%d.%m.%Y').date()
                    except ValueError:
                        errors.append(f'Zeile {row_num}: Ungültiges Anschaffungsdatum {row[3]}')
                        error_count += 1
                        continue

                device = EquipmentDeviceInstance(
                    inventory_number=inventory_number,
                    master=master,
                    serial_number=row[2] if len(row) > 2 else '',
                    purchase_date=purchase_date,
                    purchase_price=float(row[4].replace(',', '.')) if len(row) > 4 and row[4] else None,
                    location=location,
                    status=row[6] if len(row) > 6 and row[6] else 'operational',
                    condition=row[7] if len(row) > 7 and row[7] else 'good',
                    notes=row[8] if len(row) > 8 else '',
                    created_by=request.user,
                    updated_by=request.user
                )
                device.save()
                success_count += 1

            except Exception as e:
                errors.append(f'Zeile {row_num}: {str(e)}')
                error_count += 1
                logger.error(f"Device import error in row {row_num}: {e}")

        if success_count > 0:
            messages.success(request, f'{success_count} Geräte erfolgreich importiert.')
        if error_count > 0:
            messages.warning(request, f'{error_count} Fehler beim Import.')
            for error in errors[:10]:
                messages.error(request, error)

    except Exception as e:
        messages.error(request, f'Fehler beim Lesen der CSV-Datei: {str(e)}')
        logger.error(f"CSV import error: {e}")

    return redirect('equipment:import_export')


# ============================================================================
# BULK LABEL PRINTING
# ============================================================================

@login_required
def device_print_labels(request):
    """Drucke Labels für mehrere Geräte"""
    from django.shortcuts import render
    
    # IDs aus Query-Parameter holen
    ids_param = request.GET.get('ids', '')
    if not ids_param:
        messages.error(request, 'Keine Geräte ausgewählt.')
        return redirect('equipment:device_list')
    
    # IDs parsen
    try:
        device_ids = [int(id.strip()) for id in ids_param.split(',') if id.strip()]
    except ValueError:
        messages.error(request, 'Ungültige Geräte-IDs.')
        return redirect('equipment:device_list')
    
    # Geräte laden
    devices = EquipmentDeviceInstance.objects.filter(
        pk__in=device_ids
    ).select_related('master')
    
    if not devices.exists():
        messages.error(request, 'Keine Geräte gefunden.')
        return redirect('equipment:device_list')
    
    # QR-Code URLs für jedes Gerät generieren
    devices_with_qr = []
    for device in devices:
        qr_url = request.build_absolute_uri(
            reverse('equipment:device_qrcode', kwargs={'pk': device.pk})
        )
        devices_with_qr.append({
            'device': device,
            'qr_url': qr_url
        })
    
    context = {
        'devices': devices_with_qr,
    }
    
    return render(request, 'equipment/device_print_labels.html', context)


@login_required
def device_batch_qrcodes(request):
    """Generiere QR-Codes für mehrere Geräte"""
    from django.shortcuts import render

    # IDs aus Query-Parameter holen
    ids_param = request.GET.get('ids', '')
    if not ids_param:
        messages.error(request, 'Keine Geräte ausgewählt.')
        return redirect('equipment:device_list')

    # IDs parsen
    try:
        device_ids = [int(id.strip()) for id in ids_param.split(',') if id.strip()]
    except ValueError:
        messages.error(request, 'Ungültige Geräte-IDs.')
        return redirect('equipment:device_list')

    # Geräte laden
    devices = EquipmentDeviceInstance.objects.filter(
        pk__in=device_ids
    ).select_related('master')

    if not devices.exists():
        messages.error(request, 'Keine Geräte gefunden.')
        return redirect('equipment:device_list')

    # QR-Code URLs für jedes Gerät generieren
    devices_with_qr = []
    for device in devices:
        qr_url = request.build_absolute_uri(
            reverse('equipment:device_qrcode', kwargs={'pk': device.pk})
        )
        devices_with_qr.append({
            'device': device,
            'qr_url': qr_url
        })

    context = {
        'devices': devices_with_qr,
    }

    return render(request, 'equipment/device_batch_qrcodes.html', context)


@login_required
def device_batch_barcodes(request):
    """Generiere Barcodes für mehrere Geräte"""
    from django.shortcuts import render

    # IDs aus Query-Parameter holen
    ids_param = request.GET.get('ids', '')
    if not ids_param:
        messages.error(request, 'Keine Geräte ausgewählt.')
        return redirect('equipment:device_list')

    # IDs parsen
    try:
        device_ids = [int(id.strip()) for id in ids_param.split(',') if id.strip()]
    except ValueError:
        messages.error(request, 'Ungültige Geräte-IDs.')
        return redirect('equipment:device_list')

    # Geräte laden
    devices = EquipmentDeviceInstance.objects.filter(
        pk__in=device_ids
    ).select_related('master')

    if not devices.exists():
        messages.error(request, 'Keine Geräte gefunden.')
        return redirect('equipment:device_list')

    # Barcode URLs für jedes Gerät generieren
    devices_with_barcode = []
    for device in devices:
        barcode_url = request.build_absolute_uri(
            reverse('equipment:device_barcode', kwargs={'pk': device.pk})
        )
        devices_with_barcode.append({
            'device': device,
            'barcode_url': barcode_url
        })

    context = {
        'devices': devices_with_barcode,
    }

    return render(request, 'equipment/device_batch_barcodes.html', context)


# ============================================================================
# BATCH VIEWS (CHARGEN)
# ============================================================================

class EquipmentBatchListView(LoginRequiredMixin, ListView):
    """Liste aller Equipment-Chargen"""
    model = EquipmentBatch
    template_name = 'equipment/batch_list.html'
    context_object_name = 'batches'
    paginate_by = 50

    def get_queryset(self):
        queryset = EquipmentBatch.objects.select_related('master', 'location').order_by('-created_at')

        # Suchfilter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(batch_number__icontains=search) |
                Q(master__name__icontains=search) |
                Q(supplier_batch_number__icontains=search)
            )

        # Status-Filter
        status = self.request.GET.get('status')
        if status == 'expired':
            queryset = [b for b in queryset if b.is_expired()]
        elif status == 'expiring_soon':
            queryset = [b for b in queryset if b.is_expiring_soon() and not b.is_expired()]
        elif status == 'recalled':
            queryset = queryset.filter(is_recalled=True)

        return queryset


class EquipmentBatchDetailView(LoginRequiredMixin, DetailView):
    """Detail-Ansicht einer Equipment-Charge"""
    model = EquipmentBatch
    template_name = 'equipment/batch_detail.html'
    context_object_name = 'batch'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Hier können weitere Kontext-Daten hinzugefügt werden
        return context


class EquipmentBatchCreateView(LoginRequiredMixin, CreateView):
    """Neue Equipment-Charge erstellen"""
    model = EquipmentBatch
    form_class = EquipmentBatchForm
    template_name = 'equipment/batch_form.html'

    def get_initial(self):
        initial = super().get_initial()
        # Master aus GET-Parameter setzen
        master_id = self.request.GET.get('master')
        if master_id:
            initial['master'] = master_id
        return initial

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Charge wurde erfolgreich erstellt.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('equipment:batch_detail', kwargs={'pk': self.object.pk})


class EquipmentBatchUpdateView(LoginRequiredMixin, UpdateView):
    """Equipment-Charge bearbeiten"""
    model = EquipmentBatch
    form_class = EquipmentBatchForm
    template_name = 'equipment/batch_form.html'

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Charge wurde erfolgreich aktualisiert.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('equipment:batch_detail', kwargs={'pk': self.object.pk})


class EquipmentBatchDeleteView(LoginRequiredMixin, DeleteView):
    """Equipment-Charge löschen"""
    model = EquipmentBatch
    template_name = 'equipment/batch_confirm_delete.html'
    success_url = reverse_lazy('equipment:master_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, _('Charge wurde erfolgreich gelöscht.'))
        return super().delete(request, *args, **kwargs)


# ============================================================================
# INVENTUR VIEWS
# ============================================================================

from .inventory_views import (
    EquipmentInventoryListView,
    EquipmentInventoryCreateView,
    EquipmentInventoryDetailView,
    EquipmentInventoryStartView,
    EquipmentInventoryCountingView,
    EquipmentInventoryCompleteView,
    EquipmentInventoryApproveView,
    EquipmentInventoryItemUpdateView,
    EquipmentInventoryProgressView,
    EquipmentInventoryExportView,
)
