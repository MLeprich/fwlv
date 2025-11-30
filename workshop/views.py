"""
Workshop Views
Views für KFZ-Werkstatt
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.db.models import Q, Count, F, Sum
from django.utils import timezone
from django.http import HttpResponse
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import csv
import io
import logging

logger = logging.getLogger(__name__)

from .models import (
    WorkshopItemMaster,
    WorkshopToolInstance,
    WorkshopToolType,
    WorkshopItem,
    WorkshopStockMovement,
    VehicleServiceRecord,
)
from .forms import WorkshopItemMasterForm, WorkshopToolInstanceForm, VehicleServiceRecordForm, WorkshopToolTypeForm


# ============================================================================
# STAMMDATEN (MASTER DATA) VIEWS
# ============================================================================

class WorkshopMasterListView(LoginRequiredMixin, ListView):
    """Liste aller Werkstatt-Artikel-Stammdaten"""
    model = WorkshopItemMaster
    template_name = 'workshop/master_list.html'
    context_object_name = 'masters'
    paginate_by = 50

    def get_queryset(self):
        queryset = WorkshopItemMaster.objects.filter(is_active=True)

        # Suchfilter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(master_number__icontains=search) |
                Q(name__icontains=search) |
                Q(manufacturer__icontains=search) |
                Q(model__icontains=search)
            )

        return queryset.order_by('master_number')


class WorkshopMasterCreateView(LoginRequiredMixin, CreateView):
    """Neuen Werkstatt-Artikel-Stammdaten erstellen"""
    model = WorkshopItemMaster
    form_class = WorkshopItemMasterForm
    template_name = 'workshop/master_form.html'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Stammdaten wurden erfolgreich erstellt.'))
        return super().form_valid(form)


class WorkshopMasterUpdateView(LoginRequiredMixin, UpdateView):
    """Werkstatt-Artikel-Stammdaten bearbeiten"""
    model = WorkshopItemMaster
    form_class = WorkshopItemMasterForm
    template_name = 'workshop/master_form.html'

    def form_valid(self, form):
        messages.success(self.request, _('Stammdaten wurden erfolgreich aktualisiert.'))
        return super().form_valid(form)


class WorkshopMasterDetailView(LoginRequiredMixin, DetailView):
    """Detailansicht Werkstatt-Artikel-Stammdaten"""
    model = WorkshopItemMaster
    template_name = 'workshop/master_detail.html'
    context_object_name = 'master'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tool_instances'] = self.object.tool_instances.filter(is_active=True).select_related('location')
        return context


# ============================================================================
# WERKZEUG-INSTANZEN (TOOL INSTANCES) VIEWS
# ============================================================================

class WorkshopToolListView(LoginRequiredMixin, ListView):
    """Liste aller Werkzeug-Instanzen"""
    model = WorkshopToolInstance
    template_name = 'workshop/tool_list.html'
    context_object_name = 'tools'
    paginate_by = 50

    def get_queryset(self):
        queryset = WorkshopToolInstance.objects.select_related('master', 'location').filter(is_active=True)

        # Suchfilter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(inventory_number__icontains=search) |
                Q(serial_number__icontains=search) |
                Q(master__name__icontains=search)
            )

        return queryset.order_by('inventory_number')


class WorkshopToolCreateView(LoginRequiredMixin, CreateView):
    """Neue Werkzeug-Instanz erstellen"""
    model = WorkshopToolInstance
    form_class = WorkshopToolInstanceForm
    template_name = 'workshop/tool_form.html'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Werkzeug-Instanz wurde erfolgreich erstellt.'))
        return super().form_valid(form)


class WorkshopToolUpdateView(LoginRequiredMixin, UpdateView):
    """Werkzeug-Instanz bearbeiten"""
    model = WorkshopToolInstance
    form_class = WorkshopToolInstanceForm
    template_name = 'workshop/tool_form.html'

    def form_valid(self, form):
        messages.success(self.request, _('Werkzeug-Instanz wurde erfolgreich aktualisiert.'))
        return super().form_valid(form)


class WorkshopToolDetailView(LoginRequiredMixin, DetailView):
    """Detailansicht Werkzeug-Instanz"""
    model = WorkshopToolInstance
    template_name = 'workshop/tool_detail.html'
    context_object_name = 'tool'


# ============================================================================
# DASHBOARD
# ============================================================================

class WorkshopDashboardView(LoginRequiredMixin, TemplateView):
    """Dashboard für KFZ-Werkstatt"""
    template_name = 'workshop/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()

        # Master Data & Tool Instance KPIs
        context['total_masters'] = WorkshopItemMaster.objects.filter(is_active=True).count()
        context['total_tools'] = WorkshopToolInstance.objects.filter(is_active=True).count()
        context['operational_tools'] = WorkshopToolInstance.objects.filter(is_active=True, is_operational=True).count()

        # KPIs - Ersatzteile (Legacy)
        context['total_items'] = WorkshopItem.objects.filter(is_active=True).count()
        context['low_stock_count'] = WorkshopItem.objects.filter(
            is_active=True,
            workshop_status='low_stock'
        ).count()
        context['hazardous_count'] = WorkshopItem.objects.filter(
            is_active=True,
            is_hazardous=True
        ).count()

        # Abgelaufene Artikel - manuell zählen da kein expiry_date Feld
        expired_items = []
        for item in WorkshopItem.objects.filter(
            is_active=True,
            production_date__isnull=False,
            shelf_life_months__isnull=False
        ):
            if item.production_date and item.shelf_life_months:
                # Berechne Ablaufdatum: production_date + shelf_life_months
                import calendar
                year = item.production_date.year
                month = item.production_date.month + item.shelf_life_months
                # Monat/Jahr Overflow behandeln
                while month > 12:
                    month -= 12
                    year += 1
                # Tag anpassen falls ungültig (z.B. 31. in Monat mit 30 Tagen)
                day = min(item.production_date.day, calendar.monthrange(year, month)[1])
                expiry = item.production_date.replace(year=year, month=month, day=day)

                if expiry <= today:
                    expired_items.append(item)
        context['expired_count'] = len(expired_items)

        # KPIs - Services
        context['total_services'] = VehicleServiceRecord.objects.count()
        context['upcoming_services_count'] = VehicleServiceRecord.objects.filter(
            scheduled_date__gte=today,
            scheduled_date__lte=today + timedelta(days=30),
            service_status__in=['scheduled', 'in_progress']
        ).count()
        context['overdue_services_count'] = VehicleServiceRecord.objects.filter(
            scheduled_date__lt=today,
            service_status__in=['scheduled', 'in_progress', 'waiting_parts']
        ).count()
        context['in_progress_count'] = VehicleServiceRecord.objects.filter(
            service_status='in_progress'
        ).count()

        # Letzte Bewegungen
        context['recent_movements'] = WorkshopStockMovement.objects.select_related(
            'item',
            'vehicle',
            'service_record',
            'created_by'
        ).order_by('-movement_date', '-created_at')[:10]

        # Anstehende Services (nächste 14 Tage)
        context['upcoming_services'] = VehicleServiceRecord.objects.filter(
            scheduled_date__gte=today,
            scheduled_date__lte=today + timedelta(days=14),
            service_status__in=['scheduled', 'in_progress']
        ).select_related('vehicle', 'technician').order_by('scheduled_date')[:5]

        # Überfällige Services
        context['overdue_services'] = VehicleServiceRecord.objects.filter(
            scheduled_date__lt=today,
            service_status__in=['scheduled', 'in_progress', 'waiting_parts']
        ).select_related('vehicle', 'technician').order_by('scheduled_date')[:5]

        # Kritische Ersatzteile (niedriger Bestand)
        context['critical_items'] = WorkshopItem.objects.filter(
            is_active=True,
            workshop_status='low_stock'
        ).select_related('location')[:5]

        # Ablaufende Artikel (nächste 30 Tage) - manuell filtern
        expiring_items = []
        for item in WorkshopItem.objects.filter(
            is_active=True,
            production_date__isnull=False,
            shelf_life_months__isnull=False
        ).select_related('location'):
            if item.production_date and item.shelf_life_months:
                # Berechne Ablaufdatum
                import calendar
                year = item.production_date.year
                month = item.production_date.month + item.shelf_life_months
                while month > 12:
                    month -= 12
                    year += 1
                day = min(item.production_date.day, calendar.monthrange(year, month)[1])
                expiry = item.production_date.replace(year=year, month=month, day=day)

                if today <= expiry <= today + timedelta(days=30):
                    item.calculated_expiry_date = expiry  # Temporär speichern für Template
                    expiring_items.append(item)
        context['expiring_items'] = sorted(expiring_items, key=lambda x: x.calculated_expiry_date)[:5]

        # Letzte Services
        context['recent_services'] = VehicleServiceRecord.objects.select_related(
            'vehicle',
            'technician'
        ).order_by('-completed_date')[:10]

        return context


# ============================================================================
# WORKSHOP ITEM VIEWS
# ============================================================================

class WorkshopItemListView(LoginRequiredMixin, ListView):
    """Liste aller Werkstatt-Artikel"""
    model = WorkshopItem
    template_name = 'workshop/item_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        queryset = WorkshopItem.objects.select_related(
            'category',
            'supplier',
            'location',
        )

        # Suchfilter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(item_number__icontains=search) |
                Q(name__icontains=search) |
                Q(manufacturer__icontains=search) |
                Q(manufacturer_part_number__icontains=search) |
                Q(oem_number__icontains=search)
            )

        # Typ-Filter
        item_type = self.request.GET.get('type')
        if item_type:
            queryset = queryset.filter(workshop_item_type=item_type)

        # Status-Filter
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(workshop_status=status)

        # Gefahrstoff-Filter
        hazardous = self.request.GET.get('hazardous')
        if hazardous == 'yes':
            queryset = queryset.filter(is_hazardous=True)

        return queryset.order_by('workshop_item_type', 'name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_count'] = self.get_queryset().count()
        context['in_stock_count'] = self.get_queryset().filter(workshop_status='in_stock').count()
        context['low_stock_count'] = self.get_queryset().filter(workshop_status='low_stock').count()
        context['hazardous_count'] = self.get_queryset().filter(is_hazardous=True).count()
        return context


class WorkshopItemDetailView(LoginRequiredMixin, DetailView):
    """Detailansicht eines Werkstatt-Artikels"""
    model = WorkshopItem
    template_name = 'workshop/item_detail.html'
    context_object_name = 'item'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['movements'] = self.object.stock_movements.select_related(
            'vehicle',
            'service_record',
        ).order_by('-movement_date')[:10]
        context['compatible_vehicles'] = self.object.compatible_vehicles.all()
        return context


class WorkshopItemCreateView(LoginRequiredMixin, CreateView):
    """Neuen Werkstatt-Artikel anlegen"""
    model = WorkshopItem
    template_name = 'workshop/item_form.html'
    fields = [
        'item_number', 'name', 'description', 'category', 'supplier',
        'workshop_item_type', 'workshop_status',
        'manufacturer', 'manufacturer_part_number', 'oem_number',
        'viscosity', 'specification',
        'quantity', 'unit', 'min_quantity', 'max_quantity',
        'shelf_life_months', 'production_date',
        'storage_temperature_min', 'storage_temperature_max',
        'is_hazardous', 'hazard_symbols',
        'recommended_change_interval_km', 'recommended_change_interval_months',
        'unit_price', 'core_charge', 'barcode', 'notes'
    ]

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Werkstatt-Artikel wurde erfolgreich erstellt.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('workshop:item_detail', kwargs={'pk': self.object.pk})


class WorkshopItemUpdateView(LoginRequiredMixin, UpdateView):
    """Werkstatt-Artikel bearbeiten"""
    model = WorkshopItem
    template_name = 'workshop/item_form.html'
    fields = [
        'name', 'description', 'category', 'supplier',
        'workshop_item_type', 'workshop_status',
        'manufacturer', 'manufacturer_part_number', 'oem_number',
        'viscosity', 'specification',
        'quantity', 'unit', 'min_quantity', 'max_quantity',
        'shelf_life_months', 'production_date',
        'storage_temperature_min', 'storage_temperature_max',
        'is_hazardous', 'hazard_symbols',
        'recommended_change_interval_km', 'recommended_change_interval_months',
        'unit_price', 'core_charge', 'notes'
    ]

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Werkstatt-Artikel wurde erfolgreich aktualisiert.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('workshop:item_detail', kwargs={'pk': self.object.pk})


class WorkshopItemDeleteView(LoginRequiredMixin, DeleteView):
    """Werkstatt-Artikel löschen"""
    model = WorkshopItem
    template_name = 'workshop/item_confirm_delete.html'
    success_url = reverse_lazy('workshop:item_list')

    def form_valid(self, form):
        item = self.get_object()
        item_name = str(item)

        # Delete related stock movements first
        movements_count = 0
        if hasattr(item, 'stock_movements'):
            movements_count = item.stock_movements.count()
            item.stock_movements.all().delete()

        response = super().form_valid(form)

        msg = f'Werkstatt-Artikel "{item_name}" wurde gelöscht'
        if movements_count:
            msg += f' (inkl. {movements_count} Bewegungen)'
        msg += '.'
        messages.success(self.request, msg)
        return response


# ============================================================================
# STOCK MOVEMENT VIEWS
# ============================================================================

class WorkshopStockMovementListView(LoginRequiredMixin, ListView):
    """Liste aller Werkstatt-Lagerbewegungen"""
    model = WorkshopStockMovement
    template_name = 'workshop/movement_list.html'
    context_object_name = 'movements'
    paginate_by = 50

    def get_queryset(self):
        queryset = WorkshopStockMovement.objects.select_related(
            'item',
            'vehicle',
            'service_record',
            'created_by',
        )

        # Typ-Filter
        movement_type = self.request.GET.get('type')
        if movement_type:
            queryset = queryset.filter(movement_type=movement_type)

        # Fahrzeug-Filter
        vehicle_id = self.request.GET.get('vehicle')
        if vehicle_id:
            queryset = queryset.filter(vehicle_id=vehicle_id)

        # Service-Filter
        service_id = self.request.GET.get('service')
        if service_id:
            queryset = queryset.filter(service_record_id=service_id)

        return queryset.order_by('-movement_date')


class WorkshopStockMovementDetailView(LoginRequiredMixin, DetailView):
    """Detailansicht einer Werkstatt-Lagerbewegung"""
    model = WorkshopStockMovement
    template_name = 'workshop/movement_detail.html'
    context_object_name = 'movement'


class WorkshopStockMovementCreateView(LoginRequiredMixin, CreateView):
    """Neue Werkstatt-Lagerbewegung erstellen"""
    model = WorkshopStockMovement
    template_name = 'workshop/movement_form.html'
    fields = [
        'movement_type', 'movement_date', 'item', 'quantity', 'unit',
        'vehicle', 'service_record', 'vehicle_mileage',
        'core_returned', 'core_return_date',
        'reference_number', 'delivery_note',
        'unit_cost', 'total_cost', 'notes'
    ]

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Werkstatt-Lagerbewegung wurde erfolgreich erstellt.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('workshop:movement_detail', kwargs={'pk': self.object.pk})


# ============================================================================
# VEHICLE SERVICE RECORD VIEWS
# ============================================================================

class VehicleServiceRecordListView(LoginRequiredMixin, ListView):
    """Liste aller Fahrzeug-Serviceeinträge"""
    model = VehicleServiceRecord
    template_name = 'workshop/service_list.html'
    context_object_name = 'services'
    paginate_by = 50

    def get_queryset(self):
        queryset = VehicleServiceRecord.objects.select_related(
            'vehicle',
            'technician',
            'created_by',
        )

        # Service-Typ-Filter
        service_type = self.request.GET.get('type')
        if service_type:
            queryset = queryset.filter(service_type=service_type)

        # Status-Filter
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(service_status=status)

        # Fahrzeug-Filter
        vehicle_id = self.request.GET.get('vehicle')
        if vehicle_id:
            queryset = queryset.filter(vehicle_id=vehicle_id)

        # Techniker-Filter
        technician_id = self.request.GET.get('technician')
        if technician_id:
            queryset = queryset.filter(technician_id=technician_id)

        return queryset.order_by('-scheduled_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_count'] = self.get_queryset().count()
        context['scheduled_count'] = self.get_queryset().filter(service_status='scheduled').count()
        context['in_progress_count'] = self.get_queryset().filter(service_status='in_progress').count()
        context['completed_count'] = self.get_queryset().filter(service_status='completed').count()
        return context


class VehicleServiceRecordDetailView(LoginRequiredMixin, DetailView):
    """Detailansicht eines Fahrzeug-Serviceeintrags"""
    model = VehicleServiceRecord
    template_name = 'workshop/service_detail.html'
    context_object_name = 'service'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['used_items'] = self.object.used_items.select_related('item').all()
        return context


class VehicleServiceRecordCreateView(LoginRequiredMixin, CreateView):
    """Neuen Fahrzeug-Serviceeintrag erstellen"""
    model = VehicleServiceRecord
    form_class = VehicleServiceRecordForm
    template_name = 'workshop/service_form.html'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Fahrzeug-Serviceeintrag wurde erfolgreich erstellt.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('workshop:service_detail', kwargs={'pk': self.object.pk})


class VehicleServiceRecordUpdateView(LoginRequiredMixin, UpdateView):
    """Fahrzeug-Serviceeintrag bearbeiten"""
    model = VehicleServiceRecord
    form_class = VehicleServiceRecordForm
    template_name = 'workshop/service_form.html'

    def form_valid(self, form):
        messages.success(self.request, _('Fahrzeug-Serviceeintrag wurde erfolgreich aktualisiert.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('workshop:service_detail', kwargs={'pk': self.object.pk})


class VehicleServiceRecordDeleteView(LoginRequiredMixin, DeleteView):
    """Fahrzeug-Serviceeintrag löschen"""
    model = VehicleServiceRecord
    template_name = 'workshop/service_confirm_delete.html'
    success_url = reverse_lazy('workshop:service_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, _('Fahrzeug-Serviceeintrag wurde gelöscht.'))
        return super().delete(request, *args, **kwargs)


# ============================================================================
# SPECIAL VIEWS
# ============================================================================

class HazardousItemsListView(LoginRequiredMixin, ListView):
    """Liste aller Gefahrstoffe"""
    model = WorkshopItem
    template_name = 'workshop/hazardous_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        return WorkshopItem.objects.filter(
            is_hazardous=True
        ).select_related(
            'location'
        ).order_by('workshop_item_type', 'name')


class LowStockItemsListView(LoginRequiredMixin, ListView):
    """Liste aller Artikel mit niedrigem Bestand"""
    model = WorkshopItem
    template_name = 'workshop/low_stock_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        return WorkshopItem.objects.filter(
            workshop_status='low_stock'
        ).select_related(
            'location'
        ).order_by('workshop_item_type', 'name')


class ExpiredItemsListView(LoginRequiredMixin, ListView):
    """Liste aller abgelaufenen Artikel"""
    model = WorkshopItem
    template_name = 'workshop/expired_list.html'
    context_object_name = 'items'

    def get_queryset(self):
        today = timezone.now().date()
        expired_items = []

        # Manuell filtern da kein expiry_date Feld vorhanden
        for item in WorkshopItem.objects.filter(
            is_active=True,
            production_date__isnull=False,
            shelf_life_months__isnull=False
        ).select_related('location'):
            if item.production_date and item.shelf_life_months:
                import calendar
                year = item.production_date.year
                month = item.production_date.month + item.shelf_life_months
                while month > 12:
                    month -= 12
                    year += 1
                day = min(item.production_date.day, calendar.monthrange(year, month)[1])
                expiry = item.production_date.replace(year=year, month=month, day=day)

                if expiry <= today:
                    item.calculated_expiry_date = expiry
                    expired_items.append(item)

        return sorted(expired_items, key=lambda x: x.calculated_expiry_date)


class UpcomingServicesView(LoginRequiredMixin, ListView):
    """Anstehende Services (nächste 30 Tage)"""
    model = VehicleServiceRecord
    template_name = 'workshop/upcoming_services.html'
    context_object_name = 'services'

    def get_queryset(self):
        today = timezone.now().date()
        end_date = today + timezone.timedelta(days=30)

        return VehicleServiceRecord.objects.filter(
            scheduled_date__gte=today,
            scheduled_date__lte=end_date,
            service_status__in=['scheduled', 'in_progress']
        ).select_related(
            'vehicle',
            'technician'
        ).order_by('scheduled_date')


class OverdueServicesView(LoginRequiredMixin, ListView):
    """Überfällige Services"""
    model = VehicleServiceRecord
    template_name = 'workshop/overdue_services.html'
    context_object_name = 'services'

    def get_queryset(self):
        today = timezone.now().date()

        return VehicleServiceRecord.objects.filter(
            scheduled_date__lt=today,
            service_status__in=['scheduled', 'in_progress', 'waiting_parts']
        ).select_related(
            'vehicle',
            'technician'
        ).order_by('scheduled_date')


class VehicleServicesView(LoginRequiredMixin, ListView):
    """Service-Historie eines Fahrzeugs"""
    model = VehicleServiceRecord
    template_name = 'workshop/vehicle_services.html'
    context_object_name = 'services'

    def get_queryset(self):
        vehicle_id = self.kwargs.get('vehicle_id')
        return VehicleServiceRecord.objects.filter(
            vehicle_id=vehicle_id
        ).select_related(
            'technician'
        ).order_by('-scheduled_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        from vehicles.models import Vehicle
        context['vehicle'] = get_object_or_404(Vehicle, pk=self.kwargs.get('vehicle_id'))

        return context


# ============================================================================
# TYPE MANAGEMENT VIEWS (Werkzeug-/Artikeltypen)
# ============================================================================

class WorkshopToolTypeListView(LoginRequiredMixin, ListView):
    """Liste aller Werkzeug-/Artikeltypen"""
    model = WorkshopToolType
    template_name = 'workshop/tool_type_list.html'
    context_object_name = 'tool_types'

    def get_queryset(self):
        return WorkshopToolType.objects.all().order_by('category', 'sort_order', 'name')


class WorkshopToolTypeCreateView(LoginRequiredMixin, CreateView):
    """Werkzeug-/Artikeltyp erstellen"""
    model = WorkshopToolType
    form_class = WorkshopToolTypeForm
    template_name = 'workshop/tool_type_form.html'
    success_url = reverse_lazy('workshop:type_management')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Neuen Werkzeug-/Artikeltyp anlegen'
        context['submit_text'] = 'Typ speichern'
        return context

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, _('Werkzeug-/Artikeltyp erfolgreich erstellt.'))
        return super().form_valid(form)


class WorkshopToolTypeUpdateView(LoginRequiredMixin, UpdateView):
    """Werkzeug-/Artikeltyp bearbeiten"""
    model = WorkshopToolType
    form_class = WorkshopToolTypeForm
    template_name = 'workshop/tool_type_form.html'
    success_url = reverse_lazy('workshop:type_management')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Werkzeug-/Artikeltyp bearbeiten'
        context['submit_text'] = 'Änderungen speichern'
        return context

    def form_valid(self, form):
        messages.success(self.request, _('Werkzeug-/Artikeltyp erfolgreich aktualisiert.'))
        return super().form_valid(form)


class WorkshopToolTypeDeleteView(LoginRequiredMixin, DeleteView):
    """Werkzeug-/Artikeltyp löschen"""
    model = WorkshopToolType
    template_name = 'workshop/tool_type_confirm_delete.html'
    success_url = reverse_lazy('workshop:type_management')

    def delete(self, request, *args, **kwargs):
        messages.success(request, _('Werkzeug-/Artikeltyp erfolgreich gelöscht.'))
        return super().delete(request, *args, **kwargs)


# ============================================================================
# QR-CODE & BARCODE VIEWS
# ============================================================================

@login_required
def item_qrcode_view(request, pk):
    """
    QR-Code für Werkstatt-Artikel generieren und anzeigen
    """
    from django.http import HttpResponse

    item = get_object_or_404(WorkshopItem, pk=pk)

    # QR-Code als SVG generieren
    qr_svg = item.generate_qr_code()

    # Als HttpResponse mit SVG-Content-Type
    response = HttpResponse(qr_svg, content_type='image/svg+xml')
    response['Content-Disposition'] = f'inline; filename="qrcode_item_{item.item_number}.svg"'

    return response


@login_required
def item_barcode_view(request, pk):
    """
    Barcode für Werkstatt-Artikel generieren und anzeigen
    """
    from django.http import HttpResponse

    item = get_object_or_404(WorkshopItem, pk=pk)

    # Barcode als SVG generieren
    barcode_svg = item.generate_barcode()

    # Als HttpResponse mit SVG-Content-Type
    response = HttpResponse(barcode_svg, content_type='image/svg+xml')
    response['Content-Disposition'] = f'inline; filename="barcode_item_{item.item_number}.svg"'

    return response


# ============================================================================
# KALENDER VIEWS
# ============================================================================

@login_required
def workshop_calendar(request):
    """
    Kalenderansicht für Fahrzeug-Wartungen
    Zeigt wann welche Wartung geplant ist
    """
    from calendar import monthcalendar
    from datetime import datetime

    # Monat und Jahr aus Request oder aktuell
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))
    selected_day = request.GET.get('day')

    # Kalender generieren
    cal = monthcalendar(year, month)

    # Zeitraum für diesen Monat
    month_start = datetime(year, month, 1).date()
    if month == 12:
        month_end = datetime(year + 1, 1, 1).date()
    else:
        month_end = datetime(year, month + 1, 1).date()

    today = timezone.now().date()

    # Alle Wartungen mit geplanten Terminen in diesem Monat
    services_by_day = {}

    # Wartungen sammeln
    services = VehicleServiceRecord.objects.filter(
        scheduled_date__gte=month_start,
        scheduled_date__lt=month_end
    ).select_related('vehicle', 'technician').order_by('scheduled_date')

    for service in services:
        day = service.scheduled_date.day
        if day not in services_by_day:
            services_by_day[day] = []

        is_overdue = service.scheduled_date < today and service.service_status not in ['completed', 'cancelled']
        is_urgent = service.scheduled_date <= today + timedelta(days=7) and not is_overdue and service.service_status not in ['completed', 'cancelled']

        # Farbe basierend auf Status
        if service.service_status == 'completed':
            color = 'green'
        elif service.service_status == 'cancelled':
            color = 'gray'
        elif is_overdue:
            color = 'red'
        elif is_urgent:
            color = 'orange'
        elif service.service_status == 'in_progress':
            color = 'blue'
        else:
            color = 'purple'

        services_by_day[day].append({
            'service': service,
            'type': service.get_service_type_display(),
            'date': service.scheduled_date,
            'is_overdue': is_overdue,
            'is_urgent': is_urgent,
            'color': color
        })

    # Ausgewählter Tag - Standard ist heute (falls im aktuellen Monat)
    if selected_day:
        selected_day = int(selected_day)
    else:
        if today.year == year and today.month == month:
            selected_day = today.day
        else:
            # Oder den ersten Tag mit Wartungen
            selected_day = min(services_by_day.keys()) if services_by_day else 1

    # Wartungen für den ausgewählten Tag
    selected_date = datetime(year, month, selected_day).date()
    selected_services = services_by_day.get(selected_day, [])

    # Sortiere nach Datum und Status
    selected_services.sort(key=lambda x: (x['date'], x['service'].service_status))

    # Navigation
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1

    context = {
        'calendar': cal,
        'year': year,
        'month': month,
        'month_name': datetime(year, month, 1).strftime('%B'),
        'services_by_day': services_by_day,
        'selected_day': selected_day,
        'selected_date': selected_date,
        'selected_services': selected_services,
        'prev_month': prev_month,
        'prev_year': prev_year,
        'next_month': next_month,
        'next_year': next_year,
    }

    # HTMX Request - nur die Tagesdetails zurückgeben
    if request.headers.get('HX-Request'):
        return render(request, 'workshop/partials/calendar_day_detail.html', context)

    return render(request, 'workshop/calendar.html', context)


# ============================================================================
# IMPORT/EXPORT
# ============================================================================

class ImportExportView(LoginRequiredMixin, TemplateView):
    """Import/Export Übersicht"""
    template_name = 'workshop/import_export.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_masters'] = WorkshopItemMaster.objects.filter(is_active=True).count()
        context['total_tools'] = WorkshopToolInstance.objects.filter(is_active=True).count()
        context['total_movements'] = WorkshopStockMovement.objects.count()
        context['total_services'] = VehicleServiceRecord.objects.count()
        return context


# ============================================================================
# EXPORT FUNCTIONS (Excel)
# ============================================================================

@login_required
def export_masters(request):
    """Export Workshop Masters als Excel"""
    masters = WorkshopItemMaster.objects.filter(is_active=True).select_related(
        'created_by', 'updated_by'
    ).order_by('master_number')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stammdaten"

    # Header
    headers = [
        'Stammdaten-Nr.', 'Name', 'Typ', 'Hersteller', 'Modell', 'Herstellerartikelnr.',
        'Zertifizierungen', 'Gewicht (kg)', 'Gefahrstoff', 'Gefahrstoffklasse',
        'Beschreibung', 'Erstellt am', 'Aktualisiert am'
    ]

    # Header-Styling
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Daten
    for row_num, master in enumerate(masters, start=2):
        ws.cell(row=row_num, column=1, value=master.master_number)
        ws.cell(row=row_num, column=2, value=master.name)
        ws.cell(row=row_num, column=3, value=master.item_type.name if hasattr(master, 'item_type') and master.item_type else '')
        ws.cell(row=row_num, column=4, value=master.manufacturer)
        ws.cell(row=row_num, column=5, value=master.model)
        ws.cell(row=row_num, column=6, value=master.manufacturer_part_number if hasattr(master, 'manufacturer_part_number') else '')
        ws.cell(row=row_num, column=7, value=', '.join(master.certifications) if hasattr(master, 'certifications') and master.certifications else '')
        ws.cell(row=row_num, column=8, value=float(master.weight_kg) if hasattr(master, 'weight_kg') and master.weight_kg else '')
        ws.cell(row=row_num, column=9, value='Ja' if hasattr(master, 'is_hazardous') and master.is_hazardous else 'Nein')
        ws.cell(row=row_num, column=10, value=master.hazardous_class if hasattr(master, 'hazardous_class') and master.hazardous_class else '')
        ws.cell(row=row_num, column=11, value=master.description)
        ws.cell(row=row_num, column=12, value=master.created_at.strftime('%d.%m.%Y %H:%M') if master.created_at else '')
        ws.cell(row=row_num, column=13, value=master.updated_at.strftime('%d.%m.%Y %H:%M') if master.updated_at else '')

    # Spaltenbreiten
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=werkstatt_stammdaten_export.xlsx'
    wb.save(response)
    return response


@login_required
def export_tools(request):
    """Export Workshop Tools als Excel"""
    tools = WorkshopToolInstance.objects.filter(is_active=True).select_related(
        'master', 'location', 'assigned_vehicle', 'assigned_to', 'created_by', 'updated_by'
    ).order_by('inventory_number')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Werkzeug-Instanzen"

    # Header
    headers = [
        'Inventarnummer', 'Seriennummer', 'Stammdaten', 'Lagerort', 'Fahrzeug', 'Zugeordnet an',
        'Zustand', 'Einsatzbereit', 'Herstellungsdatum', 'Letzte Kalibrierung', 'Nächste Kalibrierung',
        'Notizen', 'Erstellt am', 'Aktualisiert am'
    ]

    # Header-Styling
    header_fill = PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Daten
    for row_num, tool in enumerate(tools, start=2):
        ws.cell(row=row_num, column=1, value=tool.inventory_number)
        ws.cell(row=row_num, column=2, value=tool.serial_number)
        ws.cell(row=row_num, column=3, value=tool.master.name if tool.master else '')
        ws.cell(row=row_num, column=4, value=tool.location.name if tool.location else '')
        ws.cell(row=row_num, column=5, value=str(tool.assigned_vehicle) if tool.assigned_vehicle else '')
        ws.cell(row=row_num, column=6, value=tool.assigned_to.get_full_name() if tool.assigned_to else '')
        ws.cell(row=row_num, column=7, value=tool.get_condition_display() if tool.condition and hasattr(tool, 'get_condition_display') else tool.condition if hasattr(tool, 'condition') and tool.condition else '')
        ws.cell(row=row_num, column=8, value='Ja' if tool.is_operational else 'Nein')
        ws.cell(row=row_num, column=9, value=tool.manufacturing_date.strftime('%d.%m.%Y') if hasattr(tool, 'manufacturing_date') and tool.manufacturing_date else '')
        ws.cell(row=row_num, column=10, value=tool.last_calibration_date.strftime('%d.%m.%Y') if hasattr(tool, 'last_calibration_date') and tool.last_calibration_date else '')
        ws.cell(row=row_num, column=11, value=tool.next_calibration_date.strftime('%d.%m.%Y') if hasattr(tool, 'next_calibration_date') and tool.next_calibration_date else '')
        ws.cell(row=row_num, column=12, value=tool.notes if hasattr(tool, 'notes') else '')
        ws.cell(row=row_num, column=13, value=tool.created_at.strftime('%d.%m.%Y %H:%M') if tool.created_at else '')
        ws.cell(row=row_num, column=14, value=tool.updated_at.strftime('%d.%m.%Y %H:%M') if tool.updated_at else '')

    # Spaltenbreiten
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=werkstatt_werkzeuge_export.xlsx'
    wb.save(response)
    return response


@login_required
def export_services(request):
    """Export Vehicle Service Records als Excel"""
    services = VehicleServiceRecord.objects.select_related(
        'vehicle', 'technician'
    ).order_by('-service_date')[:1000]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fahrzeug-Wartungen"

    # Header
    headers = [
        'ID', 'Wartungsdatum', 'Fahrzeug', 'Techniker', 'Wartungstyp', 'Kilometerstand',
        'Nächster Service (km)', 'Nächster Service (Datum)', 'Kosten', 'Beschreibung', 'Erstellt am'
    ]

    # Header-Styling
    header_fill = PatternFill(start_color='9C27B0', end_color='9C27B0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Daten
    for row_num, service in enumerate(services, start=2):
        ws.cell(row=row_num, column=1, value=service.id)
        ws.cell(row=row_num, column=2, value=service.service_date.strftime('%d.%m.%Y') if service.service_date else '')
        ws.cell(row=row_num, column=3, value=str(service.vehicle) if service.vehicle else '')
        ws.cell(row=row_num, column=4, value=service.technician.get_full_name() if service.technician else '')
        ws.cell(row=row_num, column=5, value=service.service_type if hasattr(service, 'service_type') else '')
        ws.cell(row=row_num, column=6, value=service.mileage if hasattr(service, 'mileage') else '')
        ws.cell(row=row_num, column=7, value=service.next_service_mileage if hasattr(service, 'next_service_mileage') else '')
        ws.cell(row=row_num, column=8, value=service.next_service_date.strftime('%d.%m.%Y') if hasattr(service, 'next_service_date') and service.next_service_date else '')
        ws.cell(row=row_num, column=9, value=float(service.cost) if hasattr(service, 'cost') and service.cost else '')
        ws.cell(row=row_num, column=10, value=service.description if hasattr(service, 'description') else '')
        ws.cell(row=row_num, column=11, value=service.created_at.strftime('%d.%m.%Y %H:%M') if service.created_at else '')

    # Spaltenbreiten
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=werkstatt_wartungen_export.xlsx'
    wb.save(response)
    return response


# ============================================================================
# TEMPLATE DOWNLOAD FUNCTIONS (CSV)
# ============================================================================

@login_required
def template_masters(request):
    """CSV-Vorlage für Stammdaten-Import"""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)

    # Header
    headers = [
        'Stammdaten-Nr.*', 'Name*', 'Typ*', 'Hersteller', 'Modell',
        'Herstellerartikelnr.', 'Gewicht (kg)', 'Gefahrstoff (Ja/Nein)',
        'Gefahrstoffklasse', 'Beschreibung'
    ]
    writer.writerow(headers)

    # Beispielzeile
    example_data = [
        'WS-2025-001', 'Drehmomentschlüssel 40-200 Nm', 'torque_wrench', 'Hazet', '6122-1CT',
        '6122-1CT', '1.2', 'Nein', '', 'Beispiel: Professioneller Drehmomentschlüssel - diese Zeile löschen!'
    ]
    writer.writerow(example_data)

    # Hinweise
    writer.writerow([])
    writer.writerow(['# HINWEISE:'])
    writer.writerow(['# Felder mit * sind Pflichtfelder'])
    writer.writerow(['# Typ: Verwenden Sie einen gültigen Typ-Code aus der Typenverwaltung'])
    writer.writerow(['# Gefahrstoff: Ja oder Nein'])

    # Response
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename=vorlage_werkstatt_stammdaten_import.csv'
    response.write('\ufeff')  # UTF-8 BOM
    response.write(output.getvalue())
    return response


@login_required
def template_tools(request):
    """CSV-Vorlage für Werkzeug-Import"""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)

    # Header
    headers = [
        'Inventarnummer*', 'Seriennummer*', 'Stammdaten-Nr.*', 'Lagerort-ID*',
        'Herstellungsdatum', 'Zustand', 'Einsatzbereit (Ja/Nein)', 'Notizen'
    ]
    writer.writerow(headers)

    # Beispielzeile
    example_data = [
        'INV-WS-001', 'SN123456', 'WS-2025-001', '1',
        '2024-01-15', 'good', 'Ja', 'Beispiel: Neues Werkzeug - diese Zeile löschen!'
    ]
    writer.writerow(example_data)

    # Hinweise
    writer.writerow([])
    writer.writerow(['# HINWEISE:'])
    writer.writerow(['# Felder mit * sind Pflichtfelder'])
    writer.writerow(['# Zustand: excellent, good, fair, poor, damaged'])
    writer.writerow(['# Einsatzbereit: Ja oder Nein'])
    writer.writerow(['# Lagerort-ID muss existieren'])
    writer.writerow(['# Stammdaten-Nr. muss existieren'])

    # Response
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename=vorlage_werkstatt_werkzeuge_import.csv'
    response.write('\ufeff')  # UTF-8 BOM
    response.write(output.getvalue())
    return response


# ============================================================================
# IMPORT FUNCTIONS (CSV)
# ============================================================================

@login_required
def import_masters(request):
    """Import Workshop Masters aus CSV"""
    if request.method != 'POST':
        messages.error(request, 'Nur POST-Anfragen erlaubt.')
        return redirect('workshop:import_export')

    if 'file' not in request.FILES:
        messages.error(request, 'Keine Datei hochgeladen.')
        return redirect('workshop:import_export')

    uploaded_file = request.FILES['file']

    try:
        # CSV-Datei lesen (UTF-8 mit BOM)
        decoded_file = uploaded_file.read().decode('utf-8-sig')
        csv_reader = csv.reader(io.StringIO(decoded_file), delimiter=';')

        success_count = 0
        error_count = 0
        errors = []

        # Header überspringen
        next(csv_reader, None)

        # Zeilen durchgehen
        for row_num, row in enumerate(csv_reader, start=2):
            # Leere Zeilen und Kommentare überspringen
            if not row or not row[0] or row[0].startswith('#'):
                continue

            # Beispielzeile überspringen
            if 'Beispiel' in str(row[-1] if row else ''):
                continue

            try:
                # Daten extrahieren (mit Index-Safety)
                master_number = row[0] if len(row) > 0 else None
                name = row[1] if len(row) > 1 else None
                item_type_code = row[2] if len(row) > 2 else None
                manufacturer = row[3] if len(row) > 3 else ''
                model = row[4] if len(row) > 4 else ''
                manufacturer_part_number = row[5] if len(row) > 5 else ''
                weight_kg = row[6] if len(row) > 6 else None
                is_hazardous_str = row[7] if len(row) > 7 else 'Nein'
                hazardous_class = row[8] if len(row) > 8 else ''
                description = row[9] if len(row) > 9 else ''

                # Validierung
                if not all([master_number, name, item_type_code]):
                    errors.append(f"Zeile {row_num}: Pflichtfelder fehlen")
                    error_count += 1
                    continue

                # Duplikatprüfung
                if WorkshopItemMaster.objects.filter(master_number=master_number).exists():
                    errors.append(f"Zeile {row_num}: Stammdaten-Nr. '{master_number}' existiert bereits")
                    error_count += 1
                    continue

                # Typ laden
                try:
                    item_type = WorkshopToolType.objects.get(code=item_type_code)
                except WorkshopToolType.DoesNotExist:
                    errors.append(f"Zeile {row_num}: Typ-Code '{item_type_code}' nicht gefunden")
                    error_count += 1
                    continue

                # Boolean-Konvertierung
                is_hazardous = is_hazardous_str.lower() in ['ja', 'yes', 'true', '1']

                # Master erstellen
                master = WorkshopItemMaster.objects.create(
                    master_number=master_number,
                    name=name,
                    item_type=item_type,
                    manufacturer=manufacturer,
                    model=model,
                    manufacturer_part_number=manufacturer_part_number if hasattr(WorkshopItemMaster, 'manufacturer_part_number') else None,
                    weight_kg=float(weight_kg) if weight_kg and hasattr(WorkshopItemMaster, 'weight_kg') else None,
                    is_hazardous=is_hazardous if hasattr(WorkshopItemMaster, 'is_hazardous') else False,
                    hazardous_class=hazardous_class if hazardous_class and hasattr(WorkshopItemMaster, 'hazardous_class') else None,
                    description=description,
                    created_by=request.user,
                    updated_by=request.user
                )

                success_count += 1

            except Exception as e:
                errors.append(f"Zeile {row_num}: Fehler - {str(e)}")
                error_count += 1
                logger.error(f"Master import error in row {row_num}: {e}")

        # Erfolgsmeldung
        if success_count > 0:
            messages.success(request, f"✅ {success_count} Stammdaten erfolgreich importiert.")

        if error_count > 0:
            messages.warning(request, f"⚠️ {error_count} Fehler beim Import.")
            for error in errors[:10]:
                messages.error(request, error)
            if len(errors) > 10:
                messages.error(request, f"... und {len(errors) - 10} weitere Fehler")

    except Exception as e:
        messages.error(request, f"Fehler beim Lesen der CSV-Datei: {str(e)}")
        logger.error(f"CSV import error: {e}")

    return redirect('workshop:import_export')


@login_required
def import_tools(request):
    """Import Workshop Tools aus CSV"""
    if request.method != 'POST':
        messages.error(request, 'Nur POST-Anfragen erlaubt.')
        return redirect('workshop:import_export')

    if 'file' not in request.FILES:
        messages.error(request, 'Keine Datei hochgeladen.')
        return redirect('workshop:import_export')

    uploaded_file = request.FILES['file']

    try:
        # CSV-Datei lesen (UTF-8 mit BOM)
        decoded_file = uploaded_file.read().decode('utf-8-sig')
        csv_reader = csv.reader(io.StringIO(decoded_file), delimiter=';')

        success_count = 0
        error_count = 0
        errors = []

        # Header überspringen
        next(csv_reader, None)

        # Zeilen durchgehen
        for row_num, row in enumerate(csv_reader, start=2):
            # Leere Zeilen und Kommentare überspringen
            if not row or not row[0] or row[0].startswith('#'):
                continue

            # Beispielzeile überspringen
            if 'Beispiel' in str(row[-1] if row else ''):
                continue

            try:
                # Daten extrahieren (mit Index-Safety)
                inventory_number = row[0] if len(row) > 0 else None
                serial_number = row[1] if len(row) > 1 else None
                master_number = row[2] if len(row) > 2 else None
                location_id = row[3] if len(row) > 3 else None
                manufacturing_date_str = row[4] if len(row) > 4 else ''
                condition = row[5] if len(row) > 5 else 'good'
                is_operational_str = row[6] if len(row) > 6 else 'Ja'
                notes = row[7] if len(row) > 7 else ''

                # Validierung
                if not all([inventory_number, serial_number, master_number, location_id]):
                    errors.append(f"Zeile {row_num}: Pflichtfelder fehlen")
                    error_count += 1
                    continue

                # Duplikatprüfung
                if WorkshopToolInstance.objects.filter(inventory_number=inventory_number).exists():
                    errors.append(f"Zeile {row_num}: Inventarnummer '{inventory_number}' existiert bereits")
                    error_count += 1
                    continue

                # Master laden
                try:
                    master = WorkshopItemMaster.objects.get(master_number=master_number)
                except WorkshopItemMaster.DoesNotExist:
                    errors.append(f"Zeile {row_num}: Stammdaten-Nr. {master_number} nicht gefunden")
                    error_count += 1
                    continue

                # Standort laden
                from locations.models import Location
                try:
                    location = Location.objects.get(id=int(location_id))
                except (Location.DoesNotExist, ValueError):
                    errors.append(f"Zeile {row_num}: Standort-ID {location_id} nicht gefunden")
                    error_count += 1
                    continue

                # Datum parsen
                manufacturing_date = None
                if manufacturing_date_str:
                    from datetime import datetime
                    try:
                        manufacturing_date = datetime.strptime(manufacturing_date_str, '%Y-%m-%d').date()
                    except ValueError:
                        try:
                            manufacturing_date = datetime.strptime(manufacturing_date_str, '%d.%m.%Y').date()
                        except ValueError:
                            pass

                # Boolean-Konvertierung
                is_operational = is_operational_str.lower() in ['ja', 'yes', 'true', '1']

                # Tool erstellen
                tool = WorkshopToolInstance.objects.create(
                    inventory_number=inventory_number,
                    serial_number=serial_number,
                    master=master,
                    location=location,
                    manufacturing_date=manufacturing_date if hasattr(WorkshopToolInstance, 'manufacturing_date') else None,
                    condition=condition if hasattr(WorkshopToolInstance, 'condition') else None,
                    is_operational=is_operational,
                    notes=notes if hasattr(WorkshopToolInstance, 'notes') else None,
                    created_by=request.user,
                    updated_by=request.user
                )

                success_count += 1

            except Exception as e:
                errors.append(f"Zeile {row_num}: Fehler - {str(e)}")
                error_count += 1
                logger.error(f"Tool import error in row {row_num}: {e}")

        # Erfolgsmeldung
        if success_count > 0:
            messages.success(request, f"✅ {success_count} Werkzeuge erfolgreich importiert.")

        if error_count > 0:
            messages.warning(request, f"⚠️ {error_count} Fehler beim Import.")
            for error in errors[:10]:
                messages.error(request, error)
            if len(errors) > 10:
                messages.error(request, f"... und {len(errors) - 10} weitere Fehler")

    except Exception as e:
        messages.error(request, f"Fehler beim Lesen der CSV-Datei: {str(e)}")
        logger.error(f"CSV import error: {e}")

    return redirect('workshop:import_export')
