"""
Clothing Views
Views für Kleiderkammer-Verwaltung
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.db.models import Q, Count, F, Sum
from django.utils import timezone
from django.http import HttpResponse
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import copy
import csv
import io
import logging

logger = logging.getLogger(__name__)

from .assignments import (
    annotate_issues,
    issued_item_ids,
    issued_items_for_person,
    open_issues,
    total_issued_quantity,
)

from .models import (
    ClothingItem,
    ClothingCategory,
    ClothingStockMovement,
    ClothingSizeAssignment,
    ClothingItemMaster,
    ClothingItemInstance,
    ClothingProductType,
)
from .forms import (
    ClothingItemForm,
    ClothingStockMovementForm,
    ClothingItemMasterForm,
    ClothingItemInstanceForm,
)


# ============================================================================
# DASHBOARD
# ============================================================================

class ClothingDashboardView(LoginRequiredMixin, TemplateView):
    """
    Kleiderkammer-Dashboard mit Übersicht und KPIs
    """
    template_name = 'clothing/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()

        # KPIs
        context['total_items'] = ClothingItem.objects.filter(is_active=True).count()
        context['psa_items'] = ClothingItem.objects.filter(is_active=True, is_psa=True).count()
        context['low_stock_count'] = ClothingItem.objects.filter(
            is_active=True,
            min_quantity__isnull=False,
            quantity__lt=F('min_quantity')
        ).count()

        # Ablaufende Zertifikate (PSA - nächste 90 Tage)
        threshold_90 = today + timedelta(days=90)
        context['expiring_certifications'] = ClothingItem.objects.filter(
            is_psa=True,
            is_active=True,
            certification_expires__lte=threshold_90,
            certification_expires__gte=today
        ).count()

        # Fällige Prüfungen
        context['inspections_due'] = ClothingItem.objects.filter(
            is_active=True,
            requires_inspection=True,
            next_inspection_date__isnull=False,
            next_inspection_date__lte=today
        ).count()

        # Persönlich ausgegebene Teile (aus den Lagerbewegungen abgeleitet)
        context['assigned_items'] = int(total_issued_quantity())

        # Pool: Bestand, der noch im Lager liegt
        context['pool_items'] = int(
            ClothingItem.objects.filter(is_active=True).aggregate(
                bestand=Sum('quantity')
            )['bestand'] or 0
        )

        # Stammdaten- und Exemplar-Anzahl
        context['master_count'] = ClothingItemMaster.objects.filter(is_active=True).count()
        context['instance_count'] = ClothingItemInstance.objects.filter(is_active=True).count()

        # Kritische Alerts
        critical_alerts = []

        # Niedrige Bestände
        if context['low_stock_count'] > 0:
            critical_alerts.append({
                'type': 'low_stock',
                'severity': 'warning',
                'icon': '📉',
                'title': f'{context["low_stock_count"]} Artikel unter Mindestbestand',
                'url': '/clothing/items/?filter=low_stock',
                'color': 'yellow'
            })

        # Ablaufende Zertifikate
        if context['expiring_certifications'] > 0:
            critical_alerts.append({
                'type': 'expiring_certs',
                'severity': 'warning',
                'icon': '⚠️',
                'title': f'{context["expiring_certifications"]} PSA-Zertifikat(e) laufen bald ab',
                'url': '/clothing/psa/expiring/',
                'color': 'orange'
            })

        # Fällige Prüfungen
        if context['inspections_due'] > 0:
            critical_alerts.append({
                'type': 'inspections',
                'severity': 'high',
                'icon': '🔍',
                'title': f'{context["inspections_due"]} Prüfung(en) überfällig',
                'url': '/clothing/inspections/due/',
                'color': 'red'
            })

        context['critical_alerts'] = critical_alerts

        # Letzte Bewegungen
        context['recent_movements'] = ClothingStockMovement.objects.select_related(
            'item',
            'created_by',
            'from_location',
            'to_location'
        ).order_by('-movement_date')[:10]

        # Aktuelles Modul für Sidebar-Highlighting
        context['current_module'] = 'clothing'

        return context


# ============================================================================
# CATEGORY VIEWS
# ============================================================================

class ClothingCategoryListView(LoginRequiredMixin, ListView):
    """Liste aller Kategorien"""
    model = ClothingCategory
    template_name = 'clothing/category_list.html'
    context_object_name = 'categories'

    def get_queryset(self):
        return ClothingCategory.objects.filter(is_active=True).annotate(
            item_count=Count('items')
        ).order_by('sort_order', 'name')


class ClothingCategoryCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Neue Kategorie erstellen"""
    model = ClothingCategory
    template_name = 'clothing/category_form.html'
    fields = ['name', 'code', 'description', 'color', 'sort_order']
    success_url = reverse_lazy('clothing:category_list')
    permission_required = 'clothing.add_clothingcategory'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, f'Kategorie "{form.instance.name}" wurde erfolgreich erstellt.')
        return super().form_valid(form)


class ClothingCategoryUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Kategorie bearbeiten"""
    model = ClothingCategory
    template_name = 'clothing/category_form.html'
    fields = ['name', 'code', 'description', 'color', 'sort_order', 'is_active']
    success_url = reverse_lazy('clothing:category_list')
    permission_required = 'clothing.change_clothingcategory'

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, f'Kategorie "{form.instance.name}" wurde erfolgreich aktualisiert.')
        return super().form_valid(form)


class ClothingCategoryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Kategorie löschen"""
    model = ClothingCategory
    template_name = 'clothing/category_confirm_delete.html'
    success_url = reverse_lazy('clothing:category_list')
    permission_required = 'clothing.delete_clothingcategory'

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.is_active = False
        self.object.save()
        messages.success(request, f'Kategorie "{self.object.name}" wurde deaktiviert.')
        return redirect(self.success_url)


# ============================================================================
# CLOTHING ITEM VIEWS
# ============================================================================

class IssuesContextMixin:
    """
    Hängt den angezeigten Artikeln ihre offenen Personen-Ausgaben als `issues` an.

    Ein Lagerartikel gehört niemandem – wer wieviel davon hat, steht in den
    Bewegungen. Ohne das würden die Listen bei jedem Artikel "Pool" behaupten,
    auch wenn Stücke bei Personen sind.
    """

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        annotate_issues(context['object_list'])
        return context


class ClothingItemListView(LoginRequiredMixin, IssuesContextMixin, ListView):
    """Liste aller Kleidungsstücke"""
    model = ClothingItem
    template_name = 'clothing/item_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        queryset = ClothingItem.objects.select_related(
            'category',
            'supplier',
            'location',
            'assigned_to'
        ).filter(is_active=True)

        # Suchfilter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(item_number__icontains=search) |
                Q(name__icontains=search) |
                Q(manufacturer__icontains=search)
            )

        # Typ-Filter
        clothing_type = self.request.GET.get('type')
        if clothing_type:
            queryset = queryset.filter(clothing_type=clothing_type)

        # Größen-Filter
        size = self.request.GET.get('size')
        if size:
            queryset = queryset.filter(size=size)

        # PSA-Filter
        is_psa = self.request.GET.get('is_psa')
        if is_psa == 'yes':
            queryset = queryset.filter(is_psa=True)

        # Zuordnungs-Filter: hat der Artikel offene Ausgaben an Personen?
        assignment_status = self.request.GET.get('assignment')
        if assignment_status == 'assigned':
            queryset = queryset.filter(pk__in=issued_item_ids())
        elif assignment_status == 'pool':
            queryset = queryset.exclude(pk__in=issued_item_ids())

        return queryset.order_by('clothing_type', 'size')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()
        context['total_count'] = queryset.count()
        context['psa_count'] = queryset.filter(is_psa=True).count()
        context['assigned_count'] = queryset.filter(pk__in=issued_item_ids()).count()
        return context


class ClothingItemDetailView(LoginRequiredMixin, DetailView):
    """Detailansicht eines Kleidungsstücks"""
    model = ClothingItem
    template_name = 'clothing/item_detail.html'
    context_object_name = 'item'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['movements'] = self.object.stock_movements.select_related(
            'from_location',
            'to_location',
            'created_by'
        ).order_by('-movement_date')[:10]

        # Wer hat von diesem Artikel aktuell noch etwas?
        from personnel.models import Person

        rows = list(open_issues(item=self.object))
        persons = Person.objects.in_bulk({row['person'] for row in rows})
        context['issues'] = [
            {
                'person': persons[row['person']],
                'quantity': row['quantity'],
                'issued_date': row['last_issue'],
            }
            for row in rows
            if row['person'] in persons
        ]

        return context


class ClothingItemCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Neues Kleidungsstück anlegen"""
    model = ClothingItem
    form_class = ClothingItemForm
    template_name = 'clothing/item_form.html'
    permission_required = 'clothing.add_clothingitem'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, f'Kleidungsstück "{form.instance.name}" wurde erfolgreich erstellt.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('clothing:item_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Neuer Artikel'
        context['submit_text'] = 'Artikel anlegen'
        return context


class ClothingItemUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Kleidungsstück bearbeiten"""
    model = ClothingItem
    form_class = ClothingItemForm
    template_name = 'clothing/item_form.html'
    permission_required = 'clothing.change_clothingitem'

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, f'Kleidungsstück "{form.instance.name}" wurde erfolgreich aktualisiert.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('clothing:item_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Bearbeiten: {self.object.name}'
        context['submit_text'] = 'Änderungen speichern'
        return context


class ClothingItemDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Kleidungsstück löschen (Soft-Delete)"""
    model = ClothingItem
    template_name = 'clothing/item_confirm_delete.html'
    success_url = reverse_lazy('clothing:item_list')
    permission_required = 'clothing.delete_clothingitem'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Warnen, wenn noch Stücke bei Personen sind
        annotate_issues([self.object])
        return context

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.is_active = False
        self.object.save()
        messages.success(request, _('Kleidungsstück wurde deaktiviert.'))
        return redirect(self.success_url)


# ============================================================================
# STOCK MOVEMENT VIEWS
# ============================================================================

class ClothingStockMovementListView(LoginRequiredMixin, ListView):
    """Liste aller Lagerbewegungen"""
    model = ClothingStockMovement
    template_name = 'clothing/movement_list.html'
    context_object_name = 'movements'
    paginate_by = 50

    def get_queryset(self):
        queryset = ClothingStockMovement.objects.select_related(
            'item',
            'from_location',
            'to_location',
            'created_by'
        )

        # Typ-Filter
        movement_type = self.request.GET.get('type')
        if movement_type:
            queryset = queryset.filter(movement_type=movement_type)

        # Personen-Filter
        person_id = self.request.GET.get('person')
        if person_id:
            queryset = queryset.filter(person_id=person_id)

        return queryset.order_by('-movement_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Alle Personen für Filter
        from personnel.models import Person
        context['all_persons'] = Person.objects.filter(is_active=True).order_by('last_name', 'first_name')

        return context


class ClothingStockMovementDetailView(LoginRequiredMixin, DetailView):
    """Detailansicht einer Lagerbewegung"""
    model = ClothingStockMovement
    template_name = 'clothing/movement_detail.html'
    context_object_name = 'movement'


class ClothingStockMovementCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Neue Lagerbewegung erstellen"""
    model = ClothingStockMovement
    form_class = ClothingStockMovementForm
    template_name = 'clothing/movement_form.html'
    permission_required = 'clothing.add_clothingstockmovement'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Lagerbewegung wurde erfolgreich erstellt.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('clothing:movement_detail', kwargs={'pk': self.object.pk})


# ============================================================================
# PSA OVERVIEW VIEWS
# ============================================================================

class PSAOverviewView(LoginRequiredMixin, IssuesContextMixin, ListView):
    """Übersicht aller PSA-Kleidung"""
    model = ClothingItem
    template_name = 'clothing/psa_overview.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        return ClothingItem.objects.filter(
            is_psa=True,
            is_active=True
        ).select_related(
            'assigned_to',
            'location'
        ).order_by('protection_level', 'clothing_type')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # PSA-Statistiken
        psa_items = self.get_queryset()
        context['total_psa_count'] = psa_items.count()
        context['expired_cert_count'] = sum(1 for item in psa_items if item.is_certification_expired())
        context['inspection_due_count'] = sum(1 for item in psa_items if item.is_inspection_due())

        return context


class PSAExpiringView(LoginRequiredMixin, IssuesContextMixin, ListView):
    """PSA mit ablaufenden Zertifikaten"""
    model = ClothingItem
    template_name = 'clothing/psa_expiring.html'
    context_object_name = 'items'

    def get_queryset(self):
        today = timezone.now().date()
        warning_date = today + timezone.timedelta(days=90)

        return ClothingItem.objects.filter(
            is_psa=True,
            is_active=True,
            certification_expires__lte=warning_date
        ).select_related(
            'assigned_to',
            'location'
        ).order_by('certification_expires')


# ============================================================================
# INSPECTION VIEWS
# ============================================================================

class InspectionListView(LoginRequiredMixin, IssuesContextMixin, ListView):
    """Liste aller prüfpflichtigen Kleidungsstücke"""
    model = ClothingItem
    template_name = 'clothing/inspection_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        return ClothingItem.objects.filter(
            requires_inspection=True,
            is_active=True
        ).select_related(
            'assigned_to',
            'location'
        ).order_by('next_inspection_date')


class InspectionDueView(LoginRequiredMixin, IssuesContextMixin, ListView):
    """Fällige Prüfungen"""
    model = ClothingItem
    template_name = 'clothing/inspection_due.html'
    context_object_name = 'items'

    def get_queryset(self):
        today = timezone.now().date()

        return ClothingItem.objects.filter(
            requires_inspection=True,
            is_active=True,
            next_inspection_date__lte=today
        ).select_related(
            'assigned_to',
            'location'
        ).order_by('next_inspection_date')


# ============================================================================
# PERSON ASSIGNMENT VIEWS
# ============================================================================

class PersonAssignmentListView(LoginRequiredMixin, ListView):
    """Übersicht aller Personenzuordnungen (aus den Lagerbewegungen abgeleitet)"""
    model = ClothingItem
    template_name = 'clothing/assignment_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        """
        Eine Zeile je (Person, Artikel) mit offener Ausgabe.

        Derselbe Artikel kann an mehrere Personen ausgegeben sein, deshalb bekommt
        jede Zeile eine eigene Kopie des Artikels. Person und Menge hängen nur als
        Anzeige-Attribute daran und werden nicht gespeichert.
        """
        from inventory_base.models import StockMovementType
        from personnel.models import Person

        rows = list(open_issues())
        if not rows:
            return []

        persons = Person.objects.in_bulk({row['person'] for row in rows})
        items = ClothingItem.objects.select_related('category', 'location').in_bulk(
            {row['item'] for row in rows}
        )

        # Letzte Ausgabe-Bewegung je (Artikel, Person) für den Detail-Link
        movements = {}
        for movement in ClothingStockMovement.objects.filter(
            movement_type=StockMovementType.OUTGOING,
            item__in=items,
            person__in=persons,
        ).order_by('movement_date'):
            movements[(movement.item_id, movement.person_id)] = movement.id

        issues = []
        for row in rows:
            person = persons.get(row['person'])
            item = items.get(row['item'])
            if person is None or item is None:
                continue

            issue = copy.copy(item)
            issue.assigned_to = person
            issue.is_personal_issue = True
            issue.issued_quantity = row['quantity']
            issue.issued_date = row['last_issue']
            issue.assignment_movement_id = movements.get((item.pk, person.pk))
            issues.append(issue)

        return sorted(
            issues,
            key=lambda issue: (
                issue.assigned_to.last_name,
                issue.assigned_to.first_name,
                issue.clothing_type,
            ),
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Personen mit Anzahl offener Ausgaben, in der Reihenfolge der Zeilen
        persons = []
        for issue in self.object_list:
            person = issue.assigned_to
            if not hasattr(person, 'item_count'):
                person.item_count = 0
                persons.append(person)
            person.item_count += 1

        context['persons_with_items'] = persons

        return context


class PersonSizeManagementView(LoginRequiredMixin, ListView):
    """Größenverwaltung für eine Person"""
    model = ClothingSizeAssignment
    template_name = 'clothing/person_sizes.html'
    context_object_name = 'size_assignments'

    def get_queryset(self):
        person_id = self.kwargs.get('person_id')
        return ClothingSizeAssignment.objects.filter(
            person_id=person_id
        ).order_by('clothing_type')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        from personnel.models import Person
        person = get_object_or_404(Person, pk=self.kwargs.get('person_id'))
        context['person'] = person

        # Aktuell an diese Person ausgegebene Kleidung
        context['assigned_items'] = issued_items_for_person(person)

        return context

# ============================================================================
# IMPORT/EXPORT
# ============================================================================

class ImportExportView(LoginRequiredMixin, TemplateView):
    """Import/Export Übersicht"""
    template_name = 'clothing/import_export.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_items'] = ClothingItem.objects.filter(is_active=True).count()
        context['total_psa'] = ClothingItem.objects.filter(is_active=True, is_psa=True).count()
        context['total_movements'] = ClothingStockMovement.objects.count()
        context['total_assigned'] = int(total_issued_quantity())
        return context


# ============================================================================
# EXPORT FUNCTIONS (Excel)
# ============================================================================

@login_required
def export_items(request):
    """Export Clothing Items als Excel"""
    items = ClothingItem.objects.filter(is_active=True).select_related(
        'category', 'location', 'created_by', 'updated_by'
    ).order_by('item_number')

    # Offene Ausgaben je Artikel: "Nachname, Vorname (2)"
    from personnel.models import Person
    rows = list(open_issues())
    persons = Person.objects.in_bulk({row['person'] for row in rows})
    ausgegeben_an = {}
    for row in rows:
        person = persons.get(row['person'])
        if person is None:
            continue
        ausgegeben_an.setdefault(row['item'], []).append(
            f"{person.get_full_name()} ({float(row['quantity']):g})"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kleidungsstücke"
    
    # Header
    headers = [
        'Artikelnummer', 'Name', 'Kategorie-ID', 'Kategorie', 'Typ', 'Größe', 'Schnitt',
        'Farbe', 'Material', 'PSA', 'Schutzlevel', 'Norm/Standard',
        'Zertifizierungsnummer', 'Zertifiziert am', 'Zertifikat läuft ab',
        'Max. Nutzungsdauer (Jahre)', 'Prüfpflicht', 'Standort', 'Hersteller',
        'Bestand', 'Einheit', 'Zugeordnet an', 'Beschreibung', 'Erstellt am', 'Aktualisiert am'
    ]

    # Header-Styling
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Daten
    for row_num, item in enumerate(items, start=2):
        ws.cell(row=row_num, column=1, value=item.item_number)
        ws.cell(row=row_num, column=2, value=item.name)
        ws.cell(row=row_num, column=3, value=item.category.id if item.category else '')
        ws.cell(row=row_num, column=4, value=item.category.name if item.category else '')
        ws.cell(row=row_num, column=5, value=item.get_clothing_type_display())
        ws.cell(row=row_num, column=6, value=item.get_size_display())
        ws.cell(row=row_num, column=7, value=item.get_gender_display())
        ws.cell(row=row_num, column=8, value=item.color)
        ws.cell(row=row_num, column=9, value=item.material)
        ws.cell(row=row_num, column=10, value='Ja' if item.is_psa else 'Nein')
        ws.cell(row=row_num, column=11, value=item.get_protection_level_display())
        ws.cell(row=row_num, column=12, value=item.norm_standard)
        ws.cell(row=row_num, column=13, value=item.certification_number)
        ws.cell(row=row_num, column=14, value=item.certification_date.strftime('%d.%m.%Y') if item.certification_date else '')
        ws.cell(row=row_num, column=15, value=item.certification_expires.strftime('%d.%m.%Y') if item.certification_expires else '')
        ws.cell(row=row_num, column=16, value=item.max_usage_years)
        ws.cell(row=row_num, column=17, value='Ja' if item.requires_inspection else 'Nein')
        ws.cell(row=row_num, column=18, value=item.location.name if item.location else '')
        ws.cell(row=row_num, column=19, value=item.manufacturer)
        ws.cell(row=row_num, column=20, value=float(item.quantity))
        ws.cell(row=row_num, column=21, value=item.unit)
        ws.cell(row=row_num, column=22, value=', '.join(ausgegeben_an.get(item.pk, [])))
        ws.cell(row=row_num, column=23, value=item.description)
        ws.cell(row=row_num, column=24, value=item.created_at.strftime('%d.%m.%Y %H:%M') if item.created_at else '')
        ws.cell(row=row_num, column=25, value=item.updated_at.strftime('%d.%m.%Y %H:%M') if item.updated_at else '')
    
    # Spaltenbreiten
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=kleidungsstuecke_export.xlsx'
    wb.save(response)
    return response


@login_required
def export_movements(request):
    """Export Clothing Stock Movements als Excel"""
    movements = ClothingStockMovement.objects.select_related(
        'item', 'from_location', 'to_location', 'person', 'created_by'
    ).order_by('-movement_date')[:1000]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lagerbewegungen"
    
    # Header
    headers = [
        'ID', 'Datum', 'Artikel', 'Bewegungstyp', 'Menge', 'Einheit',
        'Von Standort', 'Nach Standort', 'Person', 'Rückgabegrund', 'Notiz',
        'Erstellt von', 'Erstellt am'
    ]

    # Header-Styling
    header_fill = PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Daten
    for row_num, movement in enumerate(movements, start=2):
        ws.cell(row=row_num, column=1, value=movement.id)
        ws.cell(row=row_num, column=2, value=movement.movement_date.strftime('%d.%m.%Y %H:%M') if movement.movement_date else '')
        ws.cell(row=row_num, column=3, value=movement.item.name if movement.item else '')
        ws.cell(row=row_num, column=4, value=movement.get_movement_type_display())
        ws.cell(row=row_num, column=5, value=float(movement.quantity))
        ws.cell(row=row_num, column=6, value=movement.unit)
        ws.cell(row=row_num, column=7, value=movement.from_location.name if movement.from_location else '')
        ws.cell(row=row_num, column=8, value=movement.to_location.name if movement.to_location else '')
        ws.cell(row=row_num, column=9, value=movement.person.get_full_name() if movement.person else '')
        ws.cell(row=row_num, column=10, value=movement.return_reason)
        ws.cell(row=row_num, column=11, value=movement.notes)
        ws.cell(row=row_num, column=12, value=movement.created_by.get_full_name() if movement.created_by else '')
        ws.cell(row=row_num, column=13, value=movement.created_at.strftime('%d.%m.%Y %H:%M') if movement.created_at else '')
    
    # Spaltenbreiten
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=lagerbewegungen_export.xlsx'
    wb.save(response)
    return response


# ============================================================================
# TEMPLATE DOWNLOAD FUNCTIONS (CSV)
# ============================================================================

@login_required
def template_items(request):
    """CSV-Vorlage für Kleidungsstücke-Import"""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    
    # Header
    headers = [
        'Artikelnummer*', 'Name*', 'Kategorie-ID*', 'Typ*', 'Größe*', 'Schnitt',
        'Farbe', 'Material', 'PSA (Ja/Nein)', 'Schutzlevel', 'Norm/Standard',
        'Standort-ID*', 'Hersteller', 'Bestand*', 'Einheit*', 'Beschreibung'
    ]
    writer.writerow(headers)

    # Beispielzeile
    example_data = [
        'KLEID-001', 'Einsatzjacke HuPF', '1', 'jacket', 'l', 'unisex',
        'Rot/Gelb', 'Nomex', 'Ja', 'high', 'EN 469',
        '1', 'TEXPORT', '5', 'Stück', 'Beispiel: Feuerwehrjacke - diese Zeile löschen!'
    ]
    writer.writerow(example_data)
    
    # Hinweise
    writer.writerow([])
    writer.writerow(['# HINWEISE:'])
    writer.writerow(['# Felder mit * sind Pflichtfelder'])
    writer.writerow(['# Typ: jacket, pants, overall, shirt, underwear, shoes, boots, gloves, helmet, vest, coat, hat, protective, uniform, other'])
    writer.writerow(['# Größe: xxs, xs, s, m, l, xl, xxl, xxxl, xxxxl, 36-60, s36-s48 (Schuhgrößen)'])
    writer.writerow(['# Schnitt: unisex, male, female'])
    writer.writerow(['# PSA: Ja oder Nein'])
    writer.writerow(['# Schutzlevel: none, basic, enhanced, high, specialist'])
    writer.writerow(['# Einheit: Stück, Paar, Set, Packung, Karton'])
    writer.writerow(['# Kategorie-ID und Standort-ID müssen existieren; Hersteller ist Freitext'])
    
    # Response
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename=vorlage_kleidung_import.csv'
    response.write('\ufeff')  # UTF-8 BOM
    response.write(output.getvalue())
    return response


# ============================================================================
# IMPORT FUNCTIONS (CSV)
# ============================================================================

@login_required
def import_items(request):
    """Import Clothing Items aus CSV"""
    if request.method != 'POST':
        messages.error(request, 'Nur POST-Anfragen erlaubt.')
        return redirect('clothing:import_export')
    
    if 'file' not in request.FILES:
        messages.error(request, 'Keine Datei hochgeladen.')
        return redirect('clothing:import_export')
    
    uploaded_file = request.FILES['file']
    
    try:
        # CSV-Datei lesen (UTF-8 mit BOM)
        decoded_file = uploaded_file.read().decode('utf-8-sig')
        csv_reader = csv.reader(io.StringIO(decoded_file), delimiter=';')
        
        success_count = 0
        error_count = 0
        errors = []
        
        # Header überspringen
        next(csv_reader, None)
        
        # Zeilen durchgehen
        for row_num, row in enumerate(csv_reader, start=2):
            # Leere Zeilen und Kommentare überspringen
            if not row or not row[0] or row[0].startswith('#'):
                continue
            
            # Beispielzeile überspringen
            if 'Beispiel' in str(row[-1] if row else ''):
                continue
            
            try:
                # Daten extrahieren (mit Index-Safety)
                item_number = row[0] if len(row) > 0 else None
                name = row[1] if len(row) > 1 else None
                category_id = row[2] if len(row) > 2 else None
                clothing_type = row[3] if len(row) > 3 else None
                size = row[4] if len(row) > 4 else None
                gender = row[5] if len(row) > 5 else 'unisex'
                color = row[6] if len(row) > 6 else ''
                material = row[7] if len(row) > 7 else ''
                is_psa_str = row[8] if len(row) > 8 else 'Nein'
                protection_level = row[9] if len(row) > 9 else 'none'
                norm_standard = row[10] if len(row) > 10 else ''
                location_id = row[11] if len(row) > 11 else None
                manufacturer = row[12] if len(row) > 12 else ''
                quantity = row[13] if len(row) > 13 else None
                unit = row[14] if len(row) > 14 else None
                description = row[15] if len(row) > 15 else ''
                
                # Validierung
                if not all([item_number, name, category_id, clothing_type, size, location_id, quantity, unit]):
                    errors.append(f"Zeile {row_num}: Pflichtfelder fehlen")
                    error_count += 1
                    continue
                
                # Duplikatprüfung
                if ClothingItem.objects.filter(item_number=item_number).exists():
                    errors.append(f"Zeile {row_num}: Artikelnummer '{item_number}' existiert bereits")
                    error_count += 1
                    continue
                
                # Kategorie laden
                try:
                    category = ClothingCategory.objects.get(id=int(category_id))
                except (ClothingCategory.DoesNotExist, ValueError):
                    errors.append(f"Zeile {row_num}: Kategorie-ID {category_id} nicht gefunden")
                    error_count += 1
                    continue
                
                # Standort laden
                from locations.models import Location
                try:
                    location = Location.objects.get(id=int(location_id))
                except (Location.DoesNotExist, ValueError):
                    errors.append(f"Zeile {row_num}: Standort-ID {location_id} nicht gefunden")
                    error_count += 1
                    continue
                
                # Boolean-Konvertierung
                is_psa = is_psa_str.lower() in ['ja', 'yes', 'true', '1']
                
                # Item erstellen
                item = ClothingItem.objects.create(
                    item_number=item_number,
                    name=name,
                    category=category,
                    clothing_type=clothing_type,
                    size=size,
                    gender=gender,
                    color=color,
                    material=material,
                    is_psa=is_psa,
                    protection_level=protection_level,
                    norm_standard=norm_standard,
                    location=location,
                    manufacturer=manufacturer,
                    quantity=float(quantity),
                    unit=unit,
                    description=description,
                    created_by=request.user,
                    updated_by=request.user
                )
                
                success_count += 1
                
            except Exception as e:
                errors.append(f"Zeile {row_num}: Fehler - {str(e)}")
                error_count += 1
                logger.error(f"Item import error in row {row_num}: {e}")
        
        # Erfolgsmeldung
        if success_count > 0:
            messages.success(request, f"✅ {success_count} Artikel erfolgreich importiert.")
        
        if error_count > 0:
            messages.warning(request, f"⚠️ {error_count} Fehler beim Import.")
            for error in errors[:10]:
                messages.error(request, error)
            if len(errors) > 10:
                messages.error(request, f"... und {len(errors) - 10} weitere Fehler")
    
    except Exception as e:
        messages.error(request, f"Fehler beim Lesen der CSV-Datei: {str(e)}")
        logger.error(f"CSV import error: {e}")
    
    return redirect('clothing:import_export')


# ============================================================================
# CATEGORY EXPORT/IMPORT
# ============================================================================

@login_required
def export_categories(request):
    """Export Clothing Categories als Excel"""
    categories = ClothingCategory.objects.filter(is_active=True).order_by('sort_order', 'name')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kategorien"
    
    # Header
    headers = [
        'ID', 'Name', 'Code', 'Beschreibung', 'Farbe', 
        'Sortierung', 'Erstellt am', 'Aktualisiert am'
    ]
    
    # Header-Styling
    header_fill = PatternFill(start_color='9C27B0', end_color='9C27B0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Daten
    for row_num, category in enumerate(categories, start=2):
        ws.cell(row=row_num, column=1, value=category.id)
        ws.cell(row=row_num, column=2, value=category.name)
        ws.cell(row=row_num, column=3, value=category.code)
        ws.cell(row=row_num, column=4, value=category.description)
        ws.cell(row=row_num, column=5, value=category.color)
        ws.cell(row=row_num, column=6, value=category.sort_order)
        ws.cell(row=row_num, column=7, value=category.created_at.strftime('%d.%m.%Y %H:%M') if category.created_at else '')
        ws.cell(row=row_num, column=8, value=category.updated_at.strftime('%d.%m.%Y %H:%M') if category.updated_at else '')
    
    # Spaltenbreiten
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=kategorien_export.xlsx'
    wb.save(response)
    return response


@login_required
def template_categories(request):
    """CSV-Vorlage für Kategorien-Import"""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    
    # Header
    headers = [
        'Name*', 'Code*', 'Beschreibung', 'Farbe', 'Sortierung'
    ]
    writer.writerow(headers)
    
    # Beispielzeile
    example_data = [
        'PSA Einsatzkleidung', 'PSA-EINSATZ', 'Persönliche Schutzausrüstung für den Einsatz',
        '#FF5722', '10'
    ]
    writer.writerow(example_data)
    
    # Hinweise
    writer.writerow([])
    writer.writerow(['# HINWEISE:'])
    writer.writerow(['# Felder mit * sind Pflichtfelder'])
    writer.writerow(['# Code muss eindeutig sein (z.B. PSA-EINSATZ, UNIFORM, ALLGEMEIN)'])
    writer.writerow(['# Farbe: Hex-Code mit # (z.B. #FF5722, #3B82F6)'])
    writer.writerow(['# Sortierung: Numerischer Wert für Reihenfolge (Standard: 0)'])
    
    # Response
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename=vorlage_kategorien_import.csv'
    response.write('\ufeff')  # UTF-8 BOM
    response.write(output.getvalue())
    return response


@login_required
def import_categories(request):
    """Import Clothing Categories aus CSV"""
    if request.method != 'POST':
        messages.error(request, 'Nur POST-Anfragen erlaubt.')
        return redirect('clothing:import_export')

    if 'file' not in request.FILES:
        messages.error(request, 'Keine Datei hochgeladen.')
        return redirect('clothing:import_export')

    uploaded_file = request.FILES['file']

    try:
        # CSV-Datei lesen (UTF-8 mit BOM)
        decoded_file = uploaded_file.read().decode('utf-8-sig')
        csv_reader = csv.reader(io.StringIO(decoded_file), delimiter=';')

        success_count = 0
        error_count = 0
        errors = []

        # Header überspringen
        next(csv_reader, None)

        # Zeilen durchgehen
        for row_num, row in enumerate(csv_reader, start=2):
            # Leere Zeilen und Kommentare überspringen
            if not row or not row[0] or row[0].startswith('#'):
                continue

            # Beispielzeile überspringen
            if 'Beispiel' in str(row[2] if len(row) > 2 else ''):
                continue

            try:
                # Daten extrahieren (mit Index-Safety)
                name = row[0] if len(row) > 0 else None
                code = row[1] if len(row) > 1 else None
                description = row[2] if len(row) > 2 else ''
                color = row[3] if len(row) > 3 else '#3B82F6'
                sort_order = row[4] if len(row) > 4 else '0'

                # Validierung
                if not all([name, code]):
                    errors.append(f"Zeile {row_num}: Pflichtfelder fehlen (Name, Code)")
                    error_count += 1
                    continue

                # Duplikatprüfung (Name)
                if ClothingCategory.objects.filter(name=name).exists():
                    errors.append(f"Zeile {row_num}: Kategorie-Name '{name}' existiert bereits")
                    error_count += 1
                    continue

                # Duplikatprüfung (Code)
                if ClothingCategory.objects.filter(code=code).exists():
                    errors.append(f"Zeile {row_num}: Kategorie-Code '{code}' existiert bereits")
                    error_count += 1
                    continue

                # Farbe validieren
                if color and not color.startswith('#'):
                    color = '#' + color

                # Kategorie erstellen
                category = ClothingCategory.objects.create(
                    name=name,
                    code=code,
                    description=description,
                    color=color,
                    sort_order=int(sort_order) if sort_order else 0,
                    created_by=request.user,
                    updated_by=request.user
                )

                success_count += 1

            except Exception as e:
                errors.append(f"Zeile {row_num}: Fehler - {str(e)}")
                error_count += 1
                logger.error(f"Category import error in row {row_num}: {e}")

        # Erfolgsmeldung
        if success_count > 0:
            messages.success(request, f"✅ {success_count} Kategorien erfolgreich importiert.")

        if error_count > 0:
            messages.warning(request, f"⚠️ {error_count} Fehler beim Import.")
            for error in errors[:10]:
                messages.error(request, error)
            if len(errors) > 10:
                messages.error(request, f"... und {len(errors) - 10} weitere Fehler")

    except Exception as e:
        messages.error(request, f"Fehler beim Lesen der CSV-Datei: {str(e)}")
        logger.error(f"CSV import error: {e}")

    return redirect('clothing:import_export')


# ============================================================================
# BARCODE & QR-CODE VIEWS
# ============================================================================

@login_required
def item_barcode(request, pk):
    """
    Generiert Barcode für ClothingItem als SVG
    Nutzt Code128 für Artikelnummer
    """
    item = get_object_or_404(ClothingItem, pk=pk)

    try:
        svg_string = item.generate_barcode()

        if not svg_string:
            return HttpResponse("Kein Barcode verfügbar", status=400)

        return HttpResponse(svg_string, content_type='image/svg+xml')

    except Exception as e:
        logger.error(f"Barcode generation error for item {pk}: {e}")
        return HttpResponse(f"Fehler bei Barcode-Generierung: {str(e)}", status=500)


@login_required
def item_qr_code(request, pk):
    """
    Generiert QR-Code für ClothingItem als SVG
    Enthält JSON mit Artikeldetails
    """
    item = get_object_or_404(ClothingItem, pk=pk)

    try:
        svg_string = item.generate_qr_code()

        if not svg_string:
            return HttpResponse("Kein QR-Code verfügbar", status=400)

        return HttpResponse(svg_string, content_type='image/svg+xml')

    except Exception as e:
        logger.error(f"QR-Code generation error for item {pk}: {e}")
        return HttpResponse(f"Fehler bei QR-Code-Generierung: {str(e)}", status=500)


# ============================================================================
# STAMMDATEN / MASTER VIEWS
# ============================================================================

class ClothingMasterListView(LoginRequiredMixin, ListView):
    """Liste aller Kleidungs-Stammdaten (Master)"""
    model = ClothingItemMaster
    template_name = 'clothing/master_list.html'
    context_object_name = 'masters'
    paginate_by = 25

    def get_queryset(self):
        queryset = ClothingItemMaster.objects.filter(is_active=True).select_related('category').annotate(
            instance_count=Count('instances', filter=Q(instances__is_active=True))
        ).order_by('category__name', 'name')

        # Suchfilter
        search = self.request.GET.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(manufacturer__icontains=search) |
                Q(model_number__icontains=search)
            )

        # Kategorie-Filter
        category = self.request.GET.get('category', '')
        if category:
            queryset = queryset.filter(category_id=category)

        # Typ-Filter (product_type)
        product_type_id = self.request.GET.get('type', '')
        if product_type_id:
            queryset = queryset.filter(product_type_id=product_type_id)

        # PSA-Filter
        psa = self.request.GET.get('psa', '')
        if psa == 'yes':
            queryset = queryset.filter(is_psa=True)
        elif psa == 'no':
            queryset = queryset.filter(is_psa=False)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = ClothingCategory.objects.filter(is_active=True).order_by('name')
        context['producttypes'] = ClothingProductType.objects.filter(is_active=True).order_by('sort_order', 'name')
        context['search'] = self.request.GET.get('search', '')
        context['selected_category'] = self.request.GET.get('category', '')
        context['selected_type'] = self.request.GET.get('type', '')
        context['selected_psa'] = self.request.GET.get('psa', '')
        context['current_module'] = 'clothing'

        # Statistiken
        context['total_masters'] = ClothingItemMaster.objects.filter(is_active=True).count()
        context['psa_masters'] = ClothingItemMaster.objects.filter(is_active=True, is_psa=True).count()
        context['total_instances'] = ClothingItemInstance.objects.filter(is_active=True).count()

        return context


class ClothingMasterDetailView(LoginRequiredMixin, DetailView):
    """Detailansicht eines Kleidungs-Stammdatensatzes"""
    model = ClothingItemMaster
    template_name = 'clothing/master_detail.html'
    context_object_name = 'master'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        master = self.object

        # Instanzen dieses Masters
        context['instances'] = master.instances.filter(is_active=True).select_related(
            'location', 'assigned_to'
        ).order_by('-created_at')[:20]

        context['instance_count'] = master.instances.filter(is_active=True).count()
        context['assigned_count'] = master.instances.filter(is_active=True, assigned_to__isnull=False).count()
        context['available_count'] = master.instances.filter(is_active=True, assigned_to__isnull=True).count()

        context['current_module'] = 'clothing'
        return context


class ClothingMasterCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Neuen Kleidungs-Stammdatensatz erstellen"""
    model = ClothingItemMaster
    form_class = ClothingItemMasterForm
    template_name = 'clothing/master_form.html'
    success_url = reverse_lazy('clothing:master_list')
    permission_required = 'clothing.add_clothingitemmaster'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, f'Stammdaten "{form.instance.name}" erfolgreich angelegt.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Neue Stammdaten anlegen'
        context['current_module'] = 'clothing'
        return context


class ClothingMasterUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Kleidungs-Stammdaten bearbeiten"""
    model = ClothingItemMaster
    form_class = ClothingItemMasterForm
    template_name = 'clothing/master_form.html'
    success_url = reverse_lazy('clothing:master_list')
    permission_required = 'clothing.change_clothingitemmaster'

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, f'✅ Stammdaten "{form.instance.name}" erfolgreich aktualisiert.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Stammdaten bearbeiten: {self.object.name}'
        context['current_module'] = 'clothing'
        return context


class ClothingMasterDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Kleidungs-Stammdaten löschen (soft delete)"""
    model = ClothingItemMaster
    template_name = 'clothing/master_confirm_delete.html'
    success_url = reverse_lazy('clothing:master_list')
    permission_required = 'clothing.delete_clothingitemmaster'

    def delete(self, request, *args, **kwargs):
        master = self.get_object()
        # Check if instances exist
        if master.instances.filter(is_active=True).exists():
            messages.error(request, f'❌ Stammdaten "{master.name}" können nicht gelöscht werden, da noch aktive Artikel existieren.')
            return redirect('clothing:master_detail', pk=master.pk)

        # Soft delete
        master.is_active = False
        master.updated_by = request.user
        master.save()
        messages.success(request, f'✅ Stammdaten "{master.name}" erfolgreich gelöscht.')
        return redirect(self.success_url)


# ============================================================================
# EXEMPLARE / INSTANZEN (konkrete Kleidungsstücke zu einem Stammdatensatz)
# ============================================================================

class ClothingInstanceListView(LoginRequiredMixin, ListView):
    """Liste aller Exemplare (konkrete Kleidungsstücke)"""
    model = ClothingItemInstance
    template_name = 'clothing/instance_list.html'
    context_object_name = 'instances'
    paginate_by = 50

    def get_queryset(self):
        queryset = ClothingItemInstance.objects.filter(is_active=True).select_related(
            'master', 'location', 'assigned_to'
        )

        # Suchfilter
        search = self.request.GET.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(inventory_number__icontains=search) |
                Q(serial_number__icontains=search) |
                Q(master__name__icontains=search) |
                Q(color__icontains=search)
            )

        # Filter auf einen Stammdatensatz
        master_id = self.request.GET.get('master', '')
        if master_id:
            queryset = queryset.filter(master_id=master_id)

        # Größen-Filter
        size = self.request.GET.get('size', '')
        if size:
            queryset = queryset.filter(size=size)

        # Zuordnungs-Filter
        assignment = self.request.GET.get('assignment', '')
        if assignment == 'assigned':
            queryset = queryset.filter(assigned_to__isnull=False)
        elif assignment == 'pool':
            queryset = queryset.filter(assigned_to__isnull=True)

        return queryset.order_by('master__name', 'size', 'inventory_number')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        base = ClothingItemInstance.objects.filter(is_active=True)

        context['masters'] = ClothingItemMaster.objects.filter(is_active=True).order_by('name')
        context['search'] = self.request.GET.get('search', '')
        context['selected_master'] = self.request.GET.get('master', '')
        context['selected_size'] = self.request.GET.get('size', '')
        context['selected_assignment'] = self.request.GET.get('assignment', '')

        context['total_count'] = base.count()
        context['assigned_count'] = base.filter(assigned_to__isnull=False).count()
        context['pool_count'] = base.filter(assigned_to__isnull=True).count()

        context['current_module'] = 'clothing'
        return context


class ClothingInstanceDetailView(LoginRequiredMixin, DetailView):
    """Detailansicht eines Exemplars"""
    model = ClothingItemInstance
    template_name = 'clothing/instance_detail.html'
    context_object_name = 'instance'

    def get_queryset(self):
        return ClothingItemInstance.objects.select_related(
            'master', 'master__category', 'location', 'assigned_to'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_module'] = 'clothing'
        return context


class ClothingInstanceCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Neues Exemplar zu einem Stammdatensatz anlegen"""
    model = ClothingItemInstance
    form_class = ClothingItemInstanceForm
    template_name = 'clothing/instance_form.html'
    permission_required = 'clothing.add_clothingiteminstance'

    def get_initial(self):
        initial = super().get_initial()
        # Stammdatensatz aus der URL vorbelegen (?master=<pk>)
        master_id = self.request.GET.get('master')
        if master_id:
            initial['master'] = master_id
        return initial

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        messages.success(
            self.request,
            f'Exemplar "{self.object.inventory_number}" wurde erfolgreich angelegt.'
        )
        return response

    def get_success_url(self):
        return reverse_lazy('clothing:instance_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Neues Exemplar anlegen'
        context['submit_text'] = 'Exemplar anlegen'

        master_id = self.request.GET.get('master')
        if master_id:
            context['master'] = ClothingItemMaster.objects.filter(pk=master_id).first()

        context['current_module'] = 'clothing'
        return context


class ClothingInstanceUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Exemplar bearbeiten (inkl. Zuordnung zu einer Person)"""
    model = ClothingItemInstance
    form_class = ClothingItemInstanceForm
    template_name = 'clothing/instance_form.html'
    permission_required = 'clothing.change_clothingiteminstance'

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        messages.success(
            self.request,
            f'Exemplar "{self.object.inventory_number}" wurde erfolgreich aktualisiert.'
        )
        return response

    def get_success_url(self):
        return reverse_lazy('clothing:instance_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Exemplar bearbeiten: {self.object.inventory_number}'
        context['submit_text'] = 'Änderungen speichern'
        context['master'] = self.object.master
        context['current_module'] = 'clothing'
        return context


class ClothingInstanceDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Exemplar löschen (Soft-Delete)"""
    model = ClothingItemInstance
    template_name = 'clothing/instance_confirm_delete.html'
    success_url = reverse_lazy('clothing:instance_list')
    permission_required = 'clothing.delete_clothingiteminstance'
    context_object_name = 'instance'

    def form_valid(self, form):
        """Soft-Delete statt echtem Löschen"""
        instance = self.get_object()
        instance.is_active = False
        instance.updated_by = self.request.user
        instance.save()
        messages.success(self.request, f'Exemplar "{instance.inventory_number}" wurde deaktiviert.')
        return redirect(self.success_url)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_module'] = 'clothing'
        return context


# ============================================================================
# PRODUKTTYP VIEWS
# ============================================================================

class ClothingProductTypeListView(LoginRequiredMixin, ListView):
    """Liste aller Produkttypen"""
    model = ClothingProductType
    template_name = 'clothing/producttype_list.html'
    context_object_name = 'producttypes'

    def get_queryset(self):
        return ClothingProductType.objects.filter(is_active=True).annotate(
            master_count=Count('masters', filter=Q(masters__is_active=True))
        ).order_by('sort_order', 'name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_module'] = 'clothing'
        return context


class ClothingProductTypeCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Neuen Produkttyp erstellen"""
    model = ClothingProductType
    template_name = 'clothing/producttype_form.html'
    fields = ['name', 'code', 'description', 'icon', 'is_psa_typical', 'sort_order']
    success_url = reverse_lazy('clothing:producttype_list')
    permission_required = 'clothing.add_clothingproducttype'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, f'Produkttyp "{form.instance.name}" erfolgreich angelegt.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Neuer Produkttyp'
        context['current_module'] = 'clothing'
        return context


class ClothingProductTypeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Produkttyp bearbeiten"""
    model = ClothingProductType
    template_name = 'clothing/producttype_form.html'
    fields = ['name', 'code', 'description', 'icon', 'is_psa_typical', 'sort_order']
    success_url = reverse_lazy('clothing:producttype_list')
    permission_required = 'clothing.change_clothingproducttype'

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, f'Produkttyp "{form.instance.name}" erfolgreich aktualisiert.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Produkttyp bearbeiten: {self.object.name}'
        context['current_module'] = 'clothing'
        return context


class ClothingProductTypeDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Produkttyp löschen (soft delete)"""
    model = ClothingProductType
    template_name = 'clothing/producttype_confirm_delete.html'
    success_url = reverse_lazy('clothing:producttype_list')
    permission_required = 'clothing.delete_clothingproducttype'

    def delete(self, request, *args, **kwargs):
        producttype = self.get_object()
        # Check if masters use this type
        if producttype.masters.filter(is_active=True).exists():
            messages.error(request, f'Produkttyp "{producttype.name}" kann nicht gelöscht werden, da er noch verwendet wird.')
            return redirect('clothing:producttype_list')

        # Soft delete
        producttype.is_active = False
        producttype.updated_by = request.user
        producttype.save()
        messages.success(request, f'Produkttyp "{producttype.name}" erfolgreich gelöscht.')
        return redirect(self.success_url)


# ============================================================================
# INVENTUR VIEWS (aus inventory_views.py importiert)
# ============================================================================

from .inventory_views import (
    ClothingInventoryListView,
    ClothingInventoryCreateView,
    ClothingInventoryDetailView,
    ClothingInventoryStartView,
    ClothingInventoryCountingView,
    ClothingInventoryCompleteView,
    ClothingInventoryApproveView,
    ClothingInventoryItemUpdateView,
    ClothingInventoryProgressView,
    ClothingInventoryExportView,
)


# ============================================================================
# AJAX ENDPOINTS
# ============================================================================

from django.http import JsonResponse

@login_required
def get_item_locations(request, pk):
    """
    AJAX-Endpoint: Gibt Lagerorte zurück, an denen der Artikel vorhanden ist
    Wird verwendet um from_location im Bewegungsformular zu filtern
    """
    try:
        item = get_object_or_404(ClothingItem, pk=pk)

        # Der Artikel hat einen spezifischen Lagerort
        locations = []
        if item.location and item.quantity > 0:
            locations.append({
                'id': item.location.id,
                'name': item.location.name,
                'full_path': str(item.location),
                'stock': float(item.quantity)
            })

        return JsonResponse({
            'success': True,
            'locations': locations,
            'item_name': item.name,
            'item_stock': float(item.quantity)
        })
    except Exception as e:
        logger.error(f"Error getting item locations: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)
