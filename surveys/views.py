"""
Umfragen Views

ANONYMITÄT
----------
`survey_fill` ist die einzige Stelle, an der Antworten entstehen. Dort gilt:
  - kein `audit.utils.log_create()` (würde User + IP protokollieren)
  - kein Speichern der IP-Adresse
  - `SurveyResponse.user` nur bei personalisierten Umfragen
Das Modell setzt das zusätzlich in `SurveyResponse.save()` durch.
"""

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db import IntegrityError, transaction
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.generic import CreateView, DetailView, ListView, UpdateView

from .conditions import build_client_config
from .forms import (
    SurveyForm,
    SurveyQuestionForm,
    build_response_form,
    field_name_for,
    save_answers,
)
from .models import (
    Survey,
    SurveyInvitation,
    SurveyParticipation,
    SurveyQuestion,
    SurveyResponse,
    SurveyStatus,
)
from .services import (
    build_export_rows,
    build_question_statistics,
    get_missing_participants,
)


#: Obergrenze je Erzeugungs-Vorgang – verhindert, dass ein Tippfehler ("500000")
#: die Datenbank und das PDF sprengt.
MAX_INVITATIONS_PER_BATCH = 500


class ConductorRequiredMixin(PermissionRequiredMixin):
    """Nur für Umfrage-Verantwortliche"""
    permission_required = 'surveys.conduct_survey'
    raise_exception = True


class SurveyModuleMixin:
    """Markiert den Sidebar-Eintrag als aktiv (siehe includes/sidebar_nav.html)"""

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_module'] = 'surveys'
        return context


# ======================================================================
# Übersicht
# ======================================================================

class SurveyListView(LoginRequiredMixin, SurveyModuleMixin, ListView):
    """
    Übersicht der Umfragen.

    Verantwortliche sehen alle Umfragen, alle anderen nur laufende Umfragen ihrer
    Zielgruppe.
    """
    model = Survey
    template_name = 'surveys/survey_list.html'
    context_object_name = 'surveys'
    paginate_by = 20

    def get_queryset(self):
        # order_by explizit: annotate() + distinct() verlieren sonst die Meta-Ordering
        # und die Pagination wird instabil.
        queryset = Survey.objects.prefetch_related('target_groups').order_by('-created_at').annotate(
            question_count=Count('questions', distinct=True),
            answer_count=Count('responses', filter=Q(responses__is_complete=True), distinct=True),
        )

        if self.request.user.has_perm('surveys.conduct_survey'):
            status = self.request.GET.get('status')
            if status:
                queryset = queryset.filter(status=status)
            return queryset

        if not self.request.user.has_perm('surveys.participate_survey'):
            return queryset.none()

        # Teilnehmer-Sicht: nur aktive Umfragen der eigenen Zielgruppe
        group_ids = list(self.request.user.groups.values_list('pk', flat=True))
        return queryset.filter(status=SurveyStatus.ACTIVE).filter(
            Q(target_groups__isnull=True) | Q(target_groups__pk__in=group_ids)
        ).distinct()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        context['is_conductor'] = user.has_perm('surveys.conduct_survey')
        context['status_choices'] = SurveyStatus.choices
        context['current_status'] = self.request.GET.get('status', '')

        # Teilnahme-Status für die Liste (eine Query statt einer pro Umfrage)
        context['participated_ids'] = set(
            SurveyParticipation.objects
            .filter(user=user, survey__in=context['surveys'])
            .values_list('survey_id', flat=True)
        )
        return context


class SurveyDetailView(LoginRequiredMixin, ConductorRequiredMixin, SurveyModuleMixin, DetailView):
    """Bearbeitungsansicht mit Fragenliste (der eigentliche Builder)"""
    model = Survey
    template_name = 'surveys/survey_detail.html'
    context_object_name = 'survey'

    def get_queryset(self):
        return Survey.objects.prefetch_related('questions', 'target_groups')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        survey = self.object
        context['questions'] = survey.questions.all()
        context['response_count'] = survey.response_count

        if survey.is_token_mode:
            # Bei QR-Zetteln gibt es keine Zielgruppe, gegen die man „fehlt noch"
            # rechnen könnte – dort zählen offene und verbrauchte Zettel.
            invitations = survey.invitations.all()
            context['issued_count'] = invitations.count()
            context['used_count'] = invitations.filter(is_used=True).count()
        else:
            context['participation_count'] = survey.participations.count()
            context['missing_count'] = get_missing_participants(survey).count()
        return context


# ======================================================================
# Umfrage anlegen / bearbeiten
# ======================================================================

class SurveyCreateView(LoginRequiredMixin, ConductorRequiredMixin, SurveyModuleMixin, CreateView):
    model = Survey
    form_class = SurveyForm
    template_name = 'surveys/survey_form.html'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        messages.success(self.request, 'Umfrage angelegt. Fügen Sie jetzt die Fragen hinzu.')
        return response

    def get_success_url(self):
        return reverse('surveys:detail', kwargs={'pk': self.object.pk})


class SurveyUpdateView(LoginRequiredMixin, ConductorRequiredMixin, SurveyModuleMixin, UpdateView):
    model = Survey
    form_class = SurveyForm
    template_name = 'surveys/survey_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_update'] = True
        return context

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, 'Umfrage gespeichert.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('surveys:detail', kwargs={'pk': self.object.pk})


@require_POST
def survey_set_status(request, pk):
    """Umfrage freischalten oder schließen"""
    if not request.user.has_perm('surveys.conduct_survey'):
        return redirect('surveys:list')

    survey = get_object_or_404(Survey, pk=pk)
    new_status = request.POST.get('status')

    if new_status not in dict(SurveyStatus.choices):
        messages.error(request, 'Unbekannter Status.')
        return redirect('surveys:detail', pk=pk)

    if new_status == SurveyStatus.ACTIVE and not survey.questions.exists():
        messages.error(request, 'Die Umfrage enthält noch keine Fragen.')
        return redirect('surveys:detail', pk=pk)

    survey.status = new_status
    survey.updated_by = request.user
    survey.save(update_fields=['status', 'updated_by', 'updated_at'])

    labels = {
        SurveyStatus.ACTIVE: 'Umfrage ist jetzt freigeschaltet.',
        SurveyStatus.CLOSED: 'Umfrage wurde geschlossen.',
        SurveyStatus.DRAFT: 'Umfrage wurde zurück in den Entwurf gesetzt.',
    }
    messages.success(request, labels[new_status])
    return redirect('surveys:detail', pk=pk)


@require_POST
def survey_delete(request, pk):
    if not request.user.has_perm('surveys.delete_survey'):
        return redirect('surveys:list')

    survey = get_object_or_404(Survey, pk=pk)
    title = survey.title
    survey.delete()
    messages.success(request, f"Umfrage '{title}' wurde gelöscht.")
    return redirect('surveys:list')


# ======================================================================
# Fragen-Builder
# ======================================================================

def _guard_question_edit(request, survey):
    """
    Fragen dürfen nach der ersten Antwort nicht mehr verändert werden – sonst passen
    bereits gespeicherte Antworten nicht mehr zu den Fragen.
    """
    if survey.is_locked:
        messages.error(
            request,
            'Es liegen bereits Antworten vor. Die Fragen können nicht mehr geändert werden.',
        )
        return False
    return True



def _condition_builder_context(form):
    """
    Daten für die Bedingungs-Auswahl im Builder.

    Die Antwortmöglichkeiten aller in Frage kommenden Bezugsfragen wandern komplett
    ins Template – so kann Alpine die passenden Häkchen ohne weitere Server-Anfrage
    anzeigen, sobald eine Bezugsfrage gewählt wird.
    """
    selected = form['condition_values'].value() or []
    return {
        'source_options': form.condition_source_options,
        'selected_condition_values': [str(value) for value in selected],
    }


class QuestionCreateView(LoginRequiredMixin, ConductorRequiredMixin, SurveyModuleMixin, CreateView):
    model = SurveyQuestion
    form_class = SurveyQuestionForm
    template_name = 'surveys/question_form.html'

    def dispatch(self, request, *args, **kwargs):
        self.survey = get_object_or_404(Survey, pk=kwargs['survey_pk'])
        if request.user.is_authenticated and request.user.has_perm('surveys.conduct_survey'):
            if not _guard_question_edit(request, self.survey):
                return redirect('surveys:detail', pk=self.survey.pk)
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['survey'] = self.survey
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['survey'] = self.survey
        context.update(_condition_builder_context(context['form']))
        return context

    def form_valid(self, form):
        form.instance.survey = self.survey
        # Ans Ende anhängen
        last = self.survey.questions.order_by('-order').first()
        form.instance.order = (last.order + 1) if last else 0
        messages.success(self.request, 'Frage hinzugefügt.')
        return super().form_valid(form)

    def get_success_url(self):
        # Nach dem Speichern direkt die nächste Frage anlegen können
        if self.request.POST.get('save_and_new'):
            return reverse('surveys:question_create', kwargs={'survey_pk': self.survey.pk})
        return reverse('surveys:detail', kwargs={'pk': self.survey.pk})


class QuestionUpdateView(LoginRequiredMixin, ConductorRequiredMixin, SurveyModuleMixin, UpdateView):
    model = SurveyQuestion
    form_class = SurveyQuestionForm
    template_name = 'surveys/question_form.html'

    def dispatch(self, request, *args, **kwargs):
        question = get_object_or_404(SurveyQuestion, pk=kwargs['pk'])
        self.survey = question.survey
        if request.user.is_authenticated and request.user.has_perm('surveys.conduct_survey'):
            if not _guard_question_edit(request, self.survey):
                return redirect('surveys:detail', pk=self.survey.pk)
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['survey'] = self.survey
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['survey'] = self.survey
        context['is_update'] = True
        context.update(_condition_builder_context(context['form']))
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Frage gespeichert.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('surveys:detail', kwargs={'pk': self.survey.pk})


@require_POST
def question_delete(request, pk):
    if not request.user.has_perm('surveys.conduct_survey'):
        return redirect('surveys:list')

    question = get_object_or_404(SurveyQuestion, pk=pk)
    survey = question.survey
    if _guard_question_edit(request, survey):
        # Fragen, die sich auf diese hier beziehen, verlieren ihre Bedingung
        # (FK ist SET_NULL) und werden damit wieder immer angezeigt.
        dependents = list(question.dependent_questions.values_list('order', flat=True))

        question.delete()
        _renumber_questions(survey)
        messages.success(request, 'Frage gelöscht.')

        if dependents:
            positions = ', '.join(str(order + 1) for order in sorted(dependents))
            messages.warning(
                request,
                f'Die Bedingung von Frage {positions} bezog sich auf die gelöschte Frage '
                f'und wurde entfernt – diese Frage(n) werden jetzt immer angezeigt.',
            )
    return redirect('surveys:detail', pk=survey.pk)


@require_POST
def question_move(request, pk):
    """Frage eine Position nach oben oder unten schieben"""
    if not request.user.has_perm('surveys.conduct_survey'):
        return redirect('surveys:list')

    question = get_object_or_404(SurveyQuestion, pk=pk)
    survey = question.survey
    if not _guard_question_edit(request, survey):
        return redirect('surveys:detail', pk=survey.pk)

    direction = request.POST.get('direction')
    questions = list(survey.questions.all())
    index = next((i for i, q in enumerate(questions) if q.pk == question.pk), None)

    if index is not None:
        target = index - 1 if direction == 'up' else index + 1
        if 0 <= target < len(questions):
            questions[index], questions[target] = questions[target], questions[index]

            # Eine Bedingung ist nur auswertbar, wenn ihre Bezugsfrage vorher steht.
            # Einen Tausch, der das verletzt, lieber ablehnen als die Bedingung
            # stillschweigend unwirksam werden zu lassen.
            broken = _find_broken_condition(questions)
            if broken:
                dependent, source = broken
                messages.error(
                    request,
                    f'Verschieben nicht möglich: Frage „{dependent.text[:40]}" hängt von '
                    f'„{source.text[:40]}" ab, die dadurch dahinter rutschen würde. '
                    f'Bitte zuerst die Bedingung anpassen.',
                )
                return redirect('surveys:detail', pk=survey.pk)

            for position, item in enumerate(questions):
                if item.order != position:
                    item.order = position
                    item.save(update_fields=['order'])

    return redirect('surveys:detail', pk=survey.pk)


def _find_broken_condition(ordered_questions):
    """
    Erste Frage suchen, deren Bezugsfrage in dieser Reihenfolge NICHT davor steht.

    Rückgabe: (abhängige Frage, Bezugsfrage) oder None.
    """
    position_by_pk = {question.pk: index for index, question in enumerate(ordered_questions)}

    for index, question in enumerate(ordered_questions):
        source_pk = question.condition_question_id
        if not source_pk or not question.condition_values:
            continue
        source_position = position_by_pk.get(source_pk)
        if source_position is not None and source_position >= index:
            source = ordered_questions[source_position]
            return question, source

    return None


def _renumber_questions(survey):
    """Lücken in der Reihenfolge schließen (z.B. nach dem Löschen)"""
    for position, question in enumerate(survey.questions.all()):
        if question.order != position:
            question.order = position
            question.save(update_fields=['order'])


# ======================================================================
# Teilnahme
# ======================================================================


def _fill_context(survey, form):
    """
    Gemeinsamer Template-Kontext für beide Ausfüll-Wege (Login und QR-Zettel).

    `condition_config` steuert das Ein-/Ausblenden im Browser. Verbindlich ist die
    Prüfung im Formular (siehe surveys/conditions.py) – hier geht es nur um Komfort.
    """
    return {
        'survey': survey,
        'form': form,
        'field_pairs': [
            (form[name], question) for name, question in form.question_map.items()
        ],
        'condition_config': build_client_config(form.ordered_questions, field_name_for),
    }


def survey_fill(request, pk):
    """
    Fragebogen ausfüllen.

    Hier entscheidet sich, ob ein Personenbezug gespeichert wird. Antwort und
    Teilnahmevermerk werden in EINER Transaktion geschrieben, damit kein Fragebogen
    ohne Teilnahmevermerk existiert (sonst wäre eine Doppelteilnahme möglich).
    """
    if not request.user.is_authenticated:
        return redirect('core:login')

    survey = get_object_or_404(
        Survey.objects.prefetch_related('questions', 'target_groups'), pk=pk
    )
    allowed, reason = survey.can_participate(request.user)
    if not allowed:
        messages.info(request, reason)
        return redirect('surveys:list')

    if request.method == 'POST':
        form = build_response_form(survey, data=request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Teilnahmevermerk zuerst: schlägt der unique-Constraint zu
                    # (Doppelklick, zweiter Tab), wird die Antwort gar nicht erst
                    # geschrieben.
                    if survey.allow_multiple_responses:
                        SurveyParticipation.objects.get_or_create(
                            survey=survey,
                            user=request.user,
                            defaults={'participated_on': timezone.localdate()},
                        )
                    else:
                        SurveyParticipation.objects.create(
                            survey=survey,
                            user=request.user,
                            participated_on=timezone.localdate(),
                        )

                    response = SurveyResponse.objects.create(
                        survey=survey,
                        # Personenbezug NUR bei personalisierten Umfragen.
                        # SurveyResponse.save() erzwingt das zusätzlich.
                        user=None if survey.is_anonymous else request.user,
                        is_complete=True,
                    )
                    save_answers(response, form)
            except IntegrityError:
                messages.info(request, 'Sie haben an dieser Umfrage bereits teilgenommen.')
                return redirect('surveys:list')

            return redirect('surveys:thanks', pk=survey.pk)
    else:
        form = build_response_form(survey)

    context = _fill_context(survey, form)
    context['current_module'] = 'surveys'
    return render(request, 'surveys/survey_fill.html', context)


def survey_token_fill(request, token):
    """
    Fragebogen über einen Einmal-Link (QR-Zettel) ausfüllen.

    Bewusst OHNE Login-Zwang – das ist der Sinn der Zettel. Die Berechtigung steckt
    im Besitz des Tokens.

    Der Token wird in derselben Transaktion entwertet, in der die Antwort entsteht.
    `mark_used()` nutzt ein bedingtes UPDATE, damit bei zwei gleichzeitigen
    Absendungen nur genau eine durchkommt.
    """
    invitation = get_object_or_404(
        SurveyInvitation.objects.select_related('survey'), token=token
    )
    survey = invitation.survey

    def rejection(headline, text, status=403):
        return render(request, 'surveys/token_invalid.html', {
            'survey': survey,
            'headline': headline,
            'text': text,
        }, status=status)

    def already_used():
        # 410 Gone: der Zugang hat existiert, ist aber verbraucht
        return rejection(
            'Dieser Zugang wurde bereits verwendet',
            'Jeder Zettel kann nur ein einziges Mal genutzt werden – so ist '
            'sichergestellt, dass niemand doppelt abstimmt.',
            status=410,
        )

    if invitation.is_used:
        return already_used()
    if not survey.is_open:
        return rejection(
            'Diese Umfrage ist nicht geöffnet',
            'Der Zeitraum für die Teilnahme ist abgelaufen oder noch nicht gestartet.',
        )
    if not survey.questions.exists():
        return rejection(
            'Diese Umfrage enthält noch keine Fragen',
            'Bitte wenden Sie sich an die für die Umfrage verantwortliche Person.',
        )

    if request.method == 'POST':
        form = build_response_form(survey, data=request.POST)
        if form.is_valid():
            with transaction.atomic():
                if not invitation.mark_used():
                    # Parallele Absendung war schneller
                    return already_used()

                response = SurveyResponse.objects.create(
                    survey=survey,
                    # Personenbezug nur, wenn der Zettel einer Person zugeordnet ist
                    # UND die Umfrage nicht anonym ist. Beides erzwingt zusätzlich
                    # SurveyResponse.save() bzw. SurveyInvitation.save().
                    user=None if survey.is_anonymous else invitation.user,
                    is_complete=True,
                )
                save_answers(response, form)

            return render(request, 'surveys/survey_thanks.html', {
                'survey': survey,
                'base_template': 'public_base.html',
                'can_view_results': False,
                'token_mode': True,
            })
    else:
        form = build_response_form(survey)

    context = _fill_context(survey, form)
    context.update({
        # Zettel-Nutzer sind oft nicht eingeloggt – Layout ohne Sidebar
        'base_template': 'public_base.html',
        'form_action': reverse('surveys:token_fill', kwargs={'token': token}),
        'token_mode': True,
    })
    return render(request, 'surveys/survey_fill.html', context)


@require_POST
def invitation_generate(request, pk):
    """N Einmal-Zugänge erzeugen"""
    if not request.user.has_perm('surveys.conduct_survey'):
        return redirect('surveys:list')

    survey = get_object_or_404(Survey, pk=pk)

    try:
        count = int(request.POST.get('count', 0))
    except (TypeError, ValueError):
        count = 0

    if not 1 <= count <= MAX_INVITATIONS_PER_BATCH:
        messages.error(
            request,
            f'Bitte eine Anzahl zwischen 1 und {MAX_INVITATIONS_PER_BATCH} angeben.',
        )
        return redirect('surveys:invitations', pk=pk)

    label = (request.POST.get('label') or '').strip()[:100]
    if survey.is_anonymous and label:
        # Der Aufdruck wird mit dem Token gespeichert. Ein Name darin wäre exakt die
        # Zuordnung, die bei anonymen Umfragen nicht existieren darf.
        messages.warning(
            request,
            'Hinweis: Der Aufdruck wird zusammen mit dem Token gespeichert. '
            'Bitte keine Namen verwenden – die Umfrage ist anonym.',
        )

    # bulk_create umgeht SurveyInvitation.save() – deshalb hier das Token explizit
    # erzeugen und `user` bewusst nicht setzen.
    SurveyInvitation.objects.bulk_create([
        SurveyInvitation(
            survey=survey,
            token=SurveyInvitation.generate_token(),
            label=label,
        )
        for _ in range(count)
    ])

    messages.success(request, f'{count} Zugangs-Links erzeugt.')
    return redirect('surveys:invitations', pk=pk)


class SurveyInvitationsView(LoginRequiredMixin, ConductorRequiredMixin, SurveyModuleMixin, DetailView):
    """Verwaltung der Einmal-Zugänge einer Umfrage"""
    model = Survey
    template_name = 'surveys/survey_invitations.html'
    context_object_name = 'survey'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        invitations = self.object.invitations.all()
        context['invitations'] = invitations
        context['open_count'] = invitations.filter(is_used=False).count()
        context['used_count'] = invitations.filter(is_used=True).count()
        context['max_batch'] = MAX_INVITATIONS_PER_BATCH
        return context


def invitation_pdf(request, pk):
    """QR-Zettel als druckfertiges A4-PDF"""
    if not request.user.has_perm('surveys.conduct_survey'):
        return redirect('surveys:list')

    survey = get_object_or_404(Survey, pk=pk)

    invitations = survey.invitations.all()
    if request.GET.get('only') == 'open':
        invitations = invitations.filter(is_used=False)

    if not invitations.exists():
        messages.error(request, 'Es sind keine Zugangs-Links vorhanden.')
        return redirect('surveys:invitations', pk=pk)

    from .services_pdf import render_invitation_pdf

    # Absolute URLs, sonst lassen sich die QR-Codes mit dem Handy nicht öffnen
    base_url = request.build_absolute_uri('/').rstrip('/')
    pdf_bytes = render_invitation_pdf(survey, invitations, base_url)

    filename = f"umfrage_{survey.pk}_zettel_{timezone.localdate():%Y%m%d}.pdf"
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def survey_thanks(request, pk):
    """Bestätigungsseite nach dem Absenden"""
    if not request.user.is_authenticated:
        return redirect('core:login')

    survey = get_object_or_404(Survey, pk=pk)
    return render(request, 'surveys/survey_thanks.html', {
        'current_module': 'surveys',
        'survey': survey,
        'can_view_results': survey.can_view_results(request.user),
    })


# ======================================================================
# Auswertung
# ======================================================================

def survey_results(request, pk):
    """Aggregierte Auswertung"""
    if not request.user.is_authenticated:
        return redirect('core:login')

    survey = get_object_or_404(Survey.objects.prefetch_related('questions'), pk=pk)
    if not survey.can_view_results(request.user):
        messages.error(request, 'Sie dürfen die Auswertung dieser Umfrage nicht sehen.')
        return redirect('surveys:list')

    response_count = survey.response_count
    unlocked = survey.results_unlocked
    statistics = build_question_statistics(survey) if unlocked else []

    # Diagrammdaten als JSON für Chart.js (liegt bereits lokal unter static/vendor/js)
    chart_data = {
        str(entry['question'].pk): {
            'labels': entry['categories'],
            'counts': entry['counts'],
        }
        for entry in statistics
        if entry['kind'] == 'chart' and not entry['suppressed']
    }

    return render(request, 'surveys/survey_results.html', {
        'current_module': 'surveys',
        'survey': survey,
        'statistics': statistics,
        'response_count': response_count,
        'results_unlocked': unlocked,
        'missing_for_unlock': max(0, survey.min_responses_for_results - response_count),
        'chart_data': chart_data,
        'is_conductor': request.user.has_perm('surveys.conduct_survey'),
    })


class SurveyParticipantsView(LoginRequiredMixin, ConductorRequiredMixin, SurveyModuleMixin, DetailView):
    """
    Teilnahmeübersicht: wer hat teilgenommen, wer fehlt noch.

    Zeigt bewusst KEINE Antworten – bei anonymen Umfragen existiert die Verbindung
    zwischen Person und Antwort gar nicht.
    """
    model = Survey
    template_name = 'surveys/survey_participants.html'
    context_object_name = 'survey'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        survey = self.object
        context['participations'] = survey.participations.select_related('user')
        context['missing'] = get_missing_participants(survey)
        return context


def survey_export(request, pk):
    """Excel-Export der Antworten (openpyxl ist bereits im Projekt vorhanden)"""
    if not request.user.has_perm('surveys.conduct_survey'):
        return redirect('surveys:list')

    survey = get_object_or_404(Survey.objects.prefetch_related('questions'), pk=pk)

    if not survey.results_unlocked:
        messages.error(
            request,
            'Der Export ist gesperrt, solange die Mindestanzahl an Antworten nicht erreicht ist.',
        )
        return redirect('surveys:results', pk=pk)

    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    header, rows = build_export_rows(survey)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Antworten'

    sheet.append(header)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append(row)

    # Spaltenbreiten grob an den Inhalt anpassen
    for index, title in enumerate(header, start=1):
        lengths = [len(str(title))] + [len(str(row[index - 1])) for row in rows]
        sheet.column_dimensions[get_column_letter(index)].width = min(max(max(lengths) + 2, 12), 60)

    sheet.freeze_panes = 'A2'

    filename = f"umfrage_{survey.pk}_{timezone.localdate():%Y%m%d}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response
