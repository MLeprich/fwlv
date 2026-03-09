"""
Höhenrettung Views
Views für Höhenrettungs-Ausrüstungs-Verwaltung
"""

from django import forms
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.db.models import Q, Count, F, Sum
from django.utils import timezone
from django.http import HttpResponse
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import csv
import io
import logging

logger = logging.getLogger(__name__)

from .models import (
    EquipmentType,
    HeightRescueItemMaster,
    HeightRescueDeviceInstance,
    HeightRescueStockMovement,
    HeightRescueInspectionLog,
    InspectionStatus,
    HeightRescueItemType,
    HeightRescueInspectionType,
    HeightRescueInspectionAssignment,
    HeightRescueInspectionRecord,
)
from personnel.models import Person, Qualification, DutyHoursEntry, DutyHoursRequirement
from .forms import (
    EquipmentTypeForm,
    HeightRescueItemMasterForm,
    HeightRescueDeviceInstanceForm,
    HeightRescueInspectionTypeForm,
    HeightRescueInspectionAssignmentForm,
    HeightRescueInspectionRecordForm,
    HeightRescueInspectionLogForm,
    HeightRescueStockMovementForm,
)

# QR-Code und Barcode Generation
import qrcode
from qrcode.image.svg import SvgPathImage
import barcode
from barcode.writer import SVGWriter
from django.http import HttpResponse
from io import BytesIO


# ============================================================================
# DASHBOARD
# ============================================================================

def dashboard(request):
    """Dashboard für Höhenrettung mit neuer Master Data Struktur"""
    today = timezone.now().date()

    # NEU: Stammdaten und Geräte-Instanzen
    total_masters = HeightRescueItemMaster.objects.filter(is_active=True).count()
    total_devices = HeightRescueDeviceInstance.objects.filter(is_active=True).count()

    # Geräte nach Status
    operational_count = HeightRescueDeviceInstance.objects.filter(
        is_active=True,
        is_operational=True
    ).count()

    inspection_due_devices = HeightRescueDeviceInstance.objects.filter(
        is_active=True,
        next_inspection_date__isnull=False,
        next_inspection_date__lte=today + timedelta(days=30)
    ).count()

    retired_devices = HeightRescueDeviceInstance.objects.filter(
        retirement_date__isnull=False
    ).count() if HeightRescueDeviceInstance.objects.count() > 0 else 0

    context = {
        'total_masters': total_masters,
        'total_devices': total_devices,
        'operational_count': operational_count,
        'inspection_due_devices': inspection_due_devices,
        'retired_devices': retired_devices,
        'today': today,
    }

    return render(request, 'height_rescue/dashboard.html', context)


# ============================================================================
# AUSRÜSTUNGSTYPEN-VERWALTUNG
# ============================================================================

class EquipmentTypeListView(LoginRequiredMixin, ListView):
    """Liste aller Ausrüstungstypen"""
    model = EquipmentType
    template_name = 'height_rescue/equipment_type_list.html'
    context_object_name = 'types'

    def get_queryset(self):
        queryset = EquipmentType.objects.all().order_by('order', 'name')

        # Filter nach Status
        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)

        # Suche
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(code__icontains=search) |
                Q(description__icontains=search)
            )

        return queryset


class EquipmentTypeCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Neuen Ausrüstungstyp erstellen"""
    model = EquipmentType
    form_class = EquipmentTypeForm
    template_name = 'height_rescue/equipment_type_form.html'
    success_url = reverse_lazy('height_rescue:equipment_type_list')
    permission_required = 'height_rescue.add_equipmenttype'

    def form_valid(self, form):
        messages.success(self.request, f'Ausrüstungstyp "{form.instance.name}" wurde erfolgreich erstellt.')
        return super().form_valid(form)


class EquipmentTypeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Ausrüstungstyp bearbeiten"""
    model = EquipmentType
    form_class = EquipmentTypeForm
    template_name = 'height_rescue/equipment_type_form.html'
    success_url = reverse_lazy('height_rescue:equipment_type_list')
    permission_required = 'height_rescue.change_equipmenttype'

    def form_valid(self, form):
        messages.success(self.request, f'Ausrüstungstyp "{form.instance.name}" wurde erfolgreich aktualisiert.')
        return super().form_valid(form)


class EquipmentTypeDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Ausrüstungstyp löschen"""
    model = EquipmentType
    template_name = 'height_rescue/equipment_type_confirm_delete.html'
    success_url = reverse_lazy('height_rescue:equipment_type_list')
    permission_required = 'height_rescue.delete_equipmenttype'

    def delete(self, request, *args, **kwargs):
        equipment_type = self.get_object()
        messages.success(request, f'Ausrüstungstyp "{equipment_type.name}" wurde erfolgreich gelöscht.')
        return super().delete(request, *args, **kwargs)


# ============================================================================
# MASTER DATA (STAMMDATEN) VIEWS
# ============================================================================

class HeightRescueMasterListView(LoginRequiredMixin, ListView):
    """Liste aller Höhenrettungs-Stammdaten"""
    model = HeightRescueItemMaster
    template_name = 'height_rescue/master_list.html'
    context_object_name = 'masters'
    paginate_by = 50

    def get_queryset(self):
        queryset = HeightRescueItemMaster.objects.filter(is_active=True)

        # Suche
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(master_number__icontains=search) |
                Q(manufacturer__icontains=search) |
                Q(model__icontains=search)
            )

        # Typ-Filter
        item_type = self.request.GET.get('item_type')
        if item_type:
            queryset = queryset.filter(item_type=item_type)

        # Hersteller-Filter
        manufacturer = self.request.GET.get('manufacturer')
        if manufacturer:
            queryset = queryset.filter(manufacturer__icontains=manufacturer)

        # Seil-Typ Filter
        rope_type = self.request.GET.get('rope_type')
        if rope_type:
            queryset = queryset.filter(rope_type=rope_type)

        return queryset.order_by('item_type', 'name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_types'] = HeightRescueItemType.choices
        return context


class HeightRescueMasterDetailView(LoginRequiredMixin, DetailView):
    """Detail-Ansicht Höhenrettungs-Stammdaten"""
    model = HeightRescueItemMaster
    template_name = 'height_rescue/master_detail.html'
    context_object_name = 'master'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        master = self.get_object()

        # Geräte-Instanzen
        context['devices'] = master.device_instances.filter(is_active=True).select_related(
            'location',
            'assigned_vehicle',
            'assigned_to',
            'last_inspector'
        ).order_by('-created_at')

        context['total_devices'] = master.device_instances.filter(is_active=True).count()
        context['operational_devices'] = master.device_instances.filter(is_active=True, is_operational=True).count()

        return context


class HeightRescueMasterCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Neue Höhenrettungs-Stammdaten anlegen"""
    model = HeightRescueItemMaster
    form_class = HeightRescueItemMasterForm
    template_name = 'height_rescue/master_form.html'
    permission_required = 'height_rescue.add_heightrescueitemmaster'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, 'Höhenrettungs-Stammdaten erfolgreich erstellt.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('height_rescue:master_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Neue Höhenrettungs-Stammdaten'
        context['submit_text'] = 'Stammdaten speichern'
        return context


class HeightRescueMasterUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Höhenrettungs-Stammdaten bearbeiten"""
    model = HeightRescueItemMaster
    form_class = HeightRescueItemMasterForm
    template_name = 'height_rescue/master_form.html'
    permission_required = 'height_rescue.change_heightrescueitemmaster'

    def form_valid(self, form):
        messages.success(self.request, 'Höhenrettungs-Stammdaten erfolgreich aktualisiert.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('height_rescue:master_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Höhenrettungs-Stammdaten bearbeiten'
        context['submit_text'] = 'Änderungen speichern'
        return context


class HeightRescueMasterQRCodeView(LoginRequiredMixin, DetailView):
    """QR-Code für Höhenrettungs-Stammdaten"""
    model = HeightRescueItemMaster

    def get(self, request, *args, **kwargs):
        master = self.get_object()

        # URL zum Master Detail
        detail_url = request.build_absolute_uri(
            reverse_lazy('height_rescue:master_detail', kwargs={'pk': master.pk})
        )

        # QR-Code generieren
        qr = qrcode.QRCode(version=1, box_size=10, border=2)
        qr.add_data(detail_url)
        qr.make(fit=True)

        img = qr.make_image(image_factory=SvgPathImage)
        buffer = BytesIO()
        img.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='image/svg+xml')
        response['Content-Disposition'] = f'inline; filename="qr_{master.master_number}.svg"'
        return response


class HeightRescueMasterBarcodeView(LoginRequiredMixin, DetailView):
    """Barcode für Höhenrettungs-Stammdaten"""
    model = HeightRescueItemMaster

    def get(self, request, *args, **kwargs):
        master = self.get_object()

        # Barcode generieren
        code128 = barcode.get_barcode_class('code128')
        barcode_instance = code128(master.master_number, writer=SVGWriter())

        buffer = BytesIO()
        barcode_instance.write(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='image/svg+xml')
        response['Content-Disposition'] = f'inline; filename="barcode_{master.master_number}.svg"'
        return response


# ============================================================================
# DEVICE INSTANCE (GERÄTE-INSTANZ) VIEWS
# ============================================================================

class HeightRescueDeviceListView(LoginRequiredMixin, ListView):
    """Liste aller Höhenrettungsgeräte-Instanzen"""
    model = HeightRescueDeviceInstance
    template_name = 'height_rescue/device_list.html'
    context_object_name = 'devices'
    paginate_by = 50

    def get_queryset(self):
        queryset = HeightRescueDeviceInstance.objects.filter(is_active=True).select_related(
            'master',
            'location',
            'assigned_vehicle',
            'assigned_to',
            'last_inspector'
        )

        # Suche
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(inventory_number__icontains=search) |
                Q(serial_number__icontains=search) |
                Q(master__name__icontains=search)
            )

        # Master-Filter
        master = self.request.GET.get('master')
        if master:
            queryset = queryset.filter(master_id=master)

        # Lagerort-Filter
        location = self.request.GET.get('location')
        if location:
            queryset = queryset.filter(location_id=location)

        # Zustand-Filter
        condition = self.request.GET.get('condition')
        if condition:
            queryset = queryset.filter(condition=condition)

        # Einsatzbereit-Filter
        operational = self.request.GET.get('operational')
        if operational == '1':
            queryset = queryset.filter(is_operational=True)
        elif operational == '0':
            queryset = queryset.filter(is_operational=False)

        # Prüfung fällig
        if self.request.GET.get('inspection_due'):
            today = timezone.now().date()
            queryset = queryset.filter(
                next_inspection_date__isnull=False,
                next_inspection_date__lte=today + timedelta(days=30)
            )

        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        from locations.models import Location
        context = super().get_context_data(**kwargs)
        context['masters'] = HeightRescueItemMaster.objects.filter(is_active=True).order_by('name')
        context['locations'] = Location.objects.filter(is_active=True).order_by('name')
        return context


class HeightRescueDeviceDetailView(LoginRequiredMixin, DetailView):
    """Detail-Ansicht Höhenrettungsgerät-Instanz"""
    model = HeightRescueDeviceInstance
    template_name = 'height_rescue/device_detail.html'
    context_object_name = 'device'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        device = self.get_object()
        today = timezone.now().date()

        context['today'] = today
        context['is_inspection_due'] = device.is_inspection_due()
        context['is_inspection_due_soon'] = device.is_inspection_due_soon()
        context['should_be_retired'] = device.should_be_retired()

        return context


class HeightRescueDeviceCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Neues Höhenrettungsgerät-Instanz anlegen"""
    model = HeightRescueDeviceInstance
    form_class = HeightRescueDeviceInstanceForm
    template_name = 'height_rescue/device_form.html'
    permission_required = 'height_rescue.add_heightrescuedeviceinstance'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, 'Höhenrettungsgerät erfolgreich erstellt.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('height_rescue:device_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Neues Höhenrettungsgerät'
        context['submit_text'] = 'Gerät speichern'
        return context


class HeightRescueDeviceUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Höhenrettungsgerät-Instanz bearbeiten"""
    model = HeightRescueDeviceInstance
    form_class = HeightRescueDeviceInstanceForm
    template_name = 'height_rescue/device_form.html'
    permission_required = 'height_rescue.change_heightrescuedeviceinstance'

    def form_valid(self, form):
        messages.success(self.request, 'Höhenrettungsgerät erfolgreich aktualisiert.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('height_rescue:device_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Höhenrettungsgerät bearbeiten'
        context['submit_text'] = 'Änderungen speichern'
        return context


class HeightRescueDeviceQRCodeView(LoginRequiredMixin, DetailView):
    """QR-Code für Höhenrettungsgerät-Instanz"""
    model = HeightRescueDeviceInstance

    def get(self, request, *args, **kwargs):
        device = self.get_object()

        # URL zum Device Detail
        detail_url = request.build_absolute_uri(
            reverse_lazy('height_rescue:device_detail', kwargs={'pk': device.pk})
        )

        # QR-Code generieren
        qr = qrcode.QRCode(version=1, box_size=10, border=2)
        qr.add_data(detail_url)
        qr.make(fit=True)

        img = qr.make_image(image_factory=SvgPathImage)
        buffer = BytesIO()
        img.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='image/svg+xml')
        response['Content-Disposition'] = f'inline; filename="qr_{device.inventory_number}.svg"'
        return response


class HeightRescueDeviceBarcodeView(LoginRequiredMixin, DetailView):
    """Barcode für Höhenrettungsgerät-Instanz"""
    model = HeightRescueDeviceInstance

    def get(self, request, *args, **kwargs):
        device = self.get_object()

        # Barcode generieren
        code128 = barcode.get_barcode_class('code128')
        barcode_instance = code128(device.inventory_number, writer=SVGWriter())

        buffer = BytesIO()
        barcode_instance.write(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='image/svg+xml')
        response['Content-Disposition'] = f'inline; filename="barcode_{device.inventory_number}.svg"'
        return response


# ============================================================================
# LEGACY ITEM VIEWS
# ============================================================================

class HeightRescueDashboardView(LoginRequiredMixin, TemplateView):
    """Dashboard für Höhenrettung (Umgestellt auf Master/Device System)"""
    template_name = 'height_rescue/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()

        # KPIs
        context['total_items'] = HeightRescueDeviceInstance.objects.filter(is_active=True).count()

        # Prüfung fällig (innerhalb 30 Tage)
        context['inspection_due_count'] = HeightRescueDeviceInstance.objects.filter(
            is_active=True,
            next_inspection_date__isnull=False,
            next_inspection_date__lte=today + timedelta(days=30)
        ).count()

        # Ausgemustert
        context['retired_count'] = HeightRescueDeviceInstance.objects.filter(
            retirement_date__isnull=False
        ).count()

        # Prüfungstypen
        context['inspection_types_count'] = HeightRescueInspectionType.objects.filter(
            is_active=True
        ).count()

        # Letzte Bewegungen
        context['recent_movements'] = HeightRescueStockMovement.objects.select_related(
            'device',
            'created_by'
        ).order_by('-movement_date', '-created_at')[:10]

        # Kritische Prüfungen (überfällig >30 Tage)
        context['overdue_inspections'] = HeightRescueDeviceInstance.objects.filter(
            is_active=True,
            next_inspection_date__isnull=False,
            next_inspection_date__lt=today - timedelta(days=30)
        ).select_related('master')[:5]

        # Prüfungen demnächst fällig (nächste 14 Tage)
        context['upcoming_inspections'] = HeightRescueDeviceInstance.objects.filter(
            is_active=True,
            next_inspection_date__isnull=False,
            next_inspection_date__gte=today,
            next_inspection_date__lte=today + timedelta(days=14)
        ).order_by('next_inspection_date')[:5]

        # Ausrüstung kurz vor Aussonderung (Alter >80% der max. Nutzungsdauer)
        items_near_retirement = []
        for device in HeightRescueDeviceInstance.objects.filter(
            is_active=True,
            master__max_service_life_years__isnull=False,
            manufacturing_date__isnull=False
        ).select_related('master'):
            age = device.get_age_years() if hasattr(device, 'get_age_years') else None
            if age and device.master.max_service_life_years and age >= (device.master.max_service_life_years * 0.8):
                items_near_retirement.append(device)
        context['items_near_retirement'] = items_near_retirement[:5]

        # Letzte Prüfungen
        context['recent_inspections'] = HeightRescueInspectionLog.objects.select_related(
            'device',
            'inspector'
        ).order_by('-inspection_date')[:10]

        # Statistiken nach Typ (über Master-Typen)
        context['stats_by_type'] = HeightRescueDeviceInstance.objects.filter(
            is_active=True
        ).values('master__equipment_type__name').annotate(
            count=Count('id')
        ).order_by('-count')[:5]

        # ========================================================================
        # PERSONAL: Qualifikationen & Pflichtstunden
        # ========================================================================

        # Personen mit Höhenrettungs-Funktion finden
        height_rescue_personnel = Person.objects.filter(
            is_active=True,
            functions__related_module='height_rescue',
            functions__is_active=True
        ).distinct().prefetch_related('qualifications', 'duty_hours')

        # Ablaufende Qualifikationen (nächste 60 Tage oder bereits abgelaufen)
        expiring_qualifications = []
        for person in height_rescue_personnel:
            for qualification in person.qualifications.filter(is_active=True, expiry_date__isnull=False):
                days_until_expiry = (qualification.expiry_date - today).days
                if days_until_expiry <= 60:  # Läuft in 60 Tagen oder weniger ab
                    expiring_qualifications.append({
                        'person': person,
                        'qualification': qualification,
                        'days_until_expiry': days_until_expiry,
                        'is_expired': days_until_expiry < 0,
                    })

        # Nach Dringlichkeit sortieren (abgelaufen zuerst, dann nach Tagen)
        expiring_qualifications.sort(key=lambda x: (not x['is_expired'], x['days_until_expiry']))
        context['expiring_qualifications'] = expiring_qualifications[:10]

        # Fehlende Pflichtstunden für aktuelles Jahr
        current_year = today.year
        missing_duty_hours = []

        for person in height_rescue_personnel:
            # Alle Pflichtstunden-Kategorien finden, die für diese Person relevant sind
            required_categories = set()
            for function in person.functions.filter(related_module='height_rescue', is_active=True):
                for category in function.required_duty_hour_categories.all():
                    required_categories.add(category)

            # Für jede Kategorie prüfen, ob Anforderungen erfüllt sind
            for category in required_categories:
                # Anforderung für dieses Jahr finden
                try:
                    requirement = DutyHoursRequirement.objects.get(
                        category=category,
                        year=current_year,
                        is_active=True
                    )

                    # Geleistete Stunden berechnen
                    completed_hours = DutyHoursEntry.objects.filter(
                        person=person,
                        category=category,
                        date__year=current_year
                    ).aggregate(total=Sum('hours'))['total'] or 0

                    # Fehlende Stunden berechnen
                    missing_hours = requirement.required_hours - completed_hours

                    if missing_hours > 0:
                        missing_duty_hours.append({
                            'person': person,
                            'category': category,
                            'required_hours': requirement.required_hours,
                            'completed_hours': completed_hours,
                            'missing_hours': missing_hours,
                        })
                except DutyHoursRequirement.DoesNotExist:
                    # Keine Anforderung für dieses Jahr definiert
                    pass

        # Nach fehlenden Stunden sortieren (meiste fehlende Stunden zuerst)
        missing_duty_hours.sort(key=lambda x: x['missing_hours'], reverse=True)
        context['missing_duty_hours'] = missing_duty_hours[:10]

        return context


# ============================================================================
# ITEM VIEWS
# ============================================================================


# ============================================================================
# INSPECTION & RETIREMENT VIEWS (auf Device-System umgestellt)
# ============================================================================

class HeightRescueInspectionDueView(LoginRequiredMixin, ListView):
    """Liste der prüfpflichtigen Geräte-Instanzen"""
    model = HeightRescueDeviceInstance
    template_name = 'height_rescue/inspection_due.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        today = timezone.now().date()
        return HeightRescueDeviceInstance.objects.filter(
            is_active=True,
            next_inspection_date__isnull=False,
            next_inspection_date__lte=today + timedelta(days=30)
        ).select_related(
            'master',
            'location',
            'last_inspector'
        ).order_by('next_inspection_date')


class HeightRescueRetiredListView(LoginRequiredMixin, ListView):
    """Liste der ausgemusterten Geräte"""
    model = HeightRescueDeviceInstance
    template_name = 'height_rescue/retired_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        return HeightRescueDeviceInstance.objects.filter(
            retirement_date__isnull=False
        ).select_related(
            'master',
            'location',
            'last_inspector'
        ).order_by('-retirement_date')


class HeightRescueFallListView(LoginRequiredMixin, ListView):
    """Liste der Geräte, die Stürze aufgefangen haben"""
    model = HeightRescueDeviceInstance
    template_name = 'height_rescue/fall_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        return HeightRescueDeviceInstance.objects.filter(
            total_falls_arrested__gt=0
        ).select_related(
            'master',
            'location',
            'last_inspector'
        ).order_by('-total_falls_arrested')


# ============================================================================
# STOCK MOVEMENT VIEWS
# ============================================================================

class HeightRescueMovementListView(LoginRequiredMixin, ListView):
    """Liste aller Lagerbewegungen"""
    model = HeightRescueStockMovement
    template_name = 'height_rescue/movement_list.html'
    context_object_name = 'movements'
    paginate_by = 50

    def get_queryset(self):
        queryset = HeightRescueStockMovement.objects.select_related(
            'item',
            'created_by',
            'inspector'
        )

        # Typ-Filter
        movement_type = self.request.GET.get('type')
        if movement_type:
            queryset = queryset.filter(movement_type=movement_type)

        # Item-Filter
        item_id = self.request.GET.get('item')
        if item_id:
            queryset = queryset.filter(item_id=item_id)

        # Prüfung-Filter
        inspection_performed = self.request.GET.get('inspection')
        if inspection_performed == 'yes':
            queryset = queryset.filter(inspection_performed=True)
        elif inspection_performed == 'no':
            queryset = queryset.filter(inspection_performed=False)

        # Sturz-Filter
        fall_arrested = self.request.GET.get('fall')
        if fall_arrested == 'yes':
            queryset = queryset.filter(fall_arrested=True)

        return queryset.order_by('-movement_date', '-created_at')


class HeightRescueMovementDetailView(LoginRequiredMixin, DetailView):
    """Detailansicht einer Lagerbewegung"""
    model = HeightRescueStockMovement
    template_name = 'height_rescue/movement_detail.html'
    context_object_name = 'movement'


class HeightRescueMovementCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Neue Lagerbewegung erstellen"""
    model = HeightRescueStockMovement
    form_class = HeightRescueStockMovementForm
    template_name = 'height_rescue/movement_form.html'
    permission_required = 'height_rescue.add_heightrescuestockmovement'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Lagerbewegung wurde erfolgreich erstellt.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('height_rescue:movement_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        import json
        context = super().get_context_data(**kwargs)
        context['title'] = 'Neue Lagerbewegung'
        context['submit_text'] = 'Bewegung speichern'
        from .models import HeightRescueItem
        item_locations = dict(
            HeightRescueItem.objects.filter(location__isnull=False)
            .values_list('pk', 'location_id')
        )
        context['item_location_map'] = json.dumps(item_locations)
        return context


class HeightRescueInspectionListView(LoginRequiredMixin, ListView):
    """Liste aller Prüfprotokolle"""
    model = HeightRescueInspectionLog
    template_name = 'height_rescue/inspection_list.html'
    context_object_name = 'inspections'
    paginate_by = 50

    def get_queryset(self):
        return HeightRescueInspectionLog.objects.select_related(
            'item',
            'inspector'
        ).order_by('-inspection_date')


class HeightRescueInspectionLogCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Prüfprotokoll erstellen"""
    model = HeightRescueInspectionLog
    form_class = HeightRescueInspectionLogForm
    template_name = 'height_rescue/inspection_log_form.html'
    success_url = reverse_lazy('height_rescue:inspection_list')
    permission_required = 'height_rescue.add_heightrescueinspectionlog'

    def form_valid(self, form):
        # Item-Status aktualisieren wenn Prüfung bestanden
        item = form.instance.item
        if form.instance.passed:
            item.last_inspection_date = form.instance.inspection_date
            item.next_inspection_date = form.instance.next_inspection_due
            item.last_inspector = form.instance.inspector
            item.inspection_status = InspectionStatus.OK
            item.save()
        else:
            # Wenn nicht bestanden, auf FAILED setzen
            item.inspection_status = InspectionStatus.FAILED
            item.save()

        messages.success(
            self.request,
            f'Prüfprotokoll für {item.name} erfolgreich erstellt.'
        )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Prüfung dokumentieren'
        context['submit_text'] = 'Protokoll speichern'
        return context


# ============================================================================
# INSPECTION TYPE MANAGEMENT
# ============================================================================

class InspectionTypeListView(LoginRequiredMixin, ListView):
    """Liste aller Prüfungstypen"""
    model = HeightRescueInspectionType
    template_name = 'height_rescue/inspection_type_list.html'
    context_object_name = 'inspection_types'

    def get_queryset(self):
        return HeightRescueInspectionType.objects.filter(is_active=True).order_by('name')


class InspectionTypeCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Prüfungstyp erstellen"""
    model = HeightRescueInspectionType
    form_class = HeightRescueInspectionTypeForm
    template_name = 'height_rescue/inspection_type_form.html'
    success_url = reverse_lazy('height_rescue:inspection_type_list')
    permission_required = 'height_rescue.add_inspectiontype'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Prüfungstyp erfolgreich erstellt.'))
        return super().form_valid(form)


class InspectionTypeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Prüfungstyp bearbeiten"""
    model = HeightRescueInspectionType
    form_class = HeightRescueInspectionTypeForm
    template_name = 'height_rescue/inspection_type_form.html'
    success_url = reverse_lazy('height_rescue:inspection_type_list')
    permission_required = 'height_rescue.change_inspectiontype'

    def form_valid(self, form):
        messages.success(self.request, _('Prüfungstyp erfolgreich aktualisiert.'))
        return super().form_valid(form)


class InspectionTypeDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Prüfungstyp löschen"""
    model = HeightRescueInspectionType
    template_name = 'height_rescue/inspection_type_confirm_delete.html'
    success_url = reverse_lazy('height_rescue:inspection_type_list')
    permission_required = 'height_rescue.delete_inspectiontype'

    def delete(self, request, *args, **kwargs):
        messages.success(request, _('Prüfungstyp erfolgreich gelöscht.'))
        return super().delete(request, *args, **kwargs)


class InspectionManagementView(LoginRequiredMixin, TemplateView):
    """Zentrale Prüfungsverwaltung"""
    template_name = 'height_rescue/inspection_management.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Prüfungstypen
        context['inspection_types'] = HeightRescueInspectionType.objects.filter(
            is_active=True
        ).annotate(
            assignment_count=Count('item_assignments')
        ).order_by('name')

        # Statistiken
        context['total_types'] = context['inspection_types'].count()
        context['total_assignments'] = HeightRescueInspectionAssignment.objects.filter(
            is_active=True
        ).count()
        context['total_records'] = HeightRescueInspectionRecord.objects.count()

        return context


# ============================================================================
# CATEGORY VIEWS
# ============================================================================

class CategoryListView(LoginRequiredMixin, ListView):
    """Liste aller Kategorien"""
    template_name = 'height_rescue/category_list.html'
    context_object_name = 'categories'

    def get_queryset(self):
        from inventory_base.models import Category
        return Category.objects.filter(is_active=True).order_by('tree_id', 'lft')


class CategoryCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Neue Kategorie erstellen"""
    template_name = 'height_rescue/category_form.html'
    fields = ['name', 'parent', 'code', 'description']
    success_url = reverse_lazy('height_rescue:category_list')
    permission_required = 'inventory_base.add_category'

    def get_queryset(self):
        from inventory_base.models import Category
        return Category.objects.all()

    def get_form(self, form_class=None):
        from inventory_base.models import Category
        form = super().get_form(form_class)
        form.instance.__class__ = Category
        return form

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, f'Kategorie "{form.instance.name}" wurde erstellt.')
        return super().form_valid(form)


class CategoryUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Kategorie bearbeiten"""
    template_name = 'height_rescue/category_form.html'
    fields = ['name', 'parent', 'code', 'description', 'is_active']
    success_url = reverse_lazy('height_rescue:category_list')
    permission_required = 'inventory_base.change_category'

    def get_queryset(self):
        from inventory_base.models import Category
        return Category.objects.all()

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, f'Kategorie "{form.instance.name}" wurde aktualisiert.')
        return super().form_valid(form)


class CategoryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Kategorie löschen"""
    template_name = 'height_rescue/category_confirm_delete.html'
    success_url = reverse_lazy('height_rescue:category_list')
    permission_required = 'inventory_base.delete_category'

    def get_queryset(self):
        from inventory_base.models import Category
        return Category.objects.all()

    def delete(self, request, *args, **kwargs):
        category = self.get_object()
        messages.success(request, f'Kategorie "{category.name}" wurde gelöscht.')
        return super().delete(request, *args, **kwargs)


# ========================================================================
# INVENTUR VIEWS
# ========================================================================
from .inventory_views import (
    HeightRescueInventoryListView,
    HeightRescueInventoryCreateView,
    HeightRescueInventoryDetailView,
    HeightRescueInventoryStartView,
    HeightRescueInventoryCountingView,
    HeightRescueInventoryCompleteView,
    HeightRescueInventoryApproveView,
    HeightRescueInventoryItemUpdateView,
    HeightRescueInventoryProgressView,
    HeightRescueInventoryExportView,
)


# ============================================================================
# IMPORT/EXPORT
# ============================================================================

class ImportExportView(LoginRequiredMixin, TemplateView):
    """Import/Export Übersicht"""
    template_name = 'height_rescue/import_export.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_masters'] = HeightRescueItemMaster.objects.filter(is_active=True).count()
        context['total_devices'] = HeightRescueDeviceInstance.objects.filter(is_active=True).count()
        context['total_movements'] = HeightRescueStockMovement.objects.count()
        context['total_inspections'] = HeightRescueInspectionLog.objects.count()
        return context


# ============================================================================
# EXPORT FUNCTIONS (Excel)
# ============================================================================

@login_required
def export_masters(request):
    """Export Height Rescue Masters als Excel"""
    masters = HeightRescueItemMaster.objects.filter(is_active=True).select_related(
        'created_by', 'updated_by'
    ).order_by('master_number')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stammdaten"

    # Header
    headers = [
        'Stammdaten-Nr.', 'Name', 'Typ', 'Hersteller', 'Modell', 'Herstellerartikelnr.',
        'EN-Zertifizierung', 'Zertifizierungsnummer', 'Max. Nutzungsdauer (Jahre)',
        'Seiltyp', 'Seildurchmesser (mm)', 'Seillänge (m)', 'Gewicht (kg)',
        'Beschreibung', 'Erstellt am', 'Aktualisiert am'
    ]

    # Header-Styling
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Daten
    for row_num, master in enumerate(masters, start=2):
        ws.cell(row=row_num, column=1, value=master.master_number)
        ws.cell(row=row_num, column=2, value=master.name)
        ws.cell(row=row_num, column=3, value=master.get_item_type_display())
        ws.cell(row=row_num, column=4, value=master.manufacturer)
        ws.cell(row=row_num, column=5, value=master.model)
        ws.cell(row=row_num, column=6, value=master.manufacturer_part_number)
        ws.cell(row=row_num, column=7, value=master.get_en_certification_display() if master.en_certification else '')
        ws.cell(row=row_num, column=8, value=master.certification_number)
        ws.cell(row=row_num, column=9, value=master.max_service_life_years)
        ws.cell(row=row_num, column=10, value=master.get_rope_type_display() if master.rope_type else '')
        ws.cell(row=row_num, column=11, value=float(master.rope_diameter_mm) if master.rope_diameter_mm else '')
        ws.cell(row=row_num, column=12, value=float(master.rope_length_m) if master.rope_length_m else '')
        ws.cell(row=row_num, column=13, value=float(master.weight_kg) if master.weight_kg else '')
        ws.cell(row=row_num, column=14, value=master.description)
        ws.cell(row=row_num, column=15, value=master.created_at.strftime('%d.%m.%Y %H:%M') if master.created_at else '')
        ws.cell(row=row_num, column=16, value=master.updated_at.strftime('%d.%m.%Y %H:%M') if master.updated_at else '')

    # Spaltenbreiten
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=hoehenrettung_stammdaten_export.xlsx'
    wb.save(response)
    return response


@login_required
def export_devices(request):
    """Export Height Rescue Devices als Excel"""
    devices = HeightRescueDeviceInstance.objects.filter(is_active=True).select_related(
        'master', 'location', 'assigned_vehicle', 'assigned_to', 'created_by', 'updated_by'
    ).order_by('inventory_number')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Geräte-Instanzen"

    # Header
    headers = [
        'Inventarnummer', 'Seriennummer', 'Stammdaten', 'Lagerort', 'Fahrzeug', 'Zugeordnet an',
        'Zustand', 'Einsatzbereit', 'Herstellungsdatum', 'Letzte Prüfung', 'Nächste Prüfung',
        'Letzte Prüfer', 'Aussonderungsdatum', 'Notizen', 'Erstellt am', 'Aktualisiert am'
    ]

    # Header-Styling
    header_fill = PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Daten
    for row_num, device in enumerate(devices, start=2):
        ws.cell(row=row_num, column=1, value=device.inventory_number)
        ws.cell(row=row_num, column=2, value=device.serial_number)
        ws.cell(row=row_num, column=3, value=device.master.name if device.master else '')
        ws.cell(row=row_num, column=4, value=device.location.name if device.location else '')
        ws.cell(row=row_num, column=5, value=str(device.assigned_vehicle) if device.assigned_vehicle else '')
        ws.cell(row=row_num, column=6, value=device.assigned_to.get_full_name() if device.assigned_to else '')
        ws.cell(row=row_num, column=7, value=device.get_condition_display() if device.condition else '')
        ws.cell(row=row_num, column=8, value='Ja' if device.is_operational else 'Nein')
        ws.cell(row=row_num, column=9, value=device.manufacturing_date.strftime('%d.%m.%Y') if device.manufacturing_date else '')
        ws.cell(row=row_num, column=10, value=device.last_inspection_date.strftime('%d.%m.%Y') if device.last_inspection_date else '')
        ws.cell(row=row_num, column=11, value=device.next_inspection_date.strftime('%d.%m.%Y') if device.next_inspection_date else '')
        ws.cell(row=row_num, column=12, value=device.last_inspector.get_full_name() if device.last_inspector else '')
        ws.cell(row=row_num, column=13, value=device.retirement_date.strftime('%d.%m.%Y') if device.retirement_date else '')
        ws.cell(row=row_num, column=14, value=device.notes)
        ws.cell(row=row_num, column=15, value=device.created_at.strftime('%d.%m.%Y %H:%M') if device.created_at else '')
        ws.cell(row=row_num, column=16, value=device.updated_at.strftime('%d.%m.%Y %H:%M') if device.updated_at else '')

    # Spaltenbreiten
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=hoehenrettung_geraete_export.xlsx'
    wb.save(response)
    return response


@login_required
def export_inspections(request):
    """Export Height Rescue Inspection Logs als Excel"""
    inspections = HeightRescueInspectionLog.objects.select_related(
        'item', 'inspector'
    ).order_by('-inspection_date')[:1000]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prüfprotokolle"

    # Header
    headers = [
        'ID', 'Prüfdatum', 'Artikel', 'Prüfer', 'Bestanden', 'Nächste Prüfung',
        'Mängel', 'Maßnahmen', 'Notizen', 'Erstellt am'
    ]

    # Header-Styling
    header_fill = PatternFill(start_color='9C27B0', end_color='9C27B0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Daten
    for row_num, inspection in enumerate(inspections, start=2):
        ws.cell(row=row_num, column=1, value=inspection.id)
        ws.cell(row=row_num, column=2, value=inspection.inspection_date.strftime('%d.%m.%Y') if inspection.inspection_date else '')
        ws.cell(row=row_num, column=3, value=inspection.item.name if inspection.item else '')
        ws.cell(row=row_num, column=4, value=inspection.inspector.get_full_name() if inspection.inspector else '')
        ws.cell(row=row_num, column=5, value='Ja' if inspection.passed else 'Nein')
        ws.cell(row=row_num, column=6, value=inspection.next_inspection_due.strftime('%d.%m.%Y') if inspection.next_inspection_due else '')
        ws.cell(row=row_num, column=7, value=inspection.defects)
        ws.cell(row=row_num, column=8, value=inspection.actions_taken)
        ws.cell(row=row_num, column=9, value=inspection.notes)
        ws.cell(row=row_num, column=10, value=inspection.created_at.strftime('%d.%m.%Y %H:%M') if inspection.created_at else '')

    # Spaltenbreiten
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=hoehenrettung_pruefungen_export.xlsx'
    wb.save(response)
    return response


# ============================================================================
# TEMPLATE DOWNLOAD FUNCTIONS (CSV)
# ============================================================================

@login_required
def template_masters(request):
    """CSV-Vorlage für Stammdaten-Import"""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)

    # Header
    headers = [
        'Stammdaten-Nr.*', 'Name*', 'Typ*', 'Hersteller', 'Modell', 'Herstellerartikelnr.',
        'EN-Zertifizierung', 'Zertifizierungsnummer', 'Max. Nutzungsdauer (Jahre)',
        'Seiltyp', 'Seildurchmesser (mm)', 'Seillänge (m)', 'Gewicht (kg)', 'Beschreibung'
    ]
    writer.writerow(headers)

    # Beispielzeile
    example_data = [
        'HR-2025-001', 'Petzl Vertex Helm', 'helmet', 'Petzl', 'Vertex Best',
        'A010CA00', 'en_12492', 'CE-12345', '10',
        '', '', '', '0.45', 'Beispiel: Kletterhelm - diese Zeile löschen!'
    ]
    writer.writerow(example_data)

    # Hinweise
    writer.writerow([])
    writer.writerow(['# HINWEISE:'])
    writer.writerow(['# Felder mit * sind Pflichtfelder'])
    writer.writerow(['# Typ: rope, harness, helmet, carabiner, descender, ascender, pulley, anchor, fall_arrester, retrieval_device, tripod, rope_protector, webbing, chest_harness, sit_harness, work_positioning, rescue_stretcher, edge_roller, safety_line, belay_device'])
    writer.writerow(['# Seiltyp: type_a, type_b, dynamic, kernmantle, steel_cable'])
    writer.writerow(['# EN-Zertifizierung: en_361, en_362, en_363, en_795, en_813, en_892, en_1891, en_12275, en_12278, en_12841, en_12492, en_341, en_353, en_355, en_360, en_397, en_567, gs, ce, dguv'])

    # Response
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename=vorlage_hoehenrettung_stammdaten_import.csv'
    response.write('\ufeff')  # UTF-8 BOM
    response.write(output.getvalue())
    return response


@login_required
def template_devices(request):
    """CSV-Vorlage für Geräte-Import"""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)

    # Header
    headers = [
        'Inventarnummer*', 'Seriennummer*', 'Stammdaten-Nr.*', 'Lagerort-ID*',
        'Herstellungsdatum', 'Zustand', 'Einsatzbereit (Ja/Nein)', 'Notizen'
    ]
    writer.writerow(headers)

    # Beispielzeile
    example_data = [
        'INV-HR-001', 'SN123456', 'HR-2025-001', '1',
        '2024-01-15', 'good', 'Ja', 'Beispiel: Neues Gerät - diese Zeile löschen!'
    ]
    writer.writerow(example_data)

    # Hinweise
    writer.writerow([])
    writer.writerow(['# HINWEISE:'])
    writer.writerow(['# Felder mit * sind Pflichtfelder'])
    writer.writerow(['# Zustand: excellent, good, fair, poor, damaged'])
    writer.writerow(['# Einsatzbereit: Ja oder Nein'])
    writer.writerow(['# Lagerort-ID muss existieren'])
    writer.writerow(['# Stammdaten-Nr. muss existieren'])

    # Response
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename=vorlage_hoehenrettung_geraete_import.csv'
    response.write('\ufeff')  # UTF-8 BOM
    response.write(output.getvalue())
    return response


# ============================================================================
# IMPORT FUNCTIONS (CSV)
# ============================================================================

@login_required
def import_masters(request):
    """Import Height Rescue Masters aus CSV"""
    if request.method != 'POST':
        messages.error(request, 'Nur POST-Anfragen erlaubt.')
        return redirect('height_rescue:import_export')

    if 'file' not in request.FILES:
        messages.error(request, 'Keine Datei hochgeladen.')
        return redirect('height_rescue:import_export')

    uploaded_file = request.FILES['file']

    try:
        # CSV-Datei lesen (UTF-8 mit BOM)
        decoded_file = uploaded_file.read().decode('utf-8-sig')
        csv_reader = csv.reader(io.StringIO(decoded_file), delimiter=';')

        success_count = 0
        error_count = 0
        errors = []

        # Header überspringen
        next(csv_reader, None)

        # Zeilen durchgehen
        for row_num, row in enumerate(csv_reader, start=2):
            # Leere Zeilen und Kommentare überspringen
            if not row or not row[0] or row[0].startswith('#'):
                continue

            # Beispielzeile überspringen
            if 'Beispiel' in str(row[-1] if row else ''):
                continue

            try:
                # Daten extrahieren (mit Index-Safety)
                master_number = row[0] if len(row) > 0 else None
                name = row[1] if len(row) > 1 else None
                item_type = row[2] if len(row) > 2 else None
                manufacturer = row[3] if len(row) > 3 else ''
                model = row[4] if len(row) > 4 else ''
                manufacturer_part_number = row[5] if len(row) > 5 else ''
                en_certification = row[6] if len(row) > 6 else ''
                certification_number = row[7] if len(row) > 7 else ''
                max_service_life_years = row[8] if len(row) > 8 else None
                rope_type = row[9] if len(row) > 9 else ''
                rope_diameter_mm = row[10] if len(row) > 10 else None
                rope_length_m = row[11] if len(row) > 11 else None
                weight_kg = row[12] if len(row) > 12 else None
                description = row[13] if len(row) > 13 else ''

                # Validierung
                if not all([master_number, name, item_type]):
                    errors.append(f"Zeile {row_num}: Pflichtfelder fehlen")
                    error_count += 1
                    continue

                # Duplikatprüfung
                if HeightRescueItemMaster.objects.filter(master_number=master_number).exists():
                    errors.append(f"Zeile {row_num}: Stammdaten-Nr. '{master_number}' existiert bereits")
                    error_count += 1
                    continue

                # Master erstellen
                master = HeightRescueItemMaster.objects.create(
                    master_number=master_number,
                    name=name,
                    item_type=item_type,
                    manufacturer=manufacturer,
                    model=model,
                    manufacturer_part_number=manufacturer_part_number,
                    en_certification=en_certification if en_certification else None,
                    certification_number=certification_number,
                    max_service_life_years=int(max_service_life_years) if max_service_life_years else None,
                    rope_type=rope_type if rope_type else None,
                    rope_diameter_mm=float(rope_diameter_mm) if rope_diameter_mm else None,
                    rope_length_m=float(rope_length_m) if rope_length_m else None,
                    weight_kg=float(weight_kg) if weight_kg else None,
                    description=description,
                    created_by=request.user,
                    updated_by=request.user
                )

                success_count += 1

            except Exception as e:
                errors.append(f"Zeile {row_num}: Fehler - {str(e)}")
                error_count += 1
                logger.error(f"Master import error in row {row_num}: {e}")

        # Erfolgsmeldung
        if success_count > 0:
            messages.success(request, f"✅ {success_count} Stammdaten erfolgreich importiert.")

        if error_count > 0:
            messages.warning(request, f"⚠️ {error_count} Fehler beim Import.")
            for error in errors[:10]:
                messages.error(request, error)
            if len(errors) > 10:
                messages.error(request, f"... und {len(errors) - 10} weitere Fehler")

    except Exception as e:
        messages.error(request, f"Fehler beim Lesen der CSV-Datei: {str(e)}")
        logger.error(f"CSV import error: {e}")

    return redirect('height_rescue:import_export')


@login_required
def import_devices(request):
    """Import Height Rescue Devices aus CSV"""
    if request.method != 'POST':
        messages.error(request, 'Nur POST-Anfragen erlaubt.')
        return redirect('height_rescue:import_export')

    if 'file' not in request.FILES:
        messages.error(request, 'Keine Datei hochgeladen.')
        return redirect('height_rescue:import_export')

    uploaded_file = request.FILES['file']

    try:
        # CSV-Datei lesen (UTF-8 mit BOM)
        decoded_file = uploaded_file.read().decode('utf-8-sig')
        csv_reader = csv.reader(io.StringIO(decoded_file), delimiter=';')

        success_count = 0
        error_count = 0
        errors = []

        # Header überspringen
        next(csv_reader, None)

        # Zeilen durchgehen
        for row_num, row in enumerate(csv_reader, start=2):
            # Leere Zeilen und Kommentare überspringen
            if not row or not row[0] or row[0].startswith('#'):
                continue

            # Beispielzeile überspringen
            if 'Beispiel' in str(row[-1] if row else ''):
                continue

            try:
                # Daten extrahieren (mit Index-Safety)
                inventory_number = row[0] if len(row) > 0 else None
                serial_number = row[1] if len(row) > 1 else None
                master_number = row[2] if len(row) > 2 else None
                location_id = row[3] if len(row) > 3 else None
                manufacturing_date_str = row[4] if len(row) > 4 else ''
                condition = row[5] if len(row) > 5 else 'good'
                is_operational_str = row[6] if len(row) > 6 else 'Ja'
                notes = row[7] if len(row) > 7 else ''

                # Validierung
                if not all([inventory_number, serial_number, master_number, location_id]):
                    errors.append(f"Zeile {row_num}: Pflichtfelder fehlen")
                    error_count += 1
                    continue

                # Duplikatprüfung
                if HeightRescueDeviceInstance.objects.filter(inventory_number=inventory_number).exists():
                    errors.append(f"Zeile {row_num}: Inventarnummer '{inventory_number}' existiert bereits")
                    error_count += 1
                    continue

                # Master laden
                try:
                    master = HeightRescueItemMaster.objects.get(master_number=master_number)
                except HeightRescueItemMaster.DoesNotExist:
                    errors.append(f"Zeile {row_num}: Stammdaten-Nr. {master_number} nicht gefunden")
                    error_count += 1
                    continue

                # Standort laden
                from locations.models import Location
                try:
                    location = Location.objects.get(id=int(location_id))
                except (Location.DoesNotExist, ValueError):
                    errors.append(f"Zeile {row_num}: Standort-ID {location_id} nicht gefunden")
                    error_count += 1
                    continue

                # Datum parsen
                manufacturing_date = None
                if manufacturing_date_str:
                    from datetime import datetime
                    try:
                        manufacturing_date = datetime.strptime(manufacturing_date_str, '%Y-%m-%d').date()
                    except ValueError:
                        try:
                            manufacturing_date = datetime.strptime(manufacturing_date_str, '%d.%m.%Y').date()
                        except ValueError:
                            pass

                # Boolean-Konvertierung
                is_operational = is_operational_str.lower() in ['ja', 'yes', 'true', '1']

                # Device erstellen
                device = HeightRescueDeviceInstance.objects.create(
                    inventory_number=inventory_number,
                    serial_number=serial_number,
                    master=master,
                    location=location,
                    manufacturing_date=manufacturing_date,
                    condition=condition,
                    is_operational=is_operational,
                    notes=notes,
                    created_by=request.user,
                    updated_by=request.user
                )

                success_count += 1

            except Exception as e:
                errors.append(f"Zeile {row_num}: Fehler - {str(e)}")
                error_count += 1
                logger.error(f"Device import error in row {row_num}: {e}")

        # Erfolgsmeldung
        if success_count > 0:
            messages.success(request, f"✅ {success_count} Geräte erfolgreich importiert.")

        if error_count > 0:
            messages.warning(request, f"⚠️ {error_count} Fehler beim Import.")
            for error in errors[:10]:
                messages.error(request, error)
            if len(errors) > 10:
                messages.error(request, f"... und {len(errors) - 10} weitere Fehler")

    except Exception as e:
        messages.error(request, f"Fehler beim Lesen der CSV-Datei: {str(e)}")
        logger.error(f"CSV import error: {e}")

    return redirect('height_rescue:import_export')
