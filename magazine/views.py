"""
Magazine Views
Views für das Magazin-Modul (Verbrauchsmaterial)
"""

from django.views.generic import ListView, DetailView, TemplateView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db import models
from django.db.models import Q, Sum, Count, F
from django.shortcuts import redirect
from django.utils import timezone
from django.urls import reverse_lazy
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from datetime import timedelta

from .models import MagazineItem, MagazineItemMaster, MagazineStockMovement, MagazineBatch
from .forms import MagazineItemForm, MagazineItemMasterForm, MagazineStockMovementForm
from inventory_base.models import Category
from django.http import HttpResponse
import logging

logger = logging.getLogger(__name__)


# ============================================================================
# DASHBOARD
# ============================================================================

class MagazineDashboardView(LoginRequiredMixin, TemplateView):
    """Dashboard für das Magazin-Modul"""
    template_name = 'magazine/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()

        # KPIs
        context['total_items'] = MagazineItem.objects.filter(is_active=True).count()
        context['low_stock_count'] = MagazineItem.objects.filter(
            is_active=True,
            min_quantity__isnull=False,
            quantity__lt=F('min_quantity')
        ).count()
        context['reorder_needed_count'] = MagazineItem.objects.filter(
            is_active=True,
            reorder_point__isnull=False,
            quantity__lte=F('reorder_point')
        ).count()
        context['hazardous_count'] = MagazineItem.objects.filter(
            is_active=True,
            is_hazardous=True
        ).count()

        # Ablaufende Chargen (nächste 90 Tage)
        threshold = today + timedelta(days=90)
        context['expiring_batches_count'] = MagazineBatch.objects.filter(
            expiry_date__isnull=False,
            expiry_date__lte=threshold,
            expiry_date__gte=today,
            quantity_remaining__gt=0
        ).count()

        # Letzte Bewegungen
        context['recent_movements'] = MagazineStockMovement.objects.select_related(
            'item',
            'from_location',
            'to_location',
            'created_by'
        ).order_by('-movement_date')[:10]

        # Artikel mit niedrigem Bestand
        context['low_stock_items'] = MagazineItem.objects.filter(
            is_active=True,
            min_quantity__isnull=False,
            quantity__lt=F('min_quantity')
        ).select_related('category', 'location')[:10]

        # Ablaufende Chargen
        context['expiring_batches'] = MagazineBatch.objects.filter(
            expiry_date__isnull=False,
            expiry_date__lte=threshold,
            expiry_date__gte=today,
            quantity_remaining__gt=0
        ).select_related('item', 'location').order_by('expiry_date')[:10]

        return context


# ============================================================================
# MAGAZINE ITEM MASTER VIEWS (STAMMDATEN)
# ============================================================================

class MagazineMasterListView(LoginRequiredMixin, ListView):
    """Liste aller Magazin-Stammdaten"""
    model = MagazineItemMaster
    template_name = 'magazine/master_list.html'
    context_object_name = 'masters'
    paginate_by = 50

    def get_queryset(self):
        queryset = MagazineItemMaster.objects.filter(is_active=True).select_related(
            'category'
        )

        # Filter
        search = self.request.GET.get('search', '')
        item_type = self.request.GET.get('item_type', '')
        is_hazardous = self.request.GET.get('is_hazardous', '')

        if search:
            queryset = queryset.filter(
                Q(master_number__icontains=search) |
                Q(name__icontains=search) |
                Q(manufacturer__icontains=search) |
                Q(description__icontains=search)
            )

        if item_type:
            queryset = queryset.filter(item_type=item_type)

        if is_hazardous == 'yes':
            queryset = queryset.filter(is_hazardous=True)
        elif is_hazardous == 'no':
            queryset = queryset.filter(is_hazardous=False)

        return queryset.order_by('master_number')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_module'] = 'magazine'
        context['search'] = self.request.GET.get('search', '')
        context['item_type'] = self.request.GET.get('item_type', '')
        context['is_hazardous'] = self.request.GET.get('is_hazardous', '')

        # Item Types für Filter
        from .models import MagazineItemType
        context['item_types'] = MagazineItemType.choices

        return context


class MagazineMasterDetailView(LoginRequiredMixin, DetailView):
    """Detail-Ansicht für Magazin-Stammdaten"""
    model = MagazineItemMaster
    template_name = 'magazine/master_detail.html'
    context_object_name = 'master'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_module'] = 'magazine'

        # Zugehörige Chargen
        context['batches'] = self.object.batches.filter(
            quantity_remaining__gt=0
        ).select_related('location').order_by('expiry_date')

        # Gesamtbestand
        context['total_stock'] = self.object.get_total_stock()

        return context


class MagazineMasterCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Neue Magazin-Stammdaten erstellen"""
    model = MagazineItemMaster
    form_class = MagazineItemMasterForm
    template_name = 'magazine/master_form.html'
    success_url = reverse_lazy('magazine:master_list')
    permission_required = 'magazine.add_magazineitemmaster'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Stammdaten wurden erfolgreich erstellt.'))
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_module'] = 'magazine'
        context['title'] = _('Neue Stammdaten')
        return context


class MagazineMasterUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Magazin-Stammdaten bearbeiten"""
    model = MagazineItemMaster
    form_class = MagazineItemMasterForm
    template_name = 'magazine/master_form.html'
    permission_required = 'magazine.change_magazineitemmaster'

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Stammdaten wurden erfolgreich aktualisiert.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('magazine:master_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_module'] = 'magazine'
        context['title'] = _('Stammdaten bearbeiten')
        return context


from django.contrib.auth.decorators import login_required

@login_required
def master_barcode_view(request, pk):
    """Generiert Barcode für Magazin-Stammdaten als SVG"""
    from django.shortcuts import get_object_or_404
    master = get_object_or_404(MagazineItemMaster, pk=pk)

    try:
        svg_string = master.generate_barcode()

        if not svg_string:
            return HttpResponse("Kein Barcode verfügbar", status=400)

        return HttpResponse(svg_string, content_type='image/svg+xml')

    except Exception as e:
        logger.error(f"Barcode generation error for master {pk}: {e}")
        return HttpResponse(f"Fehler bei Barcode-Generierung: {str(e)}", status=500)


@login_required
def master_qr_code_view(request, pk):
    """Generiert QR-Code für Magazin-Stammdaten als SVG"""
    from django.shortcuts import get_object_or_404
    master = get_object_or_404(MagazineItemMaster, pk=pk)

    try:
        svg_string = master.generate_qr_code()

        if not svg_string:
            return HttpResponse("Kein QR-Code verfügbar", status=400)

        return HttpResponse(svg_string, content_type='image/svg+xml')

    except Exception as e:
        logger.error(f"QR-Code generation error for master {pk}: {e}")
        return HttpResponse(f"Fehler bei QR-Code-Generierung: {str(e)}", status=500)


# ============================================================================
# MAGAZINE ITEM VIEWS (LEGACY)
# ============================================================================

class MagazineItemListView(LoginRequiredMixin, ListView):
    """
    Liste aller Magazin-Artikel
    """
    model = MagazineItem
    template_name = 'magazine/item_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        queryset = MagazineItem.objects.filter(is_active=True).select_related(
            'category',
            'location',
            'supplier',
        )

        # Suchfilter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(item_number__icontains=search) |
                Q(description__icontains=search) |
                Q(barcode__icontains=search)
            )

        # Typ-Filter
        item_type = self.request.GET.get('item_type')
        if item_type:
            queryset = queryset.filter(item_type=item_type)

        # Kategorie-Filter
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category_id=category)

        # Lagerort-Filter
        location = self.request.GET.get('location')
        if location:
            queryset = queryset.filter(location_id=location)

        return queryset


class MagazineItemCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Neuen Magazin-Artikel erstellen"""
    model = MagazineItem
    form_class = MagazineItemForm
    template_name = 'magazine/item_form.html'
    permission_required = 'magazine.add_magazineitem'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Artikel wurde erfolgreich erstellt.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('magazine:item_detail', kwargs={'pk': self.object.pk})


class MagazineItemUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Magazin-Artikel bearbeiten"""
    model = MagazineItem
    form_class = MagazineItemForm
    template_name = 'magazine/item_form.html'
    permission_required = 'magazine.change_magazineitem'

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Artikel wurde erfolgreich aktualisiert.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('magazine:item_detail', kwargs={'pk': self.object.pk})


class MagazineItemDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Magazin-Artikel löschen"""
    model = MagazineItem
    template_name = 'magazine/item_confirm_delete.html'
    success_url = reverse_lazy('magazine:item_list')
    permission_required = 'magazine.delete_magazineitem'

    def form_valid(self, form):
        item = self.get_object()
        item_name = item.name

        # Delete related stock movements and batches first
        movements_count = item.stock_movements.count()
        batches_count = MagazineBatch.objects.filter(item=item).count()

        item.stock_movements.all().delete()
        MagazineBatch.objects.filter(item=item).delete()

        response = super().form_valid(form)

        msg = f'Artikel "{item_name}" wurde gelöscht'
        if movements_count or batches_count:
            msg += f' (inkl. {movements_count} Bewegungen, {batches_count} Chargen)'
        msg += '.'
        messages.success(self.request, msg)
        return response


class MagazineItemDetailView(LoginRequiredMixin, DetailView):
    """
    Detail-Ansicht für Magazin-Artikel
    """
    model = MagazineItem
    template_name = 'magazine/item_detail.html'
    context_object_name = 'item'

    def get_queryset(self):
        return MagazineItem.objects.select_related(
            'category',
            'location',
            'supplier',
            'created_by',
            'updated_by',
        ).prefetch_related(
            'stock_movements',
            'batches',
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        item = self.object

        # Letzte Lagerbewegungen
        context['recent_movements'] = item.stock_movements.select_related(
            'from_location',
            'to_location',
            'person',
            'created_by',
        ).order_by('-movement_date')[:10]

        # Aktive Chargen
        context['active_batches'] = item.batches.filter(
            quantity_remaining__gt=0
        ).select_related('location').order_by('expiry_date')

        return context


class LowStockListView(LoginRequiredMixin, ListView):
    """
    Liste der Artikel mit niedrigem Bestand
    """
    model = MagazineItem
    template_name = 'magazine/low_stock_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        queryset = MagazineItem.objects.filter(
            is_active=True,
            min_quantity__isnull=False
        ).filter(
            quantity__lt=F('min_quantity')
        ).select_related('category', 'location', 'supplier')

        # Suchfilter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(item_number__icontains=search) |
                Q(description__icontains=search)
            )

        # Kategorie-Filter
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category_id=category)

        # Sortierung
        sort = self.request.GET.get('sort', '-quantity')
        if sort:
            queryset = queryset.order_by(sort)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = Category.objects.filter(is_active=True).order_by('name')
        return context


class ReorderNeededListView(LoginRequiredMixin, ListView):
    """
    Liste der Artikel die nachbestellt werden müssen
    """
    model = MagazineItem
    template_name = 'magazine/reorder_needed_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        return MagazineItem.objects.filter(
            is_active=True,
            reorder_point__isnull=False
        ).filter(
            quantity__lte=F('reorder_point')
        ).select_related('category', 'location', 'supplier')


class HazardousItemsListView(LoginRequiredMixin, ListView):
    """
    Liste der Gefahrgut-Artikel
    """
    model = MagazineItem
    template_name = 'magazine/hazardous_items_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        queryset = MagazineItem.objects.filter(
            is_active=True,
            is_hazardous=True
        ).select_related('category', 'location', 'supplier')

        # Suchfilter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(item_number__icontains=search) |
                Q(manufacturer__icontains=search) |
                Q(description__icontains=search)
            )

        # Gefahrenklasse-Filter
        hazard_class = self.request.GET.get('hazard_class')
        if hazard_class:
            queryset = queryset.filter(hazard_class=hazard_class)

        return queryset.order_by('hazard_class', 'name')


# ============================================================================
# STOCK MOVEMENT VIEWS
# ============================================================================

class StockMovementListView(LoginRequiredMixin, ListView):
    """
    Liste aller Lagerbewegungen
    """
    model = MagazineStockMovement
    template_name = 'magazine/movement_list.html'
    context_object_name = 'movements'
    paginate_by = 50

    def get_queryset(self):
        queryset = MagazineStockMovement.objects.select_related(
            'item',
            'from_location',
            'to_location',
            'person',
        )

        # Bewegungstyp-Filter
        movement_type = self.request.GET.get('movement_type')
        if movement_type:
            queryset = queryset.filter(movement_type=movement_type)

        # Artikel-Filter
        item = self.request.GET.get('item')
        if item:
            queryset = queryset.filter(item_id=item)

        return queryset.order_by('-movement_date')


class StockMovementDetailView(LoginRequiredMixin, DetailView):
    """
    Detail-Ansicht für Lagerbewegung
    """
    model = MagazineStockMovement
    template_name = 'magazine/movement_detail.html'
    context_object_name = 'movement'

    def get_queryset(self):
        return MagazineStockMovement.objects.select_related(
            'item',
            'from_location',
            'to_location',
            'person',
            'created_by',
            'updated_by',
        )


class StockMovementCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Neue Lagerbewegung erstellen"""
    model = MagazineStockMovement
    form_class = MagazineStockMovementForm
    template_name = 'magazine/movement_form.html'
    permission_required = 'magazine.add_magazinestockmovement'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        messages.success(self.request, _('Lagerbewegung wurde erfolgreich erstellt.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('magazine:movement_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        import json
        context = super().get_context_data(**kwargs)
        item_locations = dict(
            MagazineItem.objects.filter(location__isnull=False)
            .values_list('pk', 'location_id')
        )
        context['item_location_map'] = json.dumps(item_locations)
        return context


# ============================================================================
# BATCH VIEWS
# ============================================================================

class BatchListView(LoginRequiredMixin, ListView):
    """
    Liste aller Chargen
    """
    model = MagazineBatch
    template_name = 'magazine/batch_list.html'
    context_object_name = 'batches'
    paginate_by = 50

    def get_queryset(self):
        queryset = MagazineBatch.objects.select_related(
            'item',
            'location',
        )

        # Nur aktive Chargen (nicht aufgebraucht)
        show_depleted = self.request.GET.get('show_depleted')
        if not show_depleted:
            queryset = queryset.filter(quantity_remaining__gt=0)

        # Artikel-Filter
        item = self.request.GET.get('item')
        if item:
            queryset = queryset.filter(item_id=item)

        return queryset.order_by('expiry_date', 'received_date')


class ExpiringBatchesListView(LoginRequiredMixin, ListView):
    """
    Liste der ablaufenden Chargen (nächste 90 Tage)
    """
    model = MagazineBatch
    template_name = 'magazine/expiring_batches_list.html'
    context_object_name = 'batches'
    paginate_by = 50

    def get_queryset(self):
        today = timezone.now().date()
        threshold = today + timedelta(days=90)

        return MagazineBatch.objects.filter(
            expiry_date__isnull=False,
            expiry_date__lte=threshold,
            expiry_date__gte=today,
            quantity_remaining__gt=0
        ).select_related('item', 'location').order_by('expiry_date')


# ============================================================================
# CATEGORY VIEWS
# ============================================================================

class CategoryListView(LoginRequiredMixin, ListView):
    """Liste aller Kategorien"""
    model = Category
    template_name = 'magazine/category_list.html'
    context_object_name = 'categories'

    def get_queryset(self):
        return Category.objects.filter(is_active=True, module='magazine').order_by('tree_id', 'lft')


class CategoryCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Neue Kategorie erstellen"""
    model = Category
    template_name = 'magazine/category_form.html'
    fields = ['name', 'parent', 'code', 'description']
    success_url = reverse_lazy('magazine:category_list')
    permission_required = 'inventory_base.add_category'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        form.instance.module = 'magazine'
        messages.success(self.request, f'Kategorie "{form.instance.name}" wurde erstellt.')
        return super().form_valid(form)


class CategoryUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Kategorie bearbeiten"""
    model = Category
    template_name = 'magazine/category_form.html'
    fields = ['name', 'parent', 'code', 'description', 'is_active']
    success_url = reverse_lazy('magazine:category_list')
    permission_required = 'inventory_base.change_category'

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, f'Kategorie "{form.instance.name}" wurde aktualisiert.')
        return super().form_valid(form)


class CategoryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Kategorie löschen"""
    model = Category
    template_name = 'magazine/category_confirm_delete.html'
    success_url = reverse_lazy('magazine:category_list')
    permission_required = 'inventory_base.delete_category'

    def form_valid(self, form):
        category = self.get_object()
        try:
            response = super().form_valid(form)
            messages.success(self.request, f'Kategorie "{category.name}" wurde gelöscht.')
            return response
        except models.ProtectedError:
            messages.error(
                self.request,
                f'Kategorie "{category.name}" kann nicht gelöscht werden, da sie noch von Artikeln in anderen Modulen verwendet wird.'
            )
            return redirect('magazine:category_list')


# ============================================================================
# IMPORT/EXPORT
# ============================================================================

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import csv
import io
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
import logging

logger = logging.getLogger(__name__)


class ImportExportView(LoginRequiredMixin, TemplateView):
    """Import/Export-Übersicht für Magazin"""
    template_name = 'magazine/import_export.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_items'] = MagazineItem.objects.filter(is_active=True).count()
        context['total_hazardous'] = MagazineItem.objects.filter(is_active=True, is_hazardous=True).count()
        context['total_batches'] = MagazineBatch.objects.filter(quantity_remaining__gt=0).count()
        context['total_movements'] = MagazineStockMovement.objects.count()
        return context


# ============================================================================
# EXPORT FUNCTIONS (Excel)
# ============================================================================

@login_required
def export_categories(request):
    """Export Magazine Categories als Excel"""
    categories = Category.objects.filter(is_active=True).order_by('name')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kategorien"

    headers = [
        'ID', 'Name', 'Code', 'Beschreibung', 'Übergeordnete Kategorie',
        'Sortierung', 'Erstellt am', 'Aktualisiert am'
    ]

    header_fill = PatternFill(start_color='9C27B0', end_color='9C27B0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_num, category in enumerate(categories, start=2):
        ws.cell(row=row_num, column=1, value=category.id)
        ws.cell(row=row_num, column=2, value=category.name)
        ws.cell(row=row_num, column=3, value=category.code)
        ws.cell(row=row_num, column=4, value=category.description)
        ws.cell(row=row_num, column=5, value=category.parent.name if category.parent else '')
        ws.cell(row=row_num, column=6, value=category.lft)  # Tree position for sorting
        ws.cell(row=row_num, column=7, value=category.created_at.strftime('%d.%m.%Y %H:%M') if category.created_at else '')
        ws.cell(row=row_num, column=8, value=category.updated_at.strftime('%d.%m.%Y %H:%M') if category.updated_at else '')

    # Spaltenbreiten optimieren
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="magazin_kategorien.xlsx"'
    wb.save(response)
    return response


@login_required
def export_items(request):
    """Export Magazine Items als Excel"""
    items = MagazineItem.objects.filter(is_active=True).select_related(
        'category', 'location', 'supplier'
    ).order_by('item_number')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Artikel"

    headers = [
        'Artikelnummer', 'Name', 'Beschreibung', 'Kategorie-ID', 'Kategorie', 'Artikeltyp',
        'Größe/Maße', 'Material', 'Farbe', 'Hersteller', 'Lieferant',
        'Bestand', 'Einheit', 'Min. Bestand', 'Max. Bestand', 'Bestellpunkt',
        'Standard-Bestellmenge', 'Lagerort', 'Barcode', 'Gefahrgut', 'Gefahrenklasse',
        'Verfallsdatum', 'Haltbarkeit (Monate)', 'Gewicht/Einheit (kg)', 'Volumen/Einheit (L)',
        'Min. Temp (°C)', 'Max. Temp (°C)', 'Lageranweisungen', 'Technische Spezifikationen',
        'Verwendungshinweise', 'Erstellt am', 'Aktualisiert am'
    ]

    header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_num, item in enumerate(items, start=2):
        ws.cell(row=row_num, column=1, value=item.item_number)
        ws.cell(row=row_num, column=2, value=item.name)
        ws.cell(row=row_num, column=3, value=item.description)
        ws.cell(row=row_num, column=4, value=item.category.id if item.category else '')
        ws.cell(row=row_num, column=5, value=item.category.name if item.category else '')
        ws.cell(row=row_num, column=6, value=item.get_item_type_display())
        ws.cell(row=row_num, column=7, value=item.size)
        ws.cell(row=row_num, column=8, value=item.material)
        ws.cell(row=row_num, column=9, value=item.color)
        ws.cell(row=row_num, column=10, value=item.manufacturer)
        ws.cell(row=row_num, column=11, value=item.supplier.name if item.supplier else '')
        ws.cell(row=row_num, column=12, value=float(item.quantity))
        ws.cell(row=row_num, column=13, value=item.unit)
        ws.cell(row=row_num, column=14, value=float(item.min_quantity) if item.min_quantity else '')
        ws.cell(row=row_num, column=15, value=float(item.max_quantity) if item.max_quantity else '')
        ws.cell(row=row_num, column=16, value=float(item.reorder_point) if item.reorder_point else '')
        ws.cell(row=row_num, column=17, value=float(item.standard_order_quantity) if item.standard_order_quantity else '')
        ws.cell(row=row_num, column=18, value=item.location.name if item.location else '')
        ws.cell(row=row_num, column=19, value=item.barcode)
        ws.cell(row=row_num, column=20, value='Ja' if item.is_hazardous else 'Nein')
        ws.cell(row=row_num, column=21, value=item.get_hazard_class_display())
        ws.cell(row=row_num, column=22, value='Ja' if item.has_expiry_date else 'Nein')
        ws.cell(row=row_num, column=23, value=item.shelf_life_months if item.shelf_life_months else '')
        ws.cell(row=row_num, column=24, value=float(item.weight_per_unit) if item.weight_per_unit else '')
        ws.cell(row=row_num, column=25, value=float(item.volume_per_unit) if item.volume_per_unit else '')
        ws.cell(row=row_num, column=26, value=float(item.storage_temperature_min) if item.storage_temperature_min else '')
        ws.cell(row=row_num, column=27, value=float(item.storage_temperature_max) if item.storage_temperature_max else '')
        ws.cell(row=row_num, column=28, value=item.storage_instructions)
        ws.cell(row=row_num, column=29, value=item.technical_specifications)
        ws.cell(row=row_num, column=30, value=item.usage_instructions)
        ws.cell(row=row_num, column=31, value=item.created_at.strftime('%d.%m.%Y %H:%M') if item.created_at else '')
        ws.cell(row=row_num, column=32, value=item.updated_at.strftime('%d.%m.%Y %H:%M') if item.updated_at else '')

    # Spaltenbreiten optimieren
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="magazin_artikel.xlsx"'
    wb.save(response)
    return response


@login_required
def export_batches(request):
    """Export Magazine Batches als Excel"""
    batches = MagazineBatch.objects.filter(quantity_remaining__gt=0).select_related(
        'item', 'location'
    ).order_by('expiry_date', 'batch_number')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chargen"

    headers = [
        'Artikel', 'Artikelnummer', 'Chargen-Nummer', 'Lieferanten-Chargen-Nr.',
        'Eingangsdatum', 'Verfallsdatum', 'Eingangsmenge', 'Restmenge', 'Einheit',
        'Lagerort', 'Notizen'
    ]

    header_fill = PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid')
    header_font = Font(bold=True, color='000000')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_num, batch in enumerate(batches, start=2):
        ws.cell(row=row_num, column=1, value=batch.item.name)
        ws.cell(row=row_num, column=2, value=batch.item.item_number)
        ws.cell(row=row_num, column=3, value=batch.batch_number)
        ws.cell(row=row_num, column=4, value=batch.supplier_batch_number)
        ws.cell(row=row_num, column=5, value=batch.received_date.strftime('%d.%m.%Y') if batch.received_date else '')
        ws.cell(row=row_num, column=6, value=batch.expiry_date.strftime('%d.%m.%Y') if batch.expiry_date else '')
        ws.cell(row=row_num, column=7, value=float(batch.quantity_received))
        ws.cell(row=row_num, column=8, value=float(batch.quantity_remaining))
        ws.cell(row=row_num, column=9, value=batch.item.unit)
        ws.cell(row=row_num, column=10, value=batch.location.name if batch.location else '')
        ws.cell(row=row_num, column=11, value=batch.notes)

    # Spaltenbreiten optimieren
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="magazin_chargen.xlsx"'
    wb.save(response)
    return response


@login_required
def export_movements(request):
    """Export Magazine Stock Movements als Excel (letzte 1000)"""
    movements = MagazineStockMovement.objects.select_related(
        'item', 'from_location', 'to_location', 'person', 'created_by'
    ).order_by('-movement_date')[:1000]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bewegungen"

    headers = [
        'Datum', 'Typ', 'Artikel', 'Artikelnummer', 'Menge', 'Einheit',
        'Von Lagerort', 'Nach Lagerort', 'Chargen-Nr.', 'Verfallsdatum',
        'Stückkosten', 'Gesamtkosten', 'Verwendungszweck', 'Person',
        'Empfänger', 'Lieferschein-Nr.', 'Rechnungsnummer', 'Notizen',
        'Erstellt von', 'Erstellt am'
    ]

    header_fill = PatternFill(start_color='3F51B5', end_color='3F51B5', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_num, movement in enumerate(movements, start=2):
        ws.cell(row=row_num, column=1, value=movement.movement_date.strftime('%d.%m.%Y') if movement.movement_date else '')
        ws.cell(row=row_num, column=2, value=movement.get_movement_type_display())
        ws.cell(row=row_num, column=3, value=movement.item.name)
        ws.cell(row=row_num, column=4, value=movement.item.item_number)
        ws.cell(row=row_num, column=5, value=float(movement.quantity))
        ws.cell(row=row_num, column=6, value=movement.unit)
        ws.cell(row=row_num, column=7, value=movement.from_location.name if movement.from_location else '')
        ws.cell(row=row_num, column=8, value=movement.to_location.name if movement.to_location else '')
        ws.cell(row=row_num, column=9, value=movement.batch_number)
        ws.cell(row=row_num, column=10, value=movement.expiry_date.strftime('%d.%m.%Y') if movement.expiry_date else '')
        ws.cell(row=row_num, column=11, value=float(movement.unit_cost) if movement.unit_cost else '')
        ws.cell(row=row_num, column=12, value=float(movement.total_cost) if movement.total_cost else '')
        ws.cell(row=row_num, column=13, value=movement.purpose)
        ws.cell(row=row_num, column=14, value=str(movement.person) if movement.person else '')
        ws.cell(row=row_num, column=15, value=movement.recipient_name)
        ws.cell(row=row_num, column=16, value=movement.delivery_note)
        ws.cell(row=row_num, column=17, value=movement.invoice_number)
        ws.cell(row=row_num, column=18, value=movement.notes)
        ws.cell(row=row_num, column=19, value=str(movement.created_by) if movement.created_by else '')
        ws.cell(row=row_num, column=20, value=movement.created_at.strftime('%d.%m.%Y %H:%M') if movement.created_at else '')

    # Spaltenbreiten optimieren
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="magazin_bewegungen.xlsx"'
    wb.save(response)
    return response


# ============================================================================
# TEMPLATE FUNCTIONS (CSV)
# ============================================================================

@login_required
def template_categories(request):
    """CSV-Vorlage für Kategorien-Import"""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')

    # Header
    writer.writerow([
        '# Magazin Kategorien Import-Vorlage',
    ])
    writer.writerow([
        '# Pflichtfelder sind mit * markiert',
    ])
    writer.writerow([
        '# Zeichenkodierung: UTF-8 mit BOM',
    ])
    writer.writerow([
        '# Trennzeichen: Semikolon (;)',
    ])
    writer.writerow([])

    writer.writerow([
        'Name*', 'Code', 'Beschreibung', 'Übergeordnete Kategorie (Name)'
    ])

    # Beispielzeilen
    writer.writerow([
        'Batterien', 'BAT', 'Alle Arten von Batterien', ''
    ])
    writer.writerow([
        'AA-Batterien', 'BAT-AA', 'AA Mignon Batterien', 'Batterien'
    ])
    writer.writerow([
        'AAA-Batterien', 'BAT-AAA', 'AAA Micro Batterien', 'Batterien'
    ])

    response = HttpResponse(output.getvalue().encode('utf-8-sig'), content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="magazin_kategorien_vorlage.csv"'
    return response


@login_required
def template_items(request):
    """CSV-Vorlage für Artikel-Import"""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')

    # Header mit Pflichtfeldern (*)
    writer.writerow([
        '# Magazin Artikel Import-Vorlage',
    ])
    writer.writerow([
        '# Pflichtfelder sind mit * markiert',
    ])
    writer.writerow([
        '# Zeichenkodierung: UTF-8 mit BOM',
    ])
    writer.writerow([
        '# Trennzeichen: Semikolon (;)',
    ])
    writer.writerow([
        '# Datumsformat: TT.MM.JJJJ',
    ])
    writer.writerow([])

    writer.writerow([
        'Artikelnummer*', 'Name*', 'Beschreibung', 'Kategorie-ID', 'Artikeltyp',
        'Größe/Maße', 'Material', 'Farbe', 'Hersteller', 'Lieferant',
        'Bestand*', 'Einheit*', 'Min. Bestand', 'Max. Bestand', 'Bestellpunkt',
        'Standard-Bestellmenge', 'Lagerort', 'Barcode', 'Gefahrgut (Ja/Nein)', 'Gefahrenklasse',
        'Verfallsdatum (Ja/Nein)', 'Haltbarkeit (Monate)', 'Gewicht/Einheit (kg)', 'Volumen/Einheit (L)',
        'Min. Temp (°C)', 'Max. Temp (°C)', 'Lageranweisungen', 'Technische Spezifikationen',
        'Verwendungshinweise'
    ])

    # Beispielzeile
    writer.writerow([
        'MAG-001', 'Beispiel Artikel', 'Dies ist ein Beispielartikel', '1', 'battery',
        'AA', 'Alkaline', 'Gelb', 'Duracell', 'Elektro GmbH',
        '100', 'Stück', '20', '200', '30',
        '50', 'Lager 1', '4006381333634', 'Nein', 'none',
        'Ja', '24', '0.023', '0.001',
        '5', '25', 'Kühl und trocken lagern', 'Spannung: 1.5V',
        'Für elektronische Geräte'
    ])

    response = HttpResponse(output.getvalue().encode('utf-8-sig'), content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="magazin_artikel_vorlage.csv"'
    return response


@login_required
def template_batches(request):
    """CSV-Vorlage für Chargen-Import"""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')

    # Header
    writer.writerow([
        '# Magazin Chargen Import-Vorlage',
    ])
    writer.writerow([
        '# Pflichtfelder sind mit * markiert',
    ])
    writer.writerow([
        '# Zeichenkodierung: UTF-8 mit BOM',
    ])
    writer.writerow([
        '# Trennzeichen: Semikolon (;)',
    ])
    writer.writerow([
        '# Datumsformat: TT.MM.JJJJ',
    ])
    writer.writerow([])

    writer.writerow([
        'Artikelnummer*', 'Chargen-Nummer*', 'Lieferanten-Chargen-Nr.',
        'Eingangsdatum*', 'Verfallsdatum', 'Eingangsmenge*', 'Restmenge*',
        'Lagerort-ID*', 'Notizen'
    ])

    # Beispielzeile
    writer.writerow([
        'MAG-001', 'CH-2025-001', 'SUP-12345',
        '01.01.2025', '31.12.2027', '100', '100',
        '1', 'Erste Lieferung 2025'
    ])

    response = HttpResponse(output.getvalue().encode('utf-8-sig'), content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="magazin_chargen_vorlage.csv"'
    return response


# ============================================================================
# IMPORT FUNCTIONS (CSV)
# ============================================================================

@login_required
def import_categories(request):
    """Import von Magazin-Kategorien aus CSV"""
    if request.method != 'POST' or not request.FILES.get('file'):
        messages.error(request, 'Keine Datei hochgeladen.')
        return redirect('magazine:import_export')

    csv_file = request.FILES['file']

    if not csv_file.name.endswith('.csv'):
        messages.error(request, 'Bitte eine CSV-Datei hochladen.')
        return redirect('magazine:import_export')

    try:
        decoded_file = csv_file.read().decode('utf-8-sig')
        reader = csv.reader(io.StringIO(decoded_file), delimiter=';')

        success_count = 0
        error_count = 0
        errors = []

        for row_num, row in enumerate(reader, start=1):
            # Kommentare und Header überspringen
            if not row or (row[0] and row[0].startswith('#')) or row_num <= 6:
                continue

            try:
                # Pflichtfelder
                name = row[0] if len(row) > 0 else None

                if not name:
                    errors.append(f'Zeile {row_num}: Name fehlt')
                    error_count += 1
                    continue

                # Duplikat-Check
                if Category.objects.filter(name=name, is_active=True).exists():
                    errors.append(f'Zeile {row_num}: Kategorie "{name}" existiert bereits')
                    error_count += 1
                    continue

                # Übergeordnete Kategorie
                parent = None
                if len(row) > 3 and row[3]:
                    try:
                        parent = Category.objects.get(name=row[3], is_active=True)
                    except Category.DoesNotExist:
                        errors.append(f'Zeile {row_num}: Übergeordnete Kategorie "{row[3]}" nicht gefunden')
                        error_count += 1
                        continue

                # Kategorie erstellen
                category = Category(
                    name=name,
                    code=row[1] if len(row) > 1 else '',
                    description=row[2] if len(row) > 2 else '',
                    parent=parent,
                    created_by=request.user,
                    updated_by=request.user
                )
                category.save()
                success_count += 1

            except Exception as e:
                errors.append(f'Zeile {row_num}: {str(e)}')
                error_count += 1
                logger.error(f"Category import error in row {row_num}: {e}")

        # Feedback
        if success_count > 0:
            messages.success(request, f'{success_count} Kategorien erfolgreich importiert.')
        if error_count > 0:
            messages.warning(request, f'{error_count} Fehler beim Import.')
            for error in errors[:10]:  # Nur erste 10 Fehler anzeigen
                messages.error(request, error)

    except Exception as e:
        messages.error(request, f'Fehler beim Lesen der CSV-Datei: {str(e)}')
        logger.error(f"CSV import error: {e}")

    return redirect('magazine:import_export')


@login_required
def import_items(request):
    """Import von Magazin-Artikeln aus CSV"""
    if request.method != 'POST' or not request.FILES.get('file'):
        messages.error(request, 'Keine Datei hochgeladen.')
        return redirect('magazine:import_export')

    csv_file = request.FILES['file']

    if not csv_file.name.endswith('.csv'):
        messages.error(request, 'Bitte eine CSV-Datei hochladen.')
        return redirect('magazine:import_export')

    try:
        decoded_file = csv_file.read().decode('utf-8-sig')
        reader = csv.reader(io.StringIO(decoded_file), delimiter=';')

        success_count = 0
        error_count = 0
        errors = []

        for row_num, row in enumerate(reader, start=1):
            # Kommentare und Header überspringen
            if not row or (row[0] and row[0].startswith('#')) or row_num <= 7:
                continue

            try:
                # Pflichtfelder
                item_number = row[0] if len(row) > 0 else None
                name = row[1] if len(row) > 1 else None
                quantity = row[10] if len(row) > 10 else None
                unit = row[11] if len(row) > 11 else None

                if not all([item_number, name, quantity, unit]):
                    errors.append(f'Zeile {row_num}: Pflichtfelder fehlen')
                    error_count += 1
                    continue

                # Duplikat-Check
                if MagazineItem.objects.filter(item_number=item_number).exists():
                    errors.append(f'Zeile {row_num}: Artikel {item_number} existiert bereits')
                    error_count += 1
                    continue

                # Kategorie
                category = None
                if len(row) > 3 and row[3]:
                    try:
                        category = Category.objects.get(id=int(row[3]), is_active=True)
                    except (Category.DoesNotExist, ValueError):
                        errors.append(f'Zeile {row_num}: Kategorie-ID {row[3]} nicht gefunden')
                        error_count += 1
                        continue

                # Artikel erstellen
                item = MagazineItem(
                    item_number=item_number,
                    name=name,
                    description=row[2] if len(row) > 2 else '',
                    category=category,
                    item_type=row[4] if len(row) > 4 and row[4] else 'consumable',
                    size=row[5] if len(row) > 5 else '',
                    material=row[6] if len(row) > 6 else '',
                    color=row[7] if len(row) > 7 else '',
                    manufacturer=row[8] if len(row) > 8 else '',
                    quantity=float(quantity.replace(',', '.')),
                    unit=unit,
                    min_quantity=float(row[12].replace(',', '.')) if len(row) > 12 and row[12] else None,
                    max_quantity=float(row[13].replace(',', '.')) if len(row) > 13 and row[13] else None,
                    reorder_point=float(row[14].replace(',', '.')) if len(row) > 14 and row[14] else None,
                    standard_order_quantity=float(row[15].replace(',', '.')) if len(row) > 15 and row[15] else None,
                    barcode=row[17] if len(row) > 17 else '',
                    is_hazardous=(row[18].lower() in ['ja', 'yes', '1', 'true']) if len(row) > 18 and row[18] else False,
                    has_expiry_date=(row[20].lower() in ['ja', 'yes', '1', 'true']) if len(row) > 20 and row[20] else False,
                    shelf_life_months=int(row[21]) if len(row) > 21 and row[21] else None,
                    weight_per_unit=float(row[22].replace(',', '.')) if len(row) > 22 and row[22] else None,
                    volume_per_unit=float(row[23].replace(',', '.')) if len(row) > 23 and row[23] else None,
                    storage_temperature_min=float(row[24].replace(',', '.')) if len(row) > 24 and row[24] else None,
                    storage_temperature_max=float(row[25].replace(',', '.')) if len(row) > 25 and row[25] else None,
                    storage_instructions=row[26] if len(row) > 26 else '',
                    technical_specifications=row[27] if len(row) > 27 else '',
                    usage_instructions=row[28] if len(row) > 28 else '',
                    created_by=request.user,
                    updated_by=request.user
                )
                item.save()
                success_count += 1

            except Exception as e:
                errors.append(f'Zeile {row_num}: {str(e)}')
                error_count += 1
                logger.error(f"Import error in row {row_num}: {e}")

        # Feedback
        if success_count > 0:
            messages.success(request, f'{success_count} Artikel erfolgreich importiert.')
        if error_count > 0:
            messages.warning(request, f'{error_count} Fehler beim Import.')
            for error in errors[:10]:  # Nur erste 10 Fehler anzeigen
                messages.error(request, error)

    except Exception as e:
        messages.error(request, f'Fehler beim Lesen der CSV-Datei: {str(e)}')
        logger.error(f"CSV import error: {e}")

    return redirect('magazine:import_export')


@login_required
def import_batches(request):
    """Import von Magazin-Chargen aus CSV"""
    if request.method != 'POST' or not request.FILES.get('file'):
        messages.error(request, 'Keine Datei hochgeladen.')
        return redirect('magazine:import_export')

    csv_file = request.FILES['file']

    if not csv_file.name.endswith('.csv'):
        messages.error(request, 'Bitte eine CSV-Datei hochladen.')
        return redirect('magazine:import_export')

    try:
        from datetime import datetime
        from locations.models import Location

        decoded_file = csv_file.read().decode('utf-8-sig')
        reader = csv.reader(io.StringIO(decoded_file), delimiter=';')

        success_count = 0
        error_count = 0
        errors = []

        for row_num, row in enumerate(reader, start=1):
            # Kommentare und Header überspringen
            if not row or (row[0] and row[0].startswith('#')) or row_num <= 7:
                continue

            try:
                # Pflichtfelder
                item_number = row[0] if len(row) > 0 else None
                batch_number = row[1] if len(row) > 1 else None
                received_date = row[3] if len(row) > 3 else None
                quantity_received = row[5] if len(row) > 5 else None
                quantity_remaining = row[6] if len(row) > 6 else None
                location_id = row[7] if len(row) > 7 else None

                if not all([item_number, batch_number, received_date, quantity_received, quantity_remaining, location_id]):
                    errors.append(f'Zeile {row_num}: Pflichtfelder fehlen')
                    error_count += 1
                    continue

                # Artikel finden
                try:
                    item = MagazineItem.objects.get(item_number=item_number, is_active=True)
                except MagazineItem.DoesNotExist:
                    errors.append(f'Zeile {row_num}: Artikel {item_number} nicht gefunden')
                    error_count += 1
                    continue

                # Duplikat-Check
                if MagazineBatch.objects.filter(item=item, batch_number=batch_number).exists():
                    errors.append(f'Zeile {row_num}: Charge {batch_number} für Artikel {item_number} existiert bereits')
                    error_count += 1
                    continue

                # Lagerort finden
                try:
                    location = Location.objects.get(id=int(location_id), is_active=True)
                except (Location.DoesNotExist, ValueError):
                    errors.append(f'Zeile {row_num}: Lagerort-ID {location_id} nicht gefunden')
                    error_count += 1
                    continue

                # Datum parsen
                try:
                    received_date_obj = datetime.strptime(received_date, '%d.%m.%Y').date()
                except ValueError:
                    errors.append(f'Zeile {row_num}: Ungültiges Datum {received_date}')
                    error_count += 1
                    continue

                expiry_date_obj = None
                if len(row) > 4 and row[4]:
                    try:
                        expiry_date_obj = datetime.strptime(row[4], '%d.%m.%Y').date()
                    except ValueError:
                        errors.append(f'Zeile {row_num}: Ungültiges Verfallsdatum {row[4]}')
                        error_count += 1
                        continue

                # Charge erstellen
                batch = MagazineBatch(
                    item=item,
                    batch_number=batch_number,
                    supplier_batch_number=row[2] if len(row) > 2 else '',
                    received_date=received_date_obj,
                    expiry_date=expiry_date_obj,
                    quantity_received=float(quantity_received.replace(',', '.')),
                    quantity_remaining=float(quantity_remaining.replace(',', '.')),
                    location=location,
                    notes=row[8] if len(row) > 8 else ''
                )
                batch.save()
                success_count += 1

            except Exception as e:
                errors.append(f'Zeile {row_num}: {str(e)}')
                error_count += 1
                logger.error(f"Batch import error in row {row_num}: {e}")

        # Feedback
        if success_count > 0:
            messages.success(request, f'{success_count} Chargen erfolgreich importiert.')
        if error_count > 0:
            messages.warning(request, f'{error_count} Fehler beim Import.')
            for error in errors[:10]:  # Nur erste 10 Fehler anzeigen
                messages.error(request, error)

    except Exception as e:
        messages.error(request, f'Fehler beim Lesen der CSV-Datei: {str(e)}')
        logger.error(f"CSV import error: {e}")

    return redirect('magazine:import_export')


# ============================================================================
# BARCODE & QR-CODE
# ============================================================================

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
import logging

logger = logging.getLogger(__name__)


@login_required
def item_barcode(request, pk):
    """
    Generiert Barcode für MagazineItem als SVG
    Nutzt Code128 für Artikelnummer
    """
    item = get_object_or_404(MagazineItem, pk=pk)

    try:
        svg_string = item.generate_barcode()

        if not svg_string:
            return HttpResponse("Kein Barcode verfügbar", status=400)

        return HttpResponse(svg_string, content_type='image/svg+xml')

    except Exception as e:
        logger.error(f"Barcode generation error for item {pk}: {e}")
        return HttpResponse(f"Fehler bei Barcode-Generierung: {str(e)}", status=500)


@login_required
def item_qr_code(request, pk):
    """
    Generiert QR-Code für MagazineItem als SVG
    Enthält JSON mit Artikeldetails
    """
    item = get_object_or_404(MagazineItem, pk=pk)

    try:
        svg_string = item.generate_qr_code()

        if not svg_string:
            return HttpResponse("Kein QR-Code verfügbar", status=400)

        return HttpResponse(svg_string, content_type='image/svg+xml')

    except Exception as e:
        logger.error(f"QR-Code generation error for item {pk}: {e}")
        return HttpResponse(f"Fehler bei QR-Code-Generierung: {str(e)}", status=500)


# ============================================================================
# INVENTUR VIEWS
# ============================================================================

from .inventory_views import (
    MagazineInventoryListView,
    MagazineInventoryCreateView,
    MagazineInventoryDetailView,
    MagazineInventoryStartView,
    MagazineInventoryCountingView,
    MagazineInventoryCompleteView,
    MagazineInventoryApproveView,
    MagazineInventoryItemUpdateView,
    MagazineInventoryProgressView,
    MagazineInventoryExportView,
)
