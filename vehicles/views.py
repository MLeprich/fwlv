"""
Vehicles Views
CRUD-Ansichten für Fahrzeuge und Prüfungen
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, View
from django.urls import reverse_lazy, reverse
from django.db.models import Q, Count, Max
from django.utils.translation import gettext_lazy as _
from django.utils import timezone

from .models import Vehicle, VehicleInspection, VehicleStatus, VehicleModel, VehicleType
from inventory_base.models import Supplier
from .forms import VehicleForm, VehicleInspectionForm, VehicleTypeForm
from datetime import timedelta, datetime
from calendar import monthcalendar


class VehicleDashboardView(LoginRequiredMixin, ListView):
    """
    Dashboard für Fahrzeugverwaltung mit Statistiken
    """
    model = Vehicle
    template_name = 'vehicles/dashboard.html'
    context_object_name = 'vehicles'

    def get_queryset(self):
        return Vehicle.objects.none()  # Wir brauchen keine Liste im Dashboard

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()

        # Gesamtanzahl Fahrzeuge
        context['total_vehicles'] = Vehicle.objects.filter(is_active=True).count()

        # Einsatzbereite Fahrzeuge
        context['operational_count'] = Vehicle.objects.filter(
            status=VehicleStatus.OPERATIONAL,
            is_active=True
        ).count()

        # Fahrzeuge in Wartung/Reparatur
        context['maintenance_count'] = Vehicle.objects.filter(
            status__in=[VehicleStatus.MAINTENANCE, VehicleStatus.REPAIR],
            is_active=True
        ).count()

        # HU/TÜV fällig (in den nächsten 30 Tagen)
        due_date = today + timedelta(days=30)
        context['inspection_due_count'] = Vehicle.objects.filter(
            next_inspection_date__lte=due_date,
            next_inspection_date__gte=today,
            is_active=True
        ).count()

        # Überfällige Prüfungen
        context['overdue_inspections'] = Vehicle.objects.filter(
            next_inspection_date__lt=today,
            is_active=True
        ).select_related('home_location')[:10]

        # Letzte Prüfungen
        context['recent_inspections'] = VehicleInspection.objects.select_related(
            'vehicle'
        ).order_by('-inspection_date')[:10]

        return context


class VehicleListView(LoginRequiredMixin, ListView):
    """
    Liste aller Fahrzeuge mit Such- und Filterfunktion
    """
    model = Vehicle
    template_name = 'vehicles/vehicle_list.html'
    context_object_name = 'vehicles'
    paginate_by = 50

    def get_queryset(self):
        queryset = Vehicle.objects.select_related(
            'home_location',
            'responsible_person',
            'mobile_storage_location',
            'vehicle_type'
        ).annotate(
            inspection_count=Count('inspections')
        )

        # Suchfunktion
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(call_sign__icontains=search_query) |
                Q(name__icontains=search_query) |
                Q(license_plate__icontains=search_query)
            )

        # Filter: Nur Aktive
        is_active = self.request.GET.get('is_active', '')
        if is_active == 'true':
            queryset = queryset.filter(is_active=True)
        elif is_active == 'false':
            queryset = queryset.filter(is_active=False)

        # Filter: Fahrzeugtyp (now FK-based)
        vehicle_type = self.request.GET.get('vehicle_type', '')
        if vehicle_type:
            queryset = queryset.filter(vehicle_type_id=vehicle_type)

        # Filter: Status
        status = self.request.GET.get('status', '')
        if status:
            queryset = queryset.filter(status=status)

        # Filter: Mobile Lager
        is_mobile_storage = self.request.GET.get('is_mobile_storage', '')
        if is_mobile_storage == 'true':
            queryset = queryset.filter(is_mobile_storage=True)

        # Filter: Retirement Approaching (within 90 days)
        if self.request.GET.get('retirement_approaching'):
            today = timezone.now().date()
            upcoming_date = today + timedelta(days=90)
            queryset = queryset.filter(
                planned_retirement_date__lte=upcoming_date,
                planned_retirement_date__gte=today,
                actual_retirement_date__isnull=True
            )

        # Filter: Retired
        if self.request.GET.get('retired'):
            queryset = queryset.filter(actual_retirement_date__isnull=False)

        return queryset.order_by('call_sign')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['is_active_filter'] = self.request.GET.get('is_active', '')
        context['vehicle_type_filter'] = self.request.GET.get('vehicle_type', '')
        context['status_filter'] = self.request.GET.get('status', '')
        context['is_mobile_storage_filter'] = self.request.GET.get('is_mobile_storage', '')
        context['retirement_approaching_filter'] = self.request.GET.get('retirement_approaching', '')
        context['retired_filter'] = self.request.GET.get('retired', '')

        # Für Filter-Dropdowns
        context['vehicle_types'] = VehicleType.objects.all()
        context['vehicle_statuses'] = VehicleStatus.choices

        return context


class VehicleDetailView(LoginRequiredMixin, DetailView):
    """
    Detail-Ansicht eines Fahrzeugs mit Prüfhistorie
    """
    model = Vehicle
    template_name = 'vehicles/vehicle_detail.html'
    context_object_name = 'vehicle'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Letzte Prüfungen
        context['recent_inspections'] = self.object.get_recent_inspections(limit=10)

        # Warnung: Prüfungen fällig?
        context['inspection_due'] = self.object.is_inspection_due()
        context['safety_check_due'] = self.object.is_safety_check_due()
        context['insurance_expiring'] = self.object.is_insurance_expiring()

        # Wer trug den aktuellen Funkrufnamen vor diesem Fahrzeug?
        from .call_signs import history_of_call_sign
        context['call_sign_history'] = (
            list(history_of_call_sign(self.object.call_sign)) if self.object.call_sign else []
        )

        return context


class VehicleCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """
    Fahrzeug erstellen
    """
    model = Vehicle
    form_class = VehicleForm
    template_name = 'vehicles/vehicle_form.html'
    permission_required = 'vehicles.add_vehicle'

    def form_valid(self, form):
        # Audit-Felder setzen
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user

        messages.success(
            self.request,
            _('Fahrzeug "{}" wurde erfolgreich erstellt.').format(form.instance.call_sign)
        )

        return super().form_valid(form)

    def get_success_url(self):
        return reverse('vehicles:detail', kwargs={'pk': self.object.pk})


class VehicleUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """
    Fahrzeug bearbeiten
    """
    model = Vehicle
    form_class = VehicleForm
    template_name = 'vehicles/vehicle_form.html'
    permission_required = 'vehicles.change_vehicle'

    def form_valid(self, form):
        # Audit-Feld setzen
        form.instance.updated_by = self.request.user

        messages.success(
            self.request,
            _('Fahrzeug "{}" wurde erfolgreich aktualisiert.').format(form.instance.call_sign)
        )

        return super().form_valid(form)

    def get_success_url(self):
        return reverse('vehicles:detail', kwargs={'pk': self.object.pk})


class VehicleDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """
    Fahrzeug löschen (mit Sicherheitsabfrage)
    """
    model = Vehicle
    template_name = 'vehicles/vehicle_confirm_delete.html'
    success_url = reverse_lazy('vehicles:list')
    permission_required = 'vehicles.delete_vehicle'

    def delete(self, request, *args, **kwargs):
        vehicle = self.get_object()
        messages.success(
            request,
            _('Fahrzeug "{}" wurde erfolgreich gelöscht.').format(vehicle.call_sign)
        )
        return super().delete(request, *args, **kwargs)


# ============================================================================
# VEHICLE INSPECTION VIEWS
# ============================================================================

class VehicleInspectionCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """
    Prüfung zu Fahrzeug hinzufügen
    """
    model = VehicleInspection
    form_class = VehicleInspectionForm
    template_name = 'vehicles/inspection_form.html'
    permission_required = 'vehicles.add_vehicleinspection'

    def get_initial(self):
        """Fahrzeug-ID aus URL vorausfüllen"""
        initial = super().get_initial()
        vehicle_id = self.kwargs.get('vehicle_pk')
        if vehicle_id:
            initial['vehicle'] = vehicle_id
        return initial

    def form_valid(self, form):
        # Audit-Felder setzen
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user

        # Bei bestandener Prüfung: Fahrzeug-Datum aktualisieren
        if form.instance.next_inspection_date:
            vehicle = form.instance.vehicle
            if form.instance.inspection_type == 'hu':
                vehicle.next_inspection_date = form.instance.next_inspection_date
                vehicle.save()
            elif form.instance.inspection_type == 'uvv':
                vehicle.next_safety_check_date = form.instance.next_inspection_date
                vehicle.save()

        messages.success(
            self.request,
            _('Prüfung wurde erfolgreich hinzugefügt.')
        )

        return super().form_valid(form)

    def get_success_url(self):
        return reverse('vehicles:detail', kwargs={'pk': self.object.vehicle.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Fahrzeug-Objekt für Breadcrumb/Titel
        vehicle_id = self.kwargs.get('vehicle_pk')
        if vehicle_id:
            context['vehicle'] = get_object_or_404(Vehicle, pk=vehicle_id)

        return context


class InspectionCreateDirectView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """
    Prüfung direkt anlegen (ohne vorausgewähltes Fahrzeug)
    """
    model = VehicleInspection
    form_class = VehicleInspectionForm
    template_name = 'vehicles/inspection_form.html'
    permission_required = 'vehicles.add_vehicleinspection'

    def form_valid(self, form):
        # Audit-Felder setzen
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user

        # Bei bestandener Prüfung: Fahrzeug-Datum aktualisieren
        if form.instance.next_inspection_date:
            vehicle = form.instance.vehicle
            if form.instance.inspection_type == 'hu':
                vehicle.next_inspection_date = form.instance.next_inspection_date
                vehicle.save()
            elif form.instance.inspection_type == 'uvv':
                vehicle.next_safety_check_date = form.instance.next_inspection_date
                vehicle.save()

        messages.success(
            self.request,
            _('Prüfung wurde erfolgreich erstellt.')
        )

        return super().form_valid(form)

    def get_success_url(self):
        return reverse('vehicles:inspection_detail', kwargs={'pk': self.object.pk})


class VehicleInspectionUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """
    Prüfung bearbeiten
    """
    model = VehicleInspection
    form_class = VehicleInspectionForm
    template_name = 'vehicles/inspection_form.html'
    permission_required = 'vehicles.change_vehicleinspection'

    def form_valid(self, form):
        # Audit-Feld setzen
        form.instance.updated_by = self.request.user

        # Bei bestandener Prüfung: Fahrzeug-Datum aktualisieren
        if form.instance.next_inspection_date:
            vehicle = form.instance.vehicle
            if form.instance.inspection_type == 'hu':
                vehicle.next_inspection_date = form.instance.next_inspection_date
                vehicle.save()
            elif form.instance.inspection_type == 'uvv':
                vehicle.next_safety_check_date = form.instance.next_inspection_date
                vehicle.save()

        messages.success(
            self.request,
            _('Prüfung wurde erfolgreich aktualisiert.')
        )

        return super().form_valid(form)

    def get_success_url(self):
        return reverse('vehicles:detail', kwargs={'pk': self.object.vehicle.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['vehicle'] = self.object.vehicle
        return context


class VehicleInspectionDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """
    Prüfung löschen
    """
    model = VehicleInspection
    template_name = 'vehicles/inspection_confirm_delete.html'
    permission_required = 'vehicles.delete_vehicleinspection'

    def get_success_url(self):
        return reverse('vehicles:detail', kwargs={'pk': self.object.vehicle.pk})

    def delete(self, request, *args, **kwargs):
        inspection = self.get_object()

        messages.success(
            request,
            _('Prüfung wurde erfolgreich gelöscht.')
        )

        return super().delete(request, *args, **kwargs)


class InspectionListView(LoginRequiredMixin, ListView):
    """
    Liste aller Prüfungen
    """
    model = VehicleInspection
    template_name = 'vehicles/inspection_list.html'
    context_object_name = 'inspections'
    paginate_by = 50

    def get_queryset(self):
        queryset = VehicleInspection.objects.select_related(
            'vehicle', 'inspector'
        ).order_by('-inspection_date')

        # Filter nach Fahrzeug
        vehicle_id = self.request.GET.get('vehicle')
        if vehicle_id:
            queryset = queryset.filter(vehicle_id=vehicle_id)

        # Filter nach Prüfungstyp
        inspection_type = self.request.GET.get('inspection_type')
        if inspection_type:
            queryset = queryset.filter(inspection_type=inspection_type)

        return queryset


class InspectionDetailView(LoginRequiredMixin, DetailView):
    """
    Detail-Ansicht einer Prüfung
    """
    model = VehicleInspection
    template_name = 'vehicles/inspection_detail.html'
    context_object_name = 'inspection'


@login_required
def vehicle_calendar(request):
    """
    Kalenderansicht für HU/TÜV und UVV-Prüfungen
    Zeigt wann welches Fahrzeug zur Prüfung muss
    """
    # Monat und Jahr aus Request oder aktuell
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))
    selected_day = request.GET.get('day')

    # Kalender generieren
    cal = monthcalendar(year, month)

    # Zeitraum für diesen Monat
    month_start = datetime(year, month, 1).date()
    if month == 12:
        month_end = datetime(year + 1, 1, 1).date()
    else:
        month_end = datetime(year, month + 1, 1).date()

    today = timezone.now().date()

    # Alle Fahrzeuge mit Prüfungen in diesem Monat
    inspections_by_day = {}

    # HU/TÜV Termine sammeln
    hu_vehicles = Vehicle.objects.filter(
        is_active=True,
        next_inspection_date__gte=month_start,
        next_inspection_date__lt=month_end
    ).select_related('home_location')

    for vehicle in hu_vehicles:
        day = vehicle.next_inspection_date.day
        if day not in inspections_by_day:
            inspections_by_day[day] = []

        is_overdue = vehicle.next_inspection_date < today
        is_urgent = vehicle.next_inspection_date <= today + timedelta(days=30) and not is_overdue

        inspections_by_day[day].append({
            'vehicle': vehicle,
            'type': 'HU/TÜV',
            'date': vehicle.next_inspection_date,
            'is_overdue': is_overdue,
            'is_urgent': is_urgent,
            'color': 'red' if is_overdue else 'orange' if is_urgent else 'green'
        })

    # UVV Termine sammeln
    uvv_vehicles = Vehicle.objects.filter(
        is_active=True,
        next_safety_check_date__gte=month_start,
        next_safety_check_date__lt=month_end
    ).select_related('home_location')

    for vehicle in uvv_vehicles:
        day = vehicle.next_safety_check_date.day
        if day not in inspections_by_day:
            inspections_by_day[day] = []

        is_overdue = vehicle.next_safety_check_date < today
        is_urgent = vehicle.next_safety_check_date <= today + timedelta(days=30) and not is_overdue

        inspections_by_day[day].append({
            'vehicle': vehicle,
            'type': 'UVV',
            'date': vehicle.next_safety_check_date,
            'is_overdue': is_overdue,
            'is_urgent': is_urgent,
            'color': 'red' if is_overdue else 'orange' if is_urgent else 'blue'
        })

    # Ausgewählter Tag - Standard ist heute (falls im aktuellen Monat)
    if selected_day:
        selected_day = int(selected_day)
    else:
        if today.year == year and today.month == month:
            selected_day = today.day
        else:
            # Oder den ersten Tag mit Prüfungen
            selected_day = min(inspections_by_day.keys()) if inspections_by_day else 1

    # Prüfungen für den ausgewählten Tag
    selected_date = datetime(year, month, selected_day).date()
    selected_inspections = inspections_by_day.get(selected_day, [])

    # Sortiere nach Datum und Typ
    selected_inspections.sort(key=lambda x: (x['date'], x['type']))

    # Navigation
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1

    context = {
        'calendar': cal,
        'year': year,
        'month': month,
        'month_name': datetime(year, month, 1).strftime('%B'),
        'inspections_by_day': inspections_by_day,
        'selected_day': selected_day,
        'selected_date': selected_date,
        'selected_inspections': selected_inspections,
        'prev_month': prev_month,
        'prev_year': prev_year,
        'next_month': next_month,
        'next_year': next_year,
    }

    # HTMX Request - nur die Tagesdetails zurückgeben
    if request.headers.get('HX-Request'):
        return render(request, 'vehicles/partials/calendar_day_detail.html', context)

    return render(request, 'vehicles/calendar.html', context)


from django.http import JsonResponse, HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl
from io import BytesIO


class CalendarDataView(LoginRequiredMixin, View):
    """
    JSON-Endpoint für Kalender-Daten (FullCalendar.js)
    """
    def get(self, request, *args, **kwargs):
        today = timezone.now().date()

        events = []

        # HU/TÜV Termine
        vehicles = Vehicle.objects.filter(
            next_inspection_date__isnull=False,
            is_active=True
        )

        for vehicle in vehicles:
            # Farbe basierend auf Status
            if vehicle.next_inspection_date < today:
                color = '#EF4444'  # red
            elif vehicle.next_inspection_date <= today + timedelta(days=30):
                color = '#F59E0B'  # orange
            else:
                color = '#10B981'  # green

            events.append({
                'title': f'HU: {vehicle.call_sign}',
                'start': vehicle.next_inspection_date.isoformat(),
                'url': reverse('vehicles:detail', kwargs={'pk': vehicle.pk}),
                'color': color,
                'extendedProps': {
                    'type': 'hu',
                    'vehicle_id': vehicle.pk,
                    'vehicle_name': vehicle.name
                }
            })

        # UVV Termine
        vehicles_uvv = Vehicle.objects.filter(
            next_safety_check_date__isnull=False,
            is_active=True
        )

        for vehicle in vehicles_uvv:
            if vehicle.next_safety_check_date < today:
                color = '#EF4444'
            elif vehicle.next_safety_check_date <= today + timedelta(days=30):
                color = '#F59E0B'
            else:
                color = '#3B82F6'  # blue

            events.append({
                'title': f'UVV: {vehicle.call_sign}',
                'start': vehicle.next_safety_check_date.isoformat(),
                'url': reverse('vehicles:detail', kwargs={'pk': vehicle.pk}),
                'color': color,
                'extendedProps': {
                    'type': 'uvv',
                    'vehicle_id': vehicle.pk,
                    'vehicle_name': vehicle.name
                }
            })

        return JsonResponse(events, safe=False)


# Placeholder Views für noch nicht implementierte Features
class MileageUpdateModalView(LoginRequiredMixin, DetailView):
    model = Vehicle
    template_name = 'vehicles/modals/mileage_update.html'
    context_object_name = 'vehicle'


class StatusUpdateModalView(LoginRequiredMixin, DetailView):
    model = Vehicle
    template_name = 'vehicles/modals/status_update.html'
    context_object_name = 'vehicle'


class InspectionModalView(LoginRequiredMixin, DetailView):
    model = Vehicle
    template_name = 'vehicles/modals/inspection_quick.html'
    context_object_name = 'vehicle'


class LoadingPlanView(LoginRequiredMixin, DetailView):
    model = Vehicle
    template_name = 'vehicles/loading_plan.html'
    context_object_name = 'vehicle'


class LoadingPlanEditView(LoginRequiredMixin, UpdateView):
    model = Vehicle
    template_name = 'vehicles/loading_plan_edit.html'
    fields = ['is_mobile_storage']
    context_object_name = 'vehicle'


class HandoverCreateView(LoginRequiredMixin, CreateView):
    model = Vehicle  # Placeholder - sollte eigenes HandoverModel sein
    template_name = 'vehicles/handover_form.html'
    fields = []


class HandoverDetailView(LoginRequiredMixin, DetailView):
    model = Vehicle  # Placeholder
    template_name = 'vehicles/handover_detail.html'
    context_object_name = 'handover'


# ============================================================================
# MANUFACTURER VIEWS
# ============================================================================

# Manufacturer CRUD Views entfernt - verwende /organization/suppliers/ stattdessen


# ============================================================================
# MODEL VIEWS
# ============================================================================

class VehicleModelListView(LoginRequiredMixin, ListView):
    """
    Liste aller Fahrzeugmodelle
    """
    model = VehicleModel
    template_name = 'vehicles/model_list.html'
    context_object_name = 'models'
    paginate_by = 50

    def get_queryset(self):
        queryset = VehicleModel.objects.select_related('manufacturer').annotate(
            vehicle_count=Count('vehicles')
        )

        # Suchfunktion
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(manufacturer__name__icontains=search_query)
            )

        # Filter: Hersteller
        manufacturer_id = self.request.GET.get('manufacturer', '')
        if manufacturer_id:
            queryset = queryset.filter(manufacturer_id=manufacturer_id)

        # Filter: Nur Aktive
        is_active = self.request.GET.get('is_active', '')
        if is_active == 'true':
            queryset = queryset.filter(is_active=True)
        elif is_active == 'false':
            queryset = queryset.filter(is_active=False)

        return queryset.order_by('manufacturer__name', 'name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['manufacturer_filter'] = self.request.GET.get('manufacturer', '')
        context['is_active_filter'] = self.request.GET.get('is_active', '')
        context['manufacturers'] = Supplier.objects.filter(is_active=True).order_by('name')
        return context


class VehicleModelCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """
    Modell erstellen
    """
    model = VehicleModel
    template_name = 'vehicles/model_form.html'
    fields = ['manufacturer', 'name', 'vehicle_type', 'year_from', 'year_to', 'notes', 'is_active']
    success_url = reverse_lazy('vehicles:model_list')
    permission_required = 'vehicles.add_vehiclemodel'

    def form_valid(self, form):
        messages.success(
            self.request,
            _('Modell "{}" wurde erfolgreich erstellt.').format(form.instance)
        )
        return super().form_valid(form)


class VehicleModelUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """
    Modell bearbeiten
    """
    model = VehicleModel
    template_name = 'vehicles/model_form.html'
    fields = ['manufacturer', 'name', 'vehicle_type', 'year_from', 'year_to', 'notes', 'is_active']
    success_url = reverse_lazy('vehicles:model_list')
    permission_required = 'vehicles.change_vehiclemodel'

    def form_valid(self, form):
        messages.success(
            self.request,
            _('Modell "{}" wurde erfolgreich aktualisiert.').format(form.instance)
        )
        return super().form_valid(form)


class VehicleModelDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """
    Modell löschen
    """
    model = VehicleModel
    template_name = 'vehicles/model_confirm_delete.html'
    success_url = reverse_lazy('vehicles:model_list')
    permission_required = 'vehicles.delete_vehiclemodel'

    def delete(self, request, *args, **kwargs):
        model = self.get_object()
        messages.success(
            request,
            _('Modell "{}" wurde erfolgreich gelöscht.').format(model)
        )
        return super().delete(request, *args, **kwargs)


# ============================================================================
# VEHICLE TYPE VIEWS
# ============================================================================

class VehicleTypeListView(LoginRequiredMixin, ListView):
    """
    Liste aller Fahrzeugtypen
    """
    model = VehicleType
    template_name = 'vehicles/type_list.html'
    context_object_name = 'vehicle_types'
    paginate_by = 50

    def get_queryset(self):
        queryset = VehicleType.objects.annotate(
            vehicle_count=Count('vehicles')
        )

        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(code__icontains=search_query)
            )

        is_active = self.request.GET.get('is_active', '')
        if is_active == 'true':
            queryset = queryset.filter(is_active=True)
        elif is_active == 'false':
            queryset = queryset.filter(is_active=False)

        return queryset.order_by('order', 'name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['is_active_filter'] = self.request.GET.get('is_active', '')
        return context


class VehicleTypeCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """
    Fahrzeugtyp erstellen
    """
    model = VehicleType
    form_class = VehicleTypeForm
    template_name = 'vehicles/type_form.html'
    success_url = reverse_lazy('vehicles:type_list')
    permission_required = 'vehicles.add_vehicletype'

    def form_valid(self, form):
        messages.success(
            self.request,
            _('Fahrzeugtyp "{}" wurde erfolgreich erstellt.').format(form.instance.name)
        )
        return super().form_valid(form)


class VehicleTypeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """
    Fahrzeugtyp bearbeiten
    """
    model = VehicleType
    form_class = VehicleTypeForm
    template_name = 'vehicles/type_form.html'
    success_url = reverse_lazy('vehicles:type_list')
    permission_required = 'vehicles.change_vehicletype'

    def form_valid(self, form):
        messages.success(
            self.request,
            _('Fahrzeugtyp "{}" wurde erfolgreich aktualisiert.').format(form.instance.name)
        )
        return super().form_valid(form)


class VehicleTypeDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """
    Fahrzeugtyp löschen (mit Lösch-Schutz bei zugewiesenen Fahrzeugen)
    """
    model = VehicleType
    template_name = 'vehicles/type_confirm_delete.html'
    success_url = reverse_lazy('vehicles:type_list')
    permission_required = 'vehicles.delete_vehicletype'

    def delete(self, request, *args, **kwargs):
        vehicle_type = self.get_object()
        messages.success(
            request,
            _('Fahrzeugtyp "{}" wurde erfolgreich gelöscht.').format(vehicle_type.name)
        )
        return super().delete(request, *args, **kwargs)


# ============================================================================
# IMPORT & EXPORT VIEWS
# ============================================================================

class ImportExportView(LoginRequiredMixin, View):
    """
    Import/Export Übersicht
    """
    template_name = 'vehicles/import_export.html'

    def get(self, request):
        context = {
            'vehicle_count': Vehicle.objects.count(),
            'manufacturer_count': Supplier.objects.filter(is_active=True).count(),
            'model_count': VehicleModel.objects.count(),
            'inspection_count': VehicleInspection.objects.count(),
        }
        return render(request, self.template_name, context)


class ExportVehiclesView(LoginRequiredMixin, View):
    """
    Exportiert alle Fahrzeuge als Excel
    """
    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Fahrzeuge"

        # Header Style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Headers
        headers = [
            'Funkrufname', 'Name', 'Fahrzeugtyp', 'Kennzeichen', 'Status',
            'Hersteller', 'Modell', 'Baujahr', 'FIN/VIN', 'Motornummer',
            'Kraftstoffart', 'Kilometerstand', 'Betriebsstunden',
            'Heimatstandort', 'Verantwortliche Person', 'Nächste HU',
            'Nächste UVV', 'Versicherungspolice', 'Versicherung bis', 'Aktiv'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data
        vehicles = Vehicle.objects.select_related('manufacturer', 'model', 'home_location', 'responsible_person', 'vehicle_type').all()
        for row_num, vehicle in enumerate(vehicles, 2):
            ws.cell(row=row_num, column=1, value=vehicle.call_sign)
            ws.cell(row=row_num, column=2, value=vehicle.name)
            ws.cell(row=row_num, column=3, value=vehicle.vehicle_type.name if vehicle.vehicle_type else '')
            ws.cell(row=row_num, column=4, value=vehicle.license_plate)
            ws.cell(row=row_num, column=5, value=vehicle.get_status_display())
            ws.cell(row=row_num, column=6, value=vehicle.manufacturer.name if vehicle.manufacturer else '')
            ws.cell(row=row_num, column=7, value=vehicle.model.name if vehicle.model else '')
            ws.cell(row=row_num, column=8, value=vehicle.year_of_manufacture)
            ws.cell(row=row_num, column=9, value=vehicle.vin)
            ws.cell(row=row_num, column=10, value=vehicle.engine_number)
            ws.cell(row=row_num, column=11, value=vehicle.fuel_type)
            ws.cell(row=row_num, column=12, value=vehicle.current_mileage)
            ws.cell(row=row_num, column=13, value=vehicle.engine_hours)
            ws.cell(row=row_num, column=14, value=vehicle.home_location.name if vehicle.home_location else '')
            ws.cell(row=row_num, column=15, value=str(vehicle.responsible_person) if vehicle.responsible_person else '')
            ws.cell(row=row_num, column=16, value=vehicle.next_inspection_date.strftime('%d.%m.%Y') if vehicle.next_inspection_date else '')
            ws.cell(row=row_num, column=17, value=vehicle.next_safety_check_date.strftime('%d.%m.%Y') if vehicle.next_safety_check_date else '')
            ws.cell(row=row_num, column=18, value=vehicle.insurance_policy_number)
            ws.cell(row=row_num, column=19, value=vehicle.insurance_expires.strftime('%d.%m.%Y') if vehicle.insurance_expires else '')
            ws.cell(row=row_num, column=20, value='Ja' if vehicle.is_active else 'Nein')

        # Auto-width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Response
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="fahrzeuge_export.xlsx"'
        return response


# ExportManufacturersView entfernt - verwende /organization/suppliers/ Export stattdessen


class ExportModelsView(LoginRequiredMixin, View):
    """
    Exportiert alle Modelle als Excel
    """
    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Modelle"

        # Header Style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Headers
        headers = [
            'Hersteller', 'Modellname', 'Standardtyp', 'Baujahr von',
            'Baujahr bis', 'Notizen', 'Aktiv'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data
        models = VehicleModel.objects.select_related('manufacturer', 'vehicle_type').all()
        for row_num, model in enumerate(models, 2):
            ws.cell(row=row_num, column=1, value=model.manufacturer.name)
            ws.cell(row=row_num, column=2, value=model.name)
            ws.cell(row=row_num, column=3, value=model.vehicle_type.name if model.vehicle_type else '')
            ws.cell(row=row_num, column=4, value=model.year_from)
            ws.cell(row=row_num, column=5, value=model.year_to)
            ws.cell(row=row_num, column=6, value=model.notes)
            ws.cell(row=row_num, column=7, value='Ja' if model.is_active else 'Nein')

        # Auto-width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Response
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="modelle_export.xlsx"'
        return response


class DownloadTemplateView(LoginRequiredMixin, View):
    """
    Erstellt leere Vorlagen zum Download
    """
    def get(self, request, template_type):
        wb = Workbook()
        ws = wb.active

        # Header Style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        if template_type == 'vehicles':
            ws.title = "Fahrzeuge Vorlage"
            headers = [
                'Funkrufname*', 'Name*', 'Fahrzeugtyp*', 'Kennzeichen*', 'Status',
                'Hersteller', 'Modell', 'Baujahr', 'FIN/VIN', 'Motornummer',
                'Kraftstoffart', 'Kilometerstand', 'Betriebsstunden',
                'Aktiv (Ja/Nein)'
            ]
            filename = 'fahrzeuge_vorlage.xlsx'

        elif template_type == 'models':
            ws.title = "Modelle Vorlage"
            headers = [
                'Hersteller*', 'Modellname*', 'Standardtyp', 'Baujahr von',
                'Baujahr bis', 'Notizen', 'Aktiv (Ja/Nein)'
            ]
            filename = 'modelle_vorlage.xlsx'
        else:
            return HttpResponse("Ungültiger Vorlagentyp", status=400)

        # Headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Example row
        ws.cell(row=2, column=1, value='Beispiel...')

        # Auto-width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Response
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ImportVehiclesView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    Importiert Fahrzeuge aus Excel
    """
    permission_required = 'vehicles.add_vehicle'

    def post(self, request):
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            messages.error(request, 'Bitte wählen Sie eine Datei aus.')
            return redirect('vehicles:import_export')

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            imported_count = 0
            error_count = 0
            errors = []

            # Start from row 2 (skip header)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # Skip empty rows
                    continue

                try:
                    # Parse row data
                    call_sign = row[0]
                    name = row[1]
                    vehicle_type = row[2] if len(row) > 2 else None
                    license_plate = row[3] if len(row) > 3 else None

                    # Check if vehicle already exists
                    if Vehicle.objects.filter(call_sign=call_sign).exists():
                        errors.append(f'Zeile {row_num}: Fahrzeug mit Funkrufname "{call_sign}" existiert bereits')
                        error_count += 1
                        continue

                    # Create vehicle
                    vehicle = Vehicle(
                        call_sign=call_sign,
                        name=name,
                        license_plate=license_plate,
                        is_active=True,
                        created_by=request.user,
                        updated_by=request.user
                    )

                    # Set optional fields
                    if vehicle_type:
                        # Try to match vehicle_type by name or code
                        try:
                            vt = VehicleType.objects.get(
                                Q(name__iexact=vehicle_type) | Q(code__iexact=vehicle_type)
                            )
                            vehicle.vehicle_type = vt
                        except VehicleType.DoesNotExist:
                            pass

                    vehicle.save()
                    imported_count += 1

                except Exception as e:
                    errors.append(f'Zeile {row_num}: {str(e)}')
                    error_count += 1

            # Success message
            if imported_count > 0:
                messages.success(request, f'{imported_count} Fahrzeug(e) erfolgreich importiert.')

            if error_count > 0:
                messages.warning(request, f'{error_count} Fehler beim Import. Siehe Details unten.')
                for error in errors[:10]:  # Show max 10 errors
                    messages.error(request, error)

        except Exception as e:
            messages.error(request, f'Fehler beim Lesen der Datei: {str(e)}')

        return redirect('vehicles:import_export')


# ImportManufacturersView entfernt - verwende /organization/suppliers/ Import stattdessen


class ImportModelsView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    Importiert Modelle aus Excel
    """
    permission_required = 'vehicles.add_vehiclemodel'

    def post(self, request):
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            messages.error(request, 'Bitte wählen Sie eine Datei aus.')
            return redirect('vehicles:import_export')

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            imported_count = 0
            error_count = 0
            errors = []

            # Start from row 2 (skip header)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0] or not row[1]:  # Skip empty rows or missing required fields
                    continue

                try:
                    # Parse row data
                    manufacturer_name = row[0]
                    model_name = row[1]

                    # Find manufacturer
                    try:
                        manufacturer = Supplier.objects.get(name=manufacturer_name)
                    except Supplier.DoesNotExist:
                        errors.append(f'Zeile {row_num}: Hersteller "{manufacturer_name}" nicht gefunden')
                        error_count += 1
                        continue

                    # Check if model already exists
                    if VehicleModel.objects.filter(manufacturer=manufacturer, name=model_name).exists():
                        errors.append(f'Zeile {row_num}: Modell "{model_name}" von "{manufacturer_name}" existiert bereits')
                        error_count += 1
                        continue

                    # Create model
                    model = VehicleModel(
                        manufacturer=manufacturer,
                        name=model_name,
                        notes=row[5] if len(row) > 5 and row[5] else '',
                        is_active=True if len(row) <= 6 or not row[6] or str(row[6]).lower() in ['ja', 'yes', '1', 'true'] else False
                    )

                    # Set optional fields
                    if len(row) > 2 and row[2]:
                        # Try to match vehicle_type by name or code
                        try:
                            vt = VehicleType.objects.get(
                                Q(name__iexact=str(row[2])) | Q(code__iexact=str(row[2]))
                            )
                            model.vehicle_type = vt
                        except VehicleType.DoesNotExist:
                            pass

                    if len(row) > 3 and row[3]:
                        model.year_from = int(row[3])

                    if len(row) > 4 and row[4]:
                        model.year_to = int(row[4])

                    model.save()
                    imported_count += 1

                except Exception as e:
                    errors.append(f'Zeile {row_num}: {str(e)}')
                    error_count += 1

            # Success message
            if imported_count > 0:
                messages.success(request, f'{imported_count} Modell(e) erfolgreich importiert.')

            if error_count > 0:
                messages.warning(request, f'{error_count} Fehler beim Import. Siehe Details unten.')
                for error in errors[:10]:  # Show max 10 errors
                    messages.error(request, error)

        except Exception as e:
            messages.error(request, f'Fehler beim Lesen der Datei: {str(e)}')

        return redirect('vehicles:import_export')



# ============================================================================
# LIFECYCLE MANAGEMENT VIEWS
# ============================================================================

class SetReplacementVehicleView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """
    Nachfolgefahrzeug für ausscheidendes Fahrzeug festlegen
    """
    model = Vehicle
    template_name = 'vehicles/set_replacement.html'
    fields = ['replacement_vehicle', 'actual_retirement_date', 'retirement_reason']
    permission_required = 'vehicles.change_vehicle'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Nur aktive Fahrzeuge als Nachfolger anzeigen (außer dem aktuellen)
        context['available_vehicles'] = Vehicle.objects.filter(
            is_active=True,
            actual_retirement_date__isnull=True
        ).exclude(pk=self.object.pk).order_by('call_sign')
        return context

    def form_valid(self, form):
        messages.success(
            self.request,
            _('Nachfolgefahrzeug für "{}" wurde erfolgreich festgelegt.').format(self.object.call_sign)
        )
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('vehicles:list')


class CallSignReassignView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    Funkrufnamen eines Fahrzeugs umhängen oder freigeben.

    Bewusst eine eigene Aktion statt eines Feldes im Fahrzeugformular: das Umhängen
    betrifft immer ZWEI Fahrzeuge und muss die Historie fortschreiben. Wer den
    Funkrufnamen einfach im Formular überschreibt, hinterlässt eine Historie, die lügt.
    """
    permission_required = 'vehicles.change_vehicle'
    template_name = 'vehicles/call_sign_reassign.html'

    def get_vehicle(self):
        return get_object_or_404(Vehicle, pk=self.kwargs['pk'])

    def get(self, request, *args, **kwargs):
        from .call_signs import history_of_call_sign
        from .forms import CallSignReassignForm

        vehicle = self.get_vehicle()
        return render(request, self.template_name, {
            'vehicle': vehicle,
            'form': CallSignReassignForm(vehicle=vehicle),
            'history': history_of_call_sign(vehicle.call_sign) if vehicle.call_sign else [],
        })

    def post(self, request, *args, **kwargs):
        from django.core.exceptions import ValidationError
        from .call_signs import assign_call_sign, history_of_call_sign, release_call_sign
        from .forms import CallSignReassignForm

        vehicle = self.get_vehicle()
        form = CallSignReassignForm(request.POST, vehicle=vehicle)

        if form.is_valid():
            call_sign = vehicle.call_sign
            if not call_sign:
                messages.error(request, _('Dieses Fahrzeug trägt aktuell keinen Funkrufnamen.'))
                return redirect('vehicles:detail', pk=vehicle.pk)

            try:
                if form.cleaned_data['aktion'] == 'release':
                    release_call_sign(
                        vehicle,
                        valid_to=form.cleaned_data['valid_from'],
                        reason=form.cleaned_data['reason'],
                        user=request.user,
                    )
                    messages.success(
                        request,
                        _('Funkrufname "{}" wurde freigegeben. {} fährt jetzt ohne.').format(
                            call_sign, vehicle.license_plate
                        )
                    )
                else:
                    ziel = form.cleaned_data['target_vehicle']
                    assign_call_sign(
                        call_sign,
                        ziel,
                        valid_from=form.cleaned_data['valid_from'],
                        reason=form.cleaned_data['reason'],
                        user=request.user,
                    )
                    messages.success(
                        request,
                        _('Funkrufname "{}" trägt jetzt {}. {} fährt ohne Funkrufnamen.').format(
                            call_sign, ziel.license_plate, vehicle.license_plate
                        )
                    )
                return redirect('vehicles:detail', pk=vehicle.pk)

            except ValidationError as fehler:
                form.add_error(None, fehler)

        return render(request, self.template_name, {
            'vehicle': vehicle,
            'form': form,
            'history': history_of_call_sign(vehicle.call_sign) if vehicle.call_sign else [],
        })
